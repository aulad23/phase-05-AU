import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import re
import time


def get_product_family_id(product_name):
    """Extract Product Family ID from Product Name
    Rules:
    - Remove variant after – or —  (WELLE CASHMERE BLANKET – CHARCOAL -> WELLE CASHMERE BLANKET)
    - Remove (...) parenthetical  (BRUTALIST CANDLESTICK (BLACK/WHITE) -> BRUTALIST CANDLESTICK)
    - Remove _XXX suffix          (FORM NO_001 -> FORM NO)
    - Remove .XX suffix           (WR19.26 -> WR19)
    - Remove #XX suffix           (DUNE #8 -> DUNE)
    """
    if not product_name:
        return ''
    name = product_name.strip()

    # Remove variant after – or — or  -
    for sep in ['–', '—', ' - ']:
        if sep in name:
            name = name.split(sep)[0].strip()

    # Remove (...) parenthetical
    name = re.sub(r'\s*\(.*?\)\s*', '', name).strip()

    # Remove _XXX suffix (underscore + anything at end)
    name = re.sub(r'_\S+$', '', name).strip()

    # Remove .XX suffix (dot + digits at end)
    name = re.sub(r'\.\d+$', '', name).strip()

    # Remove #XX suffix (hash + digits at end)
    name = re.sub(r'\s*#\d+$', '', name).strip()

    return name


def parse_price(text):
    """Extract first price number from text like 'Tier 1 – $16,650' or 'S: ~ $1,650'"""
    if not text:
        return ''
    # Remove ~ and whitespace around $
    text = text.replace('~', '')
    match = re.search(r'\$\s*([\d,]+(?:\.\d+)?)', text)
    if match:
        return match.group(1).replace(',', '')
    return ''


def parse_dimension_value(text):
    """Extract just the numeric value from dimension text like '15"' or '32 1/2"'"""
    if not text:
        return ''
    text = normalize_quotes(text)
    text = normalize_fractions(text)
    text = text.strip().replace('"', '').replace("'", '').strip()
    # Handle fractions like "32 1/2"
    frac_match = re.match(r'([\d]+)\s+([\d]+)/([\d]+)', text)
    if frac_match:
        whole = float(frac_match.group(1))
        numer = float(frac_match.group(2))
        denom = float(frac_match.group(3))
        return str(round(whole + numer / denom, 2))
    # Handle simple numbers (including decimals)
    num_match = re.match(r'([\d]+(?:\.\d+)?)', text)
    if num_match:
        return num_match.group(1)
    return ''


def normalize_quotes(text):
    """Normalize Unicode quote chars to standard double quote"""
    return text.replace('\u2033', '"').replace('\u201d', '"').replace('\u201c', '"').replace('\u2032', "'").replace('\u2018', "'").replace('\u2019', "'").replace('\u301d', '"').replace('\u301e', '"')


def normalize_fractions(text):
    """Convert Unicode fraction characters to decimal-friendly format"""
    frac_map = {
        '¼': ' 1/4', '½': ' 1/2', '¾': ' 3/4',
        '⅛': ' 1/8', '⅜': ' 3/8', '⅝': ' 5/8', '⅞': ' 7/8',
        '⅓': ' 1/3', '⅔': ' 2/3',
        '⅙': ' 1/6', '⅚': ' 5/6',
        '⅕': ' 1/5', '⅖': ' 2/5', '⅗': ' 3/5', '⅘': ' 4/5',
        '⅐': ' 1/7', '⅑': ' 1/9', '⅒': ' 1/10',
    }
    for char, replacement in frac_map.items():
        text = text.replace(char, replacement)
    # Clean up double spaces
    text = re.sub(r'  +', ' ', text)
    return text


def parse_dimensions(dim_text):
    """Parse dimension text into individual fields"""
    result = {
        'Weight': '', 'Width': '', 'Depth': '', 'Diameter': '',
        'Length': '', 'Height': '', 'Seat Height': '', 'Seat Depth': '',
        'Arm Height': ''
    }
    if not dim_text:
        return result

    # Normalize unicode quotes in raw HTML before processing
    dim_text = normalize_quotes(dim_text)
    dim_text = normalize_fractions(dim_text)

    # ✅ FIX 1: Split by <br>, newlines, AND pipe | character
    lines = re.split(r'<br\s*/?>|\n|\|', dim_text)
    lines = [BeautifulSoup(l, 'html.parser').get_text(strip=True) for l in lines]
    lines = [normalize_quotes(normalize_fractions(l)) for l in lines if l]

    # ✅ FIX 3: If Seat Height/Depth/Arm Height is comma-separated on same line
    # e.g. "BAR: Overall: 18" W x 19" D x 43" H, Seat Height: 30"
    # → split by comma so dimension part and seat height are separate lines
    expanded_lines = []
    for line in lines:
        if re.search(r'(Seat\s+Height|Seat\s+Depth|Arm\s+Height)', line, re.IGNORECASE):
            parts = re.split(r',\s*', line)
            expanded_lines.extend([p.strip() for p in parts if p.strip()])
        else:
            expanded_lines.append(line)
    lines = expanded_lines

    if not lines:
        return result

    for line in lines:
        # Weight (handle "approx." and similar prefixes)
        weight_match = re.search(r'Weight\s*[:\s]*(?:approx\.?\s*)?([\d,.]+)\s*(lbs?|kg|ibs?)', line, re.IGNORECASE)
        if weight_match:
            result['Weight'] = weight_match.group(1).replace(',', '')
            continue

        # Seat Height — also handle | separator inside line
        seat_h_match = re.search(r'Seat\s+Height\s*[:\s]*([^\n,|]+)', line, re.IGNORECASE)
        if seat_h_match:
            result['Seat Height'] = parse_dimension_value(seat_h_match.group(1))
            continue

        # Seat Depth
        seat_d_match = re.search(r'Seat\s+Depth\s*[:\s]*([^\n,|]+)', line, re.IGNORECASE)
        if seat_d_match:
            result['Seat Depth'] = parse_dimension_value(seat_d_match.group(1))
            continue

        # Arm Height
        arm_h_match = re.search(r'Arm\s+Height\s*[:\s]*([^\n,|]+)', line, re.IGNORECASE)
        if arm_h_match:
            result['Arm Height'] = parse_dimension_value(arm_h_match.group(1))
            continue

    # Find the first dimension line (skip Weight/Seat/Arm lines)
    first_line = ''
    for line in lines:
        line_lower = line.lower().strip()
        if any(kw in line_lower for kw in ['weight', 'seat height', 'seat depth', 'arm height']):
            continue
        if re.search(r'[\d]', line) and ('"' in line or re.search(r'\d\s*in', line, re.IGNORECASE) or "''" in line):
            first_line = line
            break

    if not first_line:
        return result

    # ✅ FIX 4: Normalize × (Unicode multiplication sign) to x
    first_line = first_line.replace('\u00d7', 'x').replace('×', 'x')

    # Normalize "in." or "in" suffix to " (e.g. 24 in. L -> 24" L, 2.71in -> 2.71")
    first_line = re.sub(r'(\d)\s*in\.?\s*', r'\1" ', first_line, flags=re.IGNORECASE)
    # Normalize '' (two single quotes) to " (e.g. 33.5'' -> 33.5")
    first_line = first_line.replace("''", '"')

    # ✅ FIX 2: Remove ANY word prefix ending with colon (Bar:, Counter:, S:, M:, L:, Overall:, etc.)
    first_line = re.sub(r'^[A-Za-z]+\s*:\s*', '', first_line, flags=re.IGNORECASE).strip()
    first_line = re.sub(r'^Overall[^:]*:\s*', '', first_line, flags=re.IGNORECASE).strip()
    # Remove any text prefix ending with ":" before first digit (e.g. "Side Table 11: 19...")
    first_line = re.sub(r'^[^"\d]*:\s*', '', first_line).strip()

    # Check for DIA / Diameter first
    dia_match = re.search(r'([\d\s/]+)"\s*DIA', first_line, re.IGNORECASE)
    if not dia_match:
        dia_match = re.search(r'DIA\s*[:\s]*([\d\s/]+)"?', first_line, re.IGNORECASE)
    if not dia_match:
        dia_match = re.search(r'Diameter\s*[:\s]*([\d\s/]+)"?', first_line, re.IGNORECASE)
    if dia_match:
        result['Diameter'] = parse_dimension_value(dia_match.group(1))
        # Remove DIA portion to avoid D being matched as Depth
        first_line = first_line[:dia_match.start()] + first_line[dia_match.end():]
        first_line = re.sub(r'^\s*x\s*', '', first_line).strip()
        first_line = re.sub(r'\s*x\s*$', '', first_line).strip()

    # Extract all measurement values with optional labels
    # Pattern 1: label AFTER value: 104" L, 32 1/2" W, 24"H
    # Pattern 2: label BEFORE value: H 2.71", D 9.33", W 11.81"
    labeled = {}
    unlabeled_vals = []

    # Try label-before-value first: H 2.71" x D 9.33" x W 11.81"
    before_parts = re.findall(r'\b([LWHDM])\s+([\d\s/.]+)"', first_line, re.IGNORECASE)
    # Try label-after-value: 104" L, 24"H
    after_parts = re.findall(r'([\d\s/.]+)"\s*([LWHDM])\b', first_line, re.IGNORECASE)

    if before_parts and len(before_parts) >= len(after_parts):
        # Label-before format
        for label, val_str in before_parts:
            parsed = parse_dimension_value(val_str)
            if not parsed:
                continue
            label = label.upper()
            if label == 'L':
                labeled['Length'] = parsed
            elif label == 'W':
                labeled['Width'] = parsed
            elif label == 'D':
                labeled['Depth'] = parsed
            elif label == 'H':
                labeled['Height'] = parsed
        # Find unlabeled values (values with " but no adjacent label)
        all_vals = re.findall(r'([\d\s/.]+)"', first_line)
        labeled_val_set = {v for _, v in before_parts}
        for v in all_vals:
            if v not in labeled_val_set:
                parsed = parse_dimension_value(v)
                if parsed:
                    unlabeled_vals.append(parsed)
    else:
        # Label-after or mixed format
        all_parts = re.findall(r'([\d\s/.]+)"\s*([LWHDM])?', first_line, re.IGNORECASE)
        for val_str, label in all_parts:
            parsed = parse_dimension_value(val_str)
            if not parsed:
                continue
            if label:
                label = label.upper()
                if label == 'L':
                    labeled['Length'] = parsed
                elif label == 'W':
                    labeled['Width'] = parsed
                elif label == 'D':
                    labeled['Depth'] = parsed
                elif label == 'H':
                    labeled['Height'] = parsed
            else:
                unlabeled_vals.append(parsed)

    # Apply labeled values first
    for k, v in labeled.items():
        result[k] = v

    has_any_label = len(labeled) > 0

    if has_any_label:
        # PARTIAL LABELS: assign unlabeled values to Width → Depth → Height
        # (skipping already assigned by labels)
        fill_order = ['Width', 'Depth', 'Height']
        fill_slots = [s for s in fill_order if s not in labeled]
        for i, val in enumerate(unlabeled_vals):
            if i < len(fill_slots):
                result[fill_slots[i]] = val
    else:
        # NO LABELS: first=Length, then Width, then Height
        if len(unlabeled_vals) == 1:
            result['Length'] = unlabeled_vals[0]
        elif len(unlabeled_vals) == 2:
            result['Length'] = unlabeled_vals[0]
            result['Width'] = unlabeled_vals[1]
        elif len(unlabeled_vals) == 3:
            result['Length'] = unlabeled_vals[0]
            result['Width'] = unlabeled_vals[1]
            result['Height'] = unlabeled_vals[2]
        elif len(unlabeled_vals) >= 4:
            result['Length'] = unlabeled_vals[0]
            result['Width'] = unlabeled_vals[1]
            result['Depth'] = unlabeled_vals[2]
            result['Height'] = unlabeled_vals[3]

    return result


def parse_com(com_html):
    """Extract first COM value (just the number) from COM section"""
    if not com_html:
        return '', ''
    text = BeautifulSoup(com_html, 'html.parser').get_text()
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    com_val = ''
    col_val = ''

    # First line is usually COM total like "21 yds"
    if lines:
        first_match = re.match(r'([\d.]+)', lines[0])
        if first_match:
            com_val = first_match.group(1)

    # Look for sqft for COL
    for line in lines:
        sqft_match = re.search(r'([\d.]+)\s*sqft', line, re.IGNORECASE)
        if sqft_match:
            col_val = sqft_match.group(1)
            break

    return com_val, col_val


def scrape_product_detail(url, session):
    """Scrape detailed product info from a single product page"""
    result = {
        'Description': '', 'Finish': '', 'Dimension_Raw': '',
        'Weight': '', 'Width': '', 'Depth': '', 'Diameter': '',
        'Length': '', 'Height': '', 'Seat Height': '', 'Seat Depth': '',
        'Arm Height': '', 'COM': '', 'COL': '', 'List Price': ''
    }

    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, 'html.parser')

        # --- Description ---
        details_div = soup.find('div', class_='product-details')
        if not details_div:
            details_div = soup.find('div', class_=lambda x: x and 'product-details' in x)

        if details_div:
            inner = details_div.find('div', class_='inner-cont')
            if inner:
                p_tag = inner.find('p')
                if p_tag:
                    result['Description'] = p_tag.get_text(strip=True)

        # --- Available Finishes ---
        finish_names = []
        aside_divs = soup.find_all('div', class_='aside')
        for aside in aside_divs:
            h3 = aside.find('h3')
            if h3 and 'finish' in h3.get_text(strip=True).lower():
                colors_div = aside.find('div', class_='colors')
                if colors_div:
                    p_tags = colors_div.find_all('p')
                    for p in p_tags:
                        txt = p.get_text(strip=True)
                        if txt:
                            finish_names.append(txt)
                break
        result['Finish'] = ', '.join(finish_names)

        # --- Standard Dimensions ---
        for aside in soup.find_all('div', class_=lambda x: x and 'aside' in x.split()):
            h3 = aside.find('h3')
            if h3 and 'dimension' in h3.get_text(strip=True).lower():
                lrg_txt = aside.find('div', class_='lrg-txt')
                if lrg_txt:
                    raw_html = normalize_quotes(str(lrg_txt))
                    result['Dimension_Raw'] = BeautifulSoup(raw_html, 'html.parser').get_text(separator='\n', strip=True)
                    dims = parse_dimensions(raw_html)
                    for k, v in dims.items():
                        result[k] = v
                break

        # --- COM ---
        for aside in soup.find_all('div', class_=lambda x: x and 'aside' in x.split()):
            h3 = aside.find('h3')
            if h3 and h3.get_text(strip=True).upper() == 'COM':
                lrg_txt = aside.find('div', class_='lrg-txt')
                if lrg_txt:
                    com_val, col_val = parse_com(str(lrg_txt))
                    result['COM'] = com_val
                    result['COL'] = col_val
                break

        # --- List Price (first one only) ---
        for aside in soup.find_all('div', class_=lambda x: x and 'aside' in x.split()):
            h3 = aside.find('h3')
            if h3 and 'list price' in h3.get_text(strip=True).lower():
                lrg_txt = aside.find('div', class_='lrg-txt')
                if lrg_txt:
                    # Get first price only
                    price_text = lrg_txt.get_text()
                    first_line = price_text.strip().split('\n')[0]
                    result['List Price'] = parse_price(first_line)
                break

    except Exception as e:
        print(f"  ERROR scraping {url}: {e}")

    return result


def main():
    # --- Configuration ---
    input_file = "stahlandband_Pulls.xlsx"
    output_file = "stahlandband_Pulls_Dtailed.xlsx"

    # --- Read input Excel ---
    print(f"Reading input file: {input_file}")
    wb_in = load_workbook(input_file, data_only=True)
    sheet_in = wb_in.active

    headers_in = [cell.value for cell in sheet_in[1]]
    print(f"Input columns: {headers_in}")

    products = []
    for row in sheet_in.iter_rows(min_row=2, values_only=True):
        product = dict(zip(headers_in, row))
        products.append(product)

    print(f"Total products to process: {len(products)}")

    # --- Setup session ---
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    })

    # --- Output headers ---
    output_headers = [
        'Product URL', 'Image URL', 'Product Name', 'SKU',
        'Product Family Id', 'Description',
        'Weight', 'Width', 'Depth', 'Diameter', 'Length', 'Height',
        'Finish', 'List Price', 'COM', 'COL',
        'Seat Height', 'Seat Depth', 'Arm Height'
    ]

    # --- Create output workbook ---
    wb_out = Workbook()
    sheet_out = wb_out.active
    sheet_out.title = "Product Details"
    sheet_out.append(output_headers)

    # --- Process each product ---
    for i, product in enumerate(products, 1):
        product_name = product.get('Product Name', '') or ''
        product_url = product.get('Product URL', '') or ''
        sku = product.get('SKU', '') or ''
        image_url = product.get('Image URL', '') or ''

        print(f"\n[{i}/{len(products)}] {product_name}")
        print(f"  URL: {product_url}")

        # Product Family Id
        family_id = get_product_family_id(product_name)

        # Scrape detail page
        detail = {}
        if product_url:
            detail = scrape_product_detail(product_url, session)
            time.sleep(1)  # Be polite, 1 second delay between requests

        # Build output row
        row = [
            product_url,
            image_url,
            product_name,
            sku,
            family_id,
            detail.get('Description', ''),
            detail.get('Weight', ''),
            detail.get('Width', ''),
            detail.get('Depth', ''),
            detail.get('Diameter', ''),
            detail.get('Length', ''),
            detail.get('Height', ''),
            detail.get('Finish', ''),
            detail.get('List Price', ''),
            detail.get('COM', ''),
            detail.get('COL', ''),
            detail.get('Seat Height', ''),
            detail.get('Seat Depth', ''),
            detail.get('Arm Height', ''),
        ]
        sheet_out.append(row)

        print(f"  Family ID: {family_id}")
        print(f"  Finish: {detail.get('Finish', 'N/A')}")
        print(f"  Dimensions: L={detail.get('Length', '')} W={detail.get('Width', '')} "
              f"H={detail.get('Height', '')} D={detail.get('Depth', '')} "
              f"DIA={detail.get('Diameter', '')}")
        print(f"  Seat Height: {detail.get('Seat Height', 'N/A')}")
        print(f"  Price: {detail.get('List Price', 'N/A')}")

    # --- Auto-adjust column widths ---
    for col_idx, header in enumerate(output_headers, 1):
        col_letter = sheet_out.cell(row=1, column=col_idx).column_letter
        max_len = len(str(header))
        for row in sheet_out.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
        sheet_out.column_dimensions[col_letter].width = max_len + 3

    # --- Bold headers ---
    from openpyxl.styles import Font
    for cell in sheet_out[1]:
        cell.font = Font(bold=True)

    # --- Save ---
    wb_out.save(output_file)
    print(f"\n{'='*60}")
    print(f"DONE! Output saved to: {output_file}")
    print(f"Total products processed: {len(products)}")


if __name__ == "__main__":
    main()