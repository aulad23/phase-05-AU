import os
import time
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse, urlencode, urlunparse, parse_qs
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ==== Base Settings ====
base_url = "https://www.alliedmaker.com"

# Categories configuration
CATEGORIES = {
    "Chandeliers": [
        "https://www.alliedmaker.com/Products/custitem_type/Chandelier"
    ],
    "Pendants": [
        "https://www.alliedmaker.com/Products/custitem_type/Pendant"
    ],
    "Sconces": [
        "https://www.alliedmaker.com/Products/custitem_type/Sconce"
    ],
    "Flush Mount": [
        "https://www.alliedmaker.com/Products/custitem_type/Flush-Mount"
    ],
    "Table Lamps": [
        "https://www.alliedmaker.com/Products/custitem_type/Table-Lamp"
    ],
    "Floor Lamps": [
        "https://www.alliedmaker.com/Products/custitem_type/Floor-Lamp"
    ],
}

# ==== Save output in same folder as script ====
script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "alliedmaker_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "AlliedMaker.xlsx")

# ==== Setup Selenium Chrome (headless) ====
options = Options()
options.add_argument("--headless=new")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--window-size=1920,1080")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# ---- Helper Functions ----
def build_page_url(base_url, page):
    parsed = urlparse(base_url)
    qs = parse_qs(parsed.query)
    qs["page"] = [str(page)]
    return urlunparse(parsed._replace(query=urlencode(qs, doseq=True)))

def parse_products_from_page(html):
    soup = BeautifulSoup(html, "html.parser")
    results = []

    # --- Check if products are found ---
    products = soup.find_all("div", class_="facets-item-cell-list")
    print(f"Found {len(products)} products on this page.")

    for p in products:
        # --- Product URL and Name ---
        try:
            name_tag = p.select_one(".facets-item-cell-list-name span[itemprop='name']")
            product_name = name_tag.get_text(strip=True) if name_tag else ""
            link_tag = p.select_one("a.facets-item-cell-list-name[href], a.facets-item-cell-list-anchor[href]")
            product_url = base_url + link_tag["href"] if link_tag else ""
        except:
            product_url = ""
            product_name = ""

        # Skip if no product name (banners usually have no proper name)
        if not product_name or len(product_name) < 3:
            continue

        # --- Image ---
        img_tag = p.find("img", class_="facets-item-cell-list-image")
        image_url = img_tag["src"].split("?")[0] if img_tag else ""

        # Skip if image is category banner or ad
        if "Commerce-category-banners" in image_url:
            continue

        # Skip if image is empty or suspiciously short
        if not image_url or len(image_url) < 10:
            continue

        results.append({
            "Product URL": product_url,
            "Image URL": image_url,
            "Product Name": product_name
        })

    return results

def scrape_category(category, links):
    all_products = []
    for link in links:
        for page in range(1, 200 + 1):  # You can increase MAX_PAGES if needed
            try:
                driver.get(build_page_url(link, page))
                time.sleep(3)  # Wait for the page to load

                products = parse_products_from_page(driver.page_source)
                if not products:
                    break
                all_products.extend(products)
            except Exception as e:
                print(f"Error on page {page}: {e}")
                break

    df = pd.DataFrame(all_products)
    if not df.empty:
        df.drop_duplicates(subset=["Product URL"], inplace=True)
    return df

# ---- Step-1 Main Functions ----
def build_step1_master_excel():
    dfs = []
    for cat, links in CATEGORIES.items():
        df = scrape_category(cat, links)
        if not df.empty:
            df.insert(0, "Category", cat)  # Insert category in the first column
            dfs.append(df)
        else:
            print(f"⚠️ No products found for category: {cat}")

    # ✅ FIX: prevent "No objects to concatenate"
    if not dfs:
        empty_cols = ["Category", "Product URL", "Image URL", "Product Name"]
        master = pd.DataFrame(columns=empty_cols)
        master.to_excel(master_output_file, index=False)
        print(f"⚠️ No data collected from any category. Empty master saved: {master_output_file}")
        return master

    master = pd.concat(dfs, ignore_index=True)
    master.to_excel(master_output_file, index=False)
    print(f"✅ Master saved: {master_output_file}")
    return master

def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    # ✅ Sheet order = CATEGORIES dict order (serial)
    ordered_categories = list(CATEGORIES.keys())

    for cat in ordered_categories:
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)

        ws["A1"] = "Brand"
        ws["B1"] = "AlliedMaker"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

        # Unfreeze panes
        ws.freeze_panes = None

    wb.save(master_sheets_file)
    print(f"✅ Final AlliedMaker.xlsx created (Category serial order): {master_sheets_file}")

def main():
    df = build_step1_master_excel()
    if df.empty:
        wb = Workbook()
        wb.save(master_sheets_file)
        print(f"⚠️ No category sheets created because no products were scraped. Empty workbook saved: {master_sheets_file}")
        return

    build_category_wise_workbook_from_df(df)

if __name__ == "__main__":
    main()
