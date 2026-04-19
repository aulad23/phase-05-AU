import time
import re
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook

# ================= CONFIG =================
MASTER_FILE = "AlliedMaker.xlsx"
SAVE_AS = "AlliedMaker_details.xlsx"

HEADLESS = True
TIMEOUT_PAGELOAD = 30
WAIT_BODY_SEC = 12
PER_PAGE_SLEEP = 2

HEADER_ROW = 4
DATA_START_ROW = 5

COLUMN_ORDER = [
    "Index", "Category", "Product URL", "Image URL", "Product Name",
    "SKU", "Product Family Id", "Description", "List Price",
    "Weight", "Width", "Depth", "Diameter", "Length", "Height", "Lamping"
]
# =========================================

def setup_driver():
    chromedriver_autoinstaller.install()
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(TIMEOUT_PAGELOAD)
    return driver


def get_header_map(ws):
    return {
        str(ws.cell(row=HEADER_ROW, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=HEADER_ROW, column=c).value
    }


def write_new_header(ws):
    for c in range(1, ws.max_column + 20):
        ws.cell(row=HEADER_ROW, column=c, value=None)

    for i, col in enumerate(COLUMN_ORDER, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=col)

    return {col: i for i, col in enumerate(COLUMN_ORDER, start=1)}


def get_last_data_row(ws, col_url):
    r = ws.max_row
    while r >= DATA_START_ROW and not ws.cell(row=r, column=col_url).value:
        r -= 1
    return r


# ---------- DIMENSION PARSER ----------
def parse_dimensions(text):
    width = depth = diameter = height = length = ""
    text = text.replace("”", '"')

    dia = re.search(r'(\d+\.?\d*)\s*"?\s*(DIA|DIAMETER)', text, re.I)
    if dia:
        diameter = dia.group(1)

    m = re.search(r'(\d+\.?\d*)\s*"?\s*L\b', text, re.I)
    if m: length = m.group(1)

    m = re.search(r'(\d+\.?\d*)\s*"?\s*W\b', text, re.I)
    if m: width = m.group(1)

    m = re.search(r'(\d+\.?\d*)\s*"?\s*D\b', text, re.I)
    if m: depth = m.group(1)

    m = re.search(r'(\d+\.?\d*)\s*"?\s*H\b', text, re.I)
    if m: height = m.group(1)

    return width, depth, diameter, height, length


# ---------- SCRAPER ----------
def extract_product_details(driver, url):
    out = {
        "SKU": "", "Description": "", "List Price": "", "Weight": "",
        "Width": "", "Depth": "", "Diameter": "",
        "Height": "", "Length": "", "Lamping": ""
    }

    try:
        driver.get(url)
        WebDriverWait(driver, WAIT_BODY_SEC).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(PER_PAGE_SLEEP)
    except:
        return out

    # SKU
    try:
        out["SKU"] = driver.find_element(
            By.CSS_SELECTOR, ".sku-spec h4"
        ).text.strip()
    except:
        pass

    # Description
    try:
        out["Description"] = driver.find_element(
            By.CSS_SELECTOR, ".product-detail-text p"
        ).text.strip()
    except:
        pass

    # Price
    try:
        price = driver.find_element(
            By.CSS_SELECTOR, ".product-views-price-lead"
        ).text.replace("+", "").strip()
        out["List Price"] = price
    except:
        pass

    # DIMENSIONS
    try:
        li = driver.find_element(
            By.XPATH, "//li[h4/a[text()='DIMENSIONS']]"
        )
        dim_text = li.find_element(By.TAG_NAME, "p").text.strip()
        w, d, dia, h, l = parse_dimensions(dim_text)
        out.update({
            "Width": w, "Depth": d, "Diameter": dia,
            "Height": h, "Length": l
        })
    except:
        pass

    # Weight
    try:
        li = driver.find_element(
            By.XPATH, "//li[h4/a[text()='WEIGHT']]"
        )
        wt = li.find_element(By.TAG_NAME, "p").text.strip()
        m = re.search(r'([\d\.]+)', wt)
        out["Weight"] = m.group(1) if m else wt
    except:
        pass

    # Lamping
    try:
        li = driver.find_element(
            By.XPATH, "//li[h4/a[text()='LAMPING']]"
        )
        out["Lamping"] = li.find_element(By.TAG_NAME, "p").text.strip()
    except:
        pass

    return out


# ---------- MAIN ----------
def main():
    wb = load_workbook(MASTER_FILE)
    driver = setup_driver()

    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            old_map = get_header_map(ws)
            new_map = write_new_header(ws)

            col_url = new_map["Product URL"]
            last_row = get_last_data_row(ws, col_url)
            if last_row < DATA_START_ROW:
                continue

            for r in range(DATA_START_ROW, last_row + 1):
                url = ws.cell(row=r, column=col_url).value
                if not url or not str(url).startswith("http"):
                    continue

                print(f"Scraping Row {r}: {url}")
                data = extract_product_details(driver, url)

                ws.cell(row=r, column=new_map["Category"], value=sheet)
                ws.cell(row=r, column=new_map["Product Family Id"],
                        value=ws.cell(row=r, column=new_map["Product Name"]).value)

                for k, v in data.items():
                    ws.cell(row=r, column=new_map[k], value=v)

            wb.save(SAVE_AS)
            print(f"Saved -> {SAVE_AS}")

    finally:
        driver.quit()
        wb.save(SAVE_AS)

    print("DONE")


if __name__ == "__main__":
    main()
