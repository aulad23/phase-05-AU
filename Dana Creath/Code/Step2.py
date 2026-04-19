"""
Dana Creath - Step 2: Product Detail Scraper
============================================
Input  : danacreath_Chandeliers.xlsx  (Step 1 এর output)
Output : danacreath_Chandeliers_FULL.xlsx

Extracts per product URL:
  - Product Family Id
  - Description
  - Weight, Width, Depth, Diameter, Height, Wattage
  - Shade Details
  - Tearsheet Link

Install: pip install playwright beautifulsoup4 openpyxl && playwright install chromium
"""

import re
import time
import random
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# ── Input / Output files ──────────────────────────────────────────────
INPUT_FILE  = "danacreath_mirrors.xlsx"
OUTPUT_FILE = "danacreath_mirrors_Final.xlsx"

# ── Product Family Id  ────────────────────────────────────────────────
def get_family_id(product_name: str) -> str:
    """
    Rule 1 → comma আছে → comma এর আগের অংশ
             e.g. "GABRIELLE, LARGE FRENCH COUNTRY ..." → "GABRIELLE"

    Rule 2 → comma নেই → "CHANDELIER" এবং তার আগের
             " LIGHT" / "-LIGHT" / " TWO-LIGHT" ইত্যাদি সরিয়ে দাও
             e.g. "WROUGHT IRON EIGHT-LIGHT CHANDELIER" → "WROUGHT IRON EIGHT"
    """
    name = product_name.strip()

    # Rule 1 – comma
    if "," in name:
        return name.split(",")[0].strip()

    # Rule 2 – remove CHANDELIER suffix + light descriptor
    # Step A: remove " CHANDELIER" and anything after
    cleaned = re.sub(r"\s+CHANDELIER.*$", "", name, flags=re.IGNORECASE).strip()

    # Step B: remove trailing "-LIGHT" / " LIGHT" / " ONE-LIGHT" / " TWO-LIGHT" etc.
    cleaned = re.sub(r"[-\s]+\w*[-]?LIGHT\s*$", "", cleaned, flags=re.IGNORECASE).strip()

    return cleaned if cleaned else name


# ── Value Cleaners ────────────────────────────────────────────────────
def clean_weight(value: str) -> str:
    """
    "APPROX. 12 LB."  → "12"
    "12 lbs"          → "12"
    "12"              → "12"
    """
    if not value:
        return ""
    # সংখ্যা (integer বা decimal) বের করো
    m = re.search(r"(\d+(?:\.\d+)?)", value)
    return m.group(1) if m else value.strip()


def clean_dimension(value: str) -> str:
    """
    '46"'      → "46"
    '6-1/2"'   → "6.5"
    '7-1/2"'   → "7.5"
    '21"'      → "21"
    '21-3/4"'  → "21.75"
    """
    if not value:
        return ""
    v = value.strip().rstrip('"').strip()

    # Pattern: whole-num/denom  e.g. "6-1/2"
    m = re.match(r"^(\d+)-(\d+)/(\d+)$", v)
    if m:
        whole = int(m.group(1))
        frac  = int(m.group(2)) / int(m.group(3))
        result = whole + frac
        # দশমিক দরকার না হলে integer ফেরত দাও
        return str(int(result)) if result == int(result) else str(round(result, 4))

    # Pattern: শুধু fraction  e.g. "1/2"
    m = re.match(r"^(\d+)/(\d+)$", v)
    if m:
        result = int(m.group(1)) / int(m.group(2))
        return str(int(result)) if result == int(result) else str(round(result, 4))

    # Pattern: plain number  e.g. "46"
    m = re.match(r"^(\d+(?:\.\d+)?)$", v)
    if m:
        return m.group(1)

    return v  # যা আছে তাই রাখো


# ── Playwright page fetch ─────────────────────────────────────────────
def get_page(url: str) -> BeautifulSoup:
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0"
        )
        page.goto(url, wait_until="networkidle", timeout=40000)
        page.wait_for_timeout(3000)
        html = page.content()
        browser.close()
    return BeautifulSoup(html, "html.parser")


# ── Parse product detail page ─────────────────────────────────────────
def parse_detail(soup: BeautifulSoup) -> dict:
    """
    HTML structure (Elementor / JetEngine):
      jet-listing-dynamic-field__content  → "Label : "  then  value
    We pair consecutive label+value divs.
    """
    data = {
        "description":   "",
        "weight":        "",
        "width":         "",
        "depth":         "",
        "diameter":      "",
        "height":        "",
        "wattage":       "",
        "shade_details": "",
        "tearsheet":     "",
    }

    # ── Tearsheet link ────────────────────────────────────────────────
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True).lower()
        if "tearsheet" in text or "tear sheet" in text or href.lower().endswith(".pdf"):
            if "tearsheet" in text or "tear" in text:
                data["tearsheet"] = href
                break

    # ── Description (Product Description field) ───────────────────────
    desc_div = soup.find("div", class_="jet-listing-dynamic-field__content",
                          string=re.compile(r"Product Description", re.I))
    if desc_div:
        data["description"] = desc_div.get_text(strip=True)\
            .replace("Product Description :", "").strip()

    # ── Key-Value pairs from JetEngine listing grid ───────────────────
    #  Each <div class="jet-listing-grid__item"> contains
    #  two  jet-listing-dynamic-field__content  divs: label & value
    for item in soup.select("div.jet-listing-grid__item"):
        contents = item.select("div.jet-listing-dynamic-field__content")
        if len(contents) < 2:
            continue
        label = contents[0].get_text(strip=True).lower().rstrip(":").strip()
        value = contents[1].get_text(strip=True)

        if "weight"     in label: data["weight"]        = clean_weight(value)
        elif "width"    in label: data["width"]         = clean_dimension(value)
        elif "depth"    in label: data["depth"]         = clean_dimension(value)
        elif "diameter" in label: data["diameter"]      = clean_dimension(value)
        elif "height"   in label: data["height"]        = clean_dimension(value)
        elif "wattage"  in label: data["wattage"]       = value
        elif "shade"    in label: data["shade_details"] = value
        elif "candle" in label: pass   # skip candle field
        elif "item #" in label or "item#" in label: pass
        elif "category" in label: pass

    # ── Also grab Description from first big content block ───────────
    if not data["description"]:
        # look for the standalone description paragraph
        for div in soup.select("div.jet-listing-dynamic-field__content"):
            txt = div.get_text(strip=True)
            if "Product Description" in txt:
                data["description"] = txt.replace("Product Description :", "").strip()
                break

    return data


# ── Read Step-1 Excel ─────────────────────────────────────────────────
def read_input_excel(path: str):
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


# ── Save output Excel ─────────────────────────────────────────────────
def save_excel(records, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chandeliers Full"

    headers = [
        "Product URL", "Image URL", "Product Name", "SKU",
        "Product Family Id",
        "Description",
        "Weight", "Width", "Depth", "Diameter", "Height",
        "Wattage", "Shade Details", "Tearsheet Link"
    ]
    ws.append(headers)

    col_widths = [55, 70, 40, 40, 30, 60, 12, 12, 12, 12, 12, 12, 30, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    for r in records:
        ws.append([
            r.get("Product URL", ""),
            r.get("Image URL", ""),
            r.get("Product Name", ""),
            r.get("SKU", ""),
            r.get("family_id", ""),
            r.get("description", ""),
            r.get("weight", ""),
            r.get("width", ""),
            r.get("depth", ""),
            r.get("diameter", ""),
            r.get("height", ""),
            r.get("wattage", ""),
            r.get("shade_details", ""),
            r.get("tearsheet", ""),
        ])

    wb.save(filename)
    print(f"\n✅  Saved {len(records)} products → '{filename}'")


# ── Main ──────────────────────────────────────────────────────────────
def main():
    if not Path(INPUT_FILE).exists():
        print(f"❌  '{INPUT_FILE}' not found! Step 1 আগে run করুন।")
        return

    print(f"📂  Reading: {INPUT_FILE}")
    rows = read_input_excel(INPUT_FILE)
    print(f"   {len(rows)} products found.\n")

    results = []
    for i, row in enumerate(rows, 1):
        url  = row.get("Product URL", "").strip()
        name = row.get("Product Name", "").strip()

        family_id = get_family_id(name) if name else ""

        print(f"[{i}/{len(rows)}] {name}")
        print(f"          Family Id → {family_id}")

        detail = {}
        if url:
            try:
                soup   = get_page(url)
                detail = parse_detail(soup)
                print(f"          ✔ scraped | "
                      f"H:{detail['height']} W:{detail['width']} "
                      f"Watt:{detail['wattage']} Tearsheet:{bool(detail['tearsheet'])}")
            except Exception as e:
                print(f"          ⚠ Error: {e}")
        else:
            print("          ⚠ No URL – skipping scrape")

        record = {**row, "family_id": family_id, **detail}
        results.append(record)

        # Polite delay
        if url and i < len(rows):
            wait = random.uniform(2, 4)
            time.sleep(wait)

    save_excel(results, OUTPUT_FILE)


if __name__ == "__main__":
    main()