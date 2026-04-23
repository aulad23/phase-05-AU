"""
SKILL: Excel Read & Write
Agent এই patterns ব্যবহার করে সব Excel কাজ করে।
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime


# ── STANDARD COLUMN ORDER ─────────────────────────────────────────────────────
STANDARD_COLUMNS = [
    "Manufacturer", "Source", "Image URL", "Product Name", "SKU",
    "Product Family Id", "Description", "Weight", "Width", "Depth",
    "Diameter", "Length", "Height", "Seat Width", "Seat Depth",
    "Seat Height", "Arm Height", "Arm Width", "List Price",
]


# ── CREATE FINAL EXCEL WITH HEADER ROWS ──────────────────────────────────────
def create_output_excel(output_path: str, brand_name: str, source_url: str,
                        columns: list = None) -> openpyxl.Workbook:
    """
    Row 1: Manufacturer name (merged A1:S1)
    Row 2: Source URL (merged A2:S2)
    Row 3: Scrape date (A3)
    Row 4: Column headers (bold)
    Row 5+: Data
    """
    cols = columns or STANDARD_COLUMNS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = brand_name[:31]

    last_col = get_column_letter(len(cols))

    # Row 1 — brand
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = brand_name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Row 2 — source
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = source_url
    ws["A2"].alignment = Alignment(horizontal="center")

    # Row 3 — date
    ws["A3"] = f"Scraped: {datetime.date.today()}"

    # Row 4 — headers
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=4, column=i, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    return wb, ws


def append_row(ws, data: dict, row_num: int, columns: list = None):
    cols = columns or STANDARD_COLUMNS
    for i, col in enumerate(cols, 1):
        ws.cell(row=row_num, column=i, value=data.get(col, ""))


def save_excel(wb: openpyxl.Workbook, path: str):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  Saved → {path}")


# ── READ STEP1 OUTPUT ─────────────────────────────────────────────────────────
def read_step1_excel(path: str) -> pd.DataFrame:
    return pd.read_excel(path)


# ── AUTO-SAVE EVERY N ROWS ────────────────────────────────────────────────────
def auto_save(wb, path: str, count: int, every: int = 10):
    if count % every == 0:
        save_excel(wb, path)
        print(f"  Auto-saved at row {count}")


# ── RESUME: ALREADY SCRAPED URLs ──────────────────────────────────────────────
def get_done_urls(output_path: str, url_col: str = "Source") -> set:
    if not Path(output_path).exists():
        return set()
    try:
        df = pd.read_excel(output_path, header=3)  # Row 4 = headers
        return set(df[url_col].dropna().astype(str).tolist())
    except Exception:
        return set()
