"""
SKILL: All Libraries — BS4, Pandas, Selenium, openpyxl, Requests
এই file-এ সব common library import + quick-setup patterns আছে।
Agent নতুন script তৈরি করার সময় এখান থেকে pattern নেয়।
"""

# ── STANDARD IMPORTS (step1 & step2 top block) ────────────────────────────────
STEP1_IMPORTS = """\
import time, random, re
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
"""

STEP1_SELENIUM_IMPORTS = """\
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
"""

STEP1_BS4_IMPORTS = """\
import requests
from bs4 import BeautifulSoup
"""

STEP2_IMPORTS = """\
import time, random, re, json
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
"""

# ── LIBRARY QUICK REFERENCE ───────────────────────────────────────────────────
LIBRARY_GUIDE = {
    "BeautifulSoup4": {
        "install": "pip install beautifulsoup4 lxml",
        "parse":   "soup = BeautifulSoup(html, 'html.parser')",
        "find":    "el = soup.select_one('div.class')",
        "findall": "els = soup.select('ul li a')",
        "text":    "el.get_text(strip=True)",
        "attr":    "el.get('href', '')",
    },
    "Pandas": {
        "install":    "pip install pandas openpyxl",
        "read_excel": "df = pd.read_excel('file.xlsx')",
        "to_excel":   "df.to_excel('out.xlsx', index=False)",
        "filter":     "df[df['col'].str.contains('x', case=False, na=False)]",
        "iterate":    "for _, row in df.iterrows(): ...",
    },
    "Selenium": {
        "install":    "pip install selenium",
        "driver":     "driver = webdriver.Chrome(options=opts)",
        "get":        "driver.get(url)",
        "find_css":   "driver.find_elements(By.CSS_SELECTOR, 'div.item')",
        "wait":       "WebDriverWait(driver, 10).until(EC.presence_of_element_located(...))",
        "scroll":     "driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')",
        "quit":       "driver.quit()",
    },
    "Requests": {
        "install": "pip install requests",
        "get":     "r = requests.get(url, headers=headers, timeout=20)",
        "json":    "data = r.json()",
        "check":   "r.raise_for_status()",
    },
    "openpyxl": {
        "install":   "pip install openpyxl",
        "new_wb":    "wb = openpyxl.Workbook(); ws = wb.active",
        "write":     "ws.cell(row=1, col=1, value='text')",
        "save":      "wb.save('output.xlsx')",
        "merge":     "ws.merge_cells('A1:Z1')",
        "bold":      "cell.font = Font(bold=True)",
    },
}


def print_guide(library: str):
    info = LIBRARY_GUIDE.get(library)
    if not info:
        print(f"Unknown library: {library}. Options: {list(LIBRARY_GUIDE.keys())}")
        return
    print(f"\n=== {library} ===")
    for k, v in info.items():
        print(f"  {k:12}: {v}")


if __name__ == "__main__":
    for lib in LIBRARY_GUIDE:
        print_guide(lib)
