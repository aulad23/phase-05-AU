# -*- coding: utf-8 -*-
# Currey Step-2 (Your Step-1 Excel format) — FINAL (FIXED)
# ✅ Dimensions accordion label mismatch fixed (fallback keywords + stronger xpath)
# ✅ Shade Details fixed: now uses Shade Top/Bottom/Height blocks first, then fallback Shade/ Shade Details
# ✅ Multi-sheet support: processes all sheets in workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os
import time
import re

# =========================
# FILE PATHS (same folder)
# =========================
script_dir = os.path.dirname(os.path.abspath(__file__))

INPUT_XLSX  = os.path.join(script_dir, "Currey.xlsx")          # ✅ your step-1 file
OUTPUT_XLSX = os.path.join(script_dir, "Currey_details.xlsx")  # ✅ step-2 output

# =========================
# EXCEL STRUCTURE (as screenshot)
# =========================
HEADER_ROW = 4
START_ROW  = 5

COL_INDEX     = 1  # A
COL_CATEGORY  = 2  # B
COL_URL       = 3  # C
COL_IMG       = 4  # D
COL_NAME      = 5  # E
COL_SKU       = 6  # F

AUTOSAVE_EVERY = 10

# =========================
# SELENIUM SETUP
# =========================
options = Options()
#options.add_argument("--headless=new")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-gpu")

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 22)

# =========================
# GENERAL HELPERS
# =========================
def js_click(el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.2)
    driver.execute_script("arguments[0].click();", el)

def find_accordion_summary(keyword):
    """
    More robust:
    - looks inside ANY descendant (not only <span>)
    - returns nearest MuiAccordionSummary-root
    """
    kw = keyword.lower().strip()
    xpath = (
        "//*[contains(@class,'MuiAccordionSummary-root')]"
        f"[.//*[contains(translate(normalize-space(.),"
        "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),"
        f"'{kw}')]]"
    )
    return wait.until(EC.presence_of_element_located((By.XPATH, xpath)))

def get_accordion_root(summary_el):
    return summary_el.find_element(By.XPATH, "./ancestor::div[contains(@class,'MuiAccordion-root')]")

def ensure_accordion_open(summary_el):
    root = get_accordion_root(summary_el)
    expanded = "Mui-expanded" in (root.get_attribute("class") or "")
    if not expanded:
        js_click(summary_el)
        time.sleep(0.9)
    return get_accordion_root(summary_el)

def ensure_accordion_closed(summary_el):
    root = get_accordion_root(summary_el)
    expanded = "Mui-expanded" in (root.get_attribute("class") or "")
    if expanded:
        js_click(summary_el)
        time.sleep(0.4)

# =========================
# PAGE DATA HELPERS
# =========================
def get_product_header(base_url, base_img, base_name, alt_name=None):
    product_url = driver.current_url or base_url

    product_name = base_name
    try:
        h1 = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "h1")))
        txt = h1.text.strip()
        if txt:
            product_name = txt
    except:
        if alt_name:
            product_name = alt_name

    image_url = base_img
    try:
        img = driver.find_element(By.CSS_SELECTOR, "picture img")
        src = img.get_attribute("src")
        if src and src.startswith("http"):
            image_url = src
    except:
        pass

    return product_url, image_url, product_name

def get_description():
    try:
        return driver.find_element(
            By.CSS_SELECTOR,
            "div.account-paragraph-s.mb-8.md\\:mb-10"
        ).text.strip()
    except:
        return ""

# =========================
# ✅ DIMENSIONS: FULL TEXT + OVERALL/WGT/CANOPY/SEAT/ARM/SHADE EXTRACTION
# =========================
def get_dimensions_full_text():
    """
    FIX:
    Some products don't have the accordion titled exactly "Dimensions".
    We'll try multiple keywords.
    """
    keywords = [
        "dimensions",
        "dimensions & weight",
        "dimension",
        "size"
    ]
    for kw in keywords:
        try:
            summ = find_accordion_summary(kw)
            root = ensure_accordion_open(summ)

            details = root.find_element(By.XPATH, ".//div[contains(@class,'MuiAccordionDetails-root')]")
            txt = details.text.strip()

            ensure_accordion_closed(summ)
            if txt:
                return txt
        except:
            continue
    return ""

def extract_value_block(full_text, label):
    """
    label এর পরে যে value লাইনগুলো আছে সেগুলো নেয়, যতক্ষণ না next label আসে।
    """
    if not isinstance(full_text, str) or not full_text.strip():
        return ""

    lines = [l.strip() for l in full_text.split("\n") if l.strip()]
    if not lines:
        return ""

    label_l = label.lower()

    known_labels = {
        "overall", "item weight", "cord", "canopy", "shade", "shade details",
        "shade top", "shade bottom", "shade height",
        "seat height", "seat width", "seat depth",
        "arm height", "arm width", "arm length"
    }

    idx = -1
    for i, ln in enumerate(lines):
        ln_l = ln.lower().rstrip(":")
        if ln_l == label_l or ln_l.startswith(label_l + " ") or ln_l.startswith(label_l + ":"):
            idx = i
            break
    if idx == -1:
        return ""

    same_line = re.sub(rf"^{re.escape(label)}\s*:?\s*", "", lines[idx], flags=re.I).strip()
    if same_line and same_line.lower() != label_l:
        return same_line

    vals = []
    for j in range(idx + 1, len(lines)):
        ln = lines[j].strip()
        if not ln:
            continue
        ln_l = ln.lower().rstrip(":")
        if ln_l in known_labels:
            break
        vals.append(ln)

        if j + 1 < len(lines):
            nxt = lines[j + 1].strip()
            if nxt and (nxt.lower().rstrip(":") in known_labels):
                break

    return " ".join(vals).strip()

def parse_dimensions_from_overall(overall_value):
    out = {"Width": "", "Depth": "", "Height": "", "Length": "", "Diameter": ""}

    if not isinstance(overall_value, str) or not overall_value.strip():
        return out

    txt = overall_value.lower()
    parts = re.findall(r'([\d\.]+)\s*"\s*([a-z\.]+)', txt)

    for v, t in parts:
        t = t.strip(".")
        if t == "w":
            out["Width"] = v
        elif t == "d":
            out["Depth"] = v
        elif t == "h":
            out["Height"] = v
        elif t == "l":
            out["Length"] = v
        elif t == "dia":
            out["Diameter"] = v

    return out

def parse_measurements(value_text):
    out = {"H": "", "W": "", "D": "", "L": "", "DIA": ""}

    if not isinstance(value_text, str) or not value_text.strip():
        return out

    txt = value_text.lower()
    parts = re.findall(r'([0-9]+(?:\.[0-9]+)?)\s*"\s*([a-z\.]+)', txt)

    for v, t in parts:
        t = t.strip(".")
        if t == "h":
            out["H"] = v
        elif t == "w":
            out["W"] = v
        elif t == "d":
            out["D"] = v
        elif t == "l":
            out["L"] = v
        elif t == "dia":
            out["DIA"] = v

    return out

def clean_number(text):
    if not isinstance(text, str):
        return ""
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", text)
    return m.group(1) if m else ""

def clean_weight_number(weight_value):
    return clean_number(weight_value)

def first_number_from_dim(text):
    """
    For Shade Top/Bottom/Height:
    - Shade Height: "5.25\"h ..." => returns 5.25
    - Shade Top: "8\"w x 3.75\"d" => returns 8  (your desired values-only style)
    """
    return clean_number(text)

def build_shade_details_one_column(dim_full):
    """
    FIX:
    1) Try Shade Top / Shade Bottom / Shade Height (best for Currey)
    2) If missing, fallback to Shade Details / Shade
    Output: "X, Y, Z"
    """
    top_val    = extract_value_block(dim_full, "Shade Top")
    bottom_val = extract_value_block(dim_full, "Shade Bottom")
    height_val = extract_value_block(dim_full, "Shade Height")

    parts = []
    if top_val:
        parts.append(first_number_from_dim(top_val))
    if bottom_val:
        parts.append(first_number_from_dim(bottom_val))
    if height_val:
        parts.append(first_number_from_dim(height_val))

    parts = [p for p in parts if p]

    if parts:
        return ", ".join(parts).strip()

    # fallback (older layout)
    shade_val = extract_value_block(dim_full, "Shade Details")
    if not shade_val:
        shade_val = extract_value_block(dim_full, "Shade")

    if not shade_val:
        return ""

    # fallback: try pull 3 numbers from whatever shade text has
    nums = re.findall(r"([0-9]+(?:\.[0-9]+)?)", shade_val)
    nums = [n for n in nums if n]
    if len(nums) >= 3:
        return ", ".join(nums[:3]).strip()
    if len(nums) > 0:
        return ", ".join(nums).strip()
    return ""

# =========================
# SPECIFICATIONS
# =========================
def get_spec_table():
    try:
        summ = find_accordion_summary("specifications")
        root = ensure_accordion_open(summ)

        txt = root.find_element(
            By.XPATH, ".//div[contains(@class,'MuiAccordionDetails-root')]"
        ).text.strip()

        ensure_accordion_closed(summ)
        return txt
    except:
        return ""

def parse_specs(spec_text):
    fields = {"Finish": "", "Color Temperature": "", "Socket Type": "", "Wattage": ""}

    if not isinstance(spec_text, str) or not spec_text.strip():
        return fields

    lines = [x.strip() for x in spec_text.split("\n") if x.strip()]

    def get_after(keyword, require_digit=False):
        kw = keyword.lower()
        idx = -1
        for i, ln in enumerate(lines):
            if kw in ln.lower():
                idx = i
                break
        if idx == -1:
            return ""
        for j in range(idx + 1, len(lines)):
            ln = lines[j].strip()
            if not ln:
                continue
            if require_digit and not re.search(r"\d", ln):
                continue
            return ln
        return ""

    fields["Finish"] = get_after("Finish", require_digit=False)
    fields["Color Temperature"] = get_after("Color Temperature", require_digit=True)

    socket_val = get_after("Socket Type", require_digit=False) or get_after("Socket", require_digit=False)
    fields["Socket Type"] = socket_val

    watt_val = get_after("Watts per Socket/Item", require_digit=True) or get_after("Wattage", require_digit=True)
    fields["Wattage"] = watt_val

    return fields

# =========================
# OUTPUT HEADERS (append after column F)
# =========================
DETAIL_HEADERS = [
    "Product Family Id", "Description", "Dimension",
    "Width", "Depth", "Diameter", "Length", "Height",
    "Weight",
    "Shade Details", "Canopy",
    "Seat Height", "Seat Width", "Seat Depth",
    "Arm Height", "Arm Width", "Arm Length",
    "Finish", "Color Temperature", "Socket Type", "Wattage",
    "Specifications"
]

def write_headers(ws):
    start_col = COL_SKU + 1  # after F
    for i, h in enumerate(DETAIL_HEADERS):
        cell = ws.cell(row=HEADER_ROW, column=start_col + i)
        cell.value = h
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

def save_copy_workbook():
    wb0 = load_workbook(INPUT_XLSX)
    wb0.save(OUTPUT_XLSX)

# =========================
# MAIN
# =========================
save_copy_workbook()

wb = load_workbook(OUTPUT_XLSX)

# ✅ MULTI-SHEET LOOP
for ws in wb.worksheets:
    print(f"\n📊 Processing sheet: {ws.title}")
    write_headers(ws)
    processed = 0
    row_num = START_ROW

    while True:
        url = ws.cell(row=row_num, column=COL_URL).value
        if not url or str(url).strip() == "":
            break

        base_url  = str(url).strip()
        base_img  = ws.cell(row=row_num, column=COL_IMG).value or ""
        base_name = ws.cell(row=row_num, column=COL_NAME).value or ""
        base_sku  = ws.cell(row=row_num, column=COL_SKU).value or ""

        print(f"\n🔗 Row {row_num} → {base_url}")

        driver.get(base_url)
        time.sleep(1.2)

        p_url, p_img, p_name = get_product_header(base_url, str(base_img), str(base_name))
        desc = get_description()
        family_id = p_name

        dim_full = get_dimensions_full_text()

        overall_val = extract_value_block(dim_full, "Overall")
        weight_val  = extract_value_block(dim_full, "Item Weight")
        canopy_val  = extract_value_block(dim_full, "Canopy")

        # ✅ FIXED Shade Details
        shade_details_one = build_shade_details_one_column(dim_full)

        # Seat/Arm blocks
        seat_h_val = extract_value_block(dim_full, "Seat Height")
        seat_w_val = extract_value_block(dim_full, "Seat Width")
        seat_d_val = extract_value_block(dim_full, "Seat Depth")

        arm_h_val  = extract_value_block(dim_full, "Arm Height")
        arm_w_val  = extract_value_block(dim_full, "Arm Width")
        arm_l_val  = extract_value_block(dim_full, "Arm Length")

        dim_vals = parse_dimensions_from_overall(overall_val)
        w_clean = clean_weight_number(weight_val)
        canopy = canopy_val

        seat_h = parse_measurements(seat_h_val).get("H") or clean_number(seat_h_val)
        seat_w = parse_measurements(seat_w_val).get("W") or clean_number(seat_w_val)
        seat_d = parse_measurements(seat_d_val).get("D") or clean_number(seat_d_val)

        arm_h  = parse_measurements(arm_h_val).get("H") or clean_number(arm_h_val)
        arm_w  = parse_measurements(arm_w_val).get("W") or clean_number(arm_w_val)
        arm_l  = parse_measurements(arm_l_val).get("L") or clean_number(arm_l_val)

        spec_text = get_spec_table()
        spec_vals = parse_specs(spec_text)

        out_vals = [
            family_id,
            desc,
            dim_full,

            dim_vals["Width"],
            dim_vals["Depth"],
            dim_vals["Diameter"],
            dim_vals["Length"],
            dim_vals["Height"],

            w_clean,

            shade_details_one,   # ✅ Shade Details as ONE column (values only)
            canopy,              # Canopy

            seat_h, seat_w, seat_d,
            arm_h, arm_w, arm_l,

            spec_vals["Finish"],
            spec_vals["Color Temperature"],
            spec_vals["Socket Type"],
            spec_vals["Wattage"],

            spec_text
        ]

        start_col = COL_SKU + 1
        for i, v in enumerate(out_vals):
            ws.cell(row=row_num, column=start_col + i).value = v

        processed += 1

        if processed % AUTOSAVE_EVERY == 0:
            wb.save(OUTPUT_XLSX)
            print(f"💾 Autosaved after {processed} rows...")

        row_num += 1

    print(f"✅ Sheet '{ws.title}' complete: {processed} products processed")

driver.quit()
wb.save(OUTPUT_XLSX)
print(f"\n✅ ALL SHEETS DONE → {OUTPUT_XLSX}")