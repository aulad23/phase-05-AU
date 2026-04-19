# -*- coding: utf-8 -*-
# Currey & Company – Step-1 Scraper
# Scraping logic preserved from CODE-A
# Input / Output system adopted from CODE-B

import os
import time
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from urllib.parse import urljoin

# =========================================================
# INPUT SYSTEM (FROM CODE-B)
# =========================================================

CATEGORIES = {
    """
    "Nightstands" :[
             "https://www.curreyandcompany.com/c/furniture/chests-nightstands/?product+type=nightstands"
        ],
        "Coffee & Cocktail Tables": [
            "https://www.curreyandcompany.com/c/furniture/tables/?product+type=cocktail+tables"
        ],
    
        "Side & End Tables": [
            "https://www.curreyandcompany.com/c/furniture/tables/?product+type=side+tables"
        ],
    
        "Dining Tables": [
            "https://www.curreyandcompany.com/c/furniture/tables/?product+type=dining+tables"
        ],
    
        "Consoles": [
            "https://www.curreyandcompany.com/c/furniture/tables/?product+type=console+tables"
        ],
    
        "Desks": [
            "https://www.curreyandcompany.com/c/furniture/desks-vanities/?"
        ],
    
        "Dressers & Chests": [
            "https://www.curreyandcompany.com/c/furniture/chests-nightstands/?product+type=chests"
        ],
           """
    
        "Cabinets": [
            "https://www.curreyandcompany.com/c/furniture/cabinets-credenzas/?"
        ],

        "Accent Tables": [
            "https://www.curreyandcompany.com/c/furniture/tables/?product+type=accent+tables"
        ],

    "Dining Chairs": [
        "https://www.curreyandcompany.com/c/furniture/dining-chairs/?"
    ],

    "Bar Stools": [
        "https://www.curreyandcompany.com/c/furniture/bar-counter-stools/?"
    ],

    "Lounge Chairs": [
        "https://www.curreyandcompany.com/c/furniture/accent-chairs/?"
    ],
    "Chandeliers": [
        "https://www.curreyandcompany.com/c/lighting/chandeliers/?"
    ],
   "Pendants": [
        "https://www.curreyandcompany.com/c/lighting/pendants/?"
    ],

    "Sconces": [
        "https://www.curreyandcompany.com/c/lighting/wall-sconces/?"
    ],

    "Flush Mount": [
        "https://www.curreyandcompany.com/c/lighting/flush-mounts/?"
    ],

    "Table Lamps": [
        "https://www.curreyandcompany.com/c/lighting/table-lamps/?"
    ],

    "Floor Lamps": [
        "https://www.curreyandcompany.com/c/lighting/floor-lamps/?"
    ],

    "Lanterns": [
        "https://www.curreyandcompany.com/c/lighting/lanterns/?"
    ],

    "Mirrors": [
        "https://www.curreyandcompany.com/c/accessories/mirrors/?"
    ],

    "Bowls & Vases": [
        "https://www.curreyandcompany.com/c/accessories/vases-jars-bowls/?"
    ],

    "Objects": [
        "https://www.curreyandcompany.com/c/accessories/objects-sculptures/?"
    ],

    "Boxes": [
        "https://www.curreyandcompany.com/c/accessories/boxes-trays/?product+type=decorative+boxes"
    ],
   "Trays": [
        "https://www.curreyandcompany.com/c/accessories/boxes-trays/?product+type=decorative+trays"
    ],


}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "currey_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "Currey.xlsx")

DOMAIN = "https://www.curreyandcompany.com"

# =========================================================
# DRIVER SETUP (UPDATED TO SHOW BROWSER WINDOW)
# =========================================================

def connect_driver():
    options = Options()
    # Removed headless mode - browser window will now be visible
    options.add_argument("--disable-blink-features=AutomationControlled")
    return webdriver.Chrome(options=options)

# =========================================================
# SCRAPING LOGIC (FROM CODE-A — UNCHANGED)
# =========================================================

def scrape_category(driver, base_url, category_name):
    all_data = []
    page = 1

    while True:
        if "?" in base_url and not base_url.endswith("?"):
            page_url = base_url + f"&page={page}"
        else:
            page_url = base_url + f"page={page}"

        driver.get(page_url)
        time.sleep(4)

        product_blocks = driver.find_elements(By.CSS_SELECTOR, "div.relative.group")
        if not product_blocks:
            break

        for p in product_blocks:
            try:
                a_tag = p.find_element(By.CSS_SELECTOR, "a[href]")
                product_url = a_tag.get_attribute("href")
                if not product_url.startswith("http"):
                    product_url = urljoin(DOMAIN, product_url)

                img_tag = p.find_element(By.CSS_SELECTOR, "img")
                img_src = img_tag.get_attribute("src")
                image_url = img_src if img_src.startswith("http") else urljoin(DOMAIN, img_src)

                name_tag = p.find_element(By.CSS_SELECTOR, "div.paragraph-3a-sm-desktop")
                product_name = name_tag.text.strip()

                sku_tag = p.find_element(By.CSS_SELECTOR, "div.paragraph-3b-sm")
                sku = sku_tag.text.strip()

                all_data.append({
                    "Category": category_name,
                    "Product URL": product_url,
                    "Image URL": image_url,
                    "Product Name": product_name,
                    "SKU": sku
                })

            except Exception:
                continue

        page += 1
        time.sleep(2)

    return all_data

# =========================================================
# OUTPUT SYSTEM (FROM CODE-B)
# =========================================================

def build_master_excel(all_products):
    if not all_products:
        df = pd.DataFrame(columns=["Category", "Product URL", "Image URL", "Product Name", "SKU"])
        df.to_excel(master_output_file, index=False)
        return df

    df = pd.DataFrame(all_products)
    df.drop_duplicates(subset=["Product URL"], inplace=True)
    df = df[["Category", "Product URL", "Image URL", "Product Name", "SKU"]]
    df.to_excel(master_output_file, index=False)
    return df

def build_category_wise_workbook(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for category in CATEGORIES.keys():
        df_cat = df[df["Category"] == category]
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=category[:31])

        ws["A1"] = "Brand"
        ws["B1"] = "Currey & Company"
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES[category])
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start_row = 4
        for col_idx, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col).font = bold

        for r_idx, row in enumerate(df_out.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        headers = {ws.cell(row=start_row, column=c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)

# =========================================================
# MAIN (FROM CODE-B STRUCTURE)
# =========================================================

def main():
    driver = connect_driver()
    all_products = []

    try:
        for category, urls in CATEGORIES.items():
            for url in urls:
                products = scrape_category(driver, url, category)
                all_products.extend(products)
    finally:
        driver.quit()

    df_master = build_master_excel(all_products)

    if not df_master.empty:
        build_category_wise_workbook(df_master)
    else:
        Workbook().save(master_sheets_file)

if __name__ == "__main__":
    main()