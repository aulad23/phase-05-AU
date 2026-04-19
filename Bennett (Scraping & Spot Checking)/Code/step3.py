import os
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from urllib.parse import urlparse, urlencode, urlunparse, parse_qs
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================
# INPUT SYSTEM (from CODE-B)
# =========================

base_domain = "https://www.bennetttothetrade.com"

CATEGORIES = {
    "Coffee & Cocktail Tables": [
        f"{base_domain}/collections/coffee-tables"
    ],
    "Dining Tables": [
        f"{base_domain}/collections/dining-tables"
    ],

    "Consoles": [
        f"{base_domain}/collections/consoles"
    ],
    "Beds & Headboards": [
        f"{base_domain}/collections/beds"
    ],

    "Desks": [
        f"{base_domain}/collections/desks"
    ],

    "Bookcases": [
        f"{base_domain}/collections/bookcases-vitrines"
    ],

    "Dressers & Chests": [
        f"{base_domain}/collections/chests"
    ],

    "Cabinets": [
        f"{base_domain}/collections/sideboards"
    ],

    "Accent Tables": [
        f"{base_domain}/collections/occasional-tables"
    ],

    "Dining Chairs": [
        f"{base_domain}/collections/chairs"
    ],

    "Benches": [
        f"{base_domain}/collections/benches"
    ],

    "Mirrors": [
        f"{base_domain}/collections/mirrors"
    ],


    # উদাহরণ (future use):
    # "Coffee Tables": [
    #     f"{base_domain}/collections/coffee-tables"
    # ],
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "bennett_master.xlsx")
category_output_file = os.path.join(script_dir, "Bennett.xlsx")

HEADERS = {"User-Agent": "Mozilla/5.0"}

# =========================
# SCRAPING LOGIC (CODE-A)
# ❗ UNCHANGED
# =========================

def scrape_collection(base_collection_url):
    page = 1
    all_data = []

    while True:
        url = f"{base_collection_url}?page={page}"
        print(f"🔎 Scraping page {page} — {url}")
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            break

        soup = BeautifulSoup(response.text, 'html.parser')
        products = soup.select('div.card-wrapper.product-card-wrapper')

        if not products:
            break

        for product in products:
            # Image
            img_tag = product.select_one('div.media img')
            image_url = None
            if img_tag and img_tag.get('src'):
                image_url = img_tag['src']
                if image_url.startswith("//"):
                    image_url = "https:" + image_url

            # Product URL, Name, SKU
            a_tag = product.select_one('h3.card__heading.h5 a.full-unstyled-link')
            product_url = name = sku = None
            if a_tag:
                href = a_tag.get('href')
                product_url = base_domain + href if href else None
                name = a_tag.get_text(strip=True)
                sku = name

            if product_url:
                all_data.append({
                    "Product URL": product_url,
                    "Image URL": image_url,
                    "Product Name": name,
                    "SKU": sku
                })

        page += 1
        time.sleep(1.5)

    return all_data

# =========================
# OUTPUT SYSTEM (from CODE-B)
# =========================

def build_master_dataframe():
    dfs = []

    for category, links in CATEGORIES.items():
        cat_products = []
        for link in links:
            data = scrape_collection(link)
            cat_products.extend(data)

        if cat_products:
            df = pd.DataFrame(cat_products)
            df.drop_duplicates(subset=["Product URL"], inplace=True)
            df.insert(0, "Category", category)
            dfs.append(df)
        else:
            print(f"⚠️ No products found for category: {category}")

    if not dfs:
        cols = ["Category", "Product URL", "Image URL", "Product Name", "SKU"]
        master = pd.DataFrame(columns=cols)
        master.to_excel(master_output_file, index=False)
        return master

    master = pd.concat(dfs, ignore_index=True)
    master.to_excel(master_output_file, index=False)
    return master

def build_category_wise_workbook(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for category in CATEGORIES.keys():
        df_cat = df[df["Category"] == category].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=category)

        ws["A1"] = "Brand"
        ws["B1"] = "Bennett"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(CATEGORIES[category])
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start_row = 4
        for col_idx, col in enumerate(df_cat.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col).font = bold

        for r_idx, row in enumerate(df_cat.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        headers = {ws.cell(row=start_row, column=c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

        ws.freeze_panes = None

    wb.save(category_output_file)

# =========================
# MAIN (from CODE-B)
# =========================

def main():
    df = build_master_dataframe()
    if df.empty:
        wb = Workbook()
        wb.save(category_output_file)
        return

    build_category_wise_workbook(df)

if __name__ == "__main__":
    main()
