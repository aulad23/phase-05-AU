import os
import time
import re
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# 🏠 Base URL
base_domain = "https://www.bennetttothetrade.com"

# 📁 File paths (UPDATED)
script_dir = os.path.dirname(os.path.abspath(__file__))
source_path = os.path.join(script_dir, "Bennett.xlsx")
final_path = os.path.join(script_dir, "Bennett_final2.xlsx")

HEADERS = {"User-Agent": "Mozilla/5.0"}

# ================================
# STEP 1: SCRAPING PRODUCT DETAILS
# ================================

print("=" * 60)
print("STEP 1: SCRAPING PRODUCT DETAILS")
print("=" * 60)

# ✅ Required columns (UNCHANGED logic; only I/O changes)
required_detail_cols = [
    "Description", "Weight", "Product Family Id", "List Price"
]


def extract_details(soup):
    desc_block = soup.select_one("div.product__description.rte.quick-add-hidden")
    description = ""
    weight = ""
    product_family_id = ""
    list_price = ""

    # Extract Product Family Id
    name_tag = soup.find("h1")
    if name_tag:
        product_family_id = name_tag.get_text(strip=True)

    # Extract Description
    if desc_block:
        description = "\n".join([p.get_text(separator="\n", strip=True) for p in desc_block.find_all("p")])

    # ✨ NEW: Check for dimensions in bold/strong tags near the price
    # Some products have dimensions like **42"x 42"x 21"H** outside description
    dimension_text = ""
    for strong_tag in soup.find_all(['strong', 'b']):
        text = strong_tag.get_text(strip=True)
        # Check if it looks like dimensions (contains x and numbers with ")
        if 'x' in text.lower() and '"' in text and re.search(r'\d+', text):
            dimension_text = text
            break

    # Append dimension text to description if found and not already in description
    if dimension_text and dimension_text not in description:
        if description:
            description = description + "\n" + dimension_text
        else:
            description = dimension_text

    # Extract List Price (try multiple selectors and extract only numbers)
    price_text = ""

    price_selectors = [
        "span.price-item.price-item--regular",
        ".price-item--regular",
        "#price-template--17007414902996__main .price__container .price__regular .price-item--regular",
        ".price__regular .price-item"
    ]

    for selector in price_selectors:
        price_block = soup.select_one(selector)
        if price_block:
            price_text = price_block.get_text(strip=True)
            break

    if price_text:
        price_match = re.search(r'([0-9,]+\.?[0-9]*)', price_text)
        if price_match:
            list_price = price_match.group(1).replace(',', '')

    return description, weight, product_family_id, list_price


# ================================
# STEP 2: EXTRACT DIMENSIONS
# ================================

print("=" * 60)
print("STEP 2: EXTRACTING DIMENSIONS")
print("=" * 60)

dimension_cols = ["Length", "Width", "Depth", "Diameter", "Height", "Seat Height", "COM"]


def extract_dimensions(text):
    length = width = depth = diameter = height = seat_height = com = ""

    if not isinstance(text, str):
        return length, width, depth, diameter, height, seat_height, com

    text = text.replace("â€³", '"').replace("Ã—", "x").replace("  ", " ").strip()

    # COM
    com_match = re.search(r'COM-?\s*([0-9\.]+)\s*(?:yards?)?', text, re.IGNORECASE)
    if com_match:
        com = com_match.group(1)
        text = re.sub(r'COM-?\s*[0-9\.]+\s*(?:yards?)?', '', text, flags=re.IGNORECASE)

    # Seat Height patterns
    seat_match1 = re.search(r'(?:\()?([0-9\.]+)"?\s*(?:H)?\s*to\s+[Ss]eat', text, re.IGNORECASE)
    if seat_match1:
        seat_height = seat_match1.group(1)

    if not seat_height:
        seat_match2 = re.search(r'Seat\s+Height[:\s]*([0-9\.]+)"?', text, re.IGNORECASE)
        if seat_match2:
            seat_height = seat_match2.group(1)

    if not seat_height:
        seat_match3 = re.search(r'\(([0-9\.]+)"?\s+to\s+seat\)', text, re.IGNORECASE)
        if seat_match3:
            seat_height = seat_match3.group(1)

    if seat_height:
        text = re.sub(r'(?:\()?[0-9\.]+"?\s*(?:H)?\s*to\s+[Ss]eat(?:\))?', '', text, flags=re.IGNORECASE)
        text = re.sub(r'Seat\s+Height[:\s]*[0-9\.]+"?', '', text, flags=re.IGNORECASE)

    # Main 3D dimensions
    if 'x' in text.lower():
        dim_part = re.search(r'([0-9\.]+)"?\s*x\s*([0-9\.]+)"?\s*x\s*([0-9\.]+)"?\s*([HhDd])?', text)
        if dim_part:
            length = dim_part.group(1)
            width = dim_part.group(2)
            height_or_depth = dim_part.group(3)
            label = dim_part.group(4)

            if label and label.upper() == 'H':
                height = height_or_depth
            elif label and label.upper() == 'D':
                depth = height_or_depth
            else:
                height = height_or_depth

    # Labeled values
    matches = re.findall(r'([0-9\.]+)\s*"?\s*([A-Za-z]*)', text)
    for num, label in matches:
        label = label.upper().strip()
        if "H" in label and not height:
            height = num
        elif "W" in label and not width:
            width = num
        elif "L" in label and not length:
            length = num
        elif "DIA" in label:
            diameter = num
        elif "D" in label and not depth:
            depth = num

    # Fallback length
    if not length:
        all_nums = re.findall(r'([0-9]+\.?[0-9]*)', text)
        if all_nums:
            length = all_nums[-1]

    # Keep numeric only
    length = re.sub(r'[^\d.]+', '', length) if length else ""
    width = re.sub(r'[^\d.]+', '', width) if width else ""
    depth = re.sub(r'[^\d.]+', '', depth) if depth else ""
    height = re.sub(r'[^\d.]+', '', height) if height else ""
    seat_height = re.sub(r'[^\d.]+', '', seat_height) if seat_height else ""
    diameter = re.sub(r'[^\d.]+', '', diameter) if diameter else ""
    com = re.sub(r'[^\d.]+', '', com) if com else ""

    return length, width, depth, diameter, height, seat_height, com


# ================================
# INPUT / OUTPUT SYSTEM (UPDATED)
# ================================

def read_sheet_as_df(ws):
    """
    Reads sheet that follows:
    A1 Brand | B1 BrandName
    A2 Link  | B2 Link(s)
    Row3 blank
    Row4 header
    Row5.. data
    """
    brand_name = ws["B1"].value or ""
    link_value = ws["B2"].value or ""

    header_row = 4
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            break
        headers.append(str(v).strip())

    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = []
        empty_row = True
        for c in range(1, len(headers) + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(v)
            if v not in (None, ""):
                empty_row = False
        if empty_row:
            continue
        data.append(row_vals)

    df = pd.DataFrame(data, columns=headers)
    return brand_name, link_value, df


def write_sheet_from_df(wb_out, sheet_name, brand_name, link_value, df):
    if sheet_name in wb_out.sheetnames:
        ws = wb_out[sheet_name]
        wb_out.remove(ws)
    ws = wb_out.create_sheet(title=sheet_name)

    bold = Font(bold=True)

    # Meta rows (exact)
    ws["A1"] = "Brand"
    ws["B1"] = brand_name
    ws["A2"] = "Link"
    ws["B2"] = link_value
    ws["B2"].alignment = Alignment(wrap_text=True)

    # Row 3 blank (do nothing)

    # Header row 4
    start_row = 4
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = bold

    # Data from row 5
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def ensure_columns(df):
    # Ensure base columns exist from your Step-1 structure
    base_cols = ["Index", "Category", "Product URL", "Image URL", "Product Name", "SKU"]
    for c in base_cols:
        if c not in df.columns:
            df[c] = ""

    # Ensure detail columns exist
    for c in required_detail_cols:
        if c not in df.columns:
            df[c] = ""

    # Ensure dimension columns exist
    for c in dimension_cols:
        if c not in df.columns:
            df[c] = ""

    return df


def process_sheet(df):
    df = ensure_columns(df)

    # Pending products (same logic)
    pending_df = df[df["Description"].isna() | (df["Description"] == "")]
    print(f"🔎 {len(pending_df)} products pending to scrape.\n")

    batch_size = 5
    processed = 0
    total = len(pending_df)

    for index, row in pending_df.iterrows():
        product_url = row["Product URL"]
        if not isinstance(product_url, str) or not product_url.startswith("http"):
            continue

        processed += 1
        print(f"🔎 [{processed}/{total}] Scraping: {product_url}")

        try:
            response = requests.get(product_url, headers=HEADERS)
            if response.status_code != 200:
                print(f"⚠️ Failed to fetch: {product_url}")
                continue

            soup = BeautifulSoup(response.text, "html.parser")
            desc, weight, product_family_id, list_price = extract_details(soup)

            df.at[index, "Description"] = desc
            df.at[index, "Weight"] = weight
            df.at[index, "Product Family Id"] = product_family_id
            df.at[index, "List Price"] = list_price

            print(f"✅ Done: {product_url}")

            if processed % batch_size == 0:
                print(f"💾 Batch complete ({processed})...\n")

            time.sleep(1.5)

        except Exception as e:
            print(f"❌ Error processing {product_url}: {e}")
            continue

    # Dimension extraction (same logic)
    print(f"🔎 Extracting dimensions from {len(df)} products...\n")
    for index, row in df.iterrows():
        description = row.get("Description", "")
        length, width, depth, diameter, height, seat_height, com = extract_dimensions(description)

        df.at[index, "Length"] = length
        df.at[index, "Width"] = width
        df.at[index, "Depth"] = depth
        df.at[index, "Diameter"] = diameter
        df.at[index, "Height"] = height
        df.at[index, "Seat Height"] = seat_height
        df.at[index, "COM"] = com

    # ✅ Reorder columns (UPDATED to match your new sheet structure)
    column_order = [
        "Index", "Category",
        "Product URL", "Image URL", "Product Name", "SKU",
        "Product Family Id", "Description", "List Price", "Weight",
        "Width", "Depth", "Diameter", "Length", "Height", "Seat Height", "COM"
    ]

    for col in column_order:
        if col not in df.columns:
            df[col] = ""

    df = df[column_order]
    return df


def main():
    if not os.path.exists(source_path):
        print(f"❌ Source file not found: {source_path}")
        return

    wb_in = load_workbook(source_path)
    wb_out = Workbook()
    # remove default sheet
    wb_out.remove(wb_out.active)

    print(f"📂 Input workbook loaded: {source_path}")
    print(f"📄 Sheets found: {len(wb_in.sheetnames)}\n")

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        print("=" * 60)
        print(f"🧾 Processing sheet: {sheet_name}")
        print("=" * 60)

        brand_name, link_value, df = read_sheet_as_df(ws)

        if df.empty:
            print(f"⚠️ No data found in sheet: {sheet_name} (skipping)")
            # still write empty structure if you want
            df = ensure_columns(pd.DataFrame())
            df = df[["Index", "Category", "Product URL", "Image URL", "Product Name", "SKU"]]
            write_sheet_from_df(wb_out, sheet_name, brand_name, link_value, df)
            continue

        df_final = process_sheet(df)

        # Write to output workbook in the same Step-1 layout
        write_sheet_from_df(wb_out, sheet_name, brand_name, link_value, df_final)

        # Save after each sheet (safer)
        wb_out.save(final_path)
        print(f"✅ Sheet saved: {sheet_name} -> {final_path}\n")

    wb_out.save(final_path)
    print("\n" + "=" * 60)
    print("🎉 PROCESS COMPLETED!")
    print("=" * 60)
    print(f"📌 Source file: {source_path}")
    print(f"📌 Final file:  {final_path}")


if __name__ == "__main__":
    main()