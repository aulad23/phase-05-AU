"""
Blue Dot — Combined Scraper
Step 1 : Shopify API /collections/{handle}/products.json  (Name, SKU, Price, Image, Description, Finish, Tags)
Step 2 : Product page  (Dimensions JSON + Product Details)
Output : Excel, one sheet per category, auto-save after each
"""

import os, re, time, json, html
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ────────────────────────────────────────────────────────────────────
MANUFACTURER = "Blu Dot"
BASE_URL     = "https://www.bludot.com"

DEMO_MODE  = True
DEMO_COUNT = 3
PAGE_SIZE  = 250
RATE_LIMIT = 0.5   # seconds between requests

OUTPUT_DIR  = "Demo" if DEMO_MODE else "Data"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "BluDot_demo.xlsx" if DEMO_MODE else "BluDot.xlsx")

HTTP_HEADERS = {
    "User-Agent"     : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

# ── CATEGORIES (from Vendor List) ─────────────────────────────────────────────
CATEGORIES = [
    {"name": "Nightstands",              "handles": ["nightstands-side-tables"]},
    {"name": "Coffee & Cocktail Tables", "handles": ["coffee-tables"]},
    {"name": "Side & End Tables",        "handles": ["side-tables"]},
    {"name": "Consoles",                 "handles": ["consoles", "media-consoles"]},
    {"name": "Beds & Headboards",        "handles": ["modern-beds"]},
    {"name": "Bookcases",                "handles": ["bookcases", "bedroom-closet-storage"]},
    {"name": "Dressers & Chests",        "handles": ["dressers"]},
    {"name": "Sofas & Loveseats",        "handles": ["living-sofas"]},
    {"name": "Sectionals",               "handles": ["sectional-sofas"]},
    {"name": "Lounge Chairs",            "handles": ["lounge-chairs"]},
    {"name": "Ottomans & Benches",       "handles": ["ottomans-benches"]},
    {"name": "Pendants",                 "handles": ["ceiling-lights"]},
    {"name": "Sconces",                  "handles": ["wall-lights"]},
    {"name": "Table Lamps",              "handles": ["table-lamps"]},
    {"name": "Floor Lamps",              "handles": ["floor-lamps"]},
    {"name": "Mirrors",                  "handles": ["mirrors"]},
    {"name": "Pillows & Throws",         "handles": ["pillows-throws"]},
    {"name": "Vases",                    "handles": ["vases-trays-catchalls"]},
    {"name": "Objects",                  "handles": ["candles-candle-holders"]},
    {"name": "Wall Decor",               "handles": ["wall-art"]},
    {"name": "Rugs",                     "handles": ["living-room-rugs"]},
]

# ── EXCEL STYLE ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BRAND_FONT  = Font(bold=True, size=12)
CENTER      = Alignment(horizontal="center")

BASE_COLUMNS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Product Family Id", "Description",
    "Weight", "Width", "Depth", "Diameter", "Height",
    "Seat Width", "Seat Depth", "Seat Height",
    "Arm Width", "Arm Height",
    "Back Width", "Back Depth", "Back Height",
    "Price", "Finish", "Materials", "Tags", "Notes",
]

# ── HELPERS ───────────────────────────────────────────────────────────────────
def _family_id(name):
    return re.split(r'[,._\-]', name, maxsplit=1)[0].strip()


def _strip_html(raw):
    """Strip HTML tags and decode entities."""
    if not raw:
        return ""
    text = re.sub(r'<br\s*/?>', '\n', raw, flags=re.I)
    text = re.sub(r'<li[^>]*>', '- ', text, flags=re.I)
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    return re.sub(r'\n{3,}', '\n\n', text).strip()


# Section title → dimension prefix for column name
_SECTION_MAP = {
    "overall": "",
    "seat"   : "Seat ",
    "arm"    : "Arm ",
    "back"   : "Back ",
}

# Abbreviation → column suffix
_ABBREV_MAP = {
    "w": "Width",
    "d": "Depth",
    "h": "Height",
    "l": "Length",
}


def _dim_col(section_title, abbrev):
    """Return standard column name, or None if unknown."""
    s = section_title.strip().lower()
    a = abbrev.strip().lower()
    prefix = _SECTION_MAP.get(s)
    if prefix is None:
        return None  # unknown section -> dynamic
    suffix = _ABBREV_MAP.get(a)
    if suffix is None:
        return None
    return (prefix + suffix).strip()


# ── SHOPIFY API ───────────────────────────────────────────────────────────────
def fetch_collection(handle):
    """Fetch all products for a Shopify collection handle."""
    products = []
    page = 1
    while True:
        url = f"{BASE_URL}/collections/{handle}/products.json?limit={PAGE_SIZE}&page={page}"
        try:
            r = requests.get(url, headers=HTTP_HEADERS, timeout=30)
            if r.status_code == 429:
                print(f"  ! 429 on {handle} page {page} — waiting 10s")
                time.sleep(10)
                continue
            if r.status_code != 200:
                print(f"  ! HTTP {r.status_code} for {handle}")
                break
            batch = r.json().get("products", [])
            if not batch:
                break
            products.extend(batch)
            if len(batch) < PAGE_SIZE:
                break
            page += 1
            time.sleep(RATE_LIMIT)
        except Exception as e:
            print(f"  ! Error fetching {handle}: {e}")
            break
    return products


def parse_listing(raw_products, cat_name):
    """Convert Shopify API product dicts to our row format."""
    rows  = []
    seen  = set()

    for p in raw_products:
        try:
            handle = p.get("handle", "")
            if not handle:
                continue

            key = handle
            if key in seen:
                continue
            seen.add(key)

            title  = p.get("title", "").strip()
            src    = f"{BASE_URL}/products/{handle}"
            images = p.get("images", [])
            img    = images[0]["src"] if images else ""
            desc   = _strip_html(p.get("body_html", ""))
            tags   = ", ".join(p.get("tags", []))

            variants = p.get("variants", [{}])
            v0       = variants[0] if variants else {}
            sku      = v0.get("sku", "").strip()
            raw_price = v0.get("price", "")
            try:
                price = f"{float(raw_price):.2f}" if raw_price else ""
            except Exception:
                price = str(raw_price)

            # Collect unique finish values across all variants
            options = p.get("options", [])
            color_idx = None
            for i, opt in enumerate(options):
                if opt.get("name", "").lower() in ("color", "finish", "colour"):
                    color_idx = i
                    break
            if color_idx is not None:
                opt_key = f"option{color_idx + 1}"
                finishes = list(dict.fromkeys(
                    v.get(opt_key, "").strip()
                    for v in variants if v.get(opt_key, "").strip()
                ))
                finish = ", ".join(finishes)
            else:
                finish = ""

            rows.append({
                "Category"        : cat_name,
                "Manufacturer"    : MANUFACTURER,
                "Source"          : src,
                "Image URL"       : img,
                "Product Name"    : title,
                "SKU"             : sku,
                "Product Family Id": _family_id(title),
                "Description"     : desc,
                "Price"           : price,
                "Finish"          : finish,
                "Tags"            : tags,
                "_handle"         : handle,
            })

            if DEMO_MODE and len(rows) >= DEMO_COUNT:
                break

        except Exception as e:
            print(f"  ! Parse error: {e}")

    return rows


def scrape_listing(cat):
    """Scrape all products for a category (one or more handles)."""
    cat_name = cat["name"]
    handles  = cat["handles"]
    rows     = []
    seen     = set()

    print(f"\n{'-'*60}")
    print(f"Category: {cat_name}  ({len(handles)} collection(s))")

    for handle in handles:
        print(f"  -> {handle}")
        raw = fetch_collection(handle)
        print(f"     API returned {len(raw)} products")

        partial = parse_listing(raw, cat_name)
        added   = 0
        for r in partial:
            k = r["_handle"]
            if k not in seen:
                seen.add(k)
                rows.append(r)
                added += 1
            if DEMO_MODE and len(rows) >= DEMO_COUNT:
                break

        print(f"     {added} new (total: {len(rows)})")
        time.sleep(RATE_LIMIT)

        if DEMO_MODE and len(rows) >= DEMO_COUNT:
            break

    print(f"  [OK] {cat_name}: {len(rows)} products")
    return rows


# ── PRODUCT PAGE ──────────────────────────────────────────────────────────────
def parse_product_page(url):
    """
    Returns dict with dimension/detail fields.
    Dims: all .pdp-product-dimensions__dimension blocks, skip --mobile to avoid duplicates.
      - --desktop blocks  -> Overall (standard cols)
      - plain blocks      -> More Dimensions: Seat/Arm/Back (standard cols) or dynamic
      - --last block      -> Weight
    Product Details: "Key: Value" bullets -> dynamic cols; rest -> Materials.
    """
    result = {}
    try:
        r = requests.get(url, headers=HTTP_HEADERS, timeout=30)
        if r.status_code == 429:
            print(f"    ! 429 on {url} — waiting 10s")
            time.sleep(10)
            r = requests.get(url, headers=HTTP_HEADERS, timeout=30)
        if r.status_code != 200:
            print(f"    ! HTTP {r.status_code} for {url}")
            return result

        soup = BeautifulSoup(r.text, "html.parser")

        # ── Dimensions ───────────────────────────────────────────────────────
        # Select all dimension blocks; skip --mobile ones to avoid duplicates
        all_dim_divs = soup.select("div.pdp-product-dimensions__dimension")
        seen_sections = set()

        for block in all_dim_divs:
            classes = block.get("class", [])
            # Skip mobile variants (they duplicate desktop/plain content)
            if "pdp-product-dimensions__dimension--mobile" in classes:
                continue

            title_el = block.select_one("div.pdp-product-dimensions__dimension-title")
            section  = title_el.get_text(strip=True) if title_el else "Overall"

            # Deduplicate sections (weight appears as both --desktop--last and plain)
            if section in seen_sections:
                continue
            seen_sections.add(section)

            script_el = block.select_one("script[type='application/json']")
            if not script_el:
                continue
            try:
                data = json.loads(script_el.string or "")
            except Exception:
                continue

            seen_abbrev_in_section = set()
            for entry in data.get("values", []):
                val  = entry.get("value")
                abbr = entry.get("abbreviation", "").lower()
                typ  = entry.get("type", "dimension")

                if val is None:
                    continue

                if typ == "weight":
                    if "Weight" not in result:
                        result["Weight"] = str(val)
                    continue

                col = _dim_col(section, abbr)
                if col:
                    if col not in result:
                        result[col] = str(val)
                    elif abbr == "d" and section.lower() == "overall" and "Diameter" not in result:
                        # second 'd' in Overall → Diameter
                        result["Diameter"] = str(val)
                else:
                    # Unknown section (Cord Length, Canopy, Leg, etc.) → dynamic
                    dyn_key = f"{section} {abbr.upper()}" if abbr else section
                    if dyn_key not in result.get("extra_dims", {}):
                        result.setdefault("extra_dims", {})[dyn_key] = str(val)

        # ── Product Details: parse key:value bullets as dynamic; rest as Materials ──
        details_el = soup.select_one("div.pdp-product-details__copy")
        if details_el:
            items = details_el.select("li")
            material_lines = []
            for li in items:
                text = li.get_text(strip=True)
                if not text:
                    continue
                # "Key: value" pattern → dynamic column
                kv = re.match(r'^([^:]{2,40}):\s*(.+)$', text)
                if kv:
                    key = kv.group(1).strip().title()
                    val = kv.group(2).strip()
                    result.setdefault("extra_dims", {})[key] = val
                else:
                    material_lines.append(text)
            if material_lines:
                result["Materials"] = " | ".join(material_lines)

    except Exception as e:
        print(f"    ! Page parse error {url}: {e}")

    return result


def enrich_products(products):
    """Fetch product page for each product and merge dimension/detail fields."""
    total = len(products)
    for i, p in enumerate(products, 1):
        url = p.get("Source", "")
        if not url:
            continue
        print(f"    Enriching {i}/{total}: {p.get('Product Name', ''[:40])}")
        details = parse_product_page(url)
        for k, v in details.items():
            if k == "extra_dims":
                p.setdefault("extra_dims", {}).update(v)
            else:
                p[k] = v
        time.sleep(RATE_LIMIT)
    return products


# ── EXCEL ─────────────────────────────────────────────────────────────────────
def write_excel(all_data):
    """Write/overwrite OUTPUT_FILE with all collected data."""
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
    else:
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

    # Collect all dynamic extra_cols across every category
    extra_cols = []
    for entry in all_data.values():
        for p in entry["rows"]:
            for k in p.get("extra_dims", {}):
                if k not in extra_cols:
                    extra_cols.append(k)

    all_cols = BASE_COLUMNS + extra_cols

    for cat_name, entry in all_data.items():
        rows    = entry["rows"]
        cat_url = entry["url"]

        if cat_name[:31] in wb.sheetnames:
            del wb[cat_name[:31]]

        ws = wb.create_sheet(title=cat_name[:31])

        # Row 1 — Brand
        ws.cell(1, 1, "Brand").font  = BRAND_FONT
        ws.cell(1, 2, MANUFACTURER).font = BRAND_FONT
        # Row 2 — Link
        ws.cell(2, 1, "Link")
        ws.cell(2, 2, cat_url)
        # Row 3 — empty
        # Row 4 — Headers
        for ci, col in enumerate(all_cols, 1):
            c = ws.cell(4, ci, col)
            c.fill      = HEADER_FILL
            c.font      = HEADER_FONT
            c.alignment = CENTER

        # Row 5+ — Data
        for ri, p in enumerate(rows, 5):
            ed = p.get("extra_dims", {})
            ws.cell(ri,  1, ri - 4)
            ws.cell(ri,  2, p.get("Category", ""))
            ws.cell(ri,  3, p.get("Manufacturer", ""))
            ws.cell(ri,  4, p.get("Source", ""))
            ws.cell(ri,  5, p.get("Image URL", ""))
            ws.cell(ri,  6, p.get("Product Name", ""))
            ws.cell(ri,  7, p.get("SKU", ""))
            ws.cell(ri,  8, p.get("Product Family Id", ""))
            ws.cell(ri,  9, p.get("Description", ""))
            ws.cell(ri, 10, p.get("Weight", ""))
            ws.cell(ri, 11, p.get("Width", ""))
            ws.cell(ri, 12, p.get("Depth", ""))
            ws.cell(ri, 13, p.get("Diameter", ""))
            ws.cell(ri, 14, p.get("Height", ""))
            ws.cell(ri, 15, p.get("Seat Width", ""))
            ws.cell(ri, 16, p.get("Seat Depth", ""))
            ws.cell(ri, 17, p.get("Seat Height", ""))
            ws.cell(ri, 18, p.get("Arm Width", ""))
            ws.cell(ri, 19, p.get("Arm Height", ""))
            ws.cell(ri, 20, p.get("Back Width", ""))
            ws.cell(ri, 21, p.get("Back Depth", ""))
            ws.cell(ri, 22, p.get("Back Height", ""))
            ws.cell(ri, 23, p.get("Price", ""))
            ws.cell(ri, 24, p.get("Finish", ""))
            ws.cell(ri, 25, p.get("Materials", ""))
            ws.cell(ri, 26, p.get("Tags", ""))
            ws.cell(ri, 27, p.get("Notes", ""))
            for ei, ec in enumerate(extra_cols, 28):
                ws.cell(ri, ei, ed.get(ec, ""))

        # Auto column width
        for col in ws.columns:
            max_w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 70)

    wb.save(OUTPUT_FILE)
    print(f"  [SAVED] -> {OUTPUT_FILE}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    mode = "DEMO (3/category)" if DEMO_MODE else "FULL RUN"
    print(f"Blue Dot Scraper  [{mode}]")
    print("=" * 60)

    all_data = {}
    total    = 0

    for cat in CATEGORIES:
        products = scrape_listing(cat)
        if not products:
            print(f"  ! No results for {cat['name']} — skipping")
            continue

        products = enrich_products(products)

        cat_url = f"{BASE_URL}/collections/{cat['handles'][0]}"
        all_data[cat["name"]] = {"rows": products, "url": cat_url}
        total += len(products)

        # Auto-save after every category
        write_excel(all_data)

    print(f"\n{'='*60}")
    print(f"Total: {total} products across {len(all_data)} categories")
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
