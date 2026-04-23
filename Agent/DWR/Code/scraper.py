"""
DWR (Design Within Reach) — Combined Scraper
Step 1 : Requests+BS4 via Search-UpdateGrid API  -> SKU, Name, Price, Image, URL
Step 2 : Product page fetch -> Description, Weight, Width, Depth, Diameter, Height
Output : Excel, one sheet per category
         Row 1: Brand | Row 2: URL | Row 3: empty | Row 4: Headers | Row 5+: Data
"""

import os
import re
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- CONFIG -----------------------------------------------------------
MANUFACTURER = "DWR"
BASE_URL     = "https://www.dwr.com"
API_URL      = "https://www.dwr.com/on/demandware.store/Sites-dwr-Site/en_US/Search-UpdateGrid"

DEMO_MODE  = False
DEMO_COUNT = 3
PAGE_SIZE  = 200
RATE_LIMIT = 1.0

OUTPUT_DIR  = "Demo" if DEMO_MODE else "Data"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "DWR_demo.xlsx" if DEMO_MODE else "DWR.xlsx")

HTTP_HEADERS = {
    "User-Agent"      : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer"         : "https://www.dwr.com/",
    "X-Requested-With": "XMLHttpRequest",
    "Accept-Language" : "en-US,en;q=0.9",
}

# --- CATEGORIES ------------------------------------------------------
CATEGORIES = [
    {"name": "Nightstands",             "links": ["https://www.dwr.com/bedroom-bedside-tables?lang=en_US"]},
    {"name": "Coffee & Cocktail Tables","links": ["https://www.dwr.com/living-accent-coffee-tables?lang=en_US"]},
    {"name": "Side & End Tables",       "links": ["https://www.dwr.com/living-drink-tables?lang=en_US",
                                                   "https://www.dwr.com/living-side-end-tables?lang=en_US"]},
    {"name": "Dining Tables",           "links": ["https://www.dwr.com/kitchen-dining-tables?lang=en_US"]},
    {"name": "Consoles",                "links": ["https://www.dwr.com/storage-credenzas-sideboards?lang=en_US",
                                                   "https://www.dwr.com/entryway-console-tables?lang=en_US",
                                                   "https://www.dwr.com/entryway-storage?lang=en_US"]},
    {"name": "Beds & Headboards",       "links": ["https://www.dwr.com/bedroom-beds?lang=en_US"]},
    {"name": "Desks",                   "links": ["https://www.dwr.com/office-desks?lang=en_US"]},
    {"name": "Bookcases",               "links": ["https://www.dwr.com/storage-shelving-bookcases?lang=en_US",
                                                   "https://www.dwr.com/office-shelving?lang=en_US"]},
    {"name": "Cabinets",                "links": ["https://www.dwr.com/storage-media?lang=en_US",
                                                   "https://www.dwr.com/office-credenzas-pedestal-storage?lang=en_US",
                                                   "https://www.dwr.com/entryway-storage?lang=en_US"]},
    {"name": "Accent Tables",           "links": ["https://www.dwr.com/bedroom-vanities?lang=en_US"]},
    {"name": "Dressers & Chests",       "links": ["https://www.dwr.com/bedroom-dressers-armoires?lang=en_US"]},
    {"name": "Bar Carts",               "links": ["https://www.dwr.com/kitchen-dining-bar-carts?lang=en_US"]},
    {"name": "Dining Chairs",           "links": ["https://www.dwr.com/kitchen-dining-chairs-benches?lang=en_US"]},
    {"name": "Bar Stools",              "links": ["https://www.dwr.com/kitchen-dining-bar-counter-stools?lang=en_US",
                                                   "https://www.dwr.com/living-benches-stools?lang=en_US"]},
    {"name": "Sofas & Loveseats",       "links": ["https://www.dwr.com/living-sofas?lang=en_US",
                                                   "https://www.dwr.com/living-sleepers?lang=en_US"]},
    {"name": "Sectionals",              "links": ["https://www.dwr.com/living-sectionals?lang=en_US"]},
    {"name": "Lounge Chairs",           "links": ["https://www.dwr.com/living-lounge-chairs?lang=en_US",
                                                   "https://www.dwr.com/living-side-chairs?lang=en_US"]},
    {"name": "Ottomans",                "links": ["https://www.dwr.com/living-ottomans?lang=en_US"]},
    {"name": "Benches",                 "links": ["https://www.dwr.com/living-benches-stools?lang=en_US"]},
    {"name": "Desk Chairs",             "links": ["https://www.dwr.com/office-chairs?lang=en_US"]},
    {"name": "Pendants",                "links": ["https://www.dwr.com/lighting-ceiling?lang=en_US"]},
    {"name": "Sconces",                 "links": ["https://www.dwr.com/lighting-wall-sconce?lang=en_US"]},
    {"name": "Table Lamps",             "links": ["https://www.dwr.com/lighting-table-lamps?lang=en_US",
                                                   "https://www.dwr.com/lighting-portable-lamps?lang=en_US"]},
    {"name": "Floor Lamps",             "links": ["https://www.dwr.com/lighting-floor?lang=en_US"]},
    {"name": "Mirrors",                 "links": ["https://www.dwr.com/accessories-living-room/constant/mirrors2?lang=en_US",
                                                   "https://www.dwr.com/accessories-mirrors?lang=en_US"]},
    {"name": "Pillows & Throws",        "links": ["https://www.dwr.com/accessories-living-room/constant/pillows-throws2?lang=en_US",
                                                   "https://www.dwr.com/accessories-living-room/constant/seat-pads-cushions-slipcovers?lang=en_US"]},
    {"name": "Vases",                   "links": ["https://www.dwr.com/accessories-living-room/constant/vases2?lang=en_US"]},
    {"name": "Objects",                 "links": ["https://www.dwr.com/accessories-living-room/constant/decorative-accessories?lang=en_US"]},
    {"name": "Baskets & Planters",      "links": ["https://www.dwr.com/accessories-living-room/constant/boxes-baskets-crates?lang=en_US"]},
    {"name": "Trays",                   "links": ["https://www.dwr.com/accessories-living-room/constant/trays-catchalls2?lang=en_US"]},
    {"name": "Wall Decor",              "links": ["https://www.dwr.com/accessories-living-room/constant/art2?lang=en_US"]},
    {"name": "Outdoor Seating",         "links": ["https://www.dwr.com/outdoor-stools?lang=en_US"]},
    {"name": "Outdoor Tables",          "links": ["https://www.dwr.com/outdoor-dining?lang=en_US",
                                                   "https://www.dwr.com/outdoor-dining-tables?lang=en_US"]},
    {"name": "Outdoor Storage",         "links": ["https://www.dwr.com/outdoor-storage?lang=en_US"]},
    {"name": "Outdoor Accessories",     "links": ["https://www.dwr.com/accessories-outdoor?lang=en_US",
                                                   "https://www.dwr.com/outdoor-lighting?lang=en_US"]},
]

# --- HELPERS ---------------------------------------------------------
def _family_id(name):
    """'Jens Chair, Leather' -> 'Jens Chair'  |  'PH5_Pendant' -> 'PH5'"""
    if not name:
        return ""
    return re.split(r'[,._\-]', name, maxsplit=1)[0].strip()


def _std_dim_key(label):
    """Map a dimension label to its standard result key, or None if it's a dynamic field."""
    ll = label.lower()
    if 'seat width'      in ll: return 'seat_width'
    if 'seat depth'      in ll: return 'seat_depth'
    if 'seat height'     in ll: return 'seat_height'
    if 'arm height'      in ll: return 'arm_height'
    if 'weight capacity' in ll: return 'weight_capacity'
    if ll == 'weight':          return 'weight'
    if 'width'           in ll: return 'width'
    if 'depth'           in ll: return 'depth'
    if 'height'          in ll: return 'height'
    if 'diameter'        in ll: return 'diameter'
    return None   # unknown field -> goes to dynamic column


# --- STEP 1: LISTING API ----------------------------------------------
def url_to_cgid(url):
    return url.replace(BASE_URL + "/", "").split("?")[0].rstrip("/")


def fetch_api(cgid, start=0, sz=PAGE_SIZE):
    params = {"cgid": cgid, "start": start, "sz": sz, "lang": "en_US"}
    try:
        r = requests.get(API_URL, params=params, headers=HTTP_HEADERS, timeout=30)
        if r.status_code == 200:
            return r.text
        print(f"    ! HTTP {r.status_code} for cgid={cgid}")
    except Exception as e:
        print(f"    ! Request error cgid={cgid}: {e}")
    return ""


def parse_listing(html, cat_name):
    products = []
    if not html:
        return products

    soup = BeautifulSoup(html, "html.parser")
    cards = soup.find_all("div", attrs={"data-pid": True})

    for card in cards:
        try:
            sku  = card.get("data-pid", "").strip()
            name = card.get("aria-label", "").strip()
            if name.lower() == "compare" or not name:
                name = ""
                tag = card.find(class_=lambda c: c and "product-name" in c)
                if tag:
                    name = tag.get_text(strip=True)

            a_tag = card.find("a", href=True)
            product_url = ""
            if a_tag:
                href = a_tag["href"].strip()
                product_url = href if href.startswith("http") else BASE_URL + href

            img = card.find("img", class_=lambda c: c and "tile-image" in c)
            image_url = ""
            if img:
                src = img.get("src") or img.get("data-src") or ""
                image_url = ("https:" + src) if src.startswith("//") else src

            price = ""
            sales = card.find("span", class_="sales")
            if sales:
                val = sales.find("span", class_="value")
                if val:
                    price = val.get("content") or val.get_text(strip=True)
            if not price:
                val = card.find("span", class_="value")
                if val:
                    price = val.get("content") or val.get_text(strip=True)

            if sku and product_url:
                products.append({
                    "Category"         : cat_name,
                    "Manufacturer"     : MANUFACTURER,
                    "Source"           : product_url,
                    "Image URL"        : image_url,
                    "Product Name"     : name,
                    "Product Family Id": _family_id(name),
                    "SKU"              : sku,
                    "Price"            : price,
                })
        except Exception as e:
            print(f"    ! Parse error: {e}")

    return products


def scrape_listing(cat):
    cat_name     = cat["name"]
    products     = []
    seen         = set()
    print(f"\n{'-'*60}")
    print(f"[Listing] {cat_name}  ({len(cat['links'])} link(s))")

    for url in cat["links"]:
        cgid  = url_to_cgid(url)
        start = 0
        print(f"  -> {cgid}")

        while True:
            sz   = DEMO_COUNT if DEMO_MODE else PAGE_SIZE
            html = fetch_api(cgid, start=start, sz=sz)
            page = parse_listing(html, cat_name)

            if not page:
                if start == 0:
                    print(f"    ! 0 results")
                break

            added = 0
            for p in page:
                key = p["SKU"] or p["Source"]
                if key not in seen:
                    seen.add(key)
                    products.append(p)
                    added += 1
                if DEMO_MODE and len(products) >= DEMO_COUNT:
                    break

            print(f"    start={start} -> {len(page)} cards, {added} new (total {len(products)})")

            if DEMO_MODE or len(page) < sz:
                break
            start += sz
            time.sleep(RATE_LIMIT)

        time.sleep(RATE_LIMIT)
        if DEMO_MODE and len(products) >= DEMO_COUNT:
            break

    return products


# --- STEP 2: PRODUCT PAGE DETAILS ------------------------------------

# Unicode fraction -> float
FRACTIONS = {
    '½': 0.5, '¼': 0.25, '¾': 0.75,
    '⅓': 1/3, '⅔': 2/3,
    '⅛': 0.125, '⅜': 0.375, '⅝': 0.625, '⅞': 0.875,
}


def _parse_dim_val(text):
    """Convert '22¾' -> '22.75', '15' -> '15'"""
    text = text.strip()
    for frac, val in FRACTIONS.items():
        if frac in text:
            num = re.sub(r'[^\d.]', '', text.replace(frac, ''))
            base = float(num) if num else 0.0
            result = base + val
            return str(int(result)) if result == int(result) else str(round(result, 3))
    m = re.search(r'[\d]+\.?[\d]*', text)
    return m.group() if m else text.strip()


def _find_section_body(soup, label):
    """Find collapse-body that follows a header/button containing label text."""
    pattern = re.compile(rf'^\s*{re.escape(label)}\s*$', re.I)
    for node in soup.find_all(string=pattern):
        parent = node.parent
        for _ in range(5):
            if parent is None:
                break
            nxt = parent.find_next_sibling()
            if nxt:
                classes = ' '.join(nxt.get('class') or [])
                if 'collapse' in classes or 'pdp-summary' in classes:
                    return nxt
            parent = parent.parent
    return None


def fetch_product_page(url):
    try:
        r = requests.get(url, headers=HTTP_HEADERS, timeout=30)
        if r.status_code == 200:
            return r.text
    except Exception as e:
        print(f"    ! Page fetch error: {e}")
    return ""


def parse_product_details(html):
    result = {
        "description": "", "assembly": "", "warranty": "",
        "weight": "", "weight_capacity": "",
        "width": "", "depth": "", "diameter": "",
        "height": "", "seat_width": "", "seat_depth": "",
        "seat_height": "", "arm_height": "",
        "materials": "", "finish": "",
        "extra_dims": {},
    }
    if not html:
        return result

    soup = BeautifulSoup(html, "html.parser")

    # ── Description: Summary section (.pdp-summary-group-col-1) ──────
    desc_parts = []
    summary_col = soup.select_one('.pdp-summary-group-col-1')
    if summary_col:
        for div in summary_col.select('div.mb-3'):
            t = div.get_text(' ', strip=True)
            if t:
                desc_parts.append(t)
        for li in summary_col.select('ul.bulleted li'):
            t = li.get_text(' ', strip=True)
            if t:
                desc_parts.append(t)
    result["description"] = ' | '.join(desc_parts)[:2000]

    # ── Summary right-side spec table (Brand, Collection, Assembly, Warranty, etc.)
    # Covers both <table> rows and <dl>/<dt>/<dd> and generic row divs
    summary_spec_pairs = []

    # Try col-2 first, then full summary collapse-body
    for container_sel in ['.pdp-summary-group-col-2', '.pdp-summary-group']:
        container = soup.select_one(container_sel)
        if container:
            # <table> rows
            for row in container.find_all('tr'):
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    summary_spec_pairs.append((
                        cells[0].get_text(strip=True),
                        cells[1].get_text(' ', strip=True),
                    ))
            # <dl> / <dt><dd>
            for dl in container.find_all('dl'):
                for dt, dd in zip(dl.find_all('dt'), dl.find_all('dd')):
                    summary_spec_pairs.append((
                        dt.get_text(strip=True),
                        dd.get_text(' ', strip=True),
                    ))
            if summary_spec_pairs:
                break

    # If still empty, scan entire page for a spec table adjacent to summary text
    if not summary_spec_pairs:
        for table in soup.select('table'):
            rows = table.find_all('tr')
            if rows and any(c.find('td') or c.find('th') for c in rows[:3]):
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) >= 2:
                        summary_spec_pairs.append((
                            cells[0].get_text(strip=True),
                            cells[1].get_text(' ', strip=True),
                        ))

    for label, value in summary_spec_pairs:
        ll = label.lower()
        val = value.strip()
        if not label or not val:
            continue
        if 'assembly' in ll:
            if not result["assembly"]:
                result["assembly"] = val
        elif 'warranty' in ll:
            if not result["warranty"]:
                result["warranty"] = val
        else:
            # Everything else (Brand, Collection, General Dimensions, etc.) → dynamic
            if label not in result["extra_dims"]:
                result["extra_dims"][label] = val

    # ── Dimensions: each variant has .dimension-title + .dimension-body
    dim_groups = []
    for title_el in soup.select('.dimension-title'):
        body_el = title_el.find_next('ul', class_='dimension-body')
        if not body_el:
            continue
        variant = title_el.get_text(strip=True)
        dims = {}
        for li in body_el.select('li'):
            text = li.get_text(strip=True)
            if ':' in text:
                lbl, val = text.rsplit(':', 1)
                lbl_clean = re.sub(r'\s*\(.*?\)', '', lbl).strip()
                dims[lbl_clean] = _parse_dim_val(val)
        if dims:
            dim_groups.append((variant, dims))

    if dim_groups:
        # First variant: map each field to a standard column or dynamic column
        for lbl, val in dim_groups[0][1].items():
            std = _std_dim_key(lbl)
            if std:
                result[std] = val
            else:
                result["extra_dims"][lbl] = val   # e.g. Shade Diameter, Cord Length, Wattage

        # Extra size variants -> dynamic column named after the variant
        for variant, dims in dim_groups[1:]:
            result["extra_dims"][variant] = ' | '.join(f"{k}: {v}" for k, v in dims.items())

    # ── Materials: .pdp-summary-materials-body ul.material-list ──────
    mat_el = soup.select_one('.pdp-summary-materials-body ul.material-list')
    if not mat_el:
        mat_el = _find_section_body(soup, 'Materials')
    if mat_el:
        items = [s.get_text(strip=True) for s in mat_el.select('li span') or mat_el.select('li')]
        result["materials"] = ', '.join(i for i in items if i)

    # ── Finish: try spec tables as fallback ───────────────────────────
    for table in soup.find_all('table'):
        for row in table.find_all('tr'):
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 2:
                lbl = cells[0].get_text(strip=True)
                val = cells[1].get_text(strip=True)
                if re.search(r'finish|color', lbl, re.I) and not result["finish"]:
                    result["finish"] = val

    return result


def enrich_products(products):
    total = len(products)
    print(f"\n[Details] Fetching product pages for {total} products...")
    for i, p in enumerate(products, 1):
        print(f"  [{i}/{total}] {p['SKU']}  {p['Source'][:70]}")
        html    = fetch_product_page(p["Source"])
        details = parse_product_details(html)
        p.update({
            "Description"    : details["description"],
            "Assembly"       : details["assembly"],
            "Warranty"       : details["warranty"],
            "Weight"         : details["weight"],
            "Weight Capacity": details["weight_capacity"],
            "Width"          : details["width"],
            "Depth"      : details["depth"],
            "Diameter"   : details["diameter"],
            "Height"     : details["height"],
            "Seat Width" : details["seat_width"],
            "Seat Depth" : details["seat_depth"],
            "Seat Height": details["seat_height"],
            "Arm Height" : details["arm_height"],
            "Finish"     : details["finish"],
            "Materials"  : details["materials"],
            "extra_dims" : details["extra_dims"],
        })
        time.sleep(RATE_LIMIT)
    return products


# --- EXCEL WRITER -----------------------------------------------------
BASE_COLUMNS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Product Family Id", "Description",
    "Weight", "Weight Capacity", "Width", "Depth", "Diameter", "Height",
    "Seat Width", "Seat Depth", "Seat Height", "Arm Height",
    "Price", "Finish", "Special Order", "Location", "Materials",
    "Assembly", "Warranty", "Tags", "Notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BRAND_FONT  = Font(bold=True, size=12)
CENTER      = Alignment(horizontal="center")


def write_excel(all_data):
    # all_data: {cat_name: {"rows": [...], "url": "..."}}
    extra_cols = []
    for entry in all_data.values():
        for p in entry["rows"]:
            for k in p.get("extra_dims", {}):
                if k not in extra_cols:
                    extra_cols.append(k)

    all_cols = BASE_COLUMNS + extra_cols

    wb = Workbook()
    wb.remove(wb.active)

    for cat_name, entry in all_data.items():
        rows     = entry["rows"]
        cat_url  = entry["url"]
        ws = wb.create_sheet(title=cat_name[:31])

        # Row 1: Brand label | Brand value
        ws.cell(1, 1, "Brand").font = BRAND_FONT
        ws.cell(1, 2, MANUFACTURER).font = BRAND_FONT

        # Row 2: Link label | Category URL
        ws.cell(2, 1, "Link")
        ws.cell(2, 2, cat_url)

        for ci, col in enumerate(all_cols, 1):
            c = ws.cell(4, ci, col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER

        for ri, p in enumerate(entry["rows"], 5):
            ed = p.get("extra_dims", {})
            ws.cell(ri,  1, ri - 4)
            ws.cell(ri,  2, p.get("Category", ""))
            ws.cell(ri,  3, p.get("Manufacturer", ""))
            ws.cell(ri,  4, p.get("Source", ""))
            ws.cell(ri,  5, p.get("Image URL", ""))
            ws.cell(ri,  6, p.get("Product Name", ""))
            ws.cell(ri,  7, p.get("SKU", ""))
            ws.cell(ri,  8, p.get("Product Family Id", ""))
            ws.cell(ri,  9, p.get("Description", ""))
            ws.cell(ri, 10, p.get("Weight", ""))
            ws.cell(ri, 11, p.get("Weight Capacity", ""))
            ws.cell(ri, 12, p.get("Width", ""))
            ws.cell(ri, 13, p.get("Depth", ""))
            ws.cell(ri, 14, p.get("Diameter", ""))
            ws.cell(ri, 15, p.get("Height", ""))
            ws.cell(ri, 16, p.get("Seat Width", ""))
            ws.cell(ri, 17, p.get("Seat Depth", ""))
            ws.cell(ri, 18, p.get("Seat Height", ""))
            ws.cell(ri, 19, p.get("Arm Height", ""))
            ws.cell(ri, 20, p.get("Price", ""))
            ws.cell(ri, 21, p.get("Finish", ""))
            ws.cell(ri, 22, p.get("Special Order", ""))
            ws.cell(ri, 23, p.get("Location", ""))
            ws.cell(ri, 24, p.get("Materials", ""))
            ws.cell(ri, 25, p.get("Assembly", ""))
            ws.cell(ri, 26, p.get("Warranty", ""))
            ws.cell(ri, 27, p.get("Tags", ""))
            ws.cell(ri, 28, p.get("Notes", ""))
            for ei, ec in enumerate(extra_cols, 29):
                ws.cell(ri, ei, ed.get(ec, ""))

        for col in ws.columns:
            max_w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 70)


    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Saved -> {OUTPUT_FILE}")


# --- MAIN -------------------------------------------------------------
def main():
    mode = "DEMO (3/category)" if DEMO_MODE else "FULL RUN"
    print(f"DWR Scraper  [{mode}]")
    print("=" * 60)

    all_data = {}
    total    = 0

    for cat in CATEGORIES:
        # Step 1 — listing
        products = scrape_listing(cat)
        if not products:
            continue

        # Step 2 — product page details
        products = enrich_products(products)

        all_data[cat["name"]] = {
            "rows": products,
            "url" : cat["links"][0],
        }
        total += len(products)
        print(f"  [OK] {cat['name']}: {len(products)} products (with details)")

        # Auto-save after every category
        write_excel(all_data)
        print(f"  [SAVED] {len(all_data)} categories saved so far ({total} products)")

    print(f"\n{'='*60}")
    print(f"Total: {total} products across {len(all_data)} categories")
    if all_data:
        print(f"[DONE] Final file -> {OUTPUT_FILE}")
    else:
        print("! No data collected.")


if __name__ == "__main__":
    main()
