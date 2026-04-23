"""
DWR — Re-run last 4 outdoor categories that got HTTP 429
Appends new sheets to existing Data/DWR.xlsx
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from scraper import (
    scrape_listing, enrich_products,
    BASE_COLUMNS, HEADER_FILL, HEADER_FONT, BRAND_FONT, CENTER,
    MANUFACTURER, BASE_URL,
)
import time
import scraper as _s

# ── Override rate limit to avoid 429 ──────────────────────────────
_s.RATE_LIMIT  = 5.0
_s.DEMO_MODE   = False
_s.PAGE_SIZE   = 200

RETRY_CATEGORIES = [
    {"name": "Pillows & Throws",    "links": ["https://www.dwr.com/accessories-living-room/constant/pillows-throws2?lang=en_US",
                                               "https://www.dwr.com/accessories-living-room/constant/seat-pads-cushions-slipcovers?lang=en_US"]},
    {"name": "Vases",               "links": ["https://www.dwr.com/accessories-living-room/constant/vases2?lang=en_US"]},
    {"name": "Objects",             "links": ["https://www.dwr.com/accessories-living-room/constant/decorative-accessories?lang=en_US"]},
    {"name": "Baskets & Planters",  "links": ["https://www.dwr.com/accessories-living-room/constant/boxes-baskets-crates?lang=en_US"]},
    {"name": "Trays",               "links": ["https://www.dwr.com/accessories-living-room/constant/trays-catchalls2?lang=en_US"]},
    {"name": "Wall Decor",          "links": ["https://www.dwr.com/accessories-living-room/constant/art2?lang=en_US"]},
    {"name": "Outdoor Seating",     "links": ["https://www.dwr.com/outdoor-stools?lang=en_US"]},
    {"name": "Outdoor Tables",      "links": ["https://www.dwr.com/outdoor-dining?lang=en_US",
                                               "https://www.dwr.com/outdoor-dining-tables?lang=en_US"]},
    {"name": "Outdoor Storage",     "links": ["https://www.dwr.com/outdoor-storage?lang=en_US"]},
    {"name": "Outdoor Accessories", "links": ["https://www.dwr.com/accessories-outdoor?lang=en_US",
                                               "https://www.dwr.com/outdoor-lighting?lang=en_US"]},
]

OUTPUT_FILE = os.path.join("Data", "DWR.xlsx")

from openpyxl import load_workbook

def append_to_excel(all_data):
    wb = load_workbook(OUTPUT_FILE)

    # Collect extra_cols already present + new ones
    extra_cols = []
    for entry in all_data.values():
        for p in entry["rows"]:
            for k in p.get("extra_dims", {}):
                if k not in extra_cols:
                    extra_cols.append(k)

    all_cols = BASE_COLUMNS + extra_cols

    for cat_name, entry in all_data.items():
        rows    = entry["rows"]
        cat_url = entry["url"]

        # Remove sheet if already exists (replace)
        if cat_name[:31] in wb.sheetnames:
            del wb[cat_name[:31]]

        ws = wb.create_sheet(title=cat_name[:31])

        ws.cell(1, 1, "Brand").font = BRAND_FONT
        ws.cell(1, 2, MANUFACTURER).font = BRAND_FONT
        ws.cell(2, 1, "Link")
        ws.cell(2, 2, cat_url)

        for ci, col in enumerate(all_cols, 1):
            c = ws.cell(4, ci, col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER

        for ri, p in enumerate(rows, 5):
            ed = p.get("extra_dims", {})
            ws.cell(ri,  1, ri - 4)
            ws.cell(ri,  2, p.get("Category", ""))
            ws.cell(ri,  3, p.get("Manufacturer", ""))
            ws.cell(ri,  4, p.get("Source", ""))
            ws.cell(ri,  5, p.get("Image URL", ""))
            ws.cell(ri,  6, p.get("Product Name", ""))
            ws.cell(ri,  7, p.get("SKU", ""))
            ws.cell(ri,  8, p.get("Product Family Id", ""))
            ws.cell(ri,  9, p.get("Description", ""))
            ws.cell(ri, 10, p.get("Weight", ""))
            ws.cell(ri, 11, p.get("Weight Capacity", ""))
            ws.cell(ri, 12, p.get("Width", ""))
            ws.cell(ri, 13, p.get("Depth", ""))
            ws.cell(ri, 14, p.get("Diameter", ""))
            ws.cell(ri, 15, p.get("Height", ""))
            ws.cell(ri, 16, p.get("Seat Width", ""))
            ws.cell(ri, 17, p.get("Seat Depth", ""))
            ws.cell(ri, 18, p.get("Seat Height", ""))
            ws.cell(ri, 19, p.get("Arm Height", ""))
            ws.cell(ri, 20, p.get("Price", ""))
            ws.cell(ri, 21, p.get("Finish", ""))
            ws.cell(ri, 22, p.get("Special Order", ""))
            ws.cell(ri, 23, p.get("Location", ""))
            ws.cell(ri, 24, p.get("Materials", ""))
            ws.cell(ri, 25, p.get("Assembly", ""))
            ws.cell(ri, 26, p.get("Warranty", ""))
            ws.cell(ri, 27, p.get("Tags", ""))
            ws.cell(ri, 28, p.get("Notes", ""))
            for ei, ec in enumerate(extra_cols, 29):
                ws.cell(ri, ei, ed.get(ec, ""))

        for col in ws.columns:
            max_w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 70)

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Appended -> {OUTPUT_FILE}")


def main():
    print("DWR Re-run — 4 Outdoor Categories  [rate_limit=3.0s]")
    print("=" * 60)

    all_data = {}
    total    = 0

    for cat in RETRY_CATEGORIES:
        print(f"\nWaiting 15s before starting {cat['name']}...")
        time.sleep(15)

        products = scrape_listing(cat)
        if not products:
            print(f"  ! Still no results for {cat['name']} — skipping")
            continue

        products = enrich_products(products)
        all_data[cat["name"]] = {"rows": products, "url": cat["links"][0]}
        total += len(products)
        print(f"  [OK] {cat['name']}: {len(products)} products")

    print(f"\n{'='*60}")
    print(f"Total new: {total} products across {len(all_data)} categories")

    if all_data:
        append_to_excel(all_data)
    else:
        print("! No data collected.")


if __name__ == "__main__":
    main()
