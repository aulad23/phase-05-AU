"""
DWR (Design Within Reach) — Step 1 Scraper
Method : Requests + BS4  via  Search-UpdateGrid API  (no Selenium)
Output : Excel, one sheet per category
         Row 1: Brand | Row 2: URL | Row 3: empty | Row 4: Headers | Row 5+: Data
"""

import os
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─── CONFIG ───────────────────────────────────────────────────────────
MANUFACTURER = "DWR"
BASE_URL     = "https://www.dwr.com"
API_URL      = "https://www.dwr.com/on/demandware.store/Sites-dwr-Site/en_US/Search-UpdateGrid"

DEMO_MODE  = True
DEMO_COUNT = 3
PAGE_SIZE  = 200
RATE_LIMIT = 1.0   # seconds between requests

OUTPUT_DIR  = "Demo" if DEMO_MODE else "Data"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "DWR_step1_demo.xlsx" if DEMO_MODE else "DWR_step1.xlsx")

HTTP_HEADERS = {
    "User-Agent"      : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer"         : "https://www.dwr.com/",
    "X-Requested-With": "XMLHttpRequest",
    "Accept-Language" : "en-US,en;q=0.9",
}

# ─── CATEGORIES (from Vendor List DWR sheet) ──────────────────────────
CATEGORIES = [
    {"name": "Nightstands",            "links": ["https://www.dwr.com/bedroom-bedside-tables?lang=en_US"]},
    {"name": "Coffee & Cocktail Tables","links": ["https://www.dwr.com/living-accent-coffee-tables?lang=en_US"]},
    {"name": "Side & End Tables",       "links": ["https://www.dwr.com/living-drink-tables?lang=en_US",
                                                   "https://www.dwr.com/living-side-end-tables?lang=en_US"]},
    {"name": "Dining Tables",           "links": ["https://www.dwr.com/kitchen-dining-tables?lang=en_US"]},
    {"name": "Consoles",                "links": ["https://www.dwr.com/storage-credenzas-sideboards?lang=en_US",
                                                   "https://www.dwr.com/entryway-console-tables?lang=en_US",
                                                   "https://www.dwr.com/entryway-storage?lang=en_US"]},
    {"name": "Beds & Headboards",       "links": ["https://www.dwr.com/bedroom-beds?lang=en_US"]},
    {"name": "Desks",                   "links": ["https://www.dwr.com/office-desks?lang=en_US"]},
    {"name": "Bookcases",               "links": ["https://www.dwr.com/storage-shelving-bookcases?lang=en_US",
                                                   "https://www.dwr.com/office-shelving?lang=en_US"]},
    {"name": "Cabinets",                "links": ["https://www.dwr.com/storage-media?lang=en_US",
                                                   "https://www.dwr.com/office-credenzas-pedestal-storage?lang=en_US",
                                                   "https://www.dwr.com/entryway-storage?lang=en_US"]},
    {"name": "Accent Tables",           "links": ["https://www.dwr.com/bedroom-vanities?lang=en_US"]},
    {"name": "Dressers & Chests",       "links": ["https://www.dwr.com/bedroom-dressers-armoires?lang=en_US"]},
    {"name": "Bar Carts",               "links": ["https://www.dwr.com/kitchen-dining-bar-carts?lang=en_US"]},
    {"name": "Dining Chairs",           "links": ["https://www.dwr.com/kitchen-dining-chairs-benches?lang=en_US"]},
    {"name": "Bar Stools",              "links": ["https://www.dwr.com/kitchen-dining-bar-counter-stools?lang=en_US",
                                                   "https://www.dwr.com/living-benches-stools?lang=en_US"]},
    {"name": "Sofas & Loveseats",       "links": ["https://www.dwr.com/living-sofas?lang=en_US",
                                                   "https://www.dwr.com/living-sleepers?lang=en_US"]},
    {"name": "Sectionals",             "links": ["https://www.dwr.com/living-sectionals?lang=en_US"]},
    {"name": "Lounge Chairs",           "links": ["https://www.dwr.com/living-lounge-chairs?lang=en_US",
                                                   "https://www.dwr.com/living-side-chairs?lang=en_US"]},
    {"name": "Ottomans",               "links": ["https://www.dwr.com/living-ottomans?lang=en_US"]},
    {"name": "Benches",                "links": ["https://www.dwr.com/living-benches-stools?lang=en_US"]},
    {"name": "Desk Chairs",            "links": ["https://www.dwr.com/office-chairs?lang=en_US"]},
    {"name": "Pendants",               "links": ["https://www.dwr.com/lighting-ceiling?lang=en_US"]},
    {"name": "Sconces",                "links": ["https://www.dwr.com/lighting-wall-sconce?lang=en_US"]},
    {"name": "Table Lamps",            "links": ["https://www.dwr.com/lighting-table-lamps?lang=en_US",
                                                   "https://www.dwr.com/lighting-portable-lamps?lang=en_US"]},
    {"name": "Floor Lamps",            "links": ["https://www.dwr.com/lighting-floor?lang=en_US"]},
    {"name": "Mirrors",                "links": ["https://www.dwr.com/accessories-living-room/constant/mirrors2?lang=en_US",
                                                   "https://www.dwr.com/accessories-mirrors?lang=en_US"]},
    {"name": "Pillows & Throws",        "links": ["https://www.dwr.com/accessories-living-room/constant/pillows-throws2?lang=en_US",
                                                   "https://www.dwr.com/accessories-living-room/constant/seat-pads-cushions-slipcovers?lang=en_US"]},
    {"name": "Vases",                  "links": ["https://www.dwr.com/accessories-living-room/constant/vases2?lang=en_US"]},
    {"name": "Objects",                "links": ["https://www.dwr.com/accessories-living-room/constant/decorative-accessories?lang=en_US"]},
    {"name": "Baskets & Planters",     "links": ["https://www.dwr.com/accessories-living-room/constant/boxes-baskets-crates?lang=en_US"]},
    {"name": "Trays",                  "links": ["https://www.dwr.com/accessories-living-room/constant/trays-catchalls2?lang=en_US"]},
    {"name": "Wall Decor",             "links": ["https://www.dwr.com/accessories-living-room/constant/art2?lang=en_US"]},
    {"name": "Outdoor Seating",        "links": ["https://www.dwr.com/outdoor-stools?lang=en_US"]},
    {"name": "Outdoor Tables",         "links": ["https://www.dwr.com/outdoor-dining?lang=en_US",
                                                   "https://www.dwr.com/outdoor-dining-tables?lang=en_US"]},
    {"name": "Outdoor Storage",        "links": ["https://www.dwr.com/outdoor-storage?lang=en_US"]},
    {"name": "Outdoor Accessories",    "links": ["https://www.dwr.com/accessories-outdoor?lang=en_US",
                                                   "https://www.dwr.com/outdoor-lighting?lang=en_US"]},
]

# ─── HELPERS ─────────────────────────────────────────────────────────
def url_to_cgid(url):
    return url.replace(BASE_URL + "/", "").split("?")[0].rstrip("/")


def fetch_api(cgid, start=0, sz=PAGE_SIZE):
    params = {"cgid": cgid, "start": start, "sz": sz, "lang": "en_US"}
    try:
        r = requests.get(API_URL, params=params, headers=HTTP_HEADERS, timeout=30)
        if r.status_code == 200:
            return r.text
        print(f"    ⚠ HTTP {r.status_code} for cgid={cgid}")
    except Exception as e:
        print(f"    ⚠ Request error cgid={cgid}: {e}")
    return ""


def parse_products(html, cat_name):
    products = []
    if not html:
        return products

    soup = BeautifulSoup(html, "html.parser")
    cards = soup.find_all("div", attrs={"data-pid": True})

    for card in cards:
        try:
            sku  = card.get("data-pid", "").strip()
            name = card.get("aria-label", "").strip()

            # Remove "Compare" noise sometimes appearing in aria-label
            if name.lower() == "compare" or not name:
                name = ""
                name_tag = card.find(class_=lambda c: c and "product-name" in c)
                if name_tag:
                    name = name_tag.get_text(strip=True)

            # Product URL
            a_tag = card.find("a", href=True)
            product_url = ""
            if a_tag:
                href = a_tag["href"].strip()
                product_url = href if href.startswith("http") else BASE_URL + href

            # Image
            img = card.find("img", class_=lambda c: c and "tile-image" in c)
            image_url = ""
            if img:
                src = img.get("src") or img.get("data-src") or ""
                image_url = ("https:" + src) if src.startswith("//") else src

            # Price
            price = ""
            sales = card.find("span", class_="sales")
            if sales:
                val = sales.find("span", class_="value")
                if val:
                    price = val.get("content") or val.get_text(strip=True)
            if not price:
                val = card.find("span", class_="value")
                if val:
                    price = val.get("content") or val.get_text(strip=True)

            if sku and product_url:
                products.append({
                    "Category"    : cat_name,
                    "Manufacturer": MANUFACTURER,
                    "Source"      : product_url,
                    "Image URL"   : image_url,
                    "Product Name": name,
                    "SKU"         : sku,
                    "Price"       : price,
                })
        except Exception as e:
            print(f"    ⚠ Parse error: {e}")

    return products


# ─── SCRAPE ONE CATEGORY ─────────────────────────────────────────────
def scrape_category(cat):
    cat_name = cat["name"]
    links    = cat["links"]
    products = []
    seen     = set()

    print(f"\n{'-'*60}")
    print(f"Category: {cat_name}  ({len(links)} link(s))")

    for url in links:
        cgid = url_to_cgid(url)
        print(f"  → {cgid}")
        start = 0

        while True:
            sz   = DEMO_COUNT if DEMO_MODE else PAGE_SIZE
            html = fetch_api(cgid, start=start, sz=sz)
            page_products = parse_products(html, cat_name)

            if not page_products:
                if start == 0:
                    print(f"    ⚠ 0 results for cgid={cgid}")
                break

            added = 0
            for p in page_products:
                key = p["SKU"] or p["Source"]
                if key not in seen:
                    seen.add(key)
                    products.append(p)
                    added += 1
                if DEMO_MODE and len(products) >= DEMO_COUNT:
                    break

            print(f"    start={start} → {len(page_products)} cards, {added} new (total: {len(products)})")

            if DEMO_MODE or len(page_products) < sz:
                break
            start += sz
            time.sleep(RATE_LIMIT)

        time.sleep(RATE_LIMIT)
        if DEMO_MODE and len(products) >= DEMO_COUNT:
            break

    print(f"  ✅ {cat_name}: {len(products)} products")
    return products


# ─── EXCEL WRITER ─────────────────────────────────────────────────────
COLUMNS = ["Index", "Category", "Manufacturer", "Source", "Image URL", "Product Name", "SKU", "Price"]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BRAND_FONT  = Font(bold=True, size=12)
CENTER      = Alignment(horizontal="center")


def write_excel(all_data):
    wb = Workbook()
    wb.remove(wb.active)

    for cat_name, rows in all_data.items():
        ws = wb.create_sheet(title=cat_name[:31])

        # Row 1 — brand
        cell = ws.cell(1, 1, MANUFACTURER)
        cell.font = BRAND_FONT
        # Row 2 — base URL
        ws.cell(2, 1, BASE_URL)
        # Row 3 — empty
        # Row 4 — headers
        for ci, col in enumerate(COLUMNS, 1):
            c = ws.cell(4, ci, col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER

        # Row 5+ — data
        for ri, p in enumerate(rows, 5):
            ws.cell(ri, 1, ri - 4)
            ws.cell(ri, 2, p.get("Category", ""))
            ws.cell(ri, 3, p.get("Manufacturer", ""))
            ws.cell(ri, 4, p.get("Source", ""))
            ws.cell(ri, 5, p.get("Image URL", ""))
            ws.cell(ri, 6, p.get("Product Name", ""))
            ws.cell(ri, 7, p.get("SKU", ""))
            ws.cell(ri, 8, p.get("Price", ""))

        # Auto column width
        for col in ws.columns:
            max_w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 70)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved → {OUTPUT_FILE}")


# ─── MAIN ─────────────────────────────────────────────────────────────
def main():
    mode = "DEMO (3/category)" if DEMO_MODE else "FULL RUN"
    print(f"DWR Step 1 Scraper  [{mode}]")
    print("=" * 60)

    all_data = {}
    total    = 0

    for cat in CATEGORIES:
        rows = scrape_category(cat)
        if rows:
            all_data[cat["name"]] = rows
            total += len(rows)

    print(f"\n{'='*60}")
    print(f"Total: {total} products across {len(all_data)} categories")

    if all_data:
        write_excel(all_data)
    else:
        print("⚠ No data collected.")


if __name__ == "__main__":
    main()
