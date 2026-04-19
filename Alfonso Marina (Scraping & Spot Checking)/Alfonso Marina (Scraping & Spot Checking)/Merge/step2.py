import os
import re
import time
from fractions import Fraction
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from bs4 import BeautifulSoup

# ================= CONFIG =================
MASTER_FILE = "AlfonsoMarina.xlsx"
SAVE_AS = "AlfonsoMarina_Details.xlsx"

HEADLESS = True
TIMEOUT_PAGELOAD = 30
WAIT_BODY_SEC = 12
PER_PAGE_SLEEP = 1.5

HEADER_ROW = 4
DATA_START_ROW = HEADER_ROW + 1

COLUMN_ORDER = [
    "Index", "Category", "Product URL", "Image URL", "Product Name",
    "SKU", "Product Family Id", "Description", "Weight",
    "Width", "Depth", "Diameter", "Height", "Finish"
]
# =========================================

# ---------- Utilities ----------
def convert_fraction_to_decimal(value):
    if not value:
        return ""
    try:
        value = value.replace('"', '').strip()
        if re.match(r'^\d+[-\s]\d+/\d+$', value):
            whole, frac = re.split('[-\s]', value)
            return float(whole) + float(Fraction(frac))
        elif re.match(r'^\d+/\d+$', value):
            return float(Fraction(value))
        elif re.match(r'^\d+(\.\d+)?$', value):
            return float(value)
    except Exception:
        return value
    return value

def setup_driver():
    chromedriver_autoinstaller.install()
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(TIMEOUT_PAGELOAD)
    return driver

# ---------- Excel Helpers ----------
def get_header_map(ws):
    return {str(ws.cell(row=HEADER_ROW, column=c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=HEADER_ROW, column=c).value}

def write_new_header(ws):
    for i, col in enumerate(COLUMN_ORDER, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=col)
    return {col: i for i, col in enumerate(COLUMN_ORDER, start=1)}

def get_last_data_row(ws, col_url):
    r = ws.max_row
    while r >= DATA_START_ROW and not ws.cell(row=r, column=col_url).value:
        r -= 1
    return r

# ---------- Scraper ----------
def scrape_product_details(driver, url):
    out = {k: "" for k in ["SKU", "Description", "Weight", "Width", "Depth", "Diameter", "Height", "Finish"]}
    if not url:
        return out

    try:
        driver.get(url)
        WebDriverWait(driver, WAIT_BODY_SEC).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(PER_PAGE_SLEEP)
    except Exception as e:
        print(f"Error loading URL: {url} -> {e}")
        return out

    soup = BeautifulSoup(driver.page_source, "html.parser")

    # --- SKU ---
    sku_tag = soup.find(string=re.compile("PRODUCT CODE", re.I))
    if sku_tag:
        match = re.search(r"PRODUCT\s*CODE\s*[:\-]\s*(.+)", sku_tag, re.I)
        if match:
            out["SKU"] = match.group(1).strip()

    # --- Description ---
    desc_header = soup.find("h2", class_="elementor-heading-title elementor-size-default",
                            string=re.compile("DETAILS", re.I))
    if desc_header:
        section = desc_header.find_next("section")
        if section:
            for tag in section.find_all(["a", "button", "script", "style"]):
                tag.decompose()
            out["Description"] = section.get_text(separator="\n", strip=True)

    # --- Finish ---
    finish_divs = soup.find_all("div", class_="elementor-widget-text-editor")
    for div in finish_divs:
        text = div.get_text(strip=True)
        if "As Shown:" in text:
            out["Finish"] = text.split("As Shown:")[-1].strip()
            break

    # --- Dimensions / Weight ---
    for tag in soup.find_all("div", class_="elementor-widget-container"):
        text = tag.get_text(strip=True)
        for key, prefix in [("Width", "W:"), ("Depth", "D:"), ("Height", "H:"), ("Diameter", "Dia:"), ("Weight", "Weight:")]:
            if text.startswith(prefix) or text.lower().startswith(prefix.lower()):
                out[key] = convert_fraction_to_decimal(text.replace(prefix, "").strip())

    return out

# ---------- Main ----------
def main():
    wb = load_workbook(MASTER_FILE)
    driver = setup_driver()
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_map = write_new_header(ws)
            col_url = header_map["Product URL"]
            last_row = get_last_data_row(ws, col_url)
            if last_row < DATA_START_ROW:
                continue

            for r in range(DATA_START_ROW, last_row + 1):
                url = ws.cell(row=r, column=col_url).value
                if not url or not str(url).startswith("http"):
                    continue

                print(f"Scraping Row {r} in sheet '{sheet_name}': {url}")
                data = scrape_product_details(driver, url)

                ws.cell(row=r, column=header_map["Product Family Id"],
                        value=ws.cell(row=r, column=header_map["Product Name"]).value)

                for k, v in data.items():
                    if k in header_map:
                        ws.cell(row=r, column=header_map[k], value=v)

            wb.save(SAVE_AS)
            print(f"Sheet '{sheet_name}' saved -> {SAVE_AS}")

    finally:
        driver.quit()
        wb.save(SAVE_AS)
    print("ALL DONE")

if __name__ == "__main__":
    main()
