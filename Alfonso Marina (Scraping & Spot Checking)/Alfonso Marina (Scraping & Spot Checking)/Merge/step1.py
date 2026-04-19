import os
import time
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ==== Categories (INPUT SYSTEM) ====
CATEGORIES = {
    "Nightstands": [
        "https://alfonsomarina.com/product-category/furniture-eng/storage-eng/nightstands-lowboys-side-chests-eng/"
    ],
    "Coffee & Cocktail Tables": [
        "https://alfonsomarina.com/product-category/furniture-eng/tables-eng/cocktail-tables-eng/"
    ],
    "Dining Tables": [
        "https://alfonsomarina.com/product-category/furniture-eng/tables-eng/dining-tables-eng/"
    ],
    "Consoles": [
        "https://alfonsomarina.com/product-category/furniture-eng/tables-eng/consoles-sofa-eng/"
    ],
    "Beds & Headboards": [
        "https://alfonsomarina.com/product-category/furniture-eng/beds-catalog-eng/"
    ],
    "Desks": [
        "https://alfonsomarina.com/product-category/furniture-eng/tables-eng/desks-wing-tables-eng/"
    ],
    "Accent Tables": [
        "https://alfonsomarina.com/product-category/furniture-eng/tables-eng/occasional-eng/"
    ],
    "Dressers & Chests": [
        "https://alfonsomarina.com/product-category/furniture-eng/storage-eng/chest-trunks-eng/",
        "https://alfonsomarina.com/product-category/furniture-eng/storage-eng/dressers-eng/"
    ],
    "Cabinets": [
        "https://alfonsomarina.com/product-category/furniture-eng/storage-eng/cabinets-bookcases-eng/",
        "https://alfonsomarina.com/product-category/furniture-eng/storage-eng/buffets-sideboards-eng/"
    ],
    "Dining Chairs": [
        "https://alfonsomarina.com/product-category/furniture-eng/seating-eng/dining-chairs-eng/"
    ],
    "Bar Stools": [
        "https://alfonsomarina.com/product-category/furniture-eng/seating-eng/bar-stools-eng/"
    ],
    "Sofas & Loveseats": [
        "https://alfonsomarina.com/product-category/furniture-eng/seating-eng/sofas-loveseats-eng/"
    ],
    "Benches": [
        "https://alfonsomarina.com/product-category/furniture-eng/seating-eng/benches-ottomans-eng/"
    ],
    "Lounge Chairs": [
        "https://alfonsomarina.com/product-category/furniture-eng/seating-eng/occasional-chairs-eng/"
    ],
    "Mirrors": [
        "https://alfonsomarina.com/product-category/furniture-eng/accessories-eng/mirrors-eng/"
    ],
    "Boxes": [
        "https://alfonsomarina.com/product-category/furniture-eng/accessories-eng/boxes-eng/"
    ],
}

# ==== Output paths (same folder as script) ====
script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "alfonsomarina_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "AlfonsoMarina.xlsx")

# ==== Selenium Setup ====
options = Options()
options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--window-size=1920,1080")

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 20)

# ---- Helper: pick real image ----
def pick_real_image(img_tag):
    if not img_tag:
        return ""

    for key in ["data-src", "data-lazy-src", "data-original", "data-srcset", "srcset", "src"]:
        val = (img_tag.get(key) or "").strip()
        if val and "1px.png" not in val:
            if "," in val:
                return val.split(",")[-1].split(" ")[0]
            return val
    return ""

# ---- Scrape one category ----
def scrape_category(category, urls):
    all_data = []

    for url in urls:
        driver.get(url)
        wait.until(EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "div.registroProducto"))
        )

        last_count = 0
        same_rounds = 0

        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2.5)

            cards = driver.find_elements(By.CSS_SELECTOR, "div.registroProducto")
            if len(cards) == last_count:
                same_rounds += 1
            else:
                same_rounds = 0

            if same_rounds >= 3:
                break
            last_count = len(cards)

        soup = BeautifulSoup(driver.page_source, "html.parser")
        products = soup.find_all("div", class_="registroProducto")

        for p in products:
            try:
                product_url = p.find("a", class_="registroImagen")["href"]
                name_tag = p.find("a", class_="registroTitulo")
                product_name = name_tag.get_text(strip=True) if name_tag else ""

                img_tag = p.find("img")
                image_url = pick_real_image(img_tag)

                if not product_name:
                    continue

                all_data.append({
                    "Category": category,
                    "Product URL": product_url,
                    "Image URL": image_url,
                    "Product Name": product_name
                })
            except:
                continue

    df = pd.DataFrame(all_data)
    if not df.empty:
        df.drop_duplicates(subset=["Product URL"], inplace=True)
    return df

# ---- Step-1 Master Excel ----
def build_step1_master_excel():
    dfs = []
    for cat, links in CATEGORIES.items():
        df = scrape_category(cat, links)
        if not df.empty:
            dfs.append(df)

    if not dfs:
        empty_cols = ["Category", "Product URL", "Image URL", "Product Name"]
        master = pd.DataFrame(columns=empty_cols)
        master.to_excel(master_output_file, index=False)
        return master

    master = pd.concat(dfs, ignore_index=True)
    master.to_excel(master_output_file, index=False)
    return master

# ---- Step-2 Category-wise Workbook ----
def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat]
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)
        ws["A1"] = "Brand"
        ws["B1"] = "Alfonso Marina"

        # Concatenate links with commas for each category
        link_text = ", ".join(CATEGORIES[cat])

        ws["A2"] = "Link"
        ws["B2"] = link_text
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)

# ---- Main ----
def main():
    df = build_step1_master_excel()
    if not df.empty:
        build_category_wise_workbook_from_df(df)
    driver.quit()

if __name__ == "__main__":
    main()
