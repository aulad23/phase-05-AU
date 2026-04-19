import re
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from pathlib import Path

INPUT_FILE = "Caracole_Dressers_Chests.xlsx"
OUTPUT_FILE = "Caracole_Dressers_Chests_details.xlsx"

BATCH_SAVE_EVERY = 5
POLITE_DELAY = 0.7

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
}

session = requests.Session()
session.headers.update(headers)

def clean_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def col_idx(headers_row, name):
    name = name.strip().lower()
    for i, h in enumerate(headers_row, start=1):
        if (h or "").strip().lower() == name:
            return i
    return None

def parse_dim_parts(dimension_line: str):
    """
    Input example:
    "Dimensions in Inches: 46.0W X 12.0D X 72.0H"
    Return dict: Width/Depth/Height/Length/Diameter
    """
    out = {"Width": "", "Depth": "", "Height": "", "Length": "", "Diameter": ""}

    if not dimension_line:
        return out

    s = clean_text(dimension_line)

    # Try to isolate inches part only
    # e.g. "Dimensions in Inches: 46.0W X 12.0D X 72.0H"
    m = re.search(r"Dimensions\s*in\s*Inches\s*:\s*(.+)", s, flags=re.I)
    if m:
        s = m.group(1)

    # Normalize separators
    s = s.replace("×", "X")
    s = re.sub(r"\s*[xX]\s*", " X ", s)

    # Match patterns like:
    # 46W, 46.0W, 12 D, 72H, 30DIA, 30 Dia., 30Diameter
    pattern = re.compile(
        r"(?P<num>\d+(?:\.\d+)?)\s*(?P<label>DIA(?:METER)?|DIAMETER|DIA\.|DIA|W|D|H|L)\b",
        flags=re.I
    )

    for m in pattern.finditer(s):
        num = m.group("num").strip()
        label = m.group("label").upper().replace(".", "")

        if label == "W":
            out["Width"] = num
        elif label == "D":
            out["Depth"] = num
        elif label == "H":
            out["Height"] = num
        elif label == "L":
            out["Length"] = num
        elif label.startswith("DIA") or label == "DIAMETER":
            out["Diameter"] = num

    return out

def extract_weight(text_blob: str):
    """
    Finds weight values in a big text area.
    Accepts: Weight, lb, lbs, Ib (common typo), etc.
    Returns string like "52 lb" (keeps unit if found).
    """
    if not text_blob:
        return ""

    t = clean_text(text_blob)

    # Example matches:
    # "Weight: 52 lb"
    # "Weight 52 lbs"
    # "Weight: 52 Ib" (typo)
    m = re.search(r"\bWeight\b\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(lb|lbs|ib)\b", t, flags=re.I)
    if m:
        num = m.group(1)
        unit = m.group(2).lower()
        # normalize "ib" -> "lb"
        if unit == "ib":
            unit = "lb"
        return f"{num} {unit}"

    # fallback: "52 lb" without word Weight
    m2 = re.search(r"\b(\d+(?:\.\d+)?)\s*(lb|lbs|ib)\b", t, flags=re.I)
    if m2:
        num = m2.group(1)
        unit = m2.group(2).lower()
        if unit == "ib":
            unit = "lb"
        return f"{num} {unit}"

    return ""

def parse_product(url: str):
    r = session.get(url, timeout=30)
    if r.status_code != 200:
        return {
            "SKU": "", "Description": "", "Dimension": "",
            "Width": "", "Depth": "", "Height": "", "Length": "", "Diameter": "",
            "Weight": "", "Features": "", "Finish": ""
        }

    soup = BeautifulSoup(r.text, "html.parser")

    # SKU
    sku = ""
    sku_div = soup.select_one("div.product__block.product__sku")
    if sku_div:
        m = re.search(r"SKU:\s*(.+)", clean_text(sku_div.get_text(" ", strip=True)), re.I)
        if m:
            sku = clean_text(m.group(1))

    # Description
    desc = ""
    desc_div = soup.select_one("div.product__description__content")
    if desc_div:
        desc = clean_text(desc_div.get_text(" ", strip=True))

    # Dimension (prefer just the inch line)
    dimension_line = ""
    for details in soup.select("details.accordion"):
        summary = details.select_one("summary")
        if summary and "DIMENSIONS" in clean_text(summary.get_text(" ", strip=True)).upper():
            body_text = clean_text((details.select_one(".accordion__content") or details).get_text(" ", strip=True))
            m = re.search(r"Dimensions in Inches:\s*([^<]+?)(?=Dimensions in Centimeters:|$)", body_text, re.I)
            if m:
                dimension_line = clean_text("Dimensions in Inches: " + m.group(1))
            else:
                dimension_line = body_text
            break

    dim_parts = parse_dim_parts(dimension_line)

    # Features
    features = []
    for details in soup.select("details.accordion"):
        summary = details.select_one("summary")
        if summary and "FEATURES" in clean_text(summary.get_text(" ", strip=True)).upper():
            for li in details.select("li"):
                t = clean_text(li.get_text(" ", strip=True))
                if t:
                    features.append(t)
            break
    features_text = " | ".join(features)

    # Finish options
    finishes = []
    for inp in soup.select('.radio__buttons input[type="radio"][name^="options[Finish]"]'):
        val = clean_text(inp.get("value", ""))
        if val:
            finishes.append(val)
    finish = " | ".join(dict.fromkeys(finishes))

    # Weight: search in whole page text + accordion content
    page_text = clean_text(soup.get_text(" ", strip=True))
    weight = extract_weight(page_text)

    return {
        "SKU": sku,
        "Description": desc,
        "Weight": weight,
        "Width": dim_parts["Width"],
        "Depth": dim_parts["Depth"],
        "Diameter": dim_parts["Diameter"],
        "Length": dim_parts["Length"],
        "Height": dim_parts["Height"],
        "Features": features_text,
        "Finish": finish,
        "Dimension": dimension_line
    }

# -----------------------------
# Read input Excel
# -----------------------------
in_wb = load_workbook(INPUT_FILE)
in_ws = in_wb.active
headers_row = [clean_text(c.value) for c in in_ws[1]]

col_url = col_idx(headers_row, "Product URL")
col_img = col_idx(headers_row, "Image URL")
col_name = col_idx(headers_row, "Product Name")

if not col_url:
    raise Exception("Input file এ 'Product URL' column পাওয়া যায়নি!")

# -----------------------------
# Setup / Resume Output
# -----------------------------
out_path = Path(OUTPUT_FILE)

if out_path.exists():
    out_wb = load_workbook(OUTPUT_FILE)
    out_ws = out_wb.active
    scraped_urls = set()
    for r in range(2, out_ws.max_row + 1):
        u = clean_text(out_ws.cell(row=r, column=1).value)  # Product URL form input excel
        if u:
            scraped_urls.add(u)
    print(f"Resume mode: already scraped {len(scraped_urls)} URLs.")
else:
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Details"

    # ✅ Final Output Order
    out_ws.append([
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
        "Product Family Id",
        "Description",
        "Weight",
        "Width",
        "Depth",
        "Diameter",
        "Length",
        "Height",
        "Features",
        "Finish",
        "Dimension"
    ])
    scraped_urls = set()
    out_wb.save(OUTPUT_FILE)
    print("Created new output file.")

# -----------------------------
# Scrape with batch save + resume
# -----------------------------
batch_count = 0
new_rows = 0

for row in range(2, in_ws.max_row + 1):
    url = clean_text(in_ws.cell(row=row, column=col_url).value)
    if not url or url in scraped_urls:
        continue

    img_url = clean_text(in_ws.cell(row=row, column=col_img).value) if col_img else ""
    pname = clean_text(in_ws.cell(row=row, column=col_name).value) if col_name else ""

    print(f"Scraping: {url}")
    data = parse_product(url)

    # ✅ Product Family Id = Product Name (as you requested) input product Name
    product_family_id = pname

    out_ws.append([
        url,
        img_url,
        pname,
        data.get("SKU", ""),
        product_family_id,
        data.get("Description", ""),
        data.get("Weight", ""),
        data.get("Width", ""),
        data.get("Depth", ""),
        data.get("Diameter", ""),
        data.get("Length", ""),
        data.get("Height", ""),
        data.get("Features", ""),
        data.get("Finish", ""),
        data.get("Dimension", "")
    ])

    scraped_urls.add(url)
    batch_count += 1
    new_rows += 1

    if batch_count >= BATCH_SAVE_EVERY:
        out_wb.save(OUTPUT_FILE)
        print(f"Saved batch ({BATCH_SAVE_EVERY}). Total new rows: {new_rows}")
        batch_count = 0

    time.sleep(POLITE_DELAY)

out_wb.save(OUTPUT_FILE)
print(f"\nDone. Total newly scraped rows this run: {new_rows}")
print(f"Excel saved as: {OUTPUT_FILE}")
