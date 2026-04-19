import os
import requests
import pandas as pd
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM
# =========================================================

VENDOR_NAME = "Pierre Frey"

CATEGORIES = {
    "Wallpapers": [
        "https://www.pierrefrey.com/en/wallpapers",
    ]
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "pierrefrey_wallpapers.xlsx")
master_sheets_file = os.path.join(script_dir, "pierrefrey_wallpapers_category.xlsx")

headers = {
    "User-Agent": "Mozilla/5.0"
}

BASE_URL = "https://www.pierrefrey.com"

# =========================================================
# SCRAPING LOGIC
# =========================================================

visited_pages = set()
all_products = []

def scrape_page(page_url, category_name):
    response = requests.get(page_url, headers=headers)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    items = soup.find_all("div", class_="resultListItem")

    for item in items:
        link_tag = item.find("a", class_="resultListItem__link")
        img_tag = item.find("img", class_="resultListItem__img")
        sup_title = item.find("div", class_="resultListItem__supTitle")
        title = item.find("div", class_="resultListItem__title")
        sub_title = item.find("div", class_="resultListItem__subTitle")

        if not link_tag or not img_tag:
            continue

        product_url = urljoin(BASE_URL, link_tag.get("href"))
        image_url = img_tag.get("src")

        product_name = ""
        if sup_title and title:
            product_name = f"{sup_title.get_text(strip=True)} {title.get_text(strip=True)}"

        sku = sub_title.get_text(strip=True) if sub_title else ""

        all_products.append({
            "Category": category_name,
            "Product URL": product_url,
            "Image URL": image_url,
            "Product Name": product_name,
            "SKU": sku
        })

    return soup

# =========================================================
# OUTPUT SYSTEM
# =========================================================

def build_master_excel():
    if not all_products:
        cols = ["Index", "Category", "Product URL", "Image URL", "Product Name", "SKU"]
        pd.DataFrame(columns=cols).to_excel(master_output_file, index=False)
        return pd.DataFrame(columns=cols)

    df = pd.DataFrame(all_products)
    df.drop_duplicates(subset=["Product URL"], inplace=True)
    df.insert(0, "Index", range(1, len(df) + 1))
    df = df[["Index", "Category", "Product URL", "Image URL", "Product Name", "SKU"]]
    df.to_excel(master_output_file, index=False)
    return df

def build_category_wise_workbook(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)

        ws["A1"] = "Brand"
        ws["B1"] = VENDOR_NAME
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        start = 4
        for j, col in enumerate(df_cat.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_cat.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers_map = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}

        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers_map["Product URL"]).value
            name_cell = ws.cell(row=r, column=headers_map["Product Name"])
            if url:
                name_cell.hyperlink = url
                name_cell.font = link_font

    wb.save(master_sheets_file)

# =========================================================
# MAIN
# =========================================================

def main():
    for category_name, urls in CATEGORIES.items():
        for idx, url in enumerate(urls):
            print(f"Scraping page 1 of category '{category_name}'")
            soup = scrape_page(url, category_name)
            visited_pages.add(url)

            while True:
                pagination_links = soup.select("ul.pagination__list a.pagination__button--num")
                new_page_found = False

                for link in pagination_links:
                    href = link.get("href")
                    full_url = urljoin(BASE_URL, href)
                    if full_url not in visited_pages:
                        print(f"Scraping {full_url}")
                        visited_pages.add(full_url)
                        soup = scrape_page(full_url, category_name)
                        new_page_found = True
                        break

                if not new_page_found:
                    break

    df = build_master_excel()
    if not df.empty:
        build_category_wise_workbook(df)

    print(f"Completed. Total products scraped: {len(df)}")

if __name__ == "__main__":
    main()
