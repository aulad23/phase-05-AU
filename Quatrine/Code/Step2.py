import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import re

# ---------------- CONFIG ----------------
INPUT_EXCEL = "quatrine_ottomans-benches.xlsx"
OUTPUT_EXCEL = "quatrine_ottomans-benches_final.xlsx"

VENDOR = "Quatrine"
CATEGORY = "ottomans-benches"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}


# ----------------------------------------


def generate_sku(index):
    """Vendor(3) + Category(2) + Sequential Index"""
    return f"{VENDOR[:3].upper()}-{CATEGORY[:2].upper()}-{index}"


def parse_dimensions(dimension_text):
    """
    Extract Width, Height, Depth, Diameter, and Weight from dimension text
    Takes ONLY the first occurrence of each dimension type
    Returns: (width, height, depth, diameter, weight)
    """
    if not dimension_text:
        return "", "", "", "", ""

    width = ""
    height = ""
    depth = ""
    diameter = ""
    weight = ""

    # Extract Weight (lbs, LBS, Ibs, IBS, Kg, kg)
    weight_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:lbs|LBS|Ibs|IBS|Kg|kg)', dimension_text, re.IGNORECASE)
    if weight_match:
        weight = weight_match.group(1)

    # Extract FIRST occurrence of standard format dimensions like: 24″W x 17″H x 20″D
    w_match = re.search(r'(\d+(?:\.\d+)?)\s*[″"]\s*W', dimension_text, re.IGNORECASE)
    h_match = re.search(r'(\d+(?:\.\d+)?)\s*[″"]\s*H', dimension_text, re.IGNORECASE)
    d_match = re.search(r'(\d+(?:\.\d+)?)\s*[″"]\s*D(?!IA)', dimension_text, re.IGNORECASE)
    dia_match = re.search(r'(\d+(?:\.\d+)?)\s*[″"]\s*DIA', dimension_text, re.IGNORECASE)

    if w_match:
        width = w_match.group(1)
    if h_match:
        height = h_match.group(1)
    if d_match:
        depth = d_match.group(1)
    if dia_match:
        diameter = dia_match.group(1)

    # Extract from text format like "Standard height: 32″" or "height: 32″"
    if not height:
        height_text_match = re.search(r'(?:Standard\s+)?height:\s*(\d+(?:\.\d+)?)\s*[″"]?', dimension_text,
                                      re.IGNORECASE)
        if height_text_match:
            height = height_text_match.group(1)

    if not depth:
        depth_text_match = re.search(r'(?:Standard\s+)?depth:\s*(\d+(?:\.\d+)?)\s*[″"]?', dimension_text, re.IGNORECASE)
        if depth_text_match:
            depth = depth_text_match.group(1)

    if not width:
        width_text_match = re.search(r'(?:Standard\s+)?width:\s*(\d+(?:\.\d+)?)\s*[″"]?', dimension_text, re.IGNORECASE)
        if width_text_match:
            width = width_text_match.group(1)

    return width, height, depth, diameter, weight


def extract_dimension_from_description(desc_div):
    """
    Extract dimension text from description div
    Handles cases where dimension might be split across multiple paragraphs
    """
    dimension = ""

    if not desc_div:
        return dimension

    # Find paragraph that starts with "Dimensions" or "Dimension"
    for p in desc_div.find_all("p"):
        text = p.get_text(strip=True)
        if text.lower().startswith("dimensions") or text.lower().startswith("dimension"):
            dimension = text

            # Check if next paragraph is part of dimensions (contains W, H, D pattern)
            next_p = p.find_next_sibling("p")
            while next_p:
                next_text = next_p.get_text(strip=True)
                # If next paragraph has dimension pattern, append it
                if re.search(r'\d+[″"]\s*[WHD]', next_text, re.IGNORECASE):
                    dimension += " " + next_text
                    next_p = next_p.find_next_sibling("p")
                else:
                    break
            break

    return dimension


def scrape_product_page(product_url):
    response = requests.get(product_url, headers=HEADERS, timeout=30)
    soup = BeautifulSoup(response.text, "html.parser")

    # -------- Description (FULL TEXT) --------
    description = ""
    dimension = ""

    desc_div = soup.find("div", class_="product-description")
    if desc_div:
        description = desc_div.get_text(separator=" ", strip=True)
        dimension = extract_dimension_from_description(desc_div)

    return description, dimension


def main():
    # -------- Read Input Excel --------
    input_wb = load_workbook(INPUT_EXCEL)
    input_ws = input_wb.active

    # -------- Output Excel --------
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Products"

    # Updated headers with separate dimension columns
    output_ws.append([
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
        "Product Family Id",
        "Description",
        "Weight",
        "Width",
        "Depth",
        "Diameter",
        "Height",
        "Dimension"
    ])

    # -------- Process Products --------
    for index, row in enumerate(input_ws.iter_rows(min_row=2, values_only=True), start=1):
        product_url, image_url, product_name = row

        print(f"Scraping ({index}): {product_url}")

        try:
            description, dimension = scrape_product_page(product_url)
            sku = generate_sku(index)
            product_family_id = product_name

            # Parse dimensions into separate fields (takes FIRST occurrence only)
            width, height, depth, diameter, weight = parse_dimensions(dimension)

            output_ws.append([
                product_url,
                image_url,
                product_name,
                sku,
                product_family_id,
                description,
                weight,
                width,
                depth,
                diameter,
                height,
                dimension  # Keep original complete dimension text
            ])
        except Exception as e:
            print(f"Error scraping {product_url}: {e}")
            # Add row with error
            output_ws.append([
                product_url,
                image_url,
                product_name,
                generate_sku(index),
                product_name,
                f"ERROR: {str(e)}",
                "", "", "", "", "", ""
            ])

    output_wb.save(OUTPUT_EXCEL)
    print("\n✅ DONE")
    print(f"📁 Output file created: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()