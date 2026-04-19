from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
import os
import re

INPUT_FILE = "sutherland_ottoman.xlsx"
OUTPUT_FILE = "sutherland_ottoman_step2.xlsx"
BATCH_SIZE = 5

# ---------------- Chrome setup ----------------
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver, 15)

# ---------------- Load input ----------------
wb_input = load_workbook(INPUT_FILE)
ws_input = wb_input.active
rows_input = list(ws_input.iter_rows(min_row=2, values_only=True))
total_products = len(rows_input)

# ---------------- Helpers ----------------
def extract_from_details(details_text: str, label: str) -> str:
    if not details_text:
        return ""
    pattern = rf"(?:^|\|\s*){re.escape(label)}\s*:\s*(.*?)(?=\s*\|\s*|$)"
    m = re.search(pattern, details_text, flags=re.IGNORECASE)
    return m.group(1).strip() if m else ""

def first_number_only(val: str) -> str:
    if not val:
        return ""
    s = val.strip()
    cut = len(s)
    if "in" in s.lower():
        cut = s.lower().find("in")
    if "/" in s:
        cut = min(cut, s.find("/"))
    left = s[:cut]
    m = re.search(r"\d+(?:\.\d+)?", left)
    return m.group(0) if m else ""

# ---------------- Output file (RESUME LOGIC) ----------------
scraped_urls = set()

if os.path.exists(OUTPUT_FILE):
    wb_output = load_workbook(OUTPUT_FILE)
    ws_output = wb_output.active

    # Already scraped URLs collect
    for r in ws_output.iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            scraped_urls.add(r[0])

    print(f"🔁 Resume mode ON — {len(scraped_urls)} products already scraped")

else:
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "Product Details"

    ws_output.append([
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
        "Product Family Id",
        "Description",
        "Details",
        "Weight",
        "Width",
        "Depth",
        "Diameter",
        "Length",
        "Height",
        "Base",
        "Com",
        "Arm Height",
        "Seat Height",
        "Cushion",
        "Finish"
    ])

# ---------------- Scraping loop ----------------
processed_count = len(scraped_urls)

for idx, row in enumerate(rows_input, start=1):
    product_url, image_url, product_name = row

    # ⏭️ SKIP if already scraped
    if product_url in scraped_urls:
        print(f"⏭️ Skipped (already scraped): {product_name}")
        continue

    product_family_id = product_name

    try:
        driver.get(product_url)
        wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "span.productDetailsHeading__sku--sku")
            )
        )

        # SKU
        try:
            sku = driver.find_element(
                By.CSS_SELECTOR,
                "span.productDetailsHeading__sku--sku"
            ).text.strip()
        except:
            sku = ""

        # -------- Description --------
        description = ""
        try:
            desc_div = driver.find_element(
                By.XPATH,
                "//span[normalize-space()='Description']/ancestor::div[contains(@class,'productDetailsAccordion__group')]//div[contains(@class,'productDetailsAccordion__details-summary')]"
            )
            description = desc_div.get_attribute("innerText").strip()
        except:
            pass

        # -------- Details --------
        details_list = []
        try:
            rows_table = driver.find_elements(
                By.CSS_SELECTOR,
                "table.productDetailsTable tbody tr"
            )
            for r in rows_table:
                k = r.find_element(By.TAG_NAME, "th").text.strip()
                v = r.find_element(By.TAG_NAME, "td").text.strip()
                details_list.append(f"{k}: {v}")
        except:
            pass

        details = " | ".join(details_list)

        # -------- Parse from details --------
        weight = extract_from_details(details, "Weight (lb)")
        width = first_number_only(extract_from_details(details, "Width"))
        depth = first_number_only(extract_from_details(details, "Depth"))
        diameter = first_number_only(extract_from_details(details, "Diameter"))
        length = first_number_only(extract_from_details(details, "Length"))
        height = first_number_only(extract_from_details(details, "Height"))
        base = extract_from_details(details, "Base Spread")
        com = extract_from_details(details, "Com") or extract_from_details(details, "COM")
        arm_height = first_number_only(extract_from_details(details, "Arm Height"))
        seat_height = first_number_only(extract_from_details(details, "Seat Height w/frame"))
        cushion = first_number_only(extract_from_details(details, "Seat Height w/cushion"))

        # -------- Finish --------
        finishes = []
        try:
            for f in driver.find_elements(By.CSS_SELECTOR, "div.suthProdDetailsSwatches__swatch-name"):
                finishes.append(f.text.strip())
        except:
            pass

        finish = ", ".join(sorted(set(finishes)))

        # -------- Save row --------
        ws_output.append([
            product_url,
            image_url,
            product_name,
            sku,
            product_family_id,
            description,
            details,
            weight,
            width,
            depth,
            diameter,
            length,
            height,
            base,
            com,
            arm_height,
            seat_height,
            cushion,
            finish
        ])

        scraped_urls.add(product_url)
        processed_count += 1

        print(f"✅ [{processed_count}] Scraped: {product_name}")

        if processed_count % BATCH_SIZE == 0:
            wb_output.save(OUTPUT_FILE)
            print("💾 Auto-saved")

    except Exception as e:
        print(f"❌ Error on {product_name}: {e}")
        continue

# ---------------- Final save ----------------
driver.quit()
wb_output.save(OUTPUT_FILE)

print("\n🎯 SCRAPING COMPLETED / RESUMED SUCCESSFULLY")
print(f"Total products scraped: {processed_count}")
print(f"File saved at: {os.path.abspath(OUTPUT_FILE)}")
