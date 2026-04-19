import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import time

# =========================================================
# INPUT SYSTEM (FROM CODE-B – FIXED URL)
# =========================================================

CATEGORIES = {
    "Lighting": [
        "https://palmerhargrave.com/shop/?post_id=138&form_id=cee1aae&queried_type=WP_Post&queried_id=138&categories[]=14"
    ],
    "Sconces": [
        "https://palmerhargrave.com/shop/?post_id=138&form_id=cee1aae&queried_type=WP_Post&queried_id=138&categories[]=15"
    ],
    "Table Lamps": [
        "https://palmerhargrave.com/shop/?post_id=138&form_id=cee1aae&queried_type=WP_Post&queried_id=138&categories[]=1"
    ],
    "Floor Lamps": [
        "https://palmerhargrave.com/shop/?post_id=138&form_id=cee1aae&queried_type=WP_Post&queried_id=138&categories[]=12"
    ],
}

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "palmerhargrave_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "PalmerHargrave.xlsx")


# =========================================================
# SCRAPING LOGIC USING SELENIUM
# =========================================================

def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("window-size=1920,1080")

    # REMOVE --headless to open the browser window for debugging
    # chrome_options.add_argument("--headless")  # REMOVED for debugging

    driver = webdriver.Chrome(options=chrome_options)
    return driver


def scroll_page(driver, delay=2):
    """Scroll the page down to load lazy-loaded content."""
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(delay)  # Wait for the new content to load

        # Check the new scroll height and compare it with the previous one
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:  # If height doesn't change, we've reached the bottom
            break
        last_height = new_height


def scrape_products_with_selenium(url):
    driver = setup_driver()
    driver.get(url)

    # Wait for a specific element (e.g., product listings) to load
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME, "e-add-post"))  # Wait until product elements are present
        )
    except Exception as e:
        print("Error waiting for page to load:", e)
        driver.quit()
        return []

    # Scroll to load more products
    scroll_page(driver)

    # Log the page source for debugging
    print(f"Page title: {driver.title}")  # This will help confirm we're on the right page
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "lxml")

    products = []

    articles = soup.find_all("article", class_="e-add-post")

    # Debugging: Print how many articles are found
    print(f"Number of articles found: {len(articles)}")

    for article in articles:
        # Product URL
        product_link = article.find("a", class_="e-add-post-image")
        product_url = product_link["href"] if product_link else ""

        # Image URL
        img = article.find("img")
        image_url = img["src"] if img else ""

        # Product Name
        title = article.select_one("h3.e-add-post-title a")
        product_name = title.get_text(strip=True) if title else ""

        # SKU
        sku_el = article.select_one("div.e-add-item_custommeta span")
        sku = sku_el.get_text(strip=True) if sku_el else ""

        products.append({
            "Product URL": product_url,
            "Image URL": image_url,
            "Product Name": product_name,
            "SKU": sku,
        })

    driver.quit()  # Close the browser after scraping
    return products


# =========================================================
# OUTPUT SYSTEM (FROM CODE-B)
# =========================================================

def build_step1_master_excel():
    all_rows = []

    for category, urls in CATEGORIES.items():
        for url in urls:
            print(f"\n{'=' * 60}")
            print(f"Scraping: {category}")
            print(f"URL: {url}")
            print(f"{'=' * 60}")

            rows = scrape_products_with_selenium(url)
            print(f"✓ Extracted {len(rows)} products from {category}")

            for r in rows:
                r["Category"] = category
                all_rows.append(r)

    cols = [
        "Category",
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
    ]

    if not all_rows:
        pd.DataFrame(columns=cols).to_excel(master_output_file, index=False)
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    df = df[cols]
    df.drop_duplicates(subset=["Product URL"], inplace=True)

    print(f"\n{'=' * 60}")
    print(f"TOTAL PRODUCTS SAVED: {len(df)}")
    print(f"{'=' * 60}\n")

    df.to_excel(master_output_file, index=False)
    return df


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat]
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat[:31])

        ws["A1"] = "Brand"
        ws["B1"] = "Palmer Hargrave"
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES[cat])
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat = df_cat.copy()
        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start = 4
        for j, col in enumerate(df_cat.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_cat.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=3).value
            cell = ws.cell(row=r, column=4)
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)


# =========================================================
# MAIN (FROM CODE-B)
# =========================================================

def main():
    df = build_step1_master_excel()
    if not df.empty:
        build_category_wise_workbook_from_df(df)


if __name__ == "__main__":
    main()
