import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# FILE PATHS
# =========================================================
script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, "PalmerHargrave.xlsx")
output_file = os.path.join(script_dir, "PalmerHargrave_detailed.xlsx")


# =========================================================
# SELENIUM SETUP
# =========================================================
def setup_driver():
    """Chrome driver setup"""
    chrome_options = Options()
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(options=chrome_options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    return driver


# =========================================================
# HELPER FUNCTIONS (SAME AS ORIGINAL)
# =========================================================
def extract_product_family_id(product_name):
    """Product Name থেকে variant part বাদ দিয়ে Product Family ID তৈরি করবে"""
    if '–' in product_name:
        return product_name.split('–')[0].strip()
    elif ' - ' in product_name:
        return product_name.split(' - ')[0].strip()
    return product_name.strip()


def extract_wattage(details_text):
    """Max Wattage থেকে শুধু wattage number extract করবে (100W)"""
    match = re.search(r'Max Wattage:\s*(\d+W)', details_text, re.IGNORECASE)
    if match:
        return match.group(1)
    return ""


def clean_dimension_value(value):
    """
    Dimension value থেকে শুধু number বের করবে
    17″ → 17
    29 1/4″ → 29 1/4
    4 1/2″ Dia. → 4 1/2
    """
    if not value:
        return ""
    cleaned = re.sub(r'[″′]', '', value).strip()
    cleaned = re.sub(r'\s*Dia\.?', '', cleaned, flags=re.IGNORECASE).strip()
    return cleaned


def extract_shade_details(notes_text):
    """
    Shade Details থেকে শুধু shade code এবং dimensions নিবে
    """
    if 'Shade:' not in notes_text:
        return ""

    shade_start = notes_text.find('Shade:')
    shade_section = notes_text[shade_start + 6:].strip()

    lines = shade_section.split('\n')
    if not lines:
        return ""

    first_line = lines[0].strip()

    patterns_to_remove = [
        r'\s*Ivory Silk.*',
        r'\s*White Silk.*',
        r'\s*Linen.*',
        r'\s*Shade is not included.*',
        r'\s*Estimated lead-time.*'
    ]

    result = first_line
    for pattern in patterns_to_remove:
        result = re.sub(pattern, '', result, flags=re.IGNORECASE)

    return result.strip()


# =========================================================
# SCRAPING LOGIC (SAME AS ORIGINAL)
# =========================================================
def scrape_product_details_selenium(url, driver):
    """Selenium দিয়ে product details scrape করবে"""
    try:
        print(f"  🌐 Loading page...")
        driver.get(url)
        time.sleep(3)

        soup = BeautifulSoup(driver.page_source, "lxml")

        details = {
            'Description': '',
            'Dimension': '',
            'Wattage': '',
            'Finish': '',
            'Shade Details': '',
            'Weight': '',
            'Width': '',
            'Depth': '',
            'Diameter': '',
            'Height': '',
            'Canopy': ''
        }

        # Details section থেকে Wattage extract
        details_section = soup.find('h2', string='Details:')
        if details_section:
            details_text_elem = details_section.find_next('div', class_='elementor-widget-container')
            if details_text_elem:
                details_text = details_text_elem.get_text(strip=True)
                details['Wattage'] = extract_wattage(details_text)
                if details['Wattage']:
                    print(f"  ✓ Wattage: {details['Wattage']}")

        # Finish extract
        finish_section = soup.find('h2', string='Finish Shown:')
        if finish_section:
            finish_elem = finish_section.find_next('div', class_='elementor-widget-container')
            if finish_elem:
                details['Finish'] = finish_elem.get_text(strip=True)
                if details['Finish']:
                    print(f"  ✓ Finish: {details['Finish']}")

        # Standard Dimensions
        dimensions_section = soup.find('h2', string='Standard Dimensions:')
        if dimensions_section:
            dim_elem = dimensions_section.find_next('div', class_='elementor-widget-container')
            if dim_elem:
                dim_text = dim_elem.get_text().strip()

                details['Dimension'] = dim_text
                print(f"  ✓ Dimension: {dim_text}")

                width_match = re.search(r'Width:\s*([\d\s/″′\-]+)', dim_text, re.IGNORECASE)
                if width_match:
                    details['Width'] = clean_dimension_value(width_match.group(1))

                height_match = re.search(r'Height:\s*([\d\s/″′\-]+)', dim_text, re.IGNORECASE)
                if height_match:
                    details['Height'] = clean_dimension_value(height_match.group(1))

                depth_match = re.search(r'Depth:\s*([\d\s/″′\-]+)', dim_text, re.IGNORECASE)
                if depth_match:
                    details['Depth'] = clean_dimension_value(depth_match.group(1))

                diameter_match = re.search(r'Diameter:\s*([\d\s/″′\-]+)', dim_text, re.IGNORECASE)
                if diameter_match:
                    details['Diameter'] = clean_dimension_value(diameter_match.group(1))

                weight_match = re.search(r'Weight:\s*([\d\s\.,lbskg]+)', dim_text, re.IGNORECASE)
                if weight_match:
                    details['Weight'] = clean_dimension_value(weight_match.group(1))

                canopy_match = re.search(r'Canopy:\s*([\d\s/″′\-]+(?:\s*Dia\.?)?)', dim_text, re.IGNORECASE)
                if canopy_match:
                    details['Canopy'] = clean_dimension_value(canopy_match.group(1))
                    print(f"  ✓ Canopy: {details['Canopy']}")

                print(f"  ✓ Individual dimensions extracted")

        # Shade Details
        notes_section = soup.find('h2', string='Notes:')
        if notes_section:
            notes_elem = notes_section.find_next('div', class_='elementor-widget-container')
            if notes_elem:
                notes_text = notes_elem.get_text()
                shade_details = extract_shade_details(notes_text)
                if shade_details:
                    details['Shade Details'] = shade_details
                    print(f"  ✓ Shade Details: {shade_details}")

        return details

    except Exception as e:
        print(f"  ❌ Error: {str(e)}")
        return {
            'Description': '',
            'Dimension': '',
            'Wattage': '',
            'Finish': '',
            'Shade Details': '',
            'Weight': '',
            'Width': '',
            'Depth': '',
            'Diameter': '',
            'Height': '',
            'Canopy': ''
        }


# =========================================================
# INPUT SYSTEM - READ FROM PalmerHargrave.xlsx
# =========================================================
def read_input_excel():
    """Input Excel theke data read korbe (1st step er format onujai)"""
    print("📂 Reading Input Excel file...")

    wb = load_workbook(input_file)
    all_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # A1: Brand, B1: Brand Name
        brand_name = ws['B1'].value

        # A2: Link, B2: URLs
        links = ws['B2'].value

        # Row 4 theke headers, Row 5 theke data
        # Headers: Index, Category, Product URL, Image URL, Product Name, SKU
        data_rows = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:  # Index column empty hole break
                break
            data_rows.append({
                'Index': row[0],
                'Category': row[1],
                'Product URL': row[2],
                'Image URL': row[3],
                'Product Name': row[4],
                'SKU': row[5]
            })

        if data_rows:
            all_data[sheet_name] = {
                'brand_name': brand_name,
                'links': links,
                'data': pd.DataFrame(data_rows)
            }
            print(f"  ✓ Loaded {len(data_rows)} products from '{sheet_name}' sheet")

    wb.close()
    return all_data


# =========================================================
# OUTPUT SYSTEM - WRITE TO PalmerHargrave_detailed.xlsx
# =========================================================
def write_output_excel(all_data):
    """Output Excel e data write korbe (1st step er format maintain kore)"""
    print("\n📝 Writing Output Excel file...")

    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for sheet_name, sheet_data in all_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        # A1: Brand, B1: Brand Name
        ws['A1'] = 'Brand'
        ws['B1'] = sheet_data['brand_name']

        # A2: Link, B2: URLs
        ws['A2'] = 'Link'
        ws['B2'] = sheet_data['links']
        ws['B2'].alignment = Alignment(wrap_text=True)

        # Row 3: Empty (skip)

        # Row 4: Headers
        df = sheet_data['data']
        headers = list(df.columns)

        for j, col in enumerate(headers, start=1):
            ws.cell(row=4, column=j, value=col).font = bold

        # Row 5 theke data
        for i, row in enumerate(df.itertuples(index=False), start=5):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # Product URL column e hyperlink add korbo (column 3 = C)
        for r in range(5, ws.max_row + 1):
            url = ws.cell(row=r, column=3).value  # Product URL
            cell = ws.cell(row=r, column=5)  # Product Name
            if url:
                cell.hyperlink = url
                cell.font = link_font

        print(f"  ✓ Wrote {len(df)} products to '{sheet_name}' sheet")

    wb.save(output_file)
    print(f"✅ Output saved: {output_file}")


# =========================================================
# MAIN EXECUTION
# =========================================================
def main():
    # Step 1: Input Excel read korbo
    all_data = read_input_excel()

    if not all_data:
        print("❌ No data found in input file!")
        return

    # Step 2: Selenium setup
    print(f"\n🔍 Starting Selenium browser...")
    driver = setup_driver()

    try:
        # Step 3: Prottek sheet er jonno scraping
        for sheet_name, sheet_data in all_data.items():
            df = sheet_data['data']

            print(f"\n{'=' * 60}")
            print(f"📊 Processing Sheet: {sheet_name}")
            print(f"🚀 Scraping details for {len(df)} products...")
            print(f"{'=' * 60}\n")

            # নতুন columns যোগ করা
            new_columns = ['Product Family Id', 'Description', 'Weight', 'Width',
                           'Depth', 'Diameter', 'Height', 'Canopy', 'Wattage',
                           'Finish', 'Shade Details', 'Dimension']

            for col in new_columns:
                df[col] = ''

            # Product Family Id তৈরি করা
            df['Product Family Id'] = df['Product Name'].apply(extract_product_family_id)

            # Prottek product er jonno details scrape korbo
            for index, row in df.iterrows():
                product_url = row['Product URL']
                product_name = row['Product Name']

                print(f"{'=' * 60}")
                print(f"[{index + 1}/{len(df)}] {product_name}")
                print(f"URL: {product_url}")

                details = scrape_product_details_selenium(product_url, driver)

                # DataFrame update korbo
                for key, value in details.items():
                    df.at[index, key] = value

                print(f"✅ Completed")

                # Rate limiting
                wait_time = 2
                print(f"⏳ Waiting {wait_time} seconds...\n")
                time.sleep(wait_time)

            # Final column order
            final_columns = [
                'Index',
                'Category',
                'Product URL',
                'Image URL',
                'Product Name',
                'SKU',
                'Product Family Id',
                'Description',
                'Weight',
                'Width',
                'Depth',
                'Diameter',
                'Height',
                'Canopy',
                'Wattage',
                'Finish',
                'Shade Details',
                'Dimension'
            ]

            df = df[final_columns]
            all_data[sheet_name]['data'] = df

    finally:
        # Browser close
        print("\n🔒 Closing browser...")
        driver.quit()

    # Step 4: Output Excel write korbo
    write_output_excel(all_data)

    print(f"\n{'=' * 60}")
    print(f"✅ সব product এর details scrape করা হয়েছে!")
    print(f"📊 Output file: {output_file}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()