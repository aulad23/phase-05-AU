# filename: tfp_bulk_detail_scraper.py
# - Uses ONE input Excel file that you specify
# - Input must have columns: Product URL | Image URL | Product Name | List Price
# - Output file name is also specified manually
# - Output: List Price + SKU + Description + Weight + Specifications + Materials + Dimensions
#           + Length/Width/Depth/Diameter/Height
#           + Seat & Arm fields
#           + COM, COL, Base
#           + Product Family Id
#           + Canopy (from Specifications)
#           + Shade Details (from Specifications + Dimensions)

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time, os, re, json, openpyxl
from openpyxl import Workbook

# ------------------- CONFIG -------------------

# Folder where script + Excel files are located
BASE_FOLDER = os.path.dirname(os.path.abspath(__file__))

# ✅ You will set these two values:
INPUT_FILE_NAME = "thefutureperfec_dining_Table.Xlsx"          # <-- Your input list file
OUTPUT_FILE_NAME = "thefutureperfect_dining_Table_Final.Xlsx"  # <-- Output details file

HEADLESS = False      # Set True if you don’t want Chrome to open
BATCH_SAVE = 5        # Save after every N rows
WAIT_TIMEOUT = 15     # Max wait time for page load

# ---------------------------------------------


def make_driver():
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,1000")
    if HEADLESS:
        opts.add_argument("--headless=new")
    return webdriver.Chrome(options=opts)


def normalize_sku(raw: str) -> str:
    if not raw:
        return ""
    raw = raw.strip().strip(",;/|")
    m = re.match(r'([A-Za-z0-9]+-[A-Za-z0-9]+)', raw)
    if m:
        return m.group(1)
    parts = re.split(r'[-\s,/|]+', raw)
    if len(parts) >= 2 and parts[0] and parts[1]:
        return f"{parts[0]}-{parts[1]}"
    return raw


def extract_sku_from_anywhere(driver):
    # 1) Try special telemetry meta tag
    try:
        script_tag = driver.find_element(By.ID, "nitro-telemetry-meta").get_attribute("innerHTML")
        m = re.search(r'"sku"\s*:\s*"([^"]+)"', script_tag)
        if m:
            return m.group(1).strip()
    except:
        pass

    # 2) Try JSON-LD structured data
    try:
        scripts = driver.find_elements(By.XPATH, "//script[@type='application/ld+json']")
        for s in scripts:
            try:
                data = json.loads(s.get_attribute("innerText") or "{}")
            except:
                continue
            data = data if isinstance(data, list) else [data]
            for obj in data:
                if isinstance(obj, dict):
                    if "sku" in obj:
                        return str(obj["sku"]).strip()
                    offer = obj.get("offers")
                    if isinstance(offer, dict) and "sku" in offer:
                        return str(offer["sku"]).strip()
    except:
        pass

    # 3) Try any element that contains the word "SKU"
    try:
        el = driver.find_element(By.XPATH, "//*[contains(translate(text(),'sku','SKU'),'SKU')]")
        m = re.search(r'SKU[:\s#-]*([A-Za-z0-9._-]+)', el.text, re.I)
        if m:
            return m.group(1).strip()
    except:
        pass

    return ""


def get_description(driver):
    # Try some common description selectors
    for sel in [
        "div.description p",
        "div.product-description p",
        "//div[contains(@class,'description')]"
    ]:
        try:
            if sel.startswith("//"):
                return driver.find_element(By.XPATH, sel).text.strip()
            else:
                return driver.find_element(By.CSS_SELECTOR, sel).text.strip()
        except:
            continue
    return ""


def get_specifications_and_dimensions(driver):
    """
    Technical Specifications block theke:
    - Full specs text (Specifications column)
    - Dimensions value (e.g. 'L 16\" x W 16\" x H 23\"' or 'Shade: Dia. 7.3\" x L 28.7\"')
    Weight ekhane direct extract korbo na,
    pore Specifications string theke ber korbo.
    """
    specs_text = ""
    dimensions_text = ""

    try:
        section = driver.find_element(
            By.CSS_SELECTOR,
            "section.border.technical[data-script='ProductTechSpecs']"
        )
        spec_divs = section.find_elements(By.CSS_SELECTOR, "div.spec")

        parts = []
        for spec in spec_divs:
            try:
                label_el = spec.find_element(By.TAG_NAME, "h6")
                label = label_el.text.strip()
            except:
                label = ""
            try:
                value_el = spec.find_element(By.TAG_NAME, "p")
                value = value_el.text.strip()
            except:
                value = ""

            if not label and not value:
                continue

            if label and value:
                parts.append(f"{label}: {value}")
            elif label:
                parts.append(label)
            else:
                parts.append(value)

            # Dimensions detect
            if label and "dimension" in label.lower() and value and not dimensions_text:
                dimensions_text = value

        specs_text = " | ".join(parts)
    except:
        pass

    return specs_text, dimensions_text


def extract_weight_from_specs(specs_text: str) -> str:
    """
    Weight sudhu Specifications string thekei ber korbo.
    Example: 'Weight: 57 Lbs' -> '57'
    """
    if not specs_text or not isinstance(specs_text, str):
        return ""

    m = re.search(r'Weight[^|]*', specs_text, re.I)
    if not m:
        return ""

    weight_block = m.group(0)
    m_num = re.search(r'([0-9]+(?:\.[0-9]+)?)', weight_block)
    if m_num:
        return m_num.group(1)

    return ""


def extract_dimension_values(dim_text: str):
    """
    Dimensions text theke:
    L -> Length
    W -> Width
    D -> Depth
    DIA./DIA/Dia./Dia -> Diameter
    H -> Height
    """
    length = width = depth = diameter = height = ""

    if not dim_text or not isinstance(dim_text, str):
        return length, width, depth, diameter, height

    t = dim_text.replace("″", '"').replace("”", '"').replace("“", '"')
    t = t.replace("′", "'").replace("’", "'").replace("‘", "'")
    t = t.replace("×", "x").replace("X", "x")
    t = re.sub(r"\s+", " ", t).strip()

    # L / Length
    m = re.search(r'\bL(?:ength)?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        length = m.group(1)

    # W / Width
    m = re.search(r'\bW(?:idth)?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        width = m.group(1)

    # D / Depth
    m = re.search(r'\bD(?:epth)?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        depth = m.group(1)

    # DIA / Dia / Dia. / DIA.
    m = re.search(r'\b(?:DIA\.?|Dia\.?|Dia|Ø)\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        diameter = m.group(1)

    # H / Height
    m = re.search(r'\bH(?:eight)?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        height = m.group(1)

    return length, width, depth, diameter, height


def extract_seat_and_arm_values(text: str):
    """
    Combined text (Specifications + Dimensions) theke:
    Seat Height / Depth / Length
    Arm Height / Width / Depth
    """
    seat_height = seat_depth = seat_length = ""
    arm_height = arm_width = arm_depth = ""

    if not text or not isinstance(text, str):
        return seat_height, seat_depth, seat_length, arm_height, arm_width, arm_depth

    t = text.replace("\n", " ")
    t = re.sub(r"\s+", " ", t).strip()

    # Seat Height
    m = re.search(r'Seat Height[:\s]*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        seat_height = m.group(1)

    # Seat Depth
    m = re.search(r'Seat Depth[:\s]*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        seat_depth = m.group(1)

    # Seat Length
    m = re.search(r'Seat Length[:\s]*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        seat_length = m.group(1)

    # Arm Height
    m = re.search(r'Arm Height[:\s]*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        arm_height = m.group(1)

    # Arm Width
    m = re.search(r'Arm Width[:\s]*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        arm_width = m.group(1)

    # Arm Depth
    m = re.search(r'Arm Depth[:\s]*([0-9]+(?:\.[0-9]+)?)', t, re.I)
    if m:
        arm_depth = m.group(1)

    return seat_height, seat_depth, seat_length, arm_height, arm_width, arm_depth


def extract_com_col_base(specs_text: str, dimensions_text: str):
    """
    Specifications theke:
    COM, COL
    Specifications + Dimensions theke:
    Base -> full 'Base: ...' block (e.g. 'Base: 8.27"')
    """
    com = col = base = ""

    # ---- COM & COL from specs ----
    if isinstance(specs_text, str):
        t_spec = specs_text.replace("\n", " ")
        t_spec = re.sub(r"\s+", " ", t_spec).strip()

        # COM
        m = re.search(r'(?:\bCOM\b[^0-9]*|Upholstery Requirements[^0-9]*)([0-9]+(?:\.[0-9]+)?)', t_spec, re.I)
        if m:
            com = m.group(1)

        # COL
        m = re.search(r'\bCOL\b[^0-9]*([0-9]+(?:\.[0-9]+)?)', t_spec, re.I)
        if m:
            col = m.group(1)

    # ---- Base from combined (Dimensions + Specs) ----
    parts = []
    if isinstance(dimensions_text, str) and dimensions_text.strip():
        parts.append(dimensions_text.strip())
    if isinstance(specs_text, str) and specs_text.strip():
        parts.append(specs_text.strip())

    if parts:
        t = " | ".join(parts)
        m = re.search(r'(Base[^|]+)', t, re.I)
        if m:
            base = m.group(1).strip()

    return com, col, base


def extract_canopy(specs_text: str) -> str:
    """
    Specifications theke Canopy size:
    Example: 'Canopy 6" Square in Matching Metal Finish' -> '6'
    """
    if not specs_text or not isinstance(specs_text, str):
        return ""

    t = specs_text.replace("\n", " ")
    t = re.sub(r"\s+", " ", t).strip()

    # Canopy line ber kora
    m = re.search(r'Canopy[^|]*', t, re.I)
    if not m:
        return ""

    canopy_block = m.group(0)
    # First numeric value before inch symbol
    m_num = re.search(r'([0-9]+(?:\.[0-9]+)?)\s*["″]', canopy_block)
    if m_num:
        return m_num.group(1)

    return ""


def extract_shade_details(specs_text: str, dimensions_text: str) -> str:
    """
    Specifications + Dimensions theke Shade Details ber korbo:
    Example: 'Shade: Dia. 6" x H 5.5"' -> 'Dia. 6" x H 5.5"'
    (Shade: chara sudhu porer part nibo)
    """
    parts = []
    if isinstance(dimensions_text, str) and dimensions_text.strip():
        parts.append(dimensions_text.strip())
    if isinstance(specs_text, str) and specs_text.strip():
        parts.append(specs_text.strip())

    if not parts:
        return ""

    t = " | ".join(parts)

    # Shade: er porer portion capture
    m = re.search(r'Shade\s*:?\s*([^|]+)', t, re.I)
    if m:
        # Group 1: 'Dia. 6" x H 5.5"' etc.
        return m.group(1).strip()

    # Fallback: jodi Shade thake but format onno rokom hoy
    m2 = re.search(r'(Shade[^|]+)', t, re.I)
    if m2:
        block = m2.group(1).strip()
        # Jodi Shade: diye start, Shade: remove kore dibo
        block = re.sub(r'^Shade\s*:?\s*', '', block, flags=re.I).strip()
        return block

    return ""


def extract_materials(specs_text: str) -> str:
    """
    Specifications theke Materials block ber korbo:
    Example: 'Materials: Hand Knotted PK 39/39, Handspun Wool and Silk. Finished with Kilim Ends.'
             -> 'Hand Knotted PK 39/39, Handspun Wool and Silk. Finished with Kilim Ends.'
    """
    if not specs_text or not isinstance(specs_text, str):
        return ""

    t = specs_text  # already ' | ' diye join kora
    m = re.search(r'(Materials[^|]+)', t, re.I)
    if not m:
        return ""

    block = m.group(1).strip()
    # Remove leading "Materials" / "Materials:"
    block = re.sub(r'^Materials\s*:?\s*', '', block, flags=re.I).strip()
    return block


def atomic_save(wb, final_path: str):
    tmp_path = final_path + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, final_path)


def scrape_excel(driver, input_path, output_path):
    print(f"\n📘 Input file: {os.path.basename(input_path)}")
    print(f"📁 Output file: {os.path.basename(output_path)}")

    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in.active

    # Read input rows (expecting: Product URL, Image URL, Product Name, List Price)
    product_rows = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        product_rows.append({
            "Product URL": row[0],
            "Image URL": row[1],
            "Product Name": row[2],
            "List Price": row[3] if len(row) > 3 else "",
        })

    if not product_rows:
        print("⚠️ No products found in input file.")
        return

    # ✅ Fresh output workbook (new schema)
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "TFP Details"

    ws_out.append([
        "Product URL",
        "Image URL",
        "Product Name",
        "Product Family Id",
        "List Price",
        "SKU",
        "Description",
        "Weight",
        "Specifications",
        "Materials",
        "Dimensions",
        "Length",
        "Width",
        "Depth",
        "Diameter",
        "Height",
        "Seat Height",
        "Seat Depth",
        "Seat Length",
        "Arm Height",
        "Arm Width",
        "Arm Depth",
        "COM",
        "COL",
        "Base",
        "Canopy",
        "Shade Details",
    ])

    since_save = 0

    for i, prod in enumerate(product_rows, 1):
        url = str(prod["Product URL"]).strip()
        print(f"({i}/{len(product_rows)}) → {url}")
        try:
            driver.get(url)
            WebDriverWait(driver, WAIT_TIMEOUT).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            time.sleep(1)
        except Exception as e:
            print("❌ Page load failed:", e)
            continue

        sku = normalize_sku(extract_sku_from_anywhere(driver))
        desc = get_description(driver)
        specs_text, dimensions_text = get_specifications_and_dimensions(driver)

        # Weight from Specifications only
        weight = extract_weight_from_specs(specs_text)

        # Main dimensions from Dimensions column
        length, width, depth, diameter, height = extract_dimension_values(dimensions_text)

        # Seat/Arm values from combined text
        combined_text = (specs_text or "") + " | " + (dimensions_text or "")
        seat_height, seat_depth, seat_length, arm_height, arm_width, arm_depth = \
            extract_seat_and_arm_values(combined_text)

        # COM, COL, Base from specs + dimensions
        com, col, base = extract_com_col_base(specs_text, dimensions_text)

        # Canopy from Specifications
        canopy = extract_canopy(specs_text)

        # Shade Details from Specifications + Dimensions
        shade_details = extract_shade_details(specs_text, dimensions_text)

        # Materials from Specifications
        materials = extract_materials(specs_text)

        product_name = prod["Product Name"]
        product_family_id = product_name  # same as Product Name

        ws_out.append([
            url,
            prod["Image URL"],
            product_name,
            product_family_id,
            prod["List Price"],
            sku,
            desc,
            weight,
            specs_text,
            materials,
            dimensions_text,
            length,
            width,
            depth,
            diameter,
            height,
            seat_height,
            seat_depth,
            seat_length,
            arm_height,
            arm_width,
            arm_depth,
            com,
            col,
            base,
            canopy,
            shade_details,
        ])

        since_save += 1
        if since_save >= BATCH_SAVE:
            atomic_save(wb_out, output_path)
            print(f"💾 Saved {since_save} rows → {os.path.basename(output_path)}")
            since_save = 0

    atomic_save(wb_out, output_path)
    print(f"✅ Finished {os.path.basename(output_path)} ({len(product_rows)} products).")


def main():
    input_path = os.path.join(BASE_FOLDER, INPUT_FILE_NAME)
    output_path = os.path.join(BASE_FOLDER, OUTPUT_FILE_NAME)

    if not os.path.exists(input_path):
        print(f"❌ Input file not found: {input_path}")
        return

    driver = make_driver()
    try:
        scrape_excel(driver, input_path, output_path)
    except Exception as e:
        print(f"❌ Error during scraping: {e}")
    finally:
        driver.quit()

    print("\n🎉 Done! Details file created.")


if __name__ == "__main__":
    main()
