"""
The Invisible Collection - Coffee Tables Scraper
==================================================
pip install selenium beautifulsoup4 openpyxl webdriver-manager
python invisible_step1.py
"""

import time
import re
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
START_URLS = [
    "https://theinvisiblecollection.com/products-category/tables/side-tables/",
    # Add more URLs here ↓
]
OUTPUT_FILE = "invisible_collection_Side_Tables.xlsx"
HEADLESS = False
INITIAL_DELAY = 10
MAX_PAGES = 20


def init_driver():
    options = Options()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=en-US")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=options
    )
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = { runtime: {} };
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3]});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
        """
    })
    return driver


def nuke_popups(driver):
    driver.execute_script("""
        try {
            var keywords = ['popup', 'modal', 'overlay', 'login', 'trade',
                            'gdpr', 'cookie', 'consent', 'newsletter', 'subscribe'];
            keywords.forEach(function(kw) {
                document.querySelectorAll(
                    '[class*="' + kw + '"], [id*="' + kw + '"]'
                ).forEach(function(el) {
                    var tag = el.tagName.toLowerCase();
                    if (tag === 'body' || tag === 'html' || tag === 'main') return;
                    if (el.querySelector && el.querySelector('.ais-Hits-item, article.hit')) return;
                    el.remove();
                });
            });
            document.querySelectorAll('div').forEach(function(el) {
                var s = window.getComputedStyle(el);
                if ((s.position === 'fixed' || s.position === 'absolute') &&
                    parseInt(s.zIndex) > 1000 &&
                    !el.querySelector('.ais-Hits-item, article.hit, .ais-Hits')) {
                    el.remove();
                }
            });
            if (document.body) document.body.style.overflow = 'auto';
            if (document.documentElement) document.documentElement.style.overflow = 'auto';
        } catch(e) {}
    """)


def debug_snapshot(driver, label="debug"):
    try:
        driver.save_screenshot(f"{label}_screenshot.png")
        print(f"    📸 {label}_screenshot.png")
    except Exception:
        pass
    try:
        with open(f"{label}_page.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
    except Exception:
        pass


def wait_for_real_products(driver, timeout=120):
    """Wait for ACTUAL products, not skeleton gray boxes."""
    print("    ⏳ Waiting for products...", flush=True)
    deadline = time.time() + timeout
    attempt = 0

    while time.time() < deadline:
        attempt += 1

        if attempt % 7 == 0:
            try:
                nuke_popups(driver)
            except Exception:
                pass

        result = driver.execute_script("""
            try {
                var imgs = document.querySelectorAll('img.pro-front-img[src*="http"]');
                if (imgs.length > 0) return {method: 'img', count: imgs.length};

                var links = document.querySelectorAll('a[data-product_name]');
                var named = 0;
                links.forEach(function(l) {
                    if (l.getAttribute('data-product_name') && l.getAttribute('data-product_name').length > 0) named++;
                });
                if (named > 0) return {method: 'data-product_name', count: named};

                var names = document.querySelectorAll('.product-name');
                var withText = 0;
                names.forEach(function(n) { if (n.textContent.trim().length > 0) withText++; });
                if (withText > 0) return {method: 'product-name', count: withText};

                return null;
            } catch(e) { return null; }
        """)

        if result:
            print(f"    ✅ {result['count']} products loaded ({result['method']})")
            time.sleep(2)
            return True

        if attempt % 15 == 0:
            print(f"    ... {attempt}s elapsed")
            try:
                driver.execute_script("window.scrollTo(0, 600);")
                time.sleep(1)
                driver.execute_script("window.scrollTo(0, 0);")
            except Exception:
                pass

        if attempt == 30:
            debug_snapshot(driver, "midwait")

        time.sleep(1)

    print("    ❌ Timeout!")
    debug_snapshot(driver, "timeout")
    raise TimeoutError("Products did not load.")


def click_next_page(driver):
    """
    Click the NEXT PAGE button in Algolia pagination.
    Returns True if clicked, False if no next page (last page).
    """
    try:
        # Clear popups BEFORE locating the button
        nuke_popups(driver)
        time.sleep(0.5)

        # Check if next button is disabled (means we're on last page)
        disabled = driver.find_elements(
            By.CSS_SELECTOR,
            ".ais-Pagination-item--nextPage.ais-Pagination-item--disabled"
        )
        if disabled:
            return False

        # Find next page button
        next_btn = driver.find_element(
            By.CSS_SELECTOR,
            ".ais-Pagination-item--nextPage a.ais-Pagination-link"
        )
        # Scroll to pagination first
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", next_btn
        )
        time.sleep(1)

        # Clear popups again right before clicking
        nuke_popups(driver)
        time.sleep(0.5)

        # Try normal click first, fall back to JS click
        try:
            next_btn.click()
        except Exception:
            print("    ⚠ Normal click blocked — using JS click fallback")
            driver.execute_script("arguments[0].click();", next_btn)

        return True

    except Exception as e:
        print(f"    ⚠ Could not click next: {e}")
        return False


def get_current_page_number(driver):
    """Get current page number from pagination."""
    try:
        selected = driver.find_element(
            By.CSS_SELECTOR,
            ".ais-Pagination-item--selected a"
        )
        return int(selected.text)
    except Exception:
        return 0


def parse_page(html):
    soup = BeautifulSoup(html, "html.parser")
    products = []

    items = soup.select(".ais-Hits-item article.hit")
    if not items:
        items = soup.select(".ais-Hits-item")
    if not items:
        items = soup.select("article.hit")

    for item in items:
        a = item.select_one("a.vsz-product-container")
        if not a:
            a = item.select_one("a[data-product_name]")
        if not a:
            if item.name == 'a' and item.get("data-product_name"):
                a = item
            else:
                continue

        product_url  = a.get("href", "").strip()
        product_name = a.get("data-product_name", "").strip()

        if not product_name:
            name_div = item.select_one(".product-name")
            product_name = name_div.get_text(strip=True) if name_div else ""

        container = item if item.name != 'a' else item.parent
        img = container.select_one("img.pro-front-img")
        if not img:
            img = container.select_one("img[src*='http']")
        image_url = ""
        if img:
            image_url = img.get("src", "") or img.get("data-src", "") or ""

        price_wrap = container.select_one(".wcpbc-price")
        if price_wrap and "Price upon request" in price_wrap.get_text():
            list_price = "0"
        else:
            amount   = container.select_one(".woocommerce-Price-amount")
            list_price = "0"
            if amount:
                raw = amount.get_text(strip=True)
                # Remove currency symbols, "From", and whitespace — keep only numbers
                list_price = re.sub(r'[£€$¥₹\s\u00a0]', '', raw)

        if product_name:
            products.append([product_url, image_url, product_name, list_price])

    return products


def scrape_all():
    print("=" * 60)
    print("🚀 Invisible Collection — Multi-URL Scraper")
    print("=" * 60)
    print(f"📋 {len(START_URLS)} URL(s) to scrape\n")

    driver = init_driver()
    all_products = []

    try:
        for url_idx, url in enumerate(START_URLS, 1):
            print(f"\n{'─' * 60}")
            print(f"🔗 URL {url_idx}/{len(START_URLS)}: {url[:80]}...")
            print(f"{'─' * 60}")

            # ── Load first page of this URL ──
            print(f"\n📖 Loading page 1...")
            driver.get(url)
            time.sleep(INITIAL_DELAY)

            try:
                nuke_popups(driver)
            except Exception:
                pass
            time.sleep(2)
            try:
                nuke_popups(driver)
            except Exception:
                pass

            try:
                wait_for_real_products(driver)
            except TimeoutError:
                print(f"    ⚠ Could not load URL {url_idx} — skipping.")
                continue

            page_num = 1
            while page_num <= MAX_PAGES:
                current = get_current_page_number(driver)
                print(f"\n📖 Scraping page {current or page_num}...")

                rows = parse_page(driver.page_source)
                print(f"    🔎 Found {len(rows)} products")

                if not rows and page_num > 1:
                    print("    No products — stopping.")
                    break

                all_products.extend(rows)
                print(f"    📊 Total so far: {len(all_products)}")

                # ── Click NEXT page ──
                print("    ➡ Clicking next page...")
                if not click_next_page(driver):
                    print("    🏁 Last page reached!")
                    break

                # Wait for new products to load after click
                time.sleep(3)
                try:
                    nuke_popups(driver)
                except Exception:
                    pass

                try:
                    wait_for_real_products(driver, timeout=60)
                except TimeoutError:
                    print(f"    ⚠ Next page timed out — stopping.")
                    break

                page_num += 1
                time.sleep(1)

            print(f"\n✅ URL {url_idx} done — {len(all_products)} total products so far")

    finally:
        driver.quit()

    # Deduplicate by product_url
    seen = set()
    unique = []
    for row in all_products:
        url = row[0]
        if url and url not in seen:
            seen.add(url)
            unique.append(row)
        elif not url:
            unique.append(row)

    if len(unique) < len(all_products):
        print(f"\n🔄 Removed {len(all_products) - len(unique)} duplicates")

    return unique


def save_excel(products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coffee Tables"

    headers = ["Product URL", "Image URL", "Product Name", "List Price"]
    header_font = Font(bold=True, name="Calibri", size=11)
    cell_font = Font(name="Calibri", size=11)
    wrap_align = Alignment(vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = wrap_align
        cell.border = thin_border

    # Data rows
    for row in products:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = cell_font
            cell.alignment = wrap_align
            cell.border = thin_border

    # Column widths
    for i, w in enumerate([65, 65, 40, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row height
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = "4472C4"

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Done! {len(products)} products → {OUTPUT_FILE}")


if __name__ == "__main__":
    data = scrape_all()
    if data:
        save_excel(data)
    else:
        print("\n❌ No products scraped.")
        print("   TIP: Set HEADLESS = False and watch the browser to debug.")