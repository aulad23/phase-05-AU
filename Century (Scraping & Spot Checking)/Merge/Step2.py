# -*- coding: utf-8 -*-
# century_step2_fill_into_master_sheets.py

import time, re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ================= CONFIG =================
MASTER_FILE = "Century.xlsx"                 # ✅ your category-wise master
SAVE_AS     = "Century_with_details.xlsx"    # ✅ output (safe)
HEADLESS    = True
# =========================================

# Excel top rows: Brand/Link then header row is at 4 (like screenshot)
HEADER_ROW = 4
DATA_START_ROW = 5

LINK_FONT = Font(color="0563C1", underline="single")

COLUMN_ORDER = [
    "Index", "Category", "Product URL", "Image URL", "Product Name", "SKU", "Product Family Id",
    "Description", "List Price", "Weight", "Width", "Depth", "Diameter", "Height",
    "Com", "Finish", "Seat Height", "Arm Height"
]

def create_driver():
    options = Options()
    options.headless = HEADLESS
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(30)
    return driver

def get_product_family_id(product_name: str) -> str:
    if not isinstance(product_name, str): return ""
    name = product_name.strip()
    if not name: return ""
    parts = re.split(r"\s*[-–—]\s*", name, maxsplit=1)
    return parts[0].strip() if parts else name

def extract_dimensions_from_accordion(driver):
    dims = {"Width":"","Depth":"","Diameter":"","Height":"","Weight":"","Seat Height":"","Arm Height":""}
    summary_elem = None
    try:
        summary_elem = driver.find_element(By.XPATH, "//summary[.//h2[contains(text(),'Dimensions')]]")
        driver.execute_script("arguments[0].click();", summary_elem)
        time.sleep(0.5)

        content_div = summary_elem.find_element(
            By.XPATH, "./following-sibling::div[contains(@class,'accordion__content')]"
        )
        html = content_div.get_attribute("innerHTML")
        lines = re.split(r"<p>|</p>", html, flags=re.I)

        for line in lines:
            clean = re.sub(r"<.*?>", "", line).strip()
            if not clean: continue

            if m := re.search(r"(?:WEIGHT|OVERALL WEIGHT):?\s*([\d\.]+)", clean, re.I): dims["Weight"]=m.group(1)
            if m := re.search(r"(?:HEIGHT|OVERALL HEIGHT):?\s*([\d\.]+)", clean, re.I): dims["Height"]=m.group(1)
            if m := re.search(r"(?:WIDTH|OVERALL WIDTH):?\s*([\d\.]+)", clean, re.I): dims["Width"]=m.group(1)
            if m := re.search(r"(?:DEPTH|OVERALL DEPTH):?\s*([\d\.]+)", clean, re.I): dims["Depth"]=m.group(1)
            if m := re.search(r"DIAMETER:?\s*([\d\.]+)", clean, re.I): dims["Diameter"]=m.group(1)

            if m := re.search(r"Seat Height:?\s*([\d\.]+)", clean, re.I): dims["Seat Height"]=m.group(1)
            if m := re.search(r"Arm Height:?\s*([\d\.]+)", clean, re.I): dims["Arm Height"]=m.group(1)

    except NoSuchElementException:
        pass
    except Exception as e:
        print("Error extracting dimensions:", e)

    return dims, summary_elem

def extract_clean_description(div_element):
    if not div_element: return ""
    try:
        html = div_element.get_attribute("innerHTML")
        parts = re.split(r"<br\s*/?>|</p>", html, flags=re.I)
        out=[]
        for p in parts:
            t=re.sub(r"<.*?>","",p).strip()
            if not t: continue
            if re.search(r"OVERALL\s+(HEIGHT|WIDTH|DEPTH)|HEIGHT|WIDTH|DEPTH|WEIGHT|DIAMETER", t, re.I):
                continue
            out.append(t)
        return " ".join(out)
    except:
        return ""

def extract_finish(div_element):
    if not div_element: return ""
    try:
        html = div_element.get_attribute("innerHTML")
        m = re.search(r"<br\s*/?>\s*Finish:\s*([^<\n\r]+)", html, re.I)
        return m.group(1).strip() if m else ""
    except:
        return ""

def extract_com(summary_elem):
    if not summary_elem: return ""
    try:
        ps = summary_elem.find_elements(By.XPATH, "./following-sibling::div[contains(@class,'accordion__content')]//p")
        for p in ps:
            t=p.text.strip()
            m=re.search(r"(?:COM|COM Fabric).*?([\d\.]+\s*\w+)", t, re.I)
            if m: return m.group(1)
    except:
        pass
    return ""

def extract_list_price(driver):
    try:
        price_elem = driver.find_element(By.CSS_SELECTOR, "div.price.price--large")
        text = price_elem.get_attribute("innerText").strip()
        text = re.sub(r"MSRP\s*\$?\s*[\d,]+", "", text, flags=re.I)
        m = re.search(r"\$\s*[\d,]+(?:\.\d+)?", text)
        price = m.group(0).strip() if m else text.strip()
        price = re.sub(r"\bUSD\b", "", price, flags=re.I).strip()
        price = re.sub(r"\s{2,}", " ", price).strip()
        return price
    except:
        return ""

def get_header_map(ws):
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            header_map[str(v).strip()] = c
    return header_map

def ensure_columns(ws):
    header_map = get_header_map(ws)
    # if header row empty or missing columns -> set header row
    missing = [c for c in COLUMN_ORDER if c not in header_map]
    if missing:
        # rewrite header row in correct order
        for i, col in enumerate(COLUMN_ORDER, start=1):
            ws.cell(row=HEADER_ROW, column=i, value=col)
        header_map = {col: i for i, col in enumerate(COLUMN_ORDER, start=1)}
    return header_map

def main():
    wb = load_workbook(MASTER_FILE)
    driver = create_driver()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # skip empty/other sheets if needed (but you can keep all)
        header_map = ensure_columns(ws)

        col_url  = header_map.get("Product URL")
        col_name = header_map.get("Product Name")
        col_sku  = header_map.get("SKU")
        col_cat  = header_map.get("Category")

        if not col_url or not col_name:
            continue

        # detect last row with data by checking Product URL column
        last_row = ws.max_row
        while last_row >= DATA_START_ROW and not ws.cell(row=last_row, column=col_url).value:
            last_row -= 1

        if last_row < DATA_START_ROW:
            continue

        print(f"\n==============================")
        print(f"📄 Sheet: {sheet_name} | Rows: {last_row - DATA_START_ROW + 1}")
        print("==============================")

        for r in range(DATA_START_ROW, last_row + 1):
            product_url = ws.cell(row=r, column=col_url).value
            product_name = ws.cell(row=r, column=col_name).value

            if not product_url or not isinstance(product_url, str) or not product_url.startswith("http"):
                continue

            # ✅ Resume: if SKU already filled -> skip
            if col_sku and ws.cell(row=r, column=col_sku).value:
                continue

            # Ensure Category cell is filled
            if col_cat and not ws.cell(row=r, column=col_cat).value:
                ws.cell(row=r, column=col_cat, value=sheet_name)

            # **No hyperlink for Product Name**, just set it normally
            name_cell = ws.cell(row=r, column=col_name)
            name_cell.value = product_name  # just set text without hyperlink
            name_cell.font = Font(bold=True)

            print(f"🔎 {sheet_name} | Row {r} -> {product_url}")

            try:
                driver.get(product_url)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                time.sleep(1)
            except TimeoutException:
                print("❌ timeout")
                continue

            # SKU
            try:
                sku = driver.find_element(By.CSS_SELECTOR, "div.hideAll span.sku").text.strip()
            except:
                sku = ""

            family = get_product_family_id(str(product_name) if product_name else "")

            # Description + Finish
            try:
                desc_div = driver.find_element(By.CSS_SELECTOR, "div.product__description.rte.quick-add-hidden")
                description = extract_clean_description(desc_div)
                finish = extract_finish(desc_div)
            except:
                description, finish = "", ""

            # Dimensions + COM
            dims, summary = extract_dimensions_from_accordion(driver)
            com = extract_com(summary)
            price = extract_list_price(driver)

            # write back to sheet
            def put(col, val):
                if col in header_map:
                    ws.cell(row=r, column=header_map[col], value=val)

            put("SKU", sku)
            put("Product Family Id", family)
            put("Description", description)
            put("List Price", price)
            put("Weight", dims.get("Weight",""))
            put("Width", dims.get("Width",""))
            put("Depth", dims.get("Depth",""))
            put("Diameter", dims.get("Diameter",""))
            put("Height", dims.get("Height",""))
            put("Com", com)
            put("Finish", finish)
            put("Seat Height", dims.get("Seat Height",""))
            put("Arm Height", dims.get("Arm Height",""))

        # save after each sheet (safe)
        wb.save(SAVE_AS)
        print(f"✅ Saved progress -> {SAVE_AS}")

    driver.quit()
    wb.save(SAVE_AS)
    print(f"\n✅ Done. Final saved: {SAVE_AS}")

if __name__ == "__main__":
    main()
