import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import re
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ================= CONFIG =================
BASE_DIR = os.getcwd()

MASTER_OUTPUT_FILE = os.path.join(BASE_DIR, "century_ALL_categories.xlsx")
MASTER_SHEETS_FILE = os.path.join(BASE_DIR, "Century.xlsx")

DEFAULT_MAX_PAGES = 200
TIMEOUT = 25

CATEGORIES = {
    "Nightstands": [
        "https://shop.centuryfurniture.com/collections/bed-room-1?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Nightstands&sort_by=title-ascending"
    ],
    "Coffee & Cocktail Tables": [
        "https://shop.centuryfurniture.com/collections/living-room?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Cocktail+Tables&sort_by=title-ascending"
    ],
    "Side & End Tables": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Side+Tables&sort_by=title-ascending"
    ],
    "Dining Tables": [
        "https://shop.centuryfurniture.com/collections/dining-room?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Rectangular+Dining+Table&filter.p.product_type=Rectangular+Dining+Tables&filter.p.product_type=Round+Dining+Tables&sort_by=title-ascending"
    ],
    "Consoles": [
        "https://shop.centuryfurniture.com/collections/dining-room?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Console+Tables&sort_by=title-ascending"
    ],
    "Beds & Headboards": [
        "https://shop.centuryfurniture.com/collections/bed-room-1?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Poster+Beds&filter.p.product_type=Sleigh+Beds&filter.p.product_type=Upholstered+Beds&filter.p.product_type=Wood+Beds&sort_by=title-ascending"
    ],
    "Desks": [
        "https://shop.centuryfurniture.com/collections/bed-room-1?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Writing+Desks&sort_by=title-ascending"
    ],
    "Bookcases": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Bookcases&sort_by=title-ascending"
    ],
    "Accent Tables": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Accent+Tables&sort_by=title-ascending"
    ],
    "Cabinets": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Credenzas&sort_by=title-ascending"
    ],
    "Dressers & Chests": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Armoires&filter.p.product_type=Bedroom+Chests&filter.p.product_type=Double+Dressers&filter.p.product_type=Occasional+Chests&filter.p.product_type=Single+Dressers&sort_by=title-ascending"
    ],
    "Dining Chairs": [
        "https://shop.centuryfurniture.com/collections/dining-room?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Dining+Chairs&sort_by=title-ascending"
    ],
    "Bar Stools": [
        "https://shop.centuryfurniture.com/collections/dining-room?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Bar+%2F+Counter+Stools&sort_by=title-ascending"
    ],
    "Sofas & Loveseats": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Leather+Sofas+%2F+Love+Seats&filter.p.product_type=Sectionals&filter.p.product_type=Sofas&sort_by=title-ascending"
    ],
    "Benches": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Benches&sort_by=title-ascending"
    ],
    "Mirrors": [
        "https://shop.centuryfurniture.com/collections/living-room?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Mirrors&sort_by=title-ascending"
    ],
    "Boxes": [
        "https://shop.centuryfurniture.com/collections/all?filter.v.price.gte=&filter.v.price.lte=&filter.p.product_type=Boxes&sort_by=title-ascending"
    ],
    "Table Lamps": [
        "https://shop.centuryfurniture.com/search?options%5Bprefix%5D=last&page=1&q=Table+Lamps"
    ],
}
# =========================================

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
})

# ---------- Step-1 helpers ----------
def build_page_url(base_url, page):
    parsed = urlparse(base_url)
    qs = parse_qs(parsed.query)
    qs["page"] = [str(page)]
    return urlunparse(parsed._replace(query=urlencode(qs, doseq=True)))

def parse_products_from_page(html):
    soup = BeautifulSoup(html, "lxml")
    results = []

    for w in soup.find_all("div", class_="card-wrapper"):
        a = w.find("a", {"id": lambda x: x and "CardLink" in x})
        if not a:
            continue

        name = a.get_text(strip=True)
        url = "https://shop.centuryfurniture.com" + a.get("href")

        img = w.find("img")
        img_url = None
        if img and img.has_attr("srcset"):
            img_url = "https:" + img["srcset"].split(",")[0].split(" ")[0].split("&width")[0]
        elif img and img.has_attr("src"):
            img_url = "https:" + img["src"].split("&width")[0]

        results.append({
            "Product URL": url,
            "Image URL": img_url,
            "Product Name": name
        })
    return results

def scrape_category(category, links):
    all_products = []
    for link in links:
        for page in range(1, DEFAULT_MAX_PAGES + 1):
            try:
                r = session.get(build_page_url(link, page), timeout=TIMEOUT)
            except requests.RequestException:
                break

            if r.status_code != 200:
                break

            products = parse_products_from_page(r.text)
            if not products:
                break

            all_products.extend(products)

    df = pd.DataFrame(all_products)
    if not df.empty:
        df.drop_duplicates(subset=["Product URL"], inplace=True)
    return df

# ---------- Step-1 master ----------
def build_step1_master_excel():
    dfs = []
    for cat, links in CATEGORIES.items():
        df = scrape_category(cat, links)
        if not df.empty:
            df.insert(0, "Category", cat)
            dfs.append(df)
        else:
            print(f"⚠️ No products found for category: {cat}")

    if not dfs:
        empty_cols = ["Category", "Product URL", "Image URL", "Product Name"]
        master = pd.DataFrame(columns=empty_cols)
        master.to_excel(MASTER_OUTPUT_FILE, index=False)
        print(f"⚠️ No data collected from any category. Empty master saved: {MASTER_OUTPUT_FILE}")
        return master

    master = pd.concat(dfs, ignore_index=True)
    master.to_excel(MASTER_OUTPUT_FILE, index=False)
    print(f"✅ Master saved: {MASTER_OUTPUT_FILE}")
    return master

# ---------- Step-2 category-wise workbook ----------
def safe_sheet_name(name):
    return re.sub(r"[\\/*?:\[\]]", " ", name)[:31]

def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    ordered_categories = list(CATEGORIES.keys())

    for cat in ordered_categories:
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=safe_sheet_name(cat))

        ws["A1"] = "Brand"
        ws["B1"] = "Century"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

        # ✅ Freeze OFF (Row 1-4 freeze হবে না)
        ws.freeze_panes = None

    wb.save(MASTER_SHEETS_FILE)
    print("✅ Final Century.xlsx created (Category serial order)")

def main():
    df = build_step1_master_excel()
    if df.empty:
        wb = Workbook()
        wb.save(MASTER_SHEETS_FILE)
        print(f"⚠️ No category sheets created because no products were scraped. Empty workbook saved: {MASTER_SHEETS_FILE}")
        return

    build_category_wise_workbook_from_df(df)

if __name__ == "__main__":
    main()
