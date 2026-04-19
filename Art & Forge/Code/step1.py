import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM (COPIED FROM CODE-B STRUCTURE)
# =========================================================

CATEGORIES = {
    "Pulls": [
        "https://artandforge.com/collections/pulls"
    ],
    "Knobs":[
        "https://artandforge.com/collections/knobs"
    ],
    "Backplates":[
          "https://artandforge.com/collections/backplates"
    ]
}

BASE_URL = "https://artandforge.com"

headers = {
    "User-Agent": "Mozilla/5.0"
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "artandforge_all_products.xlsx")
category_output_file = os.path.join(script_dir, "ArtAndForge.xlsx")

# =========================================================
# SCRAPING LOGIC (UNCHANGED FROM CODE-A)
# =========================================================

def scrape_collection(category, url):
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    products = soup.find_all("product-grid-item")
    data = []

    for product in products:
        link_tag = product.find("a", href=True)
        product_url = BASE_URL + link_tag["href"] if link_tag else ""

        img_tag = product.find("img")
        image_url = ""
        if img_tag and img_tag.get("src"):
            image_url = "https:" + img_tag["src"]

        title_tag = product.find("p", class_="product__grid__title")
        product_name = title_tag.get_text(strip=True) if title_tag else ""

        price_tag = product.find("span", class_="price")
        list_price = ""
        if price_tag:
            price_text = price_tag.get_text(strip=True)
            list_price = price_text.replace("From", "").replace("$", "").strip()

        if not product_name:
            continue

        data.append({
            "Category": category,
            "Product URL": product_url,
            "Image URL": image_url,
            "Product Name": product_name,
            "List Price": list_price
        })

    return pd.DataFrame(data)

# =========================================================
# OUTPUT SYSTEM (COPIED FROM CODE-B STRUCTURE)
# =========================================================

def build_master_excel():
    dfs = []

    for category, urls in CATEGORIES.items():
        for url in urls:
            df = scrape_collection(category, url)
            if not df.empty:
                dfs.append(df)

    if not dfs:
        cols = ["Category", "Product URL", "Image URL", "Product Name", "List Price"]
        master_df = pd.DataFrame(columns=cols)
    else:
        master_df = pd.concat(dfs, ignore_index=True)
        master_df.drop_duplicates(subset=["Product URL"], inplace=True)

    master_df.to_excel(master_output_file, index=False)
    return master_df


def build_category_wise_workbook(df):
    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    for category in CATEGORIES.keys():
        df_cat = df[df["Category"] == category]
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=category)

        ws["A1"] = "Brand"
        ws["B1"] = "Art & Forge"

        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES[category])
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start_row = 4
        for col_idx, col_name in enumerate(df_out.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = bold

        for row_idx, row in enumerate(df_out.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        headers = {ws.cell(row=start_row, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            name_cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                name_cell.hyperlink = url
                name_cell.font = link_font

    wb.save(category_output_file)

# =========================================================
# MAIN ENTRY POINT (COPIED FROM CODE-B STRUCTURE)
# =========================================================

def main():
    df = build_master_excel()
    if not df.empty:
        build_category_wise_workbook(df)


if __name__ == "__main__":
    main()
