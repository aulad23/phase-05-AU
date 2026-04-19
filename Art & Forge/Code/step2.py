import os
import re
import json
import time
import html
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode

import pandas as pd
import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter, Retry
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ========================= CONFIG =========================
# Input: আগের স্ক্রিপ্ট থেকে generated master file
INPUT_FILE = "artandforge_all_products.xlsx"

# Output directory
script_dir = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = script_dir
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Output files
MASTER_DETAILED_FILE = os.path.join(OUTPUT_DIR, "artandforge_detailed_all.xlsx")
CATEGORY_DETAILED_FILE = os.path.join(OUTPUT_DIR, "Art & Forge_Final.xlsx")

REQUEST_TIMEOUT = 25
POLITE_DELAY_SEC = 0.6
SAVE_EVERY = 10

# =================== HTTP SESSION (RETRIES) ===================
session = requests.Session()
retries = Retry(
    total=6,
    backoff_factor=0.7,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"]
)
adapter = HTTPAdapter(max_retries=retries)
session.mount("http://", adapter)
session.mount("https://", adapter)
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
})


# ========================= HELPERS =========================
def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def absolutize_img(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if url.startswith("//"):
        return "https:" + url
    if url.startswith("/"):
        return "https://artandforge.com" + url
    return url


def safe_get(url: str) -> str:
    r = session.get(url, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    return r.text


def clean_base_url(any_url: str) -> str:
    """Keep scheme+netloc+path only. Input could already have ?variant=..."""
    sp = urlsplit(any_url.strip())
    scheme = sp.scheme or "https"
    netloc = sp.netloc or "artandforge.com"
    path = (sp.path or "").rstrip("/")
    return urlunsplit((scheme, netloc, path, "", ""))


def build_variant_url(base_clean: str, variant_id) -> str:
    """base_clean + ?variant=ID (ensure unique per variant)"""
    sp = urlsplit(base_clean)
    q = [(k, v) for (k, v) in parse_qsl(sp.query, keep_blank_values=True) if k.lower() != "variant"]
    q.append(("variant", str(variant_id)))
    return urlunsplit((sp.scheme, sp.netloc, sp.path, urlencode(q), ""))


def extract_balanced_json(text: str, marker: str):
    idx = text.find(marker)
    if idx == -1:
        return None
    start = text.find("{", idx)
    if start == -1:
        return None

    depth = 0
    in_str = False
    esc = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    raw = text[start:i + 1]
                    try:
                        return json.loads(raw)
                    except Exception:
                        return None
    return None


def get_product_json_from_html(page_html: str):
    return extract_balanced_json(page_html, "window.wpdExtensionIsProductPage")


def cents_to_price(val):
    """Shopify JSON often uses cents (e.g. 7495 -> 74.95)."""
    if val is None or val == "":
        return ""
    try:
        v = float(val)
        if v >= 1000:
            return f"{v / 100:.2f}"
        return f"{v:.2f}"
    except Exception:
        return ""


def make_generated_sku(vendor: str, category: str, idx_number: int):
    v = re.sub(r"[^A-Za-z]", "", vendor or "").upper()[:3].ljust(3, "X")
    c = re.sub(r"[^A-Za-z]", "", category or "").upper()[:2].ljust(2, "X")
    return f"{v}{c}{idx_number}"


def product_family_id_from_name(product_name: str) -> str:
    # "Blair Pull - Williamsburg" -> "Blair Pull"
    name = norm(product_name)
    if " - " in name:
        return norm(name.split(" - ", 1)[0])
    return name


# ========================= DESCRIPTION & DIMENSION =========================
def split_description_and_dimension_from_desc_html(desc_html: str):
    """
    Description: first non-'Specifications' paragraph
    Dimension: li items after 'Specifications' (if present)
    """
    if not desc_html:
        return "", ""
    desc_html = html.unescape(desc_html)
    soup = BeautifulSoup(desc_html, "html.parser")

    paragraphs = [norm(p.get_text(" ", strip=True)) for p in soup.find_all("p")]
    paragraphs = [p for p in paragraphs if p]

    description_text = ""
    if paragraphs:
        for p in paragraphs:
            if "specification" not in p.lower():
                description_text = p
                break
        if not description_text:
            description_text = paragraphs[0]

    dimension_items = []
    spec_anchor = None
    for st in soup.find_all("strong"):
        if "specification" in st.get_text(" ", strip=True).lower():
            spec_anchor = st
            break

    if spec_anchor:
        start_node = spec_anchor.parent if spec_anchor.parent else spec_anchor
        node = start_node
        for _ in range(40):
            node = node.find_next_sibling()
            if node is None:
                break
            if node.name == "ul":
                for li in node.find_all("li"):
                    t = norm(li.get_text(" ", strip=True))
                    if t:
                        dimension_items.append(t)
            if node.find("strong"):
                break

    return description_text, " | ".join(dimension_items)


def extract_dimension_from_tabpanel(page_html: str) -> str:
    """
    Fallback: #tab-panel-1-1 description tab specs (<li> items)
    """
    soup = BeautifulSoup(page_html, "html.parser")
    panel = soup.find(id="tab-panel-1-1") or soup.find("div", attrs={"role": "tabpanel"})
    if not panel:
        return ""

    items = []
    for li in panel.find_all("li"):
        t = norm(li.get_text(" ", strip=True))
        if t:
            items.append(t)

    return " | ".join(items)


# ========================= DIMENSION PARSER =========================
def parse_dimension_to_fields(dimension_text: str):
    """
    Output fields: Weight Width Depth Diameter Length Height Base
    """
    fields = {"Weight": "", "Width": "", "Depth": "", "Diameter": "", "Length": "", "Height": "", "Base": ""}

    dim = norm(dimension_text)
    if not dim:
        return fields

    parts = [p.strip() for p in re.split(r"\s*\|\s*|\n|;|•", dim) if p.strip()]

    kv_pattern = re.compile(
        r"(Total Length|Total Width|Total Height|Total Depth|Total Diameter|"
        r"Bar Width|Base Width|Breadth|Length|Width|Height|Depth|Diameter|Base|"
        r"Weight|Projection|Center to Center)\s*[-:]\s*([-\d]*\.?\d+)",
        flags=re.IGNORECASE
    )

    compact_pattern = re.compile(
        r'(?P<num>[-+]?\d*\.?\d+)\s*(?:in(?:ches)?|inch|\"|\")?\s*'
        r'(?P<label>Dia|DIAM|Ø|L|W|D|H|Length|Width|Depth|Height|Diameter|Projection|Base|Weight)\b',
        flags=re.IGNORECASE
    )

    def set_field(label_raw: str, num_raw: str):
        label = (label_raw or "").strip().lower()
        val = (num_raw or "").strip()
        if not val:
            return
        if val.startswith("."):
            val = "0" + val

        if label in ("projection", "center to center"):
            return

        if label == "weight":
            if not fields["Weight"]:
                fields["Weight"] = val
        elif label in ("w", "width", "total width", "breadth", "bar width"):
            if not fields["Width"]:
                fields["Width"] = val
        elif label in ("d", "depth", "total depth"):
            if not fields["Depth"]:
                fields["Depth"] = val
        elif label in ("h", "height", "total height"):
            if not fields["Height"]:
                fields["Height"] = val
        elif label in ("dia", "diam", "ø", "diameter", "total diameter"):
            if not fields["Diameter"]:
                fields["Diameter"] = val
        elif label in ("l", "length", "total length"):
            if not fields["Length"]:
                fields["Length"] = val
        elif label in ("base", "base width"):
            if not fields["Base"]:
                fields["Base"] = val

    for p in parts:
        low = p.lower()
        if "center to center" in low:
            continue

        pairs = kv_pattern.findall(p)
        if pairs:
            for k, num in pairs:
                set_field(k, num)

        # normalize cases where Projection is glued: 0.10"Projection
        normalized = re.sub(r'(\"|\")\s*projection\b', r'\1 Projection', p, flags=re.IGNORECASE)

        tokens = compact_pattern.findall(normalized)
        if tokens:
            for num, lab in tokens:
                set_field(lab, num)

    return fields


# ========================= VARIANT PAGE EXTRACT =========================
def extract_selected_color_from_page(page_html: str) -> str:
    """Extract color from page"""
    soup = BeautifulSoup(page_html, "html.parser")
    wrapper = soup.select_one('div.selector-wrapper--color')
    if not wrapper:
        return ""
    val = wrapper.select_one('legend span[data-swapper-target]')
    if val:
        return norm(val.get_text(" ", strip=True))
    return ""


def extract_variant_image_from_meta(page_html: str) -> str:
    """Most reliable for variant image changes."""
    soup = BeautifulSoup(page_html, "html.parser")
    for sel in [
        ('meta', {'property': 'og:image'}),
        ('meta', {'name': 'og:image'}),
        ('meta', {'name': 'twitter:image'}),
        ('meta', {'property': 'twitter:image'}),
    ]:
        tag = soup.find(sel[0], sel[1])
        if tag and tag.get("content"):
            return absolutize_img(tag["content"])
    return ""


# ========================= EXCEL OUTPUT =========================
def save_category_detailed_workbook(out_rows):
    """Create category-wise workbook with exact structure from image"""
    if not out_rows:
        return

    df = pd.DataFrame(out_rows)

    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    categories = sorted(df["Category"].unique())

    for category in categories:
        df_cat = df[df["Category"] == category].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=category)

        # Row 1: Brand header
        ws["A1"] = "Brand"
        ws["B1"] = "Art & Forge"
        ws["A1"].font = bold

        # Row 2: Link header
        ws["A2"] = "Link"
        ws["B2"] = "https://artandforge.com/collections/pulls"
        ws["A2"].font = bold

        # Row 4: Column headers (exact order from image)
        headers = [
            "Index", "Category", "Product URL", "Image URL", "Product Name", "SKU",
            "Product Family", "Description", "Weight", "Width", "Depth", "Diameter",
            "Length", "Height", "Base", "List Price", "Color"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = bold

        # Add index column to dataframe
        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        # Reorder columns to match headers
        column_mapping = {
            "Index": "Index",
            "Category": "Category",
            "Product URL": "Product URL",
            "Image URL": "Image URL",
            "Product Name": "Product Name",
            "SKU": "SKU",
            "Product Family Id": "Product Family",
            "Description": "Description",
            "Weight": "Weight",
            "Width": "Width",
            "Depth": "Depth",
            "Diameter": "Diameter",
            "Length": "Length",
            "Height": "Height",
            "Base": "Base",
            "List Price": "List Price",
            "Color": "Color"
        }

        # Write data rows starting from row 5
        for row_idx, (_, row) in enumerate(df_cat.iterrows(), start=5):
            for col_idx, header in enumerate(headers, start=1):
                # Map header to dataframe column
                df_col = None
                for df_key, header_name in column_mapping.items():
                    if header_name == header:
                        df_col = df_key
                        break

                value = row.get(df_col, "") if df_col else ""
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add hyperlinks to Product URLs (column C)
        for row_idx in range(5, ws.max_row + 1):
            url_cell = ws.cell(row=row_idx, column=3)  # Column C (Product URL)
            if url_cell.value:
                url_cell.hyperlink = url_cell.value
                url_cell.font = link_font

        # Adjust column widths for better readability
        ws.column_dimensions['A'].width = 8  # Index
        ws.column_dimensions['B'].width = 12  # Category
        ws.column_dimensions['C'].width = 50  # Product URL
        ws.column_dimensions['D'].width = 50  # Image URL
        ws.column_dimensions['E'].width = 40  # Product Name
        ws.column_dimensions['F'].width = 15  # SKU
        ws.column_dimensions['G'].width = 25  # Product Family
        ws.column_dimensions['H'].width = 60  # Description

    wb.save(CATEGORY_DETAILED_FILE)
    print(f"[SAVED] Category detailed: {CATEGORY_DETAILED_FILE}")


def save_master_detailed_excel(out_rows):
    """Save master detailed Excel file"""
    out_df = pd.DataFrame(out_rows)
    if not out_df.empty:
        # Reorder columns
        out_df = out_df[[
            "Category", "Product URL", "Image URL", "Product Name", "SKU",
            "Product Family Id", "Description",
            "Weight", "Width", "Depth", "Diameter", "Length", "Height", "Base",
            "List Price", "Color", "Size", "Dimension"
        ]]
    out_df.to_excel(MASTER_DETAILED_FILE, index=False)
    print(f"[SAVED] Master detailed: {MASTER_DETAILED_FILE}")


# ========================= MAIN =========================
def main():
    print("Starting detailed scraping...")

    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        print("Please run the first scraper script first to generate the input file.")
        return

    df = pd.read_excel(INPUT_FILE)

    # Find Product URL column
    url_col = None
    for c in df.columns:
        if c.strip().lower() in ["product url", "product_url", "url", "product link", "product"]:
            url_col = c
            break
    if not url_col:
        raise ValueError("Product URL column not found.")

    # Find Category column
    cat_col = None
    for c in df.columns:
        if c.strip().lower() in ["category", "categories"]:
            cat_col = c
            break

    # Resume support
    done_urls = set()
    out_rows = []

    if os.path.exists(MASTER_DETAILED_FILE):
        try:
            prev = pd.read_excel(MASTER_DETAILED_FILE)
            out_rows.extend(prev.to_dict("records"))
            if "Product URL" in prev.columns:
                done_urls = set(prev["Product URL"].dropna().astype(str).str.strip().tolist())
            print(f"Resuming from {len(done_urls)} previously scraped products")
        except Exception:
            pass

    sku_counter = len(out_rows) + 1
    total_rows = len(df)
    processed = 0

    for idx, row in df.iterrows():
        raw_url = str(row.get(url_col, "")).strip()
        if not raw_url or raw_url.lower() == "nan":
            continue

        category = str(row.get(cat_col, "NA")).strip() if cat_col else "NA"
        base_clean = clean_base_url(raw_url)

        processed += 1
        print(f"[{processed}/{total_rows}] Processing: {base_clean}")

        try:
            base_html = safe_get(base_clean)
            product = get_product_json_from_html(base_html)
            if not product:
                print(f"  ⚠ No product JSON found")
                continue

            vendor = norm(product.get("vendor", "")) or "Unknown"

            product_name = norm(product.get("title", "")) or ""
            product_family_id = product_family_id_from_name(product_name)

            # Description + Dimension
            desc_html = product.get("description", "") or ""
            description_text, dimension_text = split_description_and_dimension_from_desc_html(desc_html)

            # Dimension fallback
            if not dimension_text:
                dimension_text = extract_dimension_from_tabpanel(base_html)

            dims = parse_dimension_to_fields(dimension_text)

            variants = product.get("variants", []) or []
            if not variants:
                print(f"  ⚠ No variants found")
                continue

            variant_count = 0
            # Process all variants
            for v in variants:
                vid = v.get("id")
                if not vid:
                    continue

                color_key = norm(v.get("option1", ""))  # Color
                size_key = norm(v.get("option2", ""))  # Size

                variant_url = build_variant_url(base_clean, vid)
                if variant_url in done_urls:
                    continue

                var_html = safe_get(variant_url)

                color = extract_selected_color_from_page(var_html) or color_key
                size = size_key

                img_url = extract_variant_image_from_meta(var_html)

                sku = norm(v.get("sku", ""))
                if not sku:
                    sku = make_generated_sku(vendor, category, sku_counter)
                    sku_counter += 1

                list_price = cents_to_price(v.get("price") if v.get("price") is not None else product.get("price"))

                out_rows.append({
                    "Category": category,
                    "Product URL": variant_url,
                    "Image URL": img_url,
                    "Product Name": product_name,
                    "SKU": sku,
                    "Product Family Id": product_family_id,
                    "Description": description_text,
                    "Weight": dims["Weight"],
                    "Width": dims["Width"],
                    "Depth": dims["Depth"],
                    "Diameter": dims["Diameter"],
                    "Length": dims["Length"],
                    "Height": dims["Height"],
                    "Base": dims["Base"],
                    "List Price": list_price,
                    "Color": color,
                    "Size": size,
                    "Dimension": dimension_text
                })
                done_urls.add(variant_url)
                variant_count += 1

                time.sleep(POLITE_DELAY_SEC)

            print(f"  ✓ Extracted {variant_count} variants")

        except Exception as e:
            print(f"  ✗ ERROR: {e}")

        # Batch save
        if len(done_urls) % SAVE_EVERY == 0 and len(done_urls) > 0:
            save_master_detailed_excel(out_rows)

    # Final save
    print("\nFinalizing output files...")
    save_master_detailed_excel(out_rows)
    save_category_detailed_workbook(out_rows)

    print(f"\n{'=' * 60}")
    print(f"DONE! Total products scraped: {len(out_rows)}")
    print(f"Output files:")
    print(f"  1. {MASTER_DETAILED_FILE}")
    print(f"  2. {CATEGORY_DETAILED_FILE}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()