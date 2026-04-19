# ==============================
# Janus et Cie – SCRAPER WITH MANUAL CLICK FALLBACK
# Automatic + Manual Hybrid Approach
# Browser restart per category
# ==============================

import time
import re
import os
import random
import pandas as pd
from urllib.parse import urljoin

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    ElementClickInterceptedException,
    StaleElementReferenceException
)

from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ================= INPUT SYSTEM =================

BASE_DIR = os.getcwd()

MASTER_OUTPUT_FILE = os.path.join(BASE_DIR, "janusetcie_ALL_categories.xlsx")
MASTER_SHEETS_FILE = os.path.join(BASE_DIR, "JanusEtCie.xlsx")

BASE_URL = "https://www.janusetcie.com"

CATEGORIES = {
    "Coffee & Cocktail Tables": [
        "https://www.janusetcie.com/residential/collections/#DX2MqZMd"
    ],
    "Side & End Tables": [
        "https://www.janusetcie.com/residential/collections/#IfLe5UFD"
    ],
    "Dining Tables": [
        "https://www.janusetcie.com/residential/collections/#hxciuOO8"
    ],

    "Consoles": [
        "https://www.janusetcie.com/residential/collections/#3EWQaR0f"
    ],

    "Bar Stools": [
        "https://www.janusetcie.com/residential/collections/#MhJtSwoC"
    ],

    "Sofas & Loveseats": [
        "https://www.janusetcie.com/residential/collections/#wLeG35j4"
    ],

    "Sectionals": [
        "https://www.janusetcie.com/residential/collections/#eVWqWLSV"
    ],

    "Lounge Chairs": [
        "https://www.janusetcie.com/residential/collections/#cqO8oSMu"
    ],

    "Ottomans": [
        "https://www.janusetcie.com/residential/collections/#jBJJwAV6"
    ],

    "Benches": [
        "https://www.janusetcie.com/residential/collections/#To77A6Jq"
    ],

    "Vases": [
        "https://www.janusetcie.com/residential/collections/#Y1ztsepq"
    ],

    "Objects": [
        "https://www.janusetcie.com/residential/collections/#NA6argsr"
    ],

    "Baskets & Planters": [
        "https://www.janusetcie.com/residential/collections/#ki6guy1L"
    ],

    "Rugs": [
        "https://www.janusetcie.com/residential/collections/#8RlRf9Bp"
    ],


}

# ================= GLOBAL DRIVER VARIABLES =================
driver = None
wait = None
actions = None


# ================= BROWSER MANAGEMENT =================

def init_driver():
    """
    Notun browser instance create kore
    """
    global driver, wait, actions

    print("\n🌐 Starting new browser instance...")

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=chrome_options
    )

    wait = WebDriverWait(driver, 80)
    actions = ActionChains(driver)

    print("   ✅ Browser ready!\n")


def close_driver():
    """
    Browser bondho kore
    """
    global driver

    if driver:
        print("\n🔴 Closing browser instance...")
        try:
            driver.quit()
            print("   ✅ Browser closed!\n")
        except:
            pass
        driver = None


# ================= HELPERS =================

def human_pause(a=0.25, b=0.75):
    time.sleep(random.uniform(a, b))


def products_count():
    return len(driver.find_elements(By.CSS_SELECTOR, "a.product-list"))


def wait_products_min(min_count=1, timeout=60):
    w = WebDriverWait(driver, timeout)
    w.until(lambda d: len(d.find_elements(By.CSS_SELECTOR, "a.product-list")) >= min_count)


def human_scroll_into_view(el):
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'center'});", el
    )
    human_pause(0.4, 0.9)


def human_click(el, tries=5):
    """
    Manual-like click:
    - scroll into view
    - mouse move + pause
    - offset click
    - fallback JS click
    """
    last = None
    for _ in range(tries):
        try:
            human_scroll_into_view(el)

            # move + pause
            actions.move_to_element(el).pause(random.uniform(0.25, 0.7)).perform()
            human_pause(0.12, 0.35)

            # offset click (more human)
            offx = random.randint(-7, 7)
            offy = random.randint(-5, 5)
            actions.move_to_element_with_offset(el, offx, offy) \
                .pause(random.uniform(0.12, 0.35)) \
                .click() \
                .perform()

            human_pause(0.5, 1.2)
            return True

        except (ElementClickInterceptedException, StaleElementReferenceException) as e:
            last = e
            human_pause(0.8, 1.6)
            continue
        except Exception as e:
            last = e
            try:
                driver.execute_script("arguments[0].click();", el)
                human_pause(0.5, 1.2)
                return True
            except Exception as e2:
                last = e2
                human_pause(0.8, 1.6)

    return False


def _first_product_fingerprint():
    cards = driver.find_elements(By.CSS_SELECTOR, "a.product-list")
    fp = []
    for a in cards[:10]:
        href = (a.get_attribute("href") or "").strip()
        if href:
            fp.append(href)
    return tuple(fp)


def get_total_pages():
    try:
        links = driver.find_elements(By.CSS_SELECTOR, "ul.filter-paging li a")
        nums = []
        for a in links:
            t = (a.text or "").strip()
            if t.isdigit():
                nums.append(int(t))
        return max(nums) if nums else 1
    except:
        return 1


def find_page_link(page_no):
    links = driver.find_elements(By.CSS_SELECTOR, "ul.filter-paging li a")
    for a in links:
        if (a.text or "").strip() == str(page_no):
            return a
    return None


# ================= IMPROVED SCROLL & WAIT FUNCTIONS =================

def aggressive_scroll_trigger():
    """
    আরো aggressive lazy load trigger
    """
    # পুরো page নিচে scroll করো
    for i in range(20):
        height = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script(f"window.scrollTo(0, {height * i / 20});")
        time.sleep(0.3)

    # একবার একদম নিচে যাও
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

    # আবার উপরে ফিরে আসো
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(2)


def wait_for_page_change(old_fingerprint, timeout=140):
    """
    আরো robust wait - multiple conditions check করে
    """
    start_time = time.time()

    while time.time() - start_time < timeout:
        try:
            # 1. Check করো products আছে কিনা
            current_count = products_count()
            if current_count == 0:
                time.sleep(1)
                continue

            # 2. Check করো fingerprint change হয়েছে কিনা
            new_fp = _first_product_fingerprint()
            if len(new_fp) > 0 and new_fp != old_fingerprint:
                return True

            # 3. Scroll trigger দাও বারবার
            driver.execute_script("window.scrollBy(0, 500);")
            time.sleep(0.5)
            driver.execute_script("window.scrollBy(0, -500);")
            time.sleep(0.5)

        except:
            pass

        time.sleep(1)

    return False


# ================= MANUAL CLICK HELPER =================

def wait_for_manual_action(page_no, timeout=300):
    """
    User কে manually click করতে দেয় এবং page change detect করে
    """
    print("\n" + "=" * 60)
    print(f"🖱️  MANUAL ACTION REQUIRED")
    print("=" * 60)
    print(f"\n👉 Please MANUALLY click on page {page_no}")
    print(f"👉 Wait for products to load completely")
    print(f"👉 Then press ENTER in this terminal\n")
    print(f"⏰ You have {timeout} seconds...")
    print("=" * 60 + "\n")

    old_fp = _first_product_fingerprint()

    # User থেকে input নাও (এটা blocking হবে)
    try:
        user_input = input("Press ENTER after you've clicked and products loaded: ")
    except:
        pass

    # Check করো products change হয়েছে কিনা
    print("\n   🔍 Verifying page change...")

    for attempt in range(10):
        try:
            new_fp = _first_product_fingerprint()
            count = products_count()

            if count > 0 and new_fp != old_fp:
                print(f"   ✅ Page change detected! Found {count} products")
                return True

            if attempt < 9:
                print(f"   ⏳ Still checking... (attempt {attempt + 1}/10)")
                time.sleep(2)
                aggressive_scroll_trigger()
        except:
            pass

    print("   ⚠️  Could not verify page change, but continuing...")
    return True


# ================= PAGINATION WITH MANUAL FALLBACK =================

def click_page(page_no, max_auto_retry=3):
    """
    HYBRID VERSION - Auto try করে, fail হলে manual mode
    """
    last_err = None

    # প্রথমে automatic try করো
    for attempt in range(1, max_auto_retry + 1):
        try:
            print(f"\n📄 Attempting page {page_no} AUTOMATICALLY (Try {attempt}/{max_auto_retry})...")

            # Old fingerprint save করো
            old_fp = _first_product_fingerprint()

            # Pagination area তে scroll করো
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            human_pause(1.5, 2.5)

            # Page link খুঁজো
            target = find_page_link(page_no)
            if target is None:
                aggressive_scroll_trigger()
                target = find_page_link(page_no)
                if target is None:
                    raise RuntimeError(f"Page link not found: {page_no}")

            # Click করো
            ok = human_click(target, tries=8)
            if not ok:
                raise RuntimeError("human_click failed on pagination link")

            print(f"   ✓ Clicked page {page_no}, waiting for products...")

            # Click করার পর wait করো
            human_pause(2.0, 3.5)

            # Aggressive scroll trigger দাও
            aggressive_scroll_trigger()

            # Products load হওয়ার জন্য wait করো
            if wait_for_page_change(old_fp, timeout=60):  # Reduced timeout for auto
                print(f"   ✅ Page {page_no} loaded AUTOMATICALLY!")

                # Final scroll trigger
                aggressive_scroll_trigger()

                final_count = products_count()
                print(f"   ✓ Found {final_count} products on page {page_no}")
                return
            else:
                raise RuntimeError("Products didn't change after clicking")

        except Exception as e:
            last_err = e
            print(f"   ✗ Auto attempt {attempt} failed: {str(e)[:100]}")

            if attempt < max_auto_retry:
                wait_time = 2.0 + attempt
                print(f"   ⏳ Waiting {wait_time:.1f}s before next auto retry...")
                time.sleep(wait_time)

    # যদি automatic fail হয়, তাহলে manual mode এ যাও
    print(f"\n❌ Automatic clicking failed after {max_auto_retry} attempts")
    print(f"🔄 Switching to MANUAL MODE...\n")

    # Browser scroll করে দাও pagination এর কাছে
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)

    # Manual action wait করো
    if wait_for_manual_action(page_no):
        # Extra scroll trigger
        aggressive_scroll_trigger()

        final_count = products_count()
        print(f"   ✓ Manual click successful! Found {final_count} products\n")
        return
    else:
        raise RuntimeError(f"Failed to load page {page_no} even with manual intervention")


def ensure_products(start_url, max_attempts=3):
    """
    IMPROVED VERSION - আরো robust initial page load
    """
    for attempt in range(1, max_attempts + 1):
        try:
            print(f"🌐 Loading {start_url} (Attempt {attempt}/{max_attempts})...")

            driver.get(start_url)
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
            human_pause(3.0, 4.5)

            # Aggressive scroll
            aggressive_scroll_trigger()

            # Check products
            for check in range(60):
                count = products_count()
                if count > 0:
                    print(f"   ✓ Found {count} products!")
                    return True

                if check % 10 == 0:
                    print(f"   ⏳ Still waiting for products... ({check}s)")
                    aggressive_scroll_trigger()

                time.sleep(1)

            # যদি products না পাও
            if attempt < max_attempts:
                print(f"   ✗ No products found, retrying...")
                continue

        except Exception as e:
            print(f"   ✗ Error: {str(e)[:100]}")
            if attempt < max_attempts:
                time.sleep(5)
                continue

    print(f"   ✗ Failed to load products after {max_attempts} attempts")
    return False


# ================= PRODUCT EXTRACTION =================

def clean_name(raw):
    if not raw:
        return ""
    raw = raw.replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in raw.split("\n")]
    lines = [ln for ln in lines if ln]

    junk = {"INDOOR", "OUTDOOR", "STOCKED ITEMS", "STOCKED"}
    for ln in lines:
        if ln.upper() in junk:
            continue
        if len(ln) >= 4:
            return ln
    return lines[0] if lines else ""


def get_name_by_hover(card):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
        time.sleep(0.2)

        actions.move_to_element(card).perform()
        time.sleep(0.35)

        try:
            el = card.find_element(By.CSS_SELECTOR, ".list-item-overlay .notranslate")
            raw = el.get_attribute("textContent") or el.text or ""
            name = clean_name(raw)
            if name:
                return name
        except:
            pass

        raw2 = card.get_attribute("innerText") or card.get_attribute("textContent") or ""
        return clean_name(raw2)
    except:
        return ""


def scrape_page(seen, category):
    rows = []
    cards = driver.find_elements(By.CSS_SELECTOR, "a.product-list")

    print(f"   🔍 Scraping {len(cards)} products from current page...")

    for a in cards:
        try:
            href = (a.get_attribute("href") or "").strip()
            product_url = href if href.startswith("http") else urljoin(BASE_URL, href)
            if not product_url or product_url in seen:
                continue
            seen.add(product_url)

            img_url = ""
            try:
                img = a.find_element(By.CSS_SELECTOR, "img")
                img_url = (img.get_attribute("src") or "").strip()
            except:
                pass

            name = get_name_by_hover(a)

            rows.append({
                "Category": category,
                "Product URL": product_url,
                "Image URL": img_url,
                "Product Name": name
            })
        except:
            continue

    print(f"   ✓ Scraped {len(rows)} new products")
    return rows


# ================= SCRAPE SINGLE CATEGORY =================

def scrape_category(category, urls, seen):
    """
    Ekta category scrape kore
    """
    category_rows = []

    print(f"\n📂 Category: {category}")
    print("-" * 60)

    for start_url in urls:
        if not ensure_products(start_url):
            print(f"   ⚠️  Skipping URL (no products found)")
            continue

        total_pages = get_total_pages()
        print(f"\n   📊 Total pages detected: {total_pages}")

        # page 1
        print(f"\n   📄 Scraping Page 1...")
        category_rows += scrape_page(seen, category)

        # next pages
        for p in range(2, total_pages + 1):
            print(f"\n   📄 Moving to Page {p}...")
            try:
                click_page(p)
                category_rows += scrape_page(seen, category)
            except Exception as e:
                print(f"   ❌ Failed to scrape page {p}: {str(e)[:100]}")

                # Ask user if they want to continue
                try:
                    cont = input(f"\n⚠️  Continue to next page? (y/n): ").lower().strip()
                    if cont != 'y':
                        break
                except:
                    break

    print(f"\n   ✅ Category '{category}' completed! Scraped {len(category_rows)} products")
    return category_rows


# ================= OUTPUT SYSTEM =================

def safe_sheet_name(name):
    return re.sub(r"[\\/*?:\[\]]", " ", name)[:31]


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat, links in CATEGORIES.items():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=safe_sheet_name(cat))

        ws["A1"] = "Brand"
        ws["B1"] = "Janus et Cie"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(links)
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start_row = 4
        for col_idx, col_name in enumerate(df_cat.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col_name).font = bold

        for row_idx, row in enumerate(df_cat.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        headers = {ws.cell(row=start_row, column=c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(MASTER_SHEETS_FILE)
    print(f"\n✅ Saved category-wise workbook: {MASTER_SHEETS_FILE}")


# ================= MAIN =================

def main():
    all_rows = []
    seen = set()

    print("\n" + "=" * 60)
    print("🚀 Starting Janus et Cie Scraper (HYBRID MODE)")
    print("=" * 60 + "\n")

    total_categories = len(CATEGORIES)
    current_cat = 0

    for category, urls in CATEGORIES.items():
        current_cat += 1

        print("\n" + "=" * 60)
        print(f"📦 Processing Category {current_cat}/{total_categories}")
        print("=" * 60)

        # Notun browser kholo
        init_driver()

        try:
            # Category scrape koro
            category_rows = scrape_category(category, urls, seen)
            all_rows += category_rows

        except Exception as e:
            print(f"\n❌ Error scraping category '{category}': {str(e)}")

        finally:
            # Browser bondho koro
            close_driver()

            # Ektu wait koro porer category er age
            if current_cat < total_categories:
                print(f"\n⏸️  Pausing 3 seconds before next category...\n")
                time.sleep(3)

    print("\n" + "=" * 60)
    print("📊 Processing results...")
    print("=" * 60 + "\n")

    df = pd.DataFrame(all_rows)
    if not df.empty:
        df = df[["Category", "Product URL", "Image URL", "Product Name"]]
        df.drop_duplicates(subset=["Product URL"], inplace=True)

        print(f"✅ Total unique products scraped: {len(df)}")
        print(f"✅ Saving to: {MASTER_OUTPUT_FILE}")

        df.to_excel(MASTER_OUTPUT_FILE, index=False)
        build_category_wise_workbook_from_df(df)
    else:
        print("⚠️  No products found!")

    print("\n" + "=" * 60)
    print("🎉 Scraping completed!")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()