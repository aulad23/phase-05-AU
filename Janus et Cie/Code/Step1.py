# ==============================
# Integrated Production Scraper
# CODE-A Scraping Logic (UNCHANGED)
# CODE-B Input / Output System
# ==============================

import time
import re
import os
import pandas as pd
from urllib.parse import urljoin

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =======================================================
# ================= INPUT SYSTEM (FROM CODE-B) ==========
# =======================================================

BASE_DIR = os.getcwd()

MASTER_OUTPUT_FILE = os.path.join(BASE_DIR, "janusetcie_ALL_categories.xlsx")
MASTER_SHEETS_FILE = os.path.join(BASE_DIR, "JanusEtCie.xlsx")

BASE_URL = "https://www.janusetcie.com"

CATEGORIES = {
    "Coffee & Cocktail Tables": [
        "https://www.janusetcie.com/residential/collections/#DX2MqZMd"
    ],
   "Side & End Tables": [
        "https://www.janusetcie.com/residential/collections/#IfLe5UFD"
    ]
}

# =======================================================
# ================= SCRAPING LOGIC (CODE-A) ==============
# ======================= UNCHANGED ======================
# =======================================================

options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 40)
actions = ActionChains(driver)


def scroll_trigger():
    for _ in range(12):
        driver.execute_script("window.scrollBy(0, 900);")
        time.sleep(0.6)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1.2)


def products_count():
    return len(driver.find_elements(By.CSS_SELECTOR, "a.product-list"))


def ensure_products(start_url):
    driver.get(start_url)
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(2)

    scroll_trigger()

    for _ in range(30):
        if products_count() > 0:
            return True
        time.sleep(1)
    return False


def get_total_pages():
    try:
        links = driver.find_elements(By.CSS_SELECTOR, "ul.filter-paging li a")
        nums = []
        for a in links:
            t = (a.text or "").strip()
            if t.isdigit():
                nums.append(int(t))
        return max(nums) if nums else 1
    except:
        return 1


def click_page(page_no):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)

    links = driver.find_elements(By.CSS_SELECTOR, "ul.filter-paging li a")
    target = None
    for a in links:
        if (a.text or "").strip() == str(page_no):
            target = a
            break

    if target is None:
        raise RuntimeError(f"Page link not found: {page_no}")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", target)
    time.sleep(2)

    scroll_trigger()

    for _ in range(25):
        if products_count() > 0:
            return
        time.sleep(1)


def clean_name(raw: str) -> str:
    if not raw:
        return ""
    raw = raw.replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in raw.split("\n")]
    lines = [ln for ln in lines if ln]

    junk = {"INDOOR", "OUTDOOR", "STOCKED ITEMS", "STOCKED"}
    for ln in lines:
        if ln.upper() in junk:
            continue
        if len(ln) >= 4:
            return ln
    return lines[0] if lines else ""


def get_name_by_hover(card) -> str:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
        time.sleep(0.2)

        actions.move_to_element(card).perform()
        time.sleep(0.35)

        try:
            el = card.find_element(By.CSS_SELECTOR, ".list-item-overlay .notranslate")
            raw = el.get_attribute("textContent") or el.text or ""
            name = clean_name(raw)
            if name:
                return name
        except:
            pass

        raw2 = card.get_attribute("innerText") or card.get_attribute("textContent") or ""
        return clean_name(raw2)

    except:
        return ""


def scrape_page(seen, category):
    rows = []
    cards = driver.find_elements(By.CSS_SELECTOR, "a.product-list")

    for a in cards:
        try:
            href = (a.get_attribute("href") or "").strip()
            product_url = href if href.startswith("http") else urljoin(BASE_URL, href)
            if not product_url or product_url in seen:
                continue
            seen.add(product_url)

            img_url = ""
            try:
                img = a.find_element(By.CSS_SELECTOR, "img")
                img_url = (img.get_attribute("src") or "").strip()
            except:
                pass

            name = get_name_by_hover(a)

            rows.append({
                "Category": category,
                "Product URL": product_url,
                "Image URL": img_url,
                "Product Name": name
            })
        except:
            continue

    return rows


# =======================================================
# ================= OUTPUT SYSTEM (FROM CODE-B) ==========
# =======================================================

def safe_sheet_name(name):
    return re.sub(r"[\\/*?:\[\]]", " ", name)[:31]


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat, links in CATEGORIES.items():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=safe_sheet_name(cat))

        ws["A1"] = "Brand"
        ws["B1"] = "Janus et Cie"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(links)
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start_row = 4
        for col_idx, col_name in enumerate(df_cat.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col_name).font = bold

        for row_idx, row in enumerate(df_cat.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        headers = {ws.cell(row=start_row, column=c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(MASTER_SHEETS_FILE)


# =======================================================
# ================= MAIN ENTRY POINT =====================
# =======================================================

def main():
    all_rows = []
    seen = set()

    for category, urls in CATEGORIES.items():
        for start_url in urls:
            ok = ensure_products(start_url)
            if not ok:
                continue

            total_pages = get_total_pages()

            all_rows += scrape_page(seen, category)

            for p in range(2, total_pages + 1):
                click_page(p)
                all_rows += scrape_page(seen, category)

    df = pd.DataFrame(all_rows)
    if not df.empty:
        df = df[["Category", "Product URL", "Image URL", "Product Name"]]
        df.drop_duplicates(subset=["Product URL"], inplace=True)

    df.to_excel(MASTER_OUTPUT_FILE, index=False)
    build_category_wise_workbook_from_df(df)

    # DO NOT quit driver (attach mode)


if __name__ == "__main__":
    main()
