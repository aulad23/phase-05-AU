import os
import time
import re
import pandas as pd
from urllib.parse import urldefrag

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# =========================
# CONFIG
# =========================
INPUT_FILE = "JanusEtCie.xlsx"
OUTPUT_FILE = "JanusEtCie_details.xlsx"
CHECKPOINT_FILE = "janusetcie_checkpoint.txt"
ERROR_LOG_FILE = "scraping_errors.txt"
BATCH_SIZE = 10

HEADER_ROW = 4
DATA_START = 5

# =========================
# Attach to existing Chrome
# =========================
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 45)


# =========================
# CAPTCHA Detection & Manual Handling
# =========================
def detect_captcha():
    """Detect if Cloudflare CAPTCHA is present"""
    try:
        # Check for common CAPTCHA indicators
        page_text = driver.page_source.lower()

        # Cloudflare CAPTCHA indicators
        if "cloudflare" in page_text and "verifying" in page_text:
            return True
        if "just a moment" in page_text:
            return True
        if "checking your browser" in page_text:
            return True

        # Check for CAPTCHA iframe
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            src = iframe.get_attribute("src") or ""
            if "captcha" in src.lower() or "challenge" in src.lower():
                return True

        # Check for CAPTCHA title
        title = driver.title.lower()
        if "just a moment" in title or "attention required" in title:
            return True

    except Exception as e:
        print(f"   ⚠️ CAPTCHA detection error: {e}")

    return False


def wait_for_manual_captcha_solve():
    """
    Detects CAPTCHA and waits for user to solve it manually.
    User presses ENTER to continue after solving.
    """
    if detect_captcha():
        print("\n" + "=" * 70)
        print("🔒 CAPTCHA DETECTED!")
        print("=" * 70)
        print("📌 Please solve the CAPTCHA manually in the browser")
        print("📌 After solving, press ENTER here to continue...")
        print("=" * 70)

        input()  # Wait for user to press Enter

        print("✅ Continuing scraping...")
        time.sleep(2)  # Small delay after solving
        return True

    return False


# =========================
# Enhanced Helpers
# =========================
def clean(txt):
    if not txt:
        return ""
    txt = txt.replace("\r", "\n")
    txt = re.sub(r"[ \t]+", " ", txt)
    txt = re.sub(r"\n+", "\n", txt)
    return txt.strip()


def page_ready():
    """Enhanced page ready check with CAPTCHA detection"""
    try:
        # Check for CAPTCHA first
        wait_for_manual_captcha_solve()

        # Wait for document ready
        wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(1.5)

        # Check again after page load (CAPTCHA might appear after redirect)
        wait_for_manual_captcha_solve()

        # Wait for key elements (multiple attempts)
        for _ in range(3):
            try:
                driver.find_element(By.CSS_SELECTOR, ".entry-title")
                break
            except:
                # Check if CAPTCHA appeared
                if detect_captcha():
                    wait_for_manual_captcha_solve()
                time.sleep(1)

        # Extra scroll to trigger lazy loading
        driver.execute_script("window.scrollTo(0, 500);")
        time.sleep(0.5)
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)

    except Exception as e:
        print(f"⚠️ Page ready warning: {e}")
        # Check one more time for CAPTCHA
        wait_for_manual_captcha_solve()
        time.sleep(2)


def get_text(css):
    try:
        return clean(driver.find_element(By.CSS_SELECTOR, css).get_attribute("textContent"))
    except:
        return ""


def get_sku():
    """Enhanced SKU extraction with multiple methods"""
    try:
        metas = driver.find_elements(By.CSS_SELECTOR, ".entry-info .item-detail-meta")
        for m in metas:
            t = clean(m.get_attribute("textContent") or "")
            if "Product Number" in t:
                t = t.replace("Product Number", "").strip()
                m1 = re.search(r"\b\d+(?:-\d+){2,}\b", t)
                if m1:
                    return m1.group(0)

        all_text = driver.find_element(By.TAG_NAME, "body").get_attribute("textContent")
        sku_match = re.search(r"Product Number[:\s]*(\d+(?:-\d+){2,})", all_text)
        if sku_match:
            return sku_match.group(1)

        url = driver.current_url
        url_parts = url.rstrip('/').split('/')
        if url_parts:
            last_part = url_parts[-1]
            sku_in_url = re.search(r"(\d+(?:-\d+){2,})", last_part)
            if sku_in_url:
                return sku_in_url.group(1)
    except Exception as e:
        print(f"   ⚠️ SKU extraction error: {e}")

    return ""


def get_description():
    """Enhanced description extraction"""
    try:
        metas = driver.find_elements(By.CSS_SELECTOR, ".entry-info .item-detail-meta")
        for m in metas:
            t = clean(m.text)
            if "Finish Shown" in t:
                continue
            if m.find_elements(By.CSS_SELECTOR, "table.item-dimensions"):
                continue
            ps = m.find_elements(By.CSS_SELECTOR, "p")
            if ps:
                desc = clean(ps[0].text)
                if desc and len(desc) > 10:
                    return desc

        try:
            desc_elem = driver.find_element(By.CSS_SELECTOR, ".product-description, .entry-description, .description")
            desc = clean(desc_elem.text)
            if desc and len(desc) > 10:
                return desc
        except:
            pass

        paragraphs = driver.find_elements(By.CSS_SELECTOR, ".entry-info p")
        for p in paragraphs:
            text = clean(p.text)
            if text and len(text) > 20 and "Finish" not in text:
                return text
    except Exception as e:
        print(f"   ⚠️ Description extraction error: {e}")

    return ""


def get_finish():
    """Enhanced finish extraction"""
    try:
        finish = clean(driver.find_element(By.CSS_SELECTOR, "p.top-color-label").text)
        if finish:
            return finish
    except:
        pass

    try:
        finish = clean(driver.find_element(By.CSS_SELECTOR, ".color-label p").text)
        if finish:
            return finish
    except:
        pass

    try:
        body_text = driver.find_element(By.TAG_NAME, "body").get_attribute("textContent")
        finish_match = re.search(r"Finish Shown[:\s]*([^\n]+)", body_text)
        if finish_match:
            return clean(finish_match.group(1))
    except:
        pass

    return ""


def get_list_price():
    """
    Extract List Price from page.
    Examples:
    - "List Price Starting At: $3,749" → "3749" or "3.749"
    - "List Price: $1,234.56" → "1234.56" or "1.234"
    """
    try:
        # Method 1: Find in item-detail-meta div
        metas = driver.find_elements(By.CSS_SELECTOR, ".item-detail-meta")
        for m in metas:
            text = clean(m.get_attribute("textContent") or "")

            # Look for "List Price Starting At:" or "List Price:"
            if "List Price" in text:
                # Extract price using regex
                # Matches: $3,749 or $1,234.56 etc.
                price_match = re.search(r"\$\s*([\d,]+(?:\.\d{2})?)", text)
                if price_match:
                    price = price_match.group(1)
                    # Remove commas: "3,749" → "3749"
                    price = price.replace(",", "")

                    # Convert to format with dot as thousand separator if needed
                    # "3749" → "3.749" (for thousands)
                    if "." not in price and len(price) >= 4:
                        # Add dot before last 3 digits: "3749" → "3.749"
                        price = price[:-3] + "." + price[-3:]

                    return price

        # Method 2: Search in entire body text
        body_text = driver.find_element(By.TAG_NAME, "body").get_attribute("textContent")

        # Look for "List Price Starting At: $X,XXX"
        price_patterns = [
            r"List Price Starting At:\s*\$\s*([\d,]+(?:\.\d{2})?)",
            r"List Price:\s*\$\s*([\d,]+(?:\.\d{2})?)",
            r"Starting at:\s*\$\s*([\d,]+(?:\.\d{2})?)",
        ]

        for pattern in price_patterns:
            match = re.search(pattern, body_text, re.IGNORECASE)
            if match:
                price = match.group(1)
                price = price.replace(",", "")

                # Format: "3749" → "3.749"
                if "." not in price and len(price) >= 4:
                    price = price[:-3] + "." + price[-3:]

                return price

        # Method 3: Look for price in HTML comments (as shown in your example)
        html_source = driver.page_source
        comment_match = re.search(r"<!--\s*Price:\s*\$\s*([\d,]+(?:\.\d{2})?)\s*-->", html_source)
        if comment_match:
            price = comment_match.group(1)
            price = price.replace(",", "")

            if "." not in price and len(price) >= 4:
                price = price[:-3] + "." + price[-3:]

            return price

    except Exception as e:
        print(f"   ⚠️ List Price extraction error: {e}")

    return ""


def get_main_image():
    """Enhanced image extraction with multiple fallbacks"""
    try:
        img = driver.find_element(By.CSS_SELECTOR, ".carousel-image-wrapper img.carousel-image")
        src = img.get_attribute("src")
        if src and "placeholder" not in src.lower():
            return src
    except:
        pass

    try:
        img = driver.find_element(By.CSS_SELECTOR, "img.carousel-image")
        src = img.get_attribute("src")
        if src and "placeholder" not in src.lower():
            return src
    except:
        pass

    try:
        img = driver.find_element(By.CSS_SELECTOR, "img.product-image, .product-gallery img")
        src = img.get_attribute("src")
        if src and "placeholder" not in src.lower():
            return src
    except:
        pass

    try:
        imgs = driver.find_elements(By.CSS_SELECTOR, "img")
        for im in imgs:
            src = im.get_attribute("src") or ""
            if "wp-content/uploads" in src and "placeholder" not in src.lower():
                return src
    except:
        pass

    return ""


def parse_dimensions():
    """Enhanced dimension parsing"""
    data = {
        "Weight": "",
        "Width": "",
        "Depth": "",
        "Diameter": "",
        "Length": "",
        "Height": "",
        "Seat Height": "",
        "Arm Height": ""
    }

    try:
        table = driver.find_element(By.CSS_SELECTOR, "table.item-dimensions")
        rows = table.find_elements(By.CSS_SELECTOR, "tr")

        for r in rows:
            tds = r.find_elements(By.CSS_SELECTOR, "td")
            if len(tds) < 2:
                continue

            key = clean(tds[0].text).upper()
            raw_val = clean(tds[1].text)

            m = re.search(r"\d+(\.\d+)?", raw_val)
            val = m.group(0) if m else ""

            if key == "WEIGHT":
                data["Weight"] = val
            elif key in ["W", "WIDTH"]:
                data["Width"] = val
            elif key == "DIAM":
                data["Diameter"] = val
            elif key in ["D", "DEPTH"]:
                data["Depth"] = val
            elif key in ["L", "LENGTH"]:
                data["Length"] = val
            elif key in ["H", "HEIGHT"]:
                data["Height"] = val
            elif key in ["SH", "SEAT HEIGHT"]:
                data["Seat Height"] = val
            elif key in ["AH", "ARM HEIGHT"]:
                data["Arm Height"] = val

    except Exception as e:
        print(f"   ⚠️ Dimension parsing error: {e}")

    return data


def get_finish_links():
    """Get all finish variation links"""
    links = []
    try:
        for a in driver.find_elements(By.CSS_SELECTOR, "a.filter.product-color"):
            href = a.get_attribute("href")
            if href:
                href, _ = urldefrag(href)
                links.append(href)
    except Exception as e:
        print(f"   ⚠️ Finish links extraction error: {e}")

    uniq = []
    seen = set()
    for u in links:
        if u not in seen:
            seen.add(u)
            uniq.append(u)
    return uniq


def log_error(url, error_msg):
    """Log errors to file"""
    try:
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} | {url} | {error_msg}\n")
    except:
        pass


# =========================
# Template Reader/Writer
# =========================
def safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", " ", str(name))[:31]


def read_step1_template_excel(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    meta = {}
    dfs = {}

    for sh in wb.sheetnames:
        ws = wb[sh]

        brand_name = ws["B1"].value or ""
        link_val = ws["B2"].value or ""

        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=HEADER_ROW, column=c).value
            if v:
                headers[str(v).strip()] = c

        required = ["index", "Category", "Product URL"]
        headers_lower = {k.lower(): v for k, v in headers.items()}
        missing = [k for k in required if k.lower() not in headers_lower]

        if missing:
            print(f"⚠️  Sheet '{sh}' skipped (missing headers): {missing}")
            continue

        rows = []
        r = DATA_START
        while True:
            idx_col = headers_lower.get("index", 1)
            cat_col = headers_lower.get("category", 2)
            purl_col = headers_lower.get("product url", 3)

            idx = ws.cell(row=r, column=idx_col).value
            cat = ws.cell(row=r, column=cat_col).value
            purl = ws.cell(row=r, column=purl_col).value

            if idx is None and cat is None and purl is None:
                break

            rows.append({
                "Index": idx,
                "Category": cat,
                "Product URL": (purl or "").strip() if isinstance(purl, str) else (purl or ""),
            })
            r += 1

        df = pd.DataFrame(rows)
        if not df.empty:
            dfs[sh] = df
            meta[sh] = {"brand_name": brand_name, "link": link_val}

    df_all = pd.concat(dfs.values(), ignore_index=True) if dfs else pd.DataFrame()
    return meta, df_all, dfs


def write_step2_output_multi_sheet(out_path: str, dfs_by_sheet: dict, meta_by_sheet: dict):
    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    for sh, df in dfs_by_sheet.items():
        if df is None or df.empty:
            continue

        ws = wb.create_sheet(title=safe_sheet_name(sh))

        brand_name = meta_by_sheet.get(sh, {}).get("brand_name", "")
        link_val = meta_by_sheet.get(sh, {}).get("link", "")

        ws["A1"] = "Brand"
        ws["A1"].font = bold
        ws["B1"] = brand_name
        ws["A2"] = "Link"
        ws["A2"].font = bold
        ws["B2"] = link_val
        ws["B2"].alignment = Alignment(wrap_text=True)

        cols = list(df.columns)
        for ci, col in enumerate(cols, start=1):
            ws.cell(row=HEADER_ROW, column=ci, value=col).font = bold

        for ri, row in enumerate(df.itertuples(index=False), start=DATA_START):
            for ci, val in enumerate(row, start=1):
                ws.cell(row=ri, column=ci, value=val)

        if "Product URL" in cols and "Product Name" in cols:
            url_c = cols.index("Product URL") + 1
            name_c = cols.index("Product Name") + 1
            for r in range(DATA_START, ws.max_row + 1):
                url = ws.cell(row=r, column=url_c).value
                nm_cell = ws.cell(row=r, column=name_c)
                if url:
                    nm_cell.hyperlink = url
                    nm_cell.font = link_font

    wb.save(out_path)
    print(f"✅ Output saved: {out_path}")


def load_existing_output(out_path: str):
    if not os.path.exists(out_path):
        return {}, set()

    wb = load_workbook(out_path, data_only=True)
    dfs = {}
    done = set()

    for sh in wb.sheetnames:
        ws = wb[sh]

        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=HEADER_ROW, column=c).value
            if v:
                headers[str(v).strip()] = c

        if "Product URL" not in headers:
            continue

        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=HEADER_ROW, column=c).value
            if v:
                cols.append(str(v).strip())

        rows = []
        r = DATA_START
        while True:
            first_cell = ws.cell(row=r, column=1).value
            url = ws.cell(row=r, column=headers["Product URL"]).value

            if first_cell is None and (url is None or str(url).strip() == ""):
                break

            row_obj = {}
            for ci, col in enumerate(cols, start=1):
                row_obj[col] = ws.cell(row=r, column=ci).value
            rows.append(row_obj)

            if url:
                done.add(str(url).strip())
            r += 1

        df = pd.DataFrame(rows)
        if not df.empty:
            dfs[sh] = df

    return dfs, done


def write_checkpoint(base_url):
    try:
        with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
            f.write(base_url or "")
    except:
        pass


# =========================
# MAIN SCRAPING
# =========================
print("\n" + "=" * 70)
print("🚀 Starting Enhanced Scraper with Manual CAPTCHA Handling")
print("=" * 70)
print("💡 If CAPTCHA appears, solve it manually and press ENTER to continue")
print("=" * 70 + "\n")

meta, df_all, dfs_in = read_step1_template_excel(INPUT_FILE)

if not dfs_in:
    raise RuntimeError("❌ No valid sheets found in input.")

dfs_out_existing, done_urls = load_existing_output(OUTPUT_FILE)
print(f"✅ Already scraped URLs: {len(done_urls)}")

dfs_out = dict(dfs_out_existing)

FINAL_COLS = [
    "Product URL", "Image URL", "Product Name", "SKU", "Product Family Id",
    "Description", "Weight", "Width", "Depth", "Diameter", "Length", "Height",
    "Seat Height", "Arm Height", "Finish", "List Price", "Category"
]

buffer_rows_by_sheet = {}
processed_since_save = 0
total_processed = 0
total_variations = 0

try:
    for sh, df_in in dfs_in.items():
        if df_in is None or df_in.empty:
            continue

        print("\n" + "=" * 70)
        print(f"📄 Sheet: {sh} | Products: {len(df_in)}")
        print("=" * 70)

        if sh not in buffer_rows_by_sheet:
            buffer_rows_by_sheet[sh] = []

        for i, row in df_in.iterrows():
            base_url = str(row.get("Product URL", "")).strip()
            category = str(row.get("Category", "")).strip()

            if not base_url:
                continue

            write_checkpoint(base_url)
            total_processed += 1

            try:
                driver.get(base_url)
                page_ready()  # Will handle CAPTCHA if it appears

                finish_links = get_finish_links()
                if not finish_links:
                    finish_links = [base_url]

                print(f"   [{i + 1}/{len(df_in)}] {base_url[:60]}... | Variations: {len(finish_links)}")

                for var_idx, url in enumerate(finish_links, 1):
                    url = (url or "").strip()
                    if not url:
                        continue

                    if url in done_urls:
                        print(f"      ⏭️  Variation {var_idx}/{len(finish_links)}: Already scraped")
                        continue

                    try:
                        driver.get(url)
                        page_ready()  # Will handle CAPTCHA if it appears

                        pname = get_text(".entry-title .notranslate")
                        sku = get_sku()
                        dims = parse_dimensions()
                        img = get_main_image()
                        desc = get_description()
                        finish = get_finish()
                        list_price = get_list_price()

                        if not pname and not sku and not img:
                            print(f"      ⚠️  Variation {var_idx}/{len(finish_links)}: No data extracted")
                            log_error(url, "No data extracted")
                            continue

                        buffer_rows_by_sheet[sh].append({
                            "Product URL": url,
                            "Image URL": img,
                            "Product Name": pname,
                            "SKU": sku,
                            "Product Family Id": pname,
                            "Description": desc,
                            "Weight": dims["Weight"],
                            "Width": dims["Width"],
                            "Depth": dims["Depth"],
                            "Diameter": dims["Diameter"],
                            "Length": dims["Length"],
                            "Height": dims["Height"],
                            "Seat Height": dims["Seat Height"],
                            "Arm Height": dims["Arm Height"],
                            "Finish": finish,
                            "List Price": list_price,
                            "Category": category
                        })

                        done_urls.add(url)
                        processed_since_save += 1
                        total_variations += 1

                        print(f"      ✅ Variation {var_idx}/{len(finish_links)}: {pname[:40] if pname else 'N/A'}...")

                    except Exception as e:
                        print(f"      ❌ Variation {var_idx}/{len(finish_links)}: Error - {str(e)[:50]}")
                        log_error(url, str(e))
                        continue

                    # Batch save
                    if processed_since_save >= BATCH_SIZE:
                        for sname, rows_buf in buffer_rows_by_sheet.items():
                            if not rows_buf:
                                continue
                            new_df = pd.DataFrame(rows_buf)

                            for c in FINAL_COLS:
                                if c not in new_df.columns:
                                    new_df[c] = ""
                            new_df = new_df[FINAL_COLS]

                            if sname in dfs_out and not dfs_out[sname].empty:
                                merged = pd.concat([dfs_out[sname], new_df], ignore_index=True)
                            else:
                                merged = new_df.copy()

                            merged = merged.drop_duplicates(subset=["Product URL"], keep="first")
                            dfs_out[sname] = merged

                            buffer_rows_by_sheet[sname] = []

                        write_step2_output_multi_sheet(OUTPUT_FILE, dfs_out, meta)
                        processed_since_save = 0
                        print(f"\n   💾 Batch saved | Total variations: {total_variations}\n")

            except Exception as e:
                print(f"   ❌ Base product error: {str(e)[:80]}")
                log_error(base_url, str(e))
                continue

    # Save remaining
    any_left = any(len(v) > 0 for v in buffer_rows_by_sheet.values())
    if any_left:
        for sname, rows_buf in buffer_rows_by_sheet.items():
            if not rows_buf:
                continue
            new_df = pd.DataFrame(rows_buf)
            for c in FINAL_COLS:
                if c not in new_df.columns:
                    new_df[c] = ""
            new_df = new_df[FINAL_COLS]

            if sname in dfs_out and not dfs_out[sname].empty:
                merged = pd.concat([dfs_out[sname], new_df], ignore_index=True)
            else:
                merged = new_df.copy()

            merged = merged.drop_duplicates(subset=["Product URL"], keep="first")
            dfs_out[sname] = merged

        write_step2_output_multi_sheet(OUTPUT_FILE, dfs_out, meta)
        print(f"\n💾 Final batch saved")

    print("\n" + "=" * 70)
    print(f"✅ SCRAPING COMPLETED")
    print(f"📊 Total base products: {total_processed}")
    print(f"📊 Total variations scraped: {total_variations}")
    print(f"📄 Check errors in: {ERROR_LOG_FILE}")
    print("=" * 70 + "\n")

finally:
    pass