"""
De Sousa Hughes - Extended Scraper (v7 Final)
Input  : desousahughes_consoles.xlsx  (Product URL, Image URL, Product Name, SKU)
Output : desousahughes_consoles_final.xlsx

Changelog v7 (dimension fixes):
  - FIX 1  (v4): Added L x D x H regex  (beauvais-console)
  - FIX 2  (v4): Force resp.encoding = 'utf-8'  (mojibake â\n")
  - FIX 3  (v5): Convert vulgar fractions ⅛¼⅜½⅝¾⅞ → decimal  (Alexander Lamont)
  - FIX 4  (v5): Strip parentheticals before parsing  (Kemizo Flip-Top)
  - FIX 5  (v5): Added L x W x H regex  (Bridger Console)
  - FIX 6  (v5): Added H x L x D regex  (Shell Console)
  - FIX 7  (v6): Added D x W x H regex  (Catena / Chelsea Nightstand)
  - FIX 8  (v6): Added W x H x D regex  (Tuy Nightstand)
  - FIX 9 : Strip size prefix "Large -" / "Small -" before parsing  (John Pomp)
  - FIX 10: W x W x H → treat second W as D  (Elsa Cocktail Table website typo)
  - FIX 11: Label-before-number format "W 29\" x L 56\" x H 14\""  (Forge Coffee Table)
  - FIX 12: Dual height "14 & 16\" H" → take first value  (Prysm Cocktail Table)
  - FIX 13: Height after DIA  e.g. "48\" DIA x 15.5\" H"  (circular tables)
  - FIX 14: Normalize "Diameter"/"Diam" → DIA  (Fourchette, Magnus, Catalina...)
  - FIX 15: H x DIA format  e.g. "19\" H x 12.5\" DIA"  (Alta Spot, Nave)
  - FIX 16: DIA x D x H full parse  e.g. "14\" DIA x 16\" D x 22\" H"
  - FIX 17: "or" multi-height e.g. '16\", 19\" or 24\" H' → first  (Soloist)
  - FIX 18: Slash multi-height e.g. "16/20/24\" H" → first  (Side Table One)
  - FIX 19: OAH → H  e.g. "32.5 OAH"  (Hohla Bar Cart)
  - FIX 20: H x D format  e.g. "22\" H X 17.25\" D"  (Zeno Side Table)
  - FIX 21: DIA range e.g. "10-15\" DIA" → take first  (Strata)
  - FIX 22: Standalone H  e.g. "21\" H +/-"  (Cascade)
  - FIX 23: T = Thickness → Depth  e.g. "28.5\" W x 28.5\" L x 1 ⅜\" T"
  - FIX 24: Unlabeled N x N x N D → Width x Height x Depth  (mirrors)
"""

import re
import time
import openpyxl
import requests
from bs4 import BeautifulSoup

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}

INPUT_FILE  = "desousahughes_Mirrors.xlsx"
OUTPUT_FILE = "desousahughes_Mirrors_final.xlsx"

DIMENSION_FIELDS = [
    "Weight", "Width", "Depth", "Diameter", "Length",
    "Height", "Seat Height", "Seat Depth", "Seat Width", "Arm Height",
]

MATERIAL_FIELDS = ["COM", "COL", "Cushion"]

COLUMNS = [
    "Product URL",
    "Image URL",
    "Product Name",
    "SKU",
    "Product Family Id",
    "Description",
    "Weight",
    "Width",
    "Depth",
    "Diameter",
    "Length",
    "Height",
    "Seat Width",
    "Seat Depth",
    "Seat Height",
    "Arm Height",
    "COM",
    "COL",
    "Cushion",
    "Finish",
    "Dimension",
    "Materials",
]


# ── File I/O ──────────────────────────────────────────────────────────────────

def read_input(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    return [dict(zip(header, row)) for row in rows[1:]]


# ── Accordion Helper ──────────────────────────────────────────────────────────

def get_accordion_text(soup: BeautifulSoup, label: str) -> str:
    """Find an accordion section by its bold label and return the answer text."""
    for section in soup.select("div.accordion div.section"):
        question = section.select_one("div.question p b")
        if question and question.get_text(strip=True).lower() == label.lower():
            answer_div = section.select_one("div.answer div.inner")
            if answer_div:
                for br in answer_div.find_all("br"):
                    br.replace_with("\n")
                text = answer_div.get_text(separator=" ").strip()
                lines = [l.strip() for l in text.splitlines()]
                lines = [l for l in lines if l]
                return "\n".join(lines)
    return ""


# ── Dimension Parser ──────────────────────────────────────────────────────────
#
#  Handles all these formats (takes FIRST size block only):
#
#  20.5" W x 23" D x 34" H 18" Seat Height
#  21" W x 25.75" D x 37.25" H  /  Seat: 18.5" H
#  Arm Height: 23.5"  /  Seat: 19.5" H x 19" D  /  22.5" W x 25.5" D x 34" H
#  30.5" W x 60" D x 34.5" H  /  Arm: 24" H  /  Seat: 18" H (seam)
#  33" W x 31" D x 27.5" H  /  Arm Height: 27.5"  /  Weight: approx 95 lbs
#  19" DIA x 3.5" D           ← DIA parsed before D
#  4 Seat - 108" W x 41.5" D x 29.75" H  /  Other Standard Sizes: 3 Seat ...
#  67" L x 17" D x 31.75" H   ← L x D x H (FIX 1)
#  16" W x 62" L x 31" H      ← W x L x H
#  56" L x 16" W x 34" H      ← L x W x H (FIX 5)
#  36" H x 48" L x 14" D      ← H x L x D (FIX 6)
#  18" D x 28" W x 25" H      ← D x W x H (FIX 7)
#  30.25" W x 26.25" H x 22" D ← W x H x D (FIX 8)
#  13 ⅜" W x 55 ⅛" L x 31 ⅞" H  ← vulgar fractions (FIX 3)
#  72" W x 20" D (37" Ext.) x 30" H  ← parenthetical (FIX 4)
#  Large - 48" W x 14" H       ← size prefix + W x H only (FIX 9)
#  48" W x 48" W x 14" H       ← W x W x H website typo (FIX 10)
#  W 29" x L 56" x H 14 ⅝"    ← label-before-number (FIX 11)
#  48" W x 24" D x 14 & 16" H  ← dual height (FIX 12)
#  48" DIA x 15.5" H             ← DIA x H (FIX 13)
#  15" Diameter x 23.5" H         ← "Diameter" word (FIX 14)
#  19" H x 12.5" DIA               ← H x DIA (FIX 15)
#  14" DIA x 16" D x 22" H         ← DIA x D x H (FIX 16)
#  16", 19" or 24" H               ← "or" multi-height (FIX 17)
#  16/20/24" H                     ← slash multi-height (FIX 18)
#  42" W x 18" D x 32.5 OAH        ← OAH → H (FIX 19)
#  22" H X 17.25" D                ← H x D (FIX 20)
#  10-15" DIA                      ← DIA range → first (FIX 21)
#  21" H +/-                       ← standalone H (FIX 22)
#  28.5" W x 28.5" L x 1 ⅜" T      ← T = Thickness → Depth (FIX 23)
#  24" x 36" x 2" D                  ← unlabeled W x H x D (FIX 24)
# ─────────────────────────────────────────────────────────────────────────────

# ── FIX 3: Vulgar fraction map ────────────────────────────────────────────────
VULGAR_FRACTIONS = {
    '⅛': 0.125,   # U+215B
    '¼': 0.25,    # U+00BC
    '⅜': 0.375,   # U+215C
    '½': 0.5,     # U+00BD
    '⅝': 0.625,   # U+215D
    '¾': 0.75,    # U+00BE
    '⅞': 0.875,   # U+215E
    '⅘': 0.8,     # U+2158
    '⅙': 0.1667,  # U+2159
    '⅔': 0.6667,  # U+2154
    '⅓': 0.3333,  # U+2153
}

def _convert_fractions(text: str) -> str:
    """
    Convert vulgar fractions (optionally preceded by a whole number) to decimals.
    e.g.  '13 ⅜"'  →  '13.375"'
          '55 ⅛"'  →  '55.125"'
          '⅞"'     →  '0.875"'
    """
    for frac_char, frac_val in VULGAR_FRACTIONS.items():
        # Whole number + space + fraction  e.g. "13 ⅜"
        text = re.sub(
            r'(\d+)\s*' + re.escape(frac_char),
            lambda m, v=frac_val: str(round(int(m.group(1)) + v, 4)),
            text
        )
        # Standalone fraction  e.g. "⅞"
        text = text.replace(frac_char, str(frac_val))
    return text


def _normalize_quotes(text: str) -> str:
    """Replace all quote-like characters with standard ASCII double quote.
    Also fixes mojibake patterns like 'â\n"' that appear when ″ is mis-decoded."""
    text = re.sub(r'â\s*"', '"', text)
    text = re.sub(r'â\s*\n\s*"', '"', text)
    for ch in ['\u2033', '\u201c', '\u201d', '\u02BA', '\uFF02']:
        text = text.replace(ch, '"')
    return text


def _strip_size_prefix(text: str) -> str:
    """
    Remove leading size labels before parsing.
    e.g. 'Large - 48" W x 14" H'  →  '48" W x 14" H'    (FIX 9)
         'Small - 30" W x 14" H'  →  '30" W x 14" H'
    """
    return re.sub(
        r'^\s*(?:Large|Small|Medium|XL|X-Large)\s*[-–]\s*',
        '', text, flags=re.IGNORECASE | re.MULTILINE
    )


def _fix_dual_height(text: str) -> str:
    """
    Replace 'N & M" H' with just 'N" H' (take first value).
    e.g. '14 & 16" H'  →  '14" H'                        (FIX 12)
    """
    return re.sub(
        r'([\d.]+)\s*&\s*[\d.]+(["\s]*H)',
        r'\1\2', text, flags=re.IGNORECASE
    )


def _strip_parentheticals(text: str) -> str:
    """
    Remove parenthetical notes that break dimension regexes.
    e.g. '72" W x 20" D (37" Ext.) x 30" H'
      →  '72" W x 20" D x 30" H'                          (FIX 4)
    """
    return re.sub(r'\s*\([^)]*\)', '', text)



def _normalize_diameter_word(text: str) -> str:
    """Normalise "Diameter" / "Diam" → "DIA".                           (FIX 14)
    e.g. '15" Diameter x 23.5" H'  →  '15" DIA x 23.5" H'
         '22" Diam x 24.5" H'      →  '22" DIA x 24.5" H'
    """
    return re.sub(r'\bDiam(?:eter)?\b', 'DIA', text, flags=re.IGNORECASE)


def _fix_multi_height(text: str) -> str:
    """Reduce multiple heights to the first value.                   (FIX 17 + 18)
    e.g. '16", 19" or 24" H'  →  '16" H'
         '16/20/24" H'         →  '16" H'
    """
    # "or" format: 16", 19" or 24" H
    text = re.sub(
        r'([\d.]+)"(?:\s*,?\s*[\d.]+")*\s*or\s*[\d.]+"\s*H',
        r'\1" H', text, flags=re.IGNORECASE
    )
    # slash format: 16/20/24" H
    text = re.sub(
        r'([\d.]+)(?:/[\d.]+)+"\s*H',
        r'\1" H', text, flags=re.IGNORECASE
    )
    return text


def _fix_oah(text: str) -> str:
    """Replace OAH (Overall Assembly Height) with '" H".               (FIX 19)
    e.g. '32.5 OAH'  →  '32.5" H'
    """
    return re.sub(r'([\d.]+)\s*OAH\b', r'\1" H', text, flags=re.IGNORECASE)


def _fix_dia_range(text: str) -> str:
    """For DIA ranges, take the first (smaller) value.                 (FIX 21)
    e.g. '10-15" DIA'  →  '10" DIA'
    """
    return re.sub(
        r'([\d.]+)-[\d.]+(["\s]*DIA)',
        r'\1\2', text, flags=re.IGNORECASE
    )


def _first_block(text: str) -> str:
    """Keep only the first size block, discard 'Other Standard Sizes' etc."""
    cutoff = re.split(
        r'\n\s*(?:Other Standard Sizes|[23456]\s*Seat\s*[-–])\s*',
        text, maxsplit=1, flags=re.IGNORECASE
    )
    return cutoff[0]


def parse_dimensions(raw: str) -> dict:
    result = {f: "" for f in DIMENSION_FIELDS}
    if not raw:
        return result

    # Apply all normalisations in order
    block = _first_block(raw)
    block = _convert_fractions(block)       # FIX 3: fractions first
    block = _normalize_quotes(block)        # FIX 2: then quotes / mojibake
    block = _strip_size_prefix(block)       # FIX 9: remove "Large - " prefix
    block = _fix_dual_height(block)         # FIX 12: "14 & 16" H" → "14" H"
    block = _normalize_diameter_word(block) # FIX 14: "Diameter"/"Diam" → DIA
    block = _fix_multi_height(block)        # FIX 17+18: multi-height → first
    block = _fix_oah(block)                 # FIX 19: OAH → H
    block = _fix_dia_range(block)           # FIX 21: DIA range → first value
    block = _strip_parentheticals(block)    # FIX 4: remove (37" Ext.) etc.

    # ── Diameter (must come before Depth) ────────────────────────────────────
    dia = re.search(r'([\d.]+)["\s]*DIA', block, re.IGNORECASE)
    if dia:
        result["Diameter"] = dia.group(1)
        # Depth after DIA  e.g. "19\" DIA x 3.5\" D"
        dep_after_dia = re.search(
            r'DIA\s*x\s*([\d.]+)["\s]*D\b', block, re.IGNORECASE
        )
        if dep_after_dia:
            result["Depth"] = dep_after_dia.group(1)
        # Height after DIA  e.g. "48\" DIA x 15.5\" H"      (FIX 13)
        # Also catches "DIA x D x H" full parse           (FIX 16)
        hgt_after_dia = re.search(
            r'DIA(?:\s*x\s*[\d.]+["\s]*D)?\s*x\s*([\d.]+)["\s]*H(?!\s*x|\w)',
            block, re.IGNORECASE
        )
        if hgt_after_dia:
            result["Height"] = hgt_after_dia.group(1)

    # ── W x D x H ────────────────────────────────────────────────────────────
    wxdxh = re.search(
        r'([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*D\s*x\s*([\d.]+)["\s]*H',
        block, re.IGNORECASE
    )
    if wxdxh:
        result["Width"]  = wxdxh.group(1)
        result["Depth"]  = wxdxh.group(2)
        result["Height"] = wxdxh.group(3)

    # ── W x L x H ────────────────────────────────────────────────────────────
    wxlxh = re.search(
        r'([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*L\s*x\s*([\d.]+)["\s]*H',
        block, re.IGNORECASE
    )
    if wxlxh:
        result["Width"]  = wxlxh.group(1)
        result["Length"] = wxlxh.group(2)
        result["Height"] = wxlxh.group(3)

    # ── L x D x H ────────────────────────────────────────────────────────────  FIX 1
    lxdxh = re.search(
        r'([\d.]+)["\s]*L\s*x\s*([\d.]+)["\s]*D\s*x\s*([\d.]+)["\s]*H',
        block, re.IGNORECASE
    )
    if lxdxh:
        result["Length"] = lxdxh.group(1)
        result["Depth"]  = lxdxh.group(2)
        result["Height"] = lxdxh.group(3)

    # ── L x W x H ────────────────────────────────────────────────────────────  FIX 5
    lxwxh = re.search(
        r'([\d.]+)["\s]*L\s*x\s*([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*H',
        block, re.IGNORECASE
    )
    if lxwxh:
        result["Length"] = lxwxh.group(1)
        result["Width"]  = lxwxh.group(2)
        result["Height"] = lxwxh.group(3)

    # ── H x L x D ────────────────────────────────────────────────────────────  FIX 6
    hxlxd = re.search(
        r'([\d.]+)["\s]*H\s*x\s*([\d.]+)["\s]*L\s*x\s*([\d.]+)["\s]*D',
        block, re.IGNORECASE
    )
    if hxlxd:
        result["Height"] = hxlxd.group(1)
        result["Length"] = hxlxd.group(2)
        result["Depth"]  = hxlxd.group(3)

    # ── D x W x H ────────────────────────────────────────────────────────────  FIX 7
    #  e.g. "18" D x 28" W x 25" H"  (Troscan Catena / Chelsea)
    dxwxh = re.search(
        r'([\d.]+)["\s]*D\s*x\s*([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*H',
        block, re.IGNORECASE
    )
    if dxwxh:
        result["Depth"]  = dxwxh.group(1)
        result["Width"]  = dxwxh.group(2)
        result["Height"] = dxwxh.group(3)

    # ── W x H x D ────────────────────────────────────────────────────────────  FIX 8
    #  e.g. "30.25" W x 26.25" H x 22" D"  (Quintus Tuy)
    wxhxd = re.search(
        r'([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*H\s*x\s*([\d.]+)["\s]*D',
        block, re.IGNORECASE
    )
    if wxhxd:
        result["Width"]  = wxhxd.group(1)
        result["Height"] = wxhxd.group(2)
        result["Depth"]  = wxhxd.group(3)

    # ── W x W x H ────────────────────────────────────────────────────────────  FIX 10
    #  Website typo: "48" W x 48" W x 14" H" — treat second W as Depth
    wxwxh = re.search(
        r'([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*H',
        block, re.IGNORECASE
    )
    if wxwxh:
        result["Width"]  = wxwxh.group(1)
        result["Depth"]  = wxwxh.group(2)
        result["Height"] = wxwxh.group(3)

    # ── Label-before-number: "W 29" x L 56" x H 14"" ─────────────────────── FIX 11
    #  e.g. "W 29" x L 56" x H 14 ⅝""  (Alexander Lamont Forge)
    label_before = re.search(
        r'W\s*([\d.]+)"\s*x\s*L\s*([\d.]+)"\s*x\s*H\s*([\d.]+)"',
        block, re.IGNORECASE
    )
    if label_before:
        result["Width"]  = label_before.group(1)
        result["Length"] = label_before.group(2)
        result["Height"] = label_before.group(3)

    # ── W x H only (no Depth in raw) ─────────────────────────────────────── FIX 9
    #  e.g. "48" W x 14" H"  (John Pomp Caldera / Warp — circular stone tables)
    #  Only fires if W and H are still empty after all 3-dim patterns above
    if not result["Width"] and not result["Height"]:
        wxh = re.search(
            r'([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*H(?!\s*x)',
            block, re.IGNORECASE
        )
        if wxh:
            result["Width"]  = wxh.group(1)
            result["Height"] = wxh.group(2)
    # ── H x DIA ──────────────────────────────────────────────────────────── FIX 15
    #  e.g. "19" H x 12.5" DIA"  (Quintus Alta Spot, Troscan Nave)
    if not result["Diameter"] and not result["Height"]:
        hxdia = re.search(
            r'([\d.]+)["\s]*H\s*x\s*([\d.]+)["\s]*DIA',
            block, re.IGNORECASE
        )
        if hxdia:
            result["Height"]   = hxdia.group(1)
            result["Diameter"] = hxdia.group(2)

    # ── H x D ────────────────────────────────────────────────────────────── FIX 20
    #  e.g. "22" H X 17.25" D"  (Kimberly Denman Zeno)
    if not result["Height"] and not result["Depth"]:
        hxd = re.search(
            r'([\d.]+)["\s]*H\s*[xX]\s*([\d.]+)["\s]*D(?!\s*[xX]|\w)',
            block, re.IGNORECASE
        )
        if hxd:
            result["Height"] = hxd.group(1)
            result["Depth"]  = hxd.group(2)

    # ── W x L x T ────────────────────────────────────────────────────────── FIX 23
    #  T = Thickness → stored as Depth
    #  e.g. "28.5" W x 28.5" L x 1 ⅜" T"  (Amulet Round, Volta)
    wxlxt = re.search(
        r'([\d.]+)["\s]*W\s*x\s*([\d.]+)["\s]*L\s*x\s*([\d.]+)["\s]*T\b',
        block, re.IGNORECASE
    )
    if wxlxt:
        result["Width"]  = wxlxt.group(1)
        result["Length"] = wxlxt.group(2)
        result["Depth"]  = wxlxt.group(3)

    # ── N x N x N D (unlabeled W x H x D) ───────────────────────────────── FIX 24
    #  e.g. "24" x 36" x 2" D"  (Dovetail, Leather Mirror)
    #  Only fires when Width still empty (no other pattern matched)
    if not result["Width"] and not result["Height"]:
        nxnxnd = re.search(
            r'([\d.]+)["\s]*x\s*([\d.]+)["\s]*x\s*([\d.]+)["\s]*D\b',
            block, re.IGNORECASE
        )
        if nxnxnd:
            result["Width"]  = nxnxnd.group(1)
            result["Height"] = nxnxnd.group(2)
            result["Depth"]  = nxnxnd.group(3)

    # ── Standalone Height  e.g. "21" H +/-" ──────────────────────────────── FIX 22
    #  Only fires when nothing else matched
    if not result["Height"]:
        sh_alone = re.search(r'([\d.]+)["\s]*H\b', block, re.IGNORECASE)
        if sh_alone:
            result["Height"] = sh_alone.group(1)

    if not result["Length"]:
        ln = re.search(r'Length\s*[:\-]?\s*([\d.]+)', block, re.IGNORECASE)
        if ln:
            result["Length"] = ln.group(1)

    # ── Weight ───────────────────────────────────────────────────────────────
    wt = re.search(r'Weight\s*[:\-]?\s*(?:approx\.?\s*)?([\d.]+)', block, re.IGNORECASE)
    if wt:
        result["Weight"] = wt.group(1)

    # ── Seat Height ──────────────────────────────────────────────────────────
    sh = (
        re.search(r'Seat\s*Height\s*[:\-]?\s*([\d.]+)', block, re.IGNORECASE)
        or re.search(r'Seat\s*[:\-]\s*([\d.]+)["\s]*H', block, re.IGNORECASE)
        or re.search(r'([\d.]+)["\s]*Seat\s*Height', block, re.IGNORECASE)
    )
    if sh:
        result["Seat Height"] = sh.group(1)

    # ── Seat Depth ───────────────────────────────────────────────────────────
    sd = re.search(
        r'Seat\s*[:\-]\s*[\d.]+["\s]*H\s*x\s*([\d.]+)["\s]*D',
        block, re.IGNORECASE
    )
    if sd:
        result["Seat Depth"] = sd.group(1)

    # ── Seat Width ───────────────────────────────────────────────────────────
    sw = re.search(r'Seat\s*Width\s*[:\-]?\s*([\d.]+)', block, re.IGNORECASE)
    if sw:
        result["Seat Width"] = sw.group(1)

    # ── Arm Height ───────────────────────────────────────────────────────────
    ah = (
        re.search(r'Arm\s*Height\s*[:\-]?\s*([\d.]+)', block, re.IGNORECASE)
        or re.search(r'Arm\s*[:\-]\s*([\d.]+)["\s]*H', block, re.IGNORECASE)
    )
    if ah:
        result["Arm Height"] = ah.group(1)

    return result


# ── Materials Parser ──────────────────────────────────────────────────────────

def parse_materials(raw: str) -> dict:
    """
    Handles all COM/COL formats:
      Format 1: COM/COL[label]: N yds / M sq ft
                e.g. 'COM/COL: 44.5 yds / 801 Sq Ft'
                     'COM/COLComplete: 14 yds / 260 sq ft'
      Format 2: COM[label]: ...text... N yds  +  COL[label]: ...text... N sq ft
                e.g. 'COM:Headboard 2.5 yds...COL:Headboard 45 sq ft'
                     'COM Seat: 0.5 yd  COL Seat: 10 sq ft'
                     'COM: 2 yds  COL 54 sq ft'   ← COL with no colon
      Format 3: Simple fallback  COM: N  /  COL: N
    Numbers only extracted when followed by yds/sq ft — avoids false matches.
    """
    result = {f: "" for f in MATERIAL_FIELDS}
    if not raw:
        return result

    # ── Format 1: COM/COL[optional label]: N yds / M sq ft ───────────────
    combined = re.search(
        r'COM\s*/\s*COL\s*\w*\s*[:\-]\s*([\d.]+)\s*yds?\s*/\s*([\d.]+)',
        raw, re.IGNORECASE
    )
    if combined:
        result["COM"] = combined.group(1)
        result["COL"] = combined.group(2)
        counts = re.findall(r'(\d+)\s+\w+\s+Cushion', raw, re.IGNORECASE)
        if counts:
            result["Cushion"] = ",".join(counts)
        return result

    # ── COM section ───────────────────────────────────────────────────────
    # Capture text after COM (+ optional label word + optional colon)
    # up to COL or end of string, then grab first "N yds" number.
    # e.g. 'COM: 2 yds'         → '2'
    #      'COM Seat: 0.5 yd'   → '0.5'
    #      'COM: yds'           → nothing (no number)
    com_section = re.search(
        r'COM\b\s*\w*\s*[:\-]?(.*?)(?=\bCOL\b|\Z)',
        raw, re.IGNORECASE | re.DOTALL
    )
    if com_section:
        com_val = re.search(r'([\d.]+)\s*yds?', com_section.group(1), re.IGNORECASE)
        if com_val:
            result["COM"] = com_val.group(1)

    # ── COL section ───────────────────────────────────────────────────────
    # After COL, skip one optional «word + colon» OR just whitespace,
    # then find the first "N sq ft".
    # e.g. 'COL: 45 sq ft'      → '45'
    #      'COL 54 sq ft'        → '54'   ← no colon
    #      'COL Seat: 10 sq ft'  → '10'   ← label+colon
    #      'COL not available'   → nothing (no sq ft)
    col_section = re.search(
        r'COL\b(?:\s*\w+\s*[:\-])?\s*(.*)',
        raw, re.IGNORECASE | re.DOTALL
    )
    if col_section:
        col_val = re.search(r'([\d.]+)\s*sq\s*ft', col_section.group(1), re.IGNORECASE)
        if col_val:
            result["COL"] = col_val.group(1)

    counts = re.findall(r'(\d+)\s+\w+\s+Cushion', raw, re.IGNORECASE)
    if counts:
        result["Cushion"] = ",".join(counts)

    return result


    # ── Format 1: COM/COL[optional label]: N yds / M sq ft ───────────────
    # Handles both 'COM/COL: 14 yds / 260 sq ft'
    # and        'COM/COLComplete: 14 yds / 260 sq ft' (no space before label)
    combined = re.search(
        r'COM\s*/\s*COL\s*\w*\s*[:\-]\s*([\d.]+)\s*yds?\s*/\s*([\d.]+)',
        raw, re.IGNORECASE
    )
    if combined:
        result["COM"] = combined.group(1)
        result["COL"] = combined.group(2)
    else:
        # ── Format 2: COM: ...label... N yds  /  COL: ...label... N sq ft ─
        # Extract the COM section (everything between COM: and COL:),
        # then find the FIRST "N yds" inside it.
        com_section = re.search(
            r'COM\s*[:\-](.*?)(?=COL\s*[:\-]|\Z)',
            raw, re.IGNORECASE | re.DOTALL
        )
        if com_section:
            com_val = re.search(r'([\d.]+)\s*yds?', com_section.group(1), re.IGNORECASE)
            if com_val:
                result["COM"] = com_val.group(1)

        col_section = re.search(r'COL\s*[:\-](.*)', raw, re.IGNORECASE | re.DOTALL)
        if col_section:
            col_val = re.search(r'([\d.]+)\s*sq\s*ft', col_section.group(1), re.IGNORECASE)
            if col_val:
                result["COL"] = col_val.group(1)

        # ── Format 3: simple fallback COM: N  /  COL: N ──────────────────
        if not result["COM"]:
            com = re.search(r'COM\s*[:\-]\s*([\d.]+)', raw, re.IGNORECASE)
            if com:
                result["COM"] = com.group(1)
        if not result["COL"]:
            col = re.search(r'COL\s*[:\-]\s*([\d.]+)', raw, re.IGNORECASE)
            if col:
                result["COL"] = col.group(1)

    counts = re.findall(r'(\d+)\s+\w+\s+Cushion', raw, re.IGNORECASE)
    if counts:
        result["Cushion"] = ",".join(counts)

    return result


# ── Page Scraper ──────────────────────────────────────────────────────────────

def scrape_product_page(url: str) -> dict:
    empty = {
        "description": "", "dimension": "", "materials": "", "finish": "",
        **{f: "" for f in DIMENSION_FIELDS},
        **{f: "" for f in MATERIAL_FIELDS},
    }
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        resp.encoding = 'utf-8'          # FIX 2: force UTF-8, prevent mojibake
    except Exception as e:
        print(f"  ⚠️  Failed: {url} — {e}")
        return empty

    soup = BeautifulSoup(resp.text, "html.parser")

    # ── Description ──────────────────────────────────────────────────────────
    description = ""
    desc_div = soup.select_one("div.desc")
    if desc_div:
        for p in desc_div.find_all("p"):
            text = p.get_text(strip=True)
            if len(text) > 40 and "Price, finish" not in text:
                description = text
                break

    # ── Raw accordion text ────────────────────────────────────────────────────
    dimension_raw = get_accordion_text(soup, "Dimensions")
    materials_raw = get_accordion_text(soup, "Materials")
    finish        = get_accordion_text(soup, "Finishes")

    # ── Parse into sub-fields ─────────────────────────────────────────────────
    dim = parse_dimensions(dimension_raw)
    mat = parse_materials(materials_raw)

    return {
        "description": description,
        "dimension":   dimension_raw,
        "materials":   materials_raw,
        "finish":      finish,
        **dim,
        **mat,
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"📂  Reading: {INPUT_FILE}")
    products = read_input(INPUT_FILE)
    print(f"   {len(products)} products found\n")

    results = []

    for i, p in enumerate(products, 1):
        url  = p.get("Product URL", "")
        name = p.get("Product Name", "")
        print(f"[{i}/{len(products)}] {name}\n   → {url}")

        extra = scrape_product_page(url)

        results.append({
            "Product URL":       url,
            "Image URL":         p.get("Image URL", ""),
            "Product Name":      name,
            "SKU":               p.get("SKU", ""),
            "Product Family Id": name,
            "Description":       extra["description"],
            "Dimension":         extra["dimension"],
            "Weight":            extra["Weight"],
            "Width":             extra["Width"],
            "Depth":             extra["Depth"],
            "Diameter":          extra["Diameter"],
            "Length":            extra["Length"],
            "Height":            extra["Height"],
            "Seat Height":       extra["Seat Height"],
            "Seat Depth":        extra["Seat Depth"],
            "Seat Width":        extra["Seat Width"],
            "Arm Height":        extra["Arm Height"],
            "Materials":         extra["materials"],
            "COM":               extra["COM"],
            "COL":               extra["COL"],
            "Cushion":           extra["Cushion"],
            "Finish":            extra["finish"],
        })

        time.sleep(0.5)

    # ── Write Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consoles"
    ws.append(COLUMNS)

    for row in results:
        ws.append([row.get(col, "") for col in COLUMNS])

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Saved {len(results)} rows → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()