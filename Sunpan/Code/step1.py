import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================
# INPUT SYSTEM (from CODE-B)
# =========================

CATEGORIES = {
    "Nightstands": [
        "https://sunpan.com/collections/nightstands"
    ],
    "Coffee & Cocktail Tables": [
        "https://sunpan.com/collections/coffee-tables"
    ],

    "Side & End Tables": [
        "https://sunpan.com/collections/end-side-tables"
    ],

    "Dining Tables": [
        "https://sunpan.com/collections/dining-tables"
    ],
    "Consoles": [
        "https://sunpan.com/collections/console-tables"
    ],

    "Beds & Headboards": [
        "https://sunpan.com/collections/beds"
    ],

    "Desks": [
        "https://sunpan.com/collections/desks"
    ],

    "Cabinets": [
        "https://sunpan.com/collections/storage-consoles-cabinets"
    ],

    "Bookcases": [
        "https://sunpan.com/collections/bookcases-shelving"
    ],

    "Dressers & Chests": [
        "https://sunpan.com/collections/dressers"
    ],

    "Bar Carts": [
        "https://sunpan.com/collections/bar-carts-and-cabinets"
    ],

    "Dining Chairs": [
        "https://sunpan.com/collections/dining-chairs"
    ],

    "Bar Stools": [
        "https://sunpan.com/collections/counter-stools"
    ],

    "Sofas & Loveseats": [
        "https://sunpan.com/collections/sofas"
    ],

    "Sectionals": [
        "https://sunpan.com/collections/sofa-chaises-sectionals"
    ],

    "Lounge Chairs": [
        "https://sunpan.com/collections/armchairs-lounge-chairs"
    ],

    "Ottomans": [
        "https://sunpan.com/collections/ottomans?&sort_by=title-ascending&filter.p.product_type=Ottomans"
    ],

    "Benches": [
        "https://sunpan.com/collections/ottomans?&sort_by=title-ascending&filter.p.product_type=Benches"
    ],

    "Desk Chairs": [
        "https://sunpan.com/collections/office-chairs"
    ],

    "Chandeliers": [
        "https://sunpan.com/collections/chandeliers"
    ],

    "Pendants": [
        "https://sunpan.com/collections/pendants"
    ],

    "Sconces": [
        "https://sunpan.com/collections/sconce"
    ],

    "Table Lamps": [
        "https://sunpan.com/collections/table-lamps"
    ],

    "Floor Lamps": [
        "https://sunpan.com/collections/floor-lamps"
    ],

    "Mirrors": [
        "https://sunpan.com/collections/floor-mirrors",
        "https://sunpan.com/collections/wall-mirrors"
    ],

    "Wall Decors": [
        "https://sunpan.com/collections/art"
    ],

    "Rugs": [
        "https://sunpan.com/collections/indoor-rugs"
    ],
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "sunpan_master.xlsx")
category_output_file = os.path.join(script_dir, "Sunpan.xlsx")


# =========================
# SCRAPING LOGIC (IMPROVED WITH AUTO-PAGINATION)
# =========================

def get_total_pages(driver, wait):
    """
    Automatically detect total number of pages from pagination
    """
    try:
        # Wait for pagination to load
        time.sleep(2)

        # Try to find all pagination links
        pagination_items = driver.find_elements(
            By.CSS_SELECTOR,
            "nav.pagination ul.pagination__list li a.pagination__item"
        )

        if not pagination_items:
            return 1

        # Extract page numbers from pagination links
        page_numbers = []
        for item in pagination_items:
            try:
                text = item.text.strip()
                if text.isdigit():
                    page_numbers.append(int(text))
            except:
                continue

        if page_numbers:
            total_pages = max(page_numbers)
            print(f"📄 Total pages detected: {total_pages}")
            return total_pages
        else:
            return 1

    except Exception as e:
        print(f"⚠️ Could not detect pagination: {e}")
        return 1


def scrape_collection(base_url):
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 15)

    all_data = []

    # First, load page 1 to detect total pages
    print(f"🔍 Loading: {base_url}")
    driver.get(base_url)
    time.sleep(3)

    # Auto-detect total pages
    total_pages = get_total_pages(driver, wait)

    # Scrape all pages
    for page in range(1, total_pages + 1):
        if page > 1:
            url = f"{base_url}?page={page}" if "?" not in base_url else f"{base_url}&page={page}"
            print(f"🔍 Loading: {url}")
            driver.get(url)
            time.sleep(3)
        else:
            url = base_url

        # Scroll to load all products
        last_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scroll_attempts = 5

        while scroll_attempts < max_scroll_attempts:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
            scroll_attempts += 1

        try:
            cards = wait.until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, "div.card-wrapper.product-card-wrapper")
                )
            )
        except:
            cards = []

        print(f"📦 Page {page}/{total_pages}: {len(cards)} products found")

        for card in cards:
            try:
                a_tag = card.find_element(By.XPATH, "./ancestor::a[1]")
                product_url = a_tag.get_attribute("href").strip()
            except:
                product_url = ""

            try:
                sku = card.find_element(
                    By.CSS_SELECTOR, "div.product__sku span.sku"
                ).text.strip()
            except:
                sku = ""

            try:
                product_name = card.find_element(
                    By.CSS_SELECTOR, "h3.card__heading.h3"
                ).text.strip()
            except:
                product_name = ""

            try:
                image_url = card.find_element(
                    By.CSS_SELECTOR, "img.card-product-image"
                ).get_attribute("src").strip()
            except:
                image_url = ""

            all_data.append({
                "Product URL": product_url,
                "Image URL": image_url,
                "Product Name": product_name,
                "SKU": sku
            })

    driver.quit()
    print(f"✅ Total products scraped: {len(all_data)}")
    return all_data


# =========================
# OUTPUT SYSTEM (from CODE-B)
# =========================

def build_master_dataframe():
    dfs = []

    for category, links in CATEGORIES.items():
        print(f"\n{'=' * 60}")
        print(f"🏷️  Scraping Category: {category}")
        print(f"{'=' * 60}")

        cat_products = []
        for link in links:
            data = scrape_collection(link)
            cat_products.extend(data)

        if cat_products:
            df = pd.DataFrame(cat_products)
            df.drop_duplicates(subset=["Product URL"], inplace=True)
            df.insert(0, "Category", category)
            dfs.append(df)
            print(f"✅ {category}: {len(df)} unique products")

    if not dfs:
        cols = ["Category", "Product URL", "Image URL", "Product Name", "SKU"]
        master = pd.DataFrame(columns=cols)
        master.to_excel(master_output_file, index=False)
        return master

    master = pd.concat(dfs, ignore_index=True)
    master.to_excel(master_output_file, index=False)
    print(f"\n✅ Master file saved: {master_output_file}")
    print(f"📊 Total products: {len(master)}")
    return master


def build_category_wise_workbook(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for category in CATEGORIES.keys():
        df_cat = df[df["Category"] == category].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=category)

        ws["A1"] = "Brand"
        ws["B1"] = "Sunpan"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(CATEGORIES[category])
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start_row = 4
        for col_idx, col in enumerate(df_cat.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col).font = bold

        for r_idx, row in enumerate(df_cat.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        headers = {
            ws.cell(row=start_row, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }

        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(category_output_file)
    print(f"✅ Category workbook saved: {category_output_file}")


# =========================
# MAIN (from CODE-B)
# =========================

def main():
    print("\n" + "=" * 60)
    print("🚀 SUNPAN SCRAPER - AUTO PAGINATION")
    print("=" * 60 + "\n")

    df = build_master_dataframe()

    if df.empty:
        wb = Workbook()
        wb.save(category_output_file)
        print("⚠️ No data found")
        return

    build_category_wise_workbook(df)

    print("\n" + "=" * 60)
    print("✅ SCRAPING COMPLETED SUCCESSFULLY!")
    print("=" * 60)


if __name__ == "__main__":
    main()