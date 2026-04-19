from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
import pandas as pd
import time
import os
import re
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# =========================
# File Path
# =========================
script_path = os.path.dirname(os.path.abspath(__file__))
input_excel = os.path.join(script_path, "Sunpan.xlsx")
output_excel = os.path.join(script_path, "Sunpan_final.xlsx")

# Columns in final order (with Index and Category added)
final_columns = [
    "Index", "Category", "Product URL", "Image URL", "Product Name", "SKU",
    "Product Family Id", "Color", "Weight", "Width", "Depth", "Diameter",
    "Length", "Height", "Arm Height", "Seat Height", "Seat Width", "Seat Depth",
    "Finish", "Base", "Seat", "Cushion", "Wattage"
]

# =========================
# Selenium Setup
# =========================
chrome_options = Options()
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.page_load_strategy = "eager"

driver = webdriver.Chrome(options=chrome_options)
driver.set_page_load_timeout(300)
wait = WebDriverWait(driver, 15)


# =========================
# Helper Functions
# =========================
def open_accordion(title_text):
    try:
        summary_tag = driver.find_element(
            By.XPATH,
            f"//h2[contains(text(),'{title_text}')]/ancestor::summary"
        )
        toggle_svg = summary_tag.find_element(By.CSS_SELECTOR, "svg.icon-caret")
        driver.execute_script("arguments[0].click();", toggle_svg)
        time.sleep(1)
    except:
        print(f"⚠ Could not open accordion: {title_text}")


def extract_table(table_css):
    try:
        table = driver.find_element(By.CSS_SELECTOR, table_css)
        rows = table.find_elements(By.TAG_NAME, "tr")
        out_lines = []
        for tr in rows:
            tds = tr.find_elements(By.TAG_NAME, "td")
            if len(tds) == 2:
                label = tds[0].text.strip()
                value = tds[1].text.strip()
                out_lines.append(f"{label}: {value}")
        return "\n".join(out_lines)
    except:
        return ""


def parse_dimensions(dim_text):
    width = depth = height = diameter = length = ""
    parts = re.findall(r"([\d\.]+)\s*([WDHLDIA]+)", dim_text, re.IGNORECASE)
    for val, key in parts:
        key = key.lower()
        if key == "w" and not width:
            width = val
        elif key == "d" and not depth:
            depth = val
        elif key == "h" and not height:
            height = val
        elif "dia" in key and not diameter:
            diameter = val
        elif key == "l" and not length:
            length = val
    return width, depth, height, diameter, length


def extract_from_dimensions_table(dim_table_text):
    data = {}
    for line in dim_table_text.split("\n"):
        if ":" not in line:
            continue
        key, val = line.split(":", 1)
        key = key.strip().lower()
        val = val.strip()

        # weight line (Net / Carton / Gross) - remove "Ib" or "lb"
        if "net weight" in key:
            data["Weight"] = re.sub(r'\s*(lb|Ib)\s*$', '', val, flags=re.IGNORECASE).strip()
        elif "carton weight" in key and "Weight" not in data:
            data["Weight"] = re.sub(r'\s*(lb|Ib)\s*$', '', val, flags=re.IGNORECASE).strip()
        elif "gross weight" in key and "Weight" not in data:
            data["Weight"] = re.sub(r'\s*(lb|Ib)\s*$', '', val, flags=re.IGNORECASE).strip()

        elif "overall dimensions" in key:
            w, d, h, dia, l = parse_dimensions(val)
            data.update({
                "Width": w, "Depth": d, "Height": h,
                "Diameter": dia, "Length": l
            })
        elif "arm height" in key:
            data["Arm Height"] = re.sub(r'\s*in\s*$', '', val, flags=re.IGNORECASE).strip()
        elif "seat height" in key:
            data["Seat Height"] = re.sub(r'\s*in\s*$', '', val, flags=re.IGNORECASE).strip()
        elif "seat width" in key:
            data["Seat Width"] = re.sub(r'\s*in\s*$', '', val, flags=re.IGNORECASE).strip()
        elif "seat depth" in key:
            data["Seat Depth"] = re.sub(r'\s*in\s*$', '', val, flags=re.IGNORECASE).strip()
    return data


def extract_from_construction_table(cons_table_text):
    data = {}
    for line in cons_table_text.split("\n"):
        if ":" not in line:
            continue
        key, val = line.split(":", 1)
        key = key.strip().lower()
        val = val.strip()
        # Finish
        if "material finish" in key:
            data["Finish"] = val
        # Base
        elif "base / legs" in key or "base finish" in key:
            if "Base" in data:
                data["Base"] += ", " + val
            else:
                data["Base"] = val
        # Seat
        elif "seat construction" in key or "seat cushion" in key or "cover fabric" in key:
            if "Seat" in data:
                data["Seat"] += ", " + val
            else:
                data["Seat"] = val
        # Cushion
        elif "back cushion" in key:
            data["Cushion"] = val
        # Wattage from construction: "Max Bulb Wattage"
        elif "max bulb wattage" in key:
            data["Wattage"] = val
    return data


def extract_color():
    # 1st priority: selected-option span (original logic)
    try:
        color_span = driver.find_element(By.CSS_SELECTOR, "span.selected-option-value--color")
        txt = color_span.text.strip()
        if txt:
            return txt
    except:
        pass

    # fallback: currently selected radio input
    try:
        selected_radio = driver.find_element(
            By.CSS_SELECTOR,
            "fieldset.js.product-form__input input[type='radio'][name='Color']:checked"
        )
        val = selected_radio.get_attribute("value")
        if val:
            return val.strip()
    except:
        pass

    return ""


def get_current_variant():
    """
    Return (variant_dict or None) for the currently selected variant.
    We try by ?variant=id first; fallback by color name (option1).
    """
    try:
        script_tag = driver.find_element(
            By.CSS_SELECTOR, "script[type='application/json'][data-product]"
        )
        prod_json = json.loads(script_tag.get_attribute("innerHTML"))
    except Exception as e:
        print(f"⚠ Could not load product JSON: {e}")
        return None

    variants = prod_json.get("variants", [])
    if not variants:
        return None

    # try via variant id from URL
    url = driver.current_url
    variant_id = None
    if "variant=" in url:
        try:
            variant_id = int(url.split("variant=")[1].split("&")[0])
        except:
            variant_id = None

    if variant_id:
        for v in variants:
            if v.get("id") == variant_id:
                return v

    # fallback: match by color / option1
    selected_color = extract_color()
    if selected_color:
        for v in variants:
            if v.get("option1") == selected_color or v.get("title") == selected_color:
                return v

    return None


def extract_image_url():
    """
    Always try to return CURRENT SELECTED VARIANT IMAGE.
    """
    # 1) Active media in gallery
    try:
        active_img = driver.find_element(
            By.CSS_SELECTOR,
            ".product__media.is-active img, .product__media-item.is-active img"
        )
        src = active_img.get_attribute("src")
        if src:
            if src.startswith("//"):
                src = "https:" + src
            return src
    except:
        pass

    # 2) Active thumbnail
    try:
        active_thumb = driver.find_element(
            By.CSS_SELECTOR,
            ".thumbnail.is-active img, .product__media-toggle.is-active img"
        )
        src = active_thumb.get_attribute("src")
        if src:
            if src.startswith("//"):
                src = "https:" + src
            return src
    except:
        pass

    # 3) From current variant JSON image
    try:
        variant = get_current_variant()
        if variant:
            image = variant.get("featured_image") or variant.get(
                "featured_media", {}
            ).get("preview_image", {})
            if isinstance(image, dict):
                src = image.get("src")
                if src:
                    if src.startswith("//"):
                        src = "https:" + src
                    return src
    except:
        pass

    # 4) Fallback: og:image
    try:
        meta = driver.find_element(By.CSS_SELECTOR, "meta[property='og:image']")
        content = meta.get_attribute("content")
        if content:
            return content.strip()
    except:
        pass

    return ""


def fill_from_variant_meta(row_index):
    """
    SKU + dimensions + construction window.variantMetafields থেকে আনার চেষ্টা।
    কিছু পেলে True return, নাহলে False.
    """
    filled_any = False

    variant = get_current_variant()
    if not variant:
        return False

    sku = variant.get("sku") or ""
    if sku:
        df.at[row_index, "SKU"] = sku
    else:
        return False

    # get window.variantMetafields from page
    try:
        vm = driver.execute_script("return window.variantMetafields || {};")
    except Exception as e:
        print(f"⚠ Could not read window.variantMetafields: {e}")
        return False

    meta = vm.get(sku)
    if not meta:
        return False

    details = meta.get("product_details", {})
    dimensions = details.get("dimensions", {}) or {}
    construction = details.get("construction", {}) or {}

    # Dimensions - clean up units
    if dimensions:
        lines = []
        for k, v in dimensions.items():
            if v:
                # Remove "lb" or "Ib" from weight-related fields
                if "weight" in k.lower():
                    v = re.sub(r'\s*(lb|Ib)\s*$', '', str(v), flags=re.IGNORECASE).strip()
                # Remove "in" from dimension fields
                elif any(dim in k.lower() for dim in ["height", "width", "depth", "arm", "seat"]):
                    v = re.sub(r'\s*in\s*$', '', str(v), flags=re.IGNORECASE).strip()
                lines.append(f"{k}: {v}")

        dim_text = "\n".join(lines)
        dim_data = extract_from_dimensions_table(dim_text)
        for k, v in dim_data.items():
            if k in final_columns and v:
                df.at[row_index, k] = v
                filled_any = True

    # Construction
    if construction:
        lines = [f"{k}: {v}" for k, v in construction.items() if v]
        cons_text = "\n".join(lines)
        cons_data = extract_from_construction_table(cons_text)
        for k, v in cons_data.items():
            if k in final_columns and v:
                df.at[row_index, k] = v
                filled_any = True

    return filled_any


# =========================
# Load Excel with Brand/Link Structure
# =========================
print("📂 Loading input Excel...")

# Load workbook to preserve Brand/Link structure
wb_input = load_workbook(input_excel)

# Process each sheet
all_sheets_data = []
sheet_metadata = {}  # Store Brand and Link for each sheet
category_order = []  # Preserve sheet order

for sheet_name in wb_input.sheetnames:
    ws = wb_input[sheet_name]

    # Extract Brand and Link from rows 1-2
    brand = ws["B1"].value or "Sunpan"
    link = ws["B2"].value or ""

    sheet_metadata[sheet_name] = {
        "brand": brand,
        "link": link
    }
    category_order.append(sheet_name)

    # Read data starting from row 4 (headers) and row 5 (data)
    df_sheet = pd.read_excel(input_excel, sheet_name=sheet_name, header=3)

    # Force Category column to be the sheet name
    df_sheet["Category"] = sheet_name

    all_sheets_data.append(df_sheet)

# Combine all sheets
df = pd.concat(all_sheets_data, ignore_index=True)

# Ensure final columns exist
for col in final_columns:
    if col not in df.columns:
        df[col] = ""

# Copy Product Name to Product Family Id
df["Product Family Id"] = df["Product Name"]

# Re-index per category
for category in category_order:
    mask = df["Category"] == category
    df.loc[mask, "Index"] = range(1, mask.sum() + 1)

print(f"✅ Loaded {len(df)} products from {len(category_order)} categories")


# =========================
# Helper: scrape 1 configuration (1 selected color / variant)
# =========================
def scrape_configuration(row_index):
    # Color
    df.at[row_index, "Color"] = extract_color()

    # First try: variantMetafields (per-variant data)
    used_variant_meta = fill_from_variant_meta(row_index)

    # If variantMetafields did NOT give us anything, fallback to old HTML logic
    if not used_variant_meta:
        # Dimensions
        open_accordion("Dimensions")
        dim_table_text = extract_table("table.tab-dimensions__table")
        if dim_table_text:
            dim_data = extract_from_dimensions_table(dim_table_text)
            for k, v in dim_data.items():
                if k in final_columns and v:
                    df.at[row_index, k] = v

        # Construction
        open_accordion("Construction")
        cons_table_text = extract_table("table.tab-construction__table")
        if cons_table_text:
            cons_data = extract_from_construction_table(cons_table_text)
            for k, v in cons_data.items():
                if k in final_columns and v:
                    df.at[row_index, k] = v

    # Variant-wise URLs (always)
    df.at[row_index, "Product URL"] = driver.current_url
    df.at[row_index, "Image URL"] = extract_image_url()

    print(f"  ✓ Saved row {row_index} ({df.at[row_index, 'Color']})")


# =========================
# Auto-save function
# =========================
def auto_save_progress():
    """Save current progress to Excel"""
    try:
        temp_df = df[final_columns].copy()
        # Re-index per category
        for category in category_order:
            mask = temp_df["Category"] == category
            if mask.any():
                temp_df.loc[mask, "Index"] = range(1, mask.sum() + 1)

        temp_df.to_excel(output_excel.replace("_final", "_progress"), index=False)
        print(f"  💾 Auto-saved progress to {output_excel.replace('_final', '_progress')}")
    except Exception as e:
        print(f"  ⚠️ Auto-save failed: {e}")


# =========================
# MAIN LOOP
# =========================
original_len = len(df)
rows_added = 0  # Track how many variation rows we've added
products_scraped = 0  # Counter for auto-save

idx = 0
while idx < len(df):
    row = df.iloc[idx]
    url = str(row["Product URL"]).strip()
    name = str(row["Product Name"])
    category = str(row.get("Category", ""))

    if not url or url.lower() == "nan":
        idx += 1
        continue

    print(f"\n[{idx + 1}/{len(df)}] 🏷️ {category} | {name}")

    # Open Product URL
    try:
        driver.get(url)
        time.sleep(2)
    except:
        print("  ⚠ Page load failed.")
        idx += 1
        continue

    # Find Color variations
    try:
        color_inputs = driver.find_elements(
            By.CSS_SELECTOR,
            "fieldset.js.product-form__input input[type='radio'][name='Color']"
        )
    except:
        color_inputs = []

    # Unique color values
    color_values = []
    for inp in color_inputs:
        val = inp.get_attribute("value")
        if val and val not in color_values:
            color_values.append(val)

    # CASE 1: No variation or only 1 color → single config
    if len(color_values) <= 1:
        scrape_configuration(idx)
        products_scraped += 1
        idx += 1

        # Auto-save every 10 products
        if products_scraped % 10 == 0:
            auto_save_progress()

        continue

    # CASE 2: Multiple variations → each color = separate row
    print(f"  🎨 Found {len(color_values)} color variations")
    base_row = df.loc[idx].copy()

    variation_rows = []  # Collect all variation data first

    for color_idx, val in enumerate(color_values):
        try:
            radio = driver.find_element(
                By.CSS_SELECTOR,
                f"fieldset.js.product-form__input input[type='radio'][name='Color'][value='{val}']"
            )
            driver.execute_script("arguments[0].click();", radio)
            time.sleep(1.5)
        except Exception as e:
            print(f"  ⚠ Could not click color '{val}': {e}")
            continue

        # For first variation, update current row
        if color_idx == 0:
            scrape_configuration(idx)
            products_scraped += 1
        else:
            # For additional variations, create new row data
            new_row = base_row.copy()

            # Scrape into a temporary location
            temp_index = len(df)
            df.loc[temp_index] = new_row
            scrape_configuration(temp_index)
            products_scraped += 1

            # Store the scraped row
            variation_rows.append(df.loc[temp_index].copy())

            # Remove temporary row
            df.drop(temp_index, inplace=True)

    # Insert all variation rows immediately after current product
    if variation_rows:
        # Reset index to make insertion easier
        df.reset_index(drop=True, inplace=True)

        # Insert each variation row after the current product
        for i, var_row in enumerate(variation_rows):
            insert_position = idx + 1 + i
            # Split dataframe and insert
            df1 = df.iloc[:insert_position]
            df2 = df.iloc[insert_position:]
            df = pd.concat([df1, pd.DataFrame([var_row]), df2], ignore_index=True)

        rows_added += len(variation_rows)
        idx += len(variation_rows)  # Skip the newly added variation rows

    # Auto-save every 10 products
    if products_scraped % 10 == 0:
        auto_save_progress()

    idx += 1

# =========================
# SAVE OUTPUT with Brand/Link Structure
# =========================
print("\n💾 Saving output Excel...")

# Reorder columns
df = df[final_columns]

# Re-index per category after adding new rows
for category in category_order:
    mask = df["Category"] == category
    df.loc[mask, "Index"] = range(1, mask.sum() + 1)

# Create output workbook
from openpyxl import Workbook

wb_output = Workbook()
wb_output.remove(wb_output.active)

bold = Font(bold=True)
link_font = Font(color="0563C1", underline="single")

# Process categories in original sheet order
for category in category_order:
    df_cat = df[df["Category"] == category].copy()

    if df_cat.empty:
        continue

    # Create sheet
    ws = wb_output.create_sheet(title=category)

    # Add Brand and Link (from metadata)
    metadata = sheet_metadata.get(category, {"brand": "Sunpan", "link": ""})
    ws["A1"] = "Brand"
    ws["B1"] = metadata["brand"]
    ws["A2"] = "Link"
    ws["B2"] = metadata["link"]
    ws["B2"].alignment = Alignment(wrap_text=True)

    # Row 3 empty

    # Headers in row 4
    for col_idx, col_name in enumerate(final_columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = bold

    # Data from row 5
    for r_idx, row_data in enumerate(df_cat.itertuples(index=False), start=5):
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Add hyperlinks to Product Name
    product_url_col = final_columns.index("Product URL") + 1
    product_name_col = final_columns.index("Product Name") + 1

    for r in range(5, ws.max_row + 1):
        url = ws.cell(row=r, column=product_url_col).value
        name_cell = ws.cell(row=r, column=product_name_col)
        if url:
            name_cell.hyperlink = url
            name_cell.font = link_font

    print(f"  ✓ Saved sheet: {category} ({len(df_cat)} products)")

wb_output.save(output_excel)

print(f"\n✅ DONE — {len(df)} products saved to {output_excel}")
print(f"📊 {len(category_order)} categories processed")
print(f"🔢 {products_scraped} products scraped (including variations)")

driver.quit()