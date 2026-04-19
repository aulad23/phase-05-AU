# =========================================================
# INTERLUDE HOME SCRAPER (SELENIUM DIRECT EXTRACTION)
# - FIX: Now extracts directly from Selenium elements
# - No more BeautifulSoup (was missing lazy loaded items)
# - Captures ALL 53 products correctly
# - Works in WINDOW OPEN (headed) mode
# =========================================================

import os
import time
import pandas as pd

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM
# =========================================================

CATEGORIES = {
    "Nightstands": [
        "https://interludehome.com/ih-collections/bedroom/bedside.html",
    ],
    "Coffee & Cocktail Tables": [
        "https://interludehome.com/ih-collections/living/cocktail-tables.html",
    ],
    "Side & End Tables": [
        "https://interludehome.com/ih-collections/living/occasional-tables.html",
        "https://interludehome.com/ih-collections/living/drink-tables.html",
        "https://interludehome.com/ih-collections/living/game-tables.html",
    ],

    "Consoles": [
        "https://interludehome.com/ih-collections/dining/cabinets-consoles.html",
        "https://interludehome.com/ih-collections/living/cabinets-consoles.html"
    ],

    "Beds & Headboards": [
        "https://interludehome.com/ih-collections/bedroom/beds.html",
    ],

    "Desks": [
        "https://interludehome.com/ih-collections/office/desks.html",
    ],

    "Dressers & Chests": [
        "https://interludehome.com/ih-collections/bedroom/dressers-chests.html",
    ],

    "Bar Carts": [
        "https://interludehome.com/ih-collections/dining/bar-cabinets-carts.html",
    ],

    "Bar Stools": [
        "https://interludehome.com/ih-collections/dining/counter-stools.html",
        "https://interludehome.com/ih-collections/dining/bar-stools.html"
    ],
    "Sofas & Loveseats": [
        "https://interludehome.com/ih-collections/living/sofas.html",
    ],

    "Sectionals": [
        "https://interludehome.com/ih-collections/living/sectionals.html",
    ],

    "Lounge Chairs": [
        "https://interludehome.com/ih-collections/living/upholstered-chairs.html",
        "https://interludehome.com/ih-collections/living/occasional-chairs.html"
    ],

    "Ottomans & Benches": [
        "https://interludehome.com/ih-collections/bedroom/benchesottomansstools5002.html",
        "https://interludehome.com/ih-collections/living/benchesottomansstools4871.html"
    ],
     "Desk Chairs": [
        "https://interludehome.com/ih-collections/office/desk-chairs.html",
    ],
    "Mirrors": [
        "https://interludehome.com/ih-collections/decor/mirrors.html",
    ],

    "Vases": [
        "https://interludehome.com/ih-collections/decor/vessels-bowls.html",
    ],

    "Objects": [
        "https://interludehome.com/ih-collections/decor/objets.html",
    ],

    "Rugs": [
        "https://interludehome.com/ih-collections/decor/rugs.html",
    ],

    "Wall Decor": [
        "https://interludehome.com/ih-collections/decor/wall-decor.html",
    ],

}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "interlude_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "InterludeHome.xlsx")


# =========================================================
# DRIVER SETUP (WINDOW OPEN MODE)
# =========================================================

def setup_driver():
    chrome_options = Options()
    # ❌ headless REMOVED (window open)
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    )
    return webdriver.Chrome(options=chrome_options)


# =========================================================
# SCRAPING LOGIC (LAZY LOAD FIXED)
# =========================================================

def scroll_and_load_all_products(driver, url):
    """Scroll and return loaded product elements directly (not page_source)"""
    driver.get(url)

    time.sleep(8)  # allow JS + first render

    last_count = 0
    stable_rounds = 0
    max_attempts = 20  # prevent infinite loop

    for attempt in range(max_attempts):
        items = driver.find_elements(
            By.CSS_SELECTOR, "li.item.product.product-item"
        )
        current_count = len(items)

        print(f"Attempt {attempt + 1}: Loaded products: {current_count}")

        if current_count == last_count:
            stable_rounds += 1
        else:
            stable_rounds = 0

        # Exit only after 5 stable rounds
        if stable_rounds >= 5:
            print(f"✓ Finished loading. Total products: {current_count}")
            break

        last_count = current_count

        # Scroll to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(5)

        # Scroll back up a bit to trigger lazy load
        driver.execute_script("window.scrollBy(0, -300);")
        time.sleep(2)

        # Scroll down again
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(5)

    # Final wait to ensure all loaded
    time.sleep(3)

    # Return the actual elements, not page_source
    return driver.find_elements(By.CSS_SELECTOR, "li.item.product.product-item")


def extract_product_data(product_elements):
    """Extract data directly from Selenium elements (not BeautifulSoup)"""
    products = []

    for item in product_elements:
        try:
            # Product URL
            try:
                product_url = item.find_element(
                    By.CSS_SELECTOR, "a.product.photo.product-item-photo"
                ).get_attribute("href")
            except:
                product_url = ""

            # Image URL
            try:
                image_url = item.find_element(
                    By.CSS_SELECTOR, "img.product-image-photo"
                ).get_attribute("src")
            except:
                image_url = ""

            # Product Name
            try:
                product_name = item.find_element(
                    By.CSS_SELECTOR, "a.product-item-link"
                ).text.strip()
            except:
                product_name = ""

            # SKU
            sku = ""
            try:
                sku_div = item.find_element(By.CSS_SELECTOR, "div.list_sku div")
                sku = sku_div.text.strip()
            except:
                pass

            # List Price
            list_price = ""
            try:
                price_span = item.find_element(By.CSS_SELECTOR, "span.price")
                list_price = price_span.text.replace("$", "").replace(",", "").strip()
            except:
                pass

            products.append(
                {
                    "Product URL": product_url,
                    "Image URL": image_url,
                    "Product Name": product_name,
                    "SKU": sku,
                    "List Price": list_price,
                }
            )

        except Exception as e:
            print(f"Error extracting product: {e}")
            continue

    return products


# =========================================================
# OUTPUT SYSTEM
# =========================================================

def build_step1_master_excel(driver):
    all_rows = []

    for category, urls in CATEGORIES.items():
        for url in urls:
            print(f"\n{'=' * 60}")
            print(f"Scraping: {category}")
            print(f"URL: {url}")
            print(f"{'=' * 60}")

            product_elements = scroll_and_load_all_products(driver, url)
            rows = extract_product_data(product_elements)

            print(f"✓ Extracted {len(rows)} products from {category}")

            for r in rows:
                r["Category"] = category
                all_rows.append(r)

    cols = [
        "Category",
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
        "List Price",
    ]

    if not all_rows:
        pd.DataFrame(columns=cols).to_excel(master_output_file, index=False)
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    df = df[cols]
    df.drop_duplicates(subset=["Product URL"], inplace=True)

    print(f"\n{'=' * 60}")
    print(f"TOTAL PRODUCTS SAVED: {len(df)}")
    print(f"{'=' * 60}\n")

    df.to_excel(master_output_file, index=False)
    return df


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat]
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat[:31])

        ws["A1"] = "Brand"
        ws["B1"] = "Interlude Home"
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES[cat])
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat = df_cat.copy()
        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start = 4
        for j, col in enumerate(df_cat.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_cat.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=3).value
            cell = ws.cell(row=r, column=4)
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)


# =========================================================
# MAIN
# =========================================================

def main():
    driver = setup_driver()
    try:
        df = build_step1_master_excel(driver)
        if not df.empty:
            build_category_wise_workbook_from_df(df)
    finally:
        driver.quit()


if __name__ == "__main__":
    main()