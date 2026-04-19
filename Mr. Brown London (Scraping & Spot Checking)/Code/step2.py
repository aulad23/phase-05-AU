# =========================================================
# STEP 2 – FINAL (SYSTEM UNCHANGED)
# Only change: Input column positions (D/E/F) mapping
# D4 = Image URL, E4 = Product Name, F4 = SKU
# =========================================================

import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import time
import random
from openpyxl import load_workbook
from openpyxl.styles import Font
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ============================
# CONFIG
# ============================

INPUT_FILE = "MrBrownHome.xlsx"
OUTPUT_FILE = "MrBrownHome_Final.xlsx"

BRAND_NAME = "Mr. Brown London"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}

DELAY_RANGE = (1.5, 3.5)

# ============================
# REQUESTS SESSION (retry)
# ============================
session = requests.Session()
retries = Retry(
    total=4,
    backoff_factor=0.8,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"]
)
adapter = HTTPAdapter(max_retries=retries)
session.mount("http://", adapter)
session.mount("https://", adapter)

# ============================
# HELPER FUNCTIONS
# ============================

def fetch_specification(url: str) -> str:
    """
    Keep function name same, but reduce missing by fallback:
    - short description
    - description tab
    - additional info / attributes table
    """
    try:
        r = session.get(url, headers=HEADERS, timeout=25)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        parts = []

        # 1) Short description
        short_div = soup.find("div", class_="woocommerce-product-details__short-description")
        if short_div:
            parts.append(str(short_div))

        # 2) Description tab (sometimes specs are here)
        tab_desc = soup.select_one("#tab-description") or soup.select_one("div.woocommerce-Tabs-panel--description")
        if tab_desc:
            parts.append(str(tab_desc))

        # 3) Additional information / attributes
        add_info = soup.select_one("#tab-additional_information") or soup.select_one(
            "div.woocommerce-Tabs-panel--additional_information"
        )
        if add_info:
            parts.append(str(add_info))

        attr_table = soup.select_one("table.woocommerce-product-attributes")
        if attr_table:
            parts.append(str(attr_table))

        return "\n".join(parts).strip()

    except Exception as e:
        print(f"❌ {url} => {e}")
        return ""


def clean_text(t):
    if not isinstance(t, str):
        return ""
    return re.sub(r"\s+", " ", t.strip())


def extract_description(html):
    """
    Extract first paragraph that looks like a true description.
    Skip paragraphs containing Finish/Finishes/Dimension/Dimensions/Overall
    """
    soup = BeautifulSoup(html, "html.parser")
    for p in soup.find_all("p"):
        txt = p.get_text(" ", strip=True)
        if not re.search(r"\b(Finish|Finishes|Dimension|Dimensions|Overall|Overall Dimensions)\b", txt, re.I):
            return clean_text(txt)
    return ""


def extract_field(html, field):
    """
    Extract Finish / Seat exactly as is
    Fix: Finish vs Finishes (plural)
    """
    soup = BeautifulSoup(html, "html.parser")

    # ✅ Fix for Finish/Finishes
    if field.lower() == "finish":
        pattern = r"\b(Finish|Finishes)\b"
    else:
        pattern = rf"\b{re.escape(field)}s?\b"

    for p in soup.find_all("p"):
        txt = p.get_text(" ", strip=True)
        if re.search(pattern, txt, re.I):
            txt = re.sub(pattern + r"\s*[:\-]?\s*", "", txt, flags=re.I)
            return txt.strip()

    # Fallback: sometimes attributes table has "Finish"
    for row in soup.select("table.woocommerce-product-attributes tr"):
        th = row.find("th")
        td = row.find("td")
        if th and td and re.search(pattern, th.get_text(" ", strip=True), re.I):
            return td.get_text(" ", strip=True).strip()

    return ""


def extract_com(spec_text):
    """
    Extract COM from lines starting with COM:
    Fallback: COM without colon
    """
    if not spec_text:
        return ""

    lines = spec_text.splitlines()
    for line in lines:
        line = line.strip()
        if line.upper().startswith("COM:"):
            return line.split(":", 1)[1].strip()

    m = re.search(r"\bCOM\b\s*[:\-]?\s*([^\n]+)", spec_text, re.I)
    if m:
        return m.group(1).strip()

    return ""


def extract_dimensions_from_spec(spec_text):
    """
    Extract Width, Depth, Height, Diameter from specification text.
    Handles:
    - 24W x 18D x 30H
    - Width: 24
    """
    dims = {"Width": "", "Depth": "", "Height": "", "Diameter": ""}
    if not spec_text:
        return dims

    t = spec_text.replace("×", "x").replace("X", "x").replace("″", "").replace('"', "")
    t = t.replace("’", "").replace("'", "")
    t = re.sub(r"\b(Overall|Dimensions|Overall Dimensions)\s*[:\-]?\s*", "", t, flags=re.I)

    lines = t.splitlines()
    for line in lines:
        parts = [p.strip() for p in line.split("x")]
        for p in parts:
            m = re.search(r'([0-9]+(?:\.[0-9]+)?)\s*W\b|\bW(?:idth)?[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', p, re.I)
            if m and not dims["Width"]:
                dims["Width"] = m.group(1) if m.group(1) else m.group(2)

            m = re.search(r'([0-9]+(?:\.[0-9]+)?)\s*D\b|\bD(?:epth)?[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', p, re.I)
            if m and not dims["Depth"]:
                dims["Depth"] = m.group(1) if m.group(1) else m.group(2)

            m = re.search(r'([0-9]+(?:\.[0-9]+)?)\s*H\b|\bH(?:eight)?[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', p, re.I)
            if m and not dims["Height"]:
                dims["Height"] = m.group(1) if m.group(1) else m.group(2)

            m = re.search(r'([0-9]+(?:\.[0-9]+)?)\s*(?:Dia|Diameter)\b|\b(?:Dia|Diameter)[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', p, re.I)
            if m and not dims["Diameter"]:
                dims["Diameter"] = m.group(1) if m.group(1) else m.group(2)

    # Label fallback
    if not dims["Width"]:
        m = re.search(r'\bWidth\b\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
        if m: dims["Width"] = m.group(1)
    if not dims["Depth"]:
        m = re.search(r'\bDepth\b\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
        if m: dims["Depth"] = m.group(1)
    if not dims["Height"]:
        m = re.search(r'\bHeight\b\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
        if m: dims["Height"] = m.group(1)
    if not dims["Diameter"]:
        m = re.search(r'\b(Dia|Diameter)\b\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)', t, re.I)
        if m: dims["Diameter"] = m.group(2)

    return dims


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # ✅ SYSTEM UNCHANGED
    # ✅ Only mapping fix:
    # C = Product URL, D = Image URL, E = Product Name, F = SKU (row 4 data)
    cols = [str(c).strip() for c in df.columns]

    if len(cols) < 3:
        df.columns = cols
        return df

    # C column -> Product URL
    if "Product URL" not in cols:
        cols[2] = "Product URL"

    # D/E/F columns
    if len(cols) >= 4:
        cols[3] = "Image URL"
    if len(cols) >= 5:
        cols[4] = "Product Name"
    if len(cols) >= 6:
        cols[5] = "SKU"

    df.columns = cols
    return df


# ============================
# PROCESS ONE CATEGORY SHEET
# ============================

def process_sheet(df: pd.DataFrame, category_name: str) -> pd.DataFrame:
    df = normalize_columns(df)

    # ✅ Only real URLs (removes "Product URL" garbage row)
    df["Product URL"] = df["Product URL"].astype(str).str.strip()
    df = df[df["Product URL"].str.match(r"^https?://", na=False)].reset_index(drop=True)

    output_rows = []

    for idx, row in df.iterrows():
        url = row["Product URL"]
        print(f"🔎 {category_name} → {idx+1}/{len(df)}")

        html = fetch_specification(url)
        time.sleep(random.uniform(*DELAY_RANGE))

        spec_text = BeautifulSoup(html, "html.parser").get_text("\n", strip=True) if html else ""
        dims = extract_dimensions_from_spec(spec_text)

        output_rows.append({
            "Field No.": idx + 1,
            "Category": category_name,
            "Product URL": url,
            "Image URL": row.get("Image URL", ""),
            "Product Name": row.get("Product Name", ""),
            "SKU": row.get("SKU", ""),
            "Product Family Id": row.get("Product Name", ""),
            "Description": extract_description(html),
            "Width": dims["Width"],
            "Depth": dims["Depth"],
            "Diameter": dims["Diameter"],
            "Height": dims["Height"],
            "Finish": extract_field(html, "Finish"),
            "Seat": extract_field(html, "Seat"),
            "Com": extract_com(spec_text)
        })

    return pd.DataFrame(output_rows)


# ============================
# MAIN
# ============================

def main():
    wb_in = load_workbook(INPUT_FILE)
    writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")

    for sheet in wb_in.sheetnames:
        ws_in = wb_in[sheet]

        # ✅ Category Link from B2 (unchanged system)
        category_link = str(ws_in["B2"].value or "").strip()

        # ✅ Header on Excel row 3, data starts row 4 (C4)
        df = pd.read_excel(INPUT_FILE, sheet_name=sheet, header=2)

        out_df = process_sheet(df, sheet)
        out_df.to_excel(writer, sheet_name=sheet, index=False, startrow=3)

        ws_out = writer.book[sheet]

        ws_out["A1"] = "Brand"
        ws_out["B1"] = BRAND_NAME
        ws_out["A2"] = "Category Link"
        ws_out["B2"] = category_link

        # Bold column headers (row 4)
        for cell in ws_out[4]:
            cell.font = Font(bold=True)

    writer.close()
    print(f"\n✅ FINAL OUTPUT READY → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
