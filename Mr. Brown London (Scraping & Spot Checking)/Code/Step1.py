# =========================================================
# Integrated Scraper (FINAL)
# - Scraping logic: FROM CODE-A (UNCHANGED)
# - Input / Output system: FROM CODE-B
# - SKU SYSTEM: SD_SKU AUTO-GENERATION (FINAL)
# =========================================================

import os
import re
import time
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM
# =========================================================

VENDOR_NAME = "Mr Brown Home"

CATEGORIES = {
    "Nightstands": [
        "https://mrbrownhome.com/products_by_category/furniture/chests-bedsides/",
    ],
   "Dining Tables": [
        "https://mrbrownhome.com/products_by_category/furniture/dining-table-sets/",
    ],
    "Coffee & Cocktail Tables": [
        "https://mrbrownhome.com/products_by_category/coffee-tables/",
    ],

    "Side & End Tables": [
        "https://mrbrownhome.com/products_by_category/tables/side-tables/",
    ],
    "Consoles": [
        "https://mrbrownhome.com/products_by_category/furniture/occasional-tables/console-tables/",
    ],

    "Beds & Headboards": [
        "https://mrbrownhome.com/products_by_category/furniture/bed/",
    ],

    "Desks": [
        "https://mrbrownhome.com/products_by_category/furniture/desks/",
    ],

    "Bookshelves": [
        "https://mrbrownhome.com/products_by_category/furniture/bookshelves/",
    ],

    "Cabinets": [
        "https://mrbrownhome.com/products_by_category/furniture/cabinets/",
    ],

    "Dining Chairs": [
        "https://mrbrownhome.com/products_by_category/furniture/seating/dining-chairs/",
    ],

    "Bar Stools": [
        "https://mrbrownhome.com/products_by_category/furniture/seating/bar-counter-stools/",
    ],

    "Sofas & Loveseats": [
        "https://mrbrownhome.com/products_by_category/furniture/seating/sofas-loveseats/",
    ],

    "Benches": [
        "https://mrbrownhome.com/products_by_category/furniture/seating/benches-and-daybeds/",
    ],

    "Ottomans": [
        "https://mrbrownhome.com/products_by_category/furniture/seating/ottomans-stools/",
    ],

    "Chandeliers": [
        "https://mrbrownhome.com/products_by_category/lighting/chandeliers/",
    ],

    "Sconces": [
        "https://mrbrownhome.com/products_by_category/lighting/wall-lamps/",
    ],
    "Table Lamps": [
        "https://mrbrownhome.com/products_by_category/lighting/table-lamps/",
    ],

    "Floor Lamps": [
        "https://mrbrownhome.com/products_by_category/lighting/floor-lamps/",
    ],

    "Mirrors": [
        "https://mrbrownhome.com/products_by_category/mirrors/",
    ],

    "Objects": [
        "https://mrbrownhome.com/products_by_category/accessories/decorative-accessories/",
    ],

    "Rugs": [
        "https://mrbrownhome.com/products_by_category/accessories/rugs/",
    ],

}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "mrbrown_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "MrBrownHome.xlsx")

DEFAULT_WAIT = 20

# =========================================================
# DRIVER
# =========================================================

def init_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1400,900")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

# =========================================================
# SCRAPING LOGIC (UNCHANGED)
# =========================================================

def smooth_scroll_to_bottom(driver, pause=0.8, max_tries=15):
    last_height = driver.execute_script("return document.body.scrollHeight")
    tries = 0
    while tries < max_tries:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            tries += 1
        else:
            tries = 0
            last_height = new_height


def parse_product_card(li_element):
    product_url = image_url = product_name = ""

    try:
        a = li_element.find_element(
            By.CSS_SELECTOR,
            "a.woocommerce-LoopProduct-link.woocommerce-loop-product__link"
        )
        product_url = a.get_attribute("href") or ""
        try:
            title_el = a.find_element(By.CSS_SELECTOR, "h2.woocommerce-loop-product__title")
            product_name = title_el.text.strip()
        except Exception:
            pass
    except Exception:
        pass

    try:
        img = li_element.find_element(By.CSS_SELECTOR, "img")
        src = img.get_attribute("src") or ""
        if src:
            image_url = src
        else:
            srcset = img.get_attribute("srcset") or ""
            if srcset:
                image_url = srcset.split(",")[0].strip().split(" ")[0]
    except Exception:
        pass

    return {
        "Product URL": product_url,
        "Image URL": image_url,
        "Product Name": product_name,
        "SKU": None
    }


def scrape_category_with_pagination(driver, base_url):
    page_num = 1
    results = []

    while True:
        url = base_url if page_num == 1 else f"{base_url}page/{page_num}/"
        driver.get(url)

        try:
            WebDriverWait(driver, DEFAULT_WAIT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "ul.products.columns-4"))
            )
        except Exception:
            break

        smooth_scroll_to_bottom(driver)

        items = driver.find_elements(
            By.CSS_SELECTOR,
            "ul.products.columns-4 > li.product.type-product"
        )
        if not items:
            break

        for li in items:
            try:
                data = parse_product_card(li)
                if data["Product URL"]:
                    results.append(data)
            except Exception:
                continue

        page_num += 1
        time.sleep(1)

    return results

# =========================================================
# SKU GENERATION SYSTEM (SD_SKU FINAL)
# =========================================================

def generate_sku(vendor_name: str, category_name: str, index: int) -> str:
    vendor_code = re.sub(r"[^A-Z]", "", vendor_name.upper())[:3]
    category_code = re.sub(r"[^A-Z]", "", category_name.upper())[:2]
    return f"{vendor_code}{category_code}{index}"

# =========================================================
# OUTPUT SYSTEM
# =========================================================

def build_step1_master_excel():
    driver = init_driver()
    all_rows = []

    try:
        for category, urls in CATEGORIES.items():
            product_index = 1
            for url in urls:
                rows = scrape_category_with_pagination(driver, url)
                for r in rows:
                    r["Category"] = category
                    r["SKU"] = generate_sku(VENDOR_NAME, category, product_index)
                    product_index += 1
                    all_rows.append(r)
    finally:
        driver.quit()

    if not all_rows:
        cols = ["Category", "Product URL", "Image URL", "Product Name", "SKU"]
        pd.DataFrame(columns=cols).to_excel(master_output_file, index=False)
        return pd.DataFrame(columns=cols)

    df = pd.DataFrame(all_rows)
    df.drop_duplicates(subset=["Product URL"], inplace=True)
    df = df[["Category", "Product URL", "Image URL", "Product Name", "SKU"]]
    df.to_excel(master_output_file, index=False)
    return df


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)

        ws["A1"] = "Brand"
        ws["B1"] = VENDOR_NAME
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start = 4
        for j, col in enumerate(df_cat.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_cat.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {
            ws.cell(row=start, column=j).value: j
            for j in range(1, ws.max_column + 1)
        }

        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            name_cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                name_cell.hyperlink = url
                name_cell.font = link_font

    wb.save(master_sheets_file)

# =========================================================
# MAIN
# =========================================================

def main():
    df = build_step1_master_excel()
    if df.empty:
        Workbook().save(master_sheets_file)
        return
    build_category_wise_workbook_from_df(df)


if __name__ == "__main__":
    main()
