"""
Step 2 Scraper - Siemon & Salazar  (FIXED)
Input:  chandeliers_products.xlsx  (Product URL, Image URL, Product Name, SKU)
Output: chandeliers_products_full.xlsx  (same file updated with extra columns)

New columns added:
  Product Family ID | Description | Weight | Width | Depth |
  Diameter | Height | Finish | Socket | Canopy

Requirements:
  pip install playwright openpyxl
  playwright install chromium
"""

import asyncio
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

INPUT_FILE  = "SiemonSalazar_Sconces.xlsx"
OUTPUT_FILE = "SiemonSalazar_Sconces_Final.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_family_id(product_name: str) -> str:
    """Product Family ID = Product Name as-is."""
    return product_name.strip() if product_name else ""


def parse_dimensions(dim_str: str) -> dict:
    """
    Handles multiple formats:
      - Labeled:  "H: 9.0" W: 6.0" D: 6.5""
      - WxHxD:    "6.0" x 9.0" x 6.5""   (unlabeled, assumes W x H x D)
      - Ø + H:    "4.5" Ø, 4.5" H"  or  "18.0 Ø, 1.6" H"
    """
    result = {"Height": "", "Width": "", "Depth": "", "Diameter": ""}
    dim_str = dim_str.replace("\u00d8", "Ø").replace("\u2019", '"').replace("\u201d", '"')

    # ── Try labeled tokens first (H/W/D/Dia/Ø before or after value) ──
    tokens = re.findall(
        r'(?:(H|W|D|Dia\.?|Diam\.?|Ø)\s*[:\.]?\s*(\d+(?:\.\d+)?)\s*"?)'
        r'|(?:(\d+(?:\.\d+)?)\s*"?\s*(H|W|D|Dia\.?|Diam\.?|DIA\.?|Ø))',
        dim_str, re.IGNORECASE
    )
    label_map = {
        "h": "Height", "w": "Width", "d": "Depth",
        "dia": "Diameter", "dia.": "Diameter",
        "diam": "Diameter", "diam.": "Diameter",
        "ø": "Diameter",
    }
    found_labeled = False
    for token in tokens:
        lbl_first, val_first, val_second, lbl_second = token
        if lbl_first and val_first:
            key = label_map.get(lbl_first.lower().rstrip("."), "")
            if key and not result[key]:
                result[key] = f'{val_first}"'
                found_labeled = True
        elif val_second and lbl_second:
            key = label_map.get(lbl_second.lower().rstrip("."), "")
            if key and not result[key]:
                result[key] = f'{val_second}"'
                found_labeled = True

    if found_labeled:
        return result

    # ── FIX BUG 1: Handle unlabeled "N x N x N" format (W x H x D) ──
    vals = re.findall(r'(\d+(?:\.\d+)?)\s*"?', dim_str)
    if len(vals) == 3:
        result["Width"]  = f'{vals[0]}"'
        result["Height"] = f'{vals[1]}"'
        result["Depth"]  = f'{vals[2]}"'
    elif len(vals) == 2:
        # Could be Ø + H or W x H — check for Ø
        if "Ø" in dim_str or "ø" in dim_str.lower():
            result["Diameter"] = f'{vals[0]}"'
            result["Height"]   = f'{vals[1]}"'
        else:
            result["Width"]  = f'{vals[0]}"'
            result["Height"] = f'{vals[1]}"'
    elif len(vals) == 1:
        if "Ø" in dim_str or "ø" in dim_str.lower():
            result["Diameter"] = f'{vals[0]}"'
        else:
            result["Height"] = f'{vals[0]}"'

    return result


def parse_weight(full_text: str) -> str:
    """Extract just the first weight number (no 'lbs' suffix)."""
    # Match "Product Weight: ..." or "weight*: ..." line
    m = re.search(
        r'(?:Total\s+)?Product\s*[Ww]eight\*?\s*:\s*(.+?)(?:\n|$)',
        full_text, re.IGNORECASE
    )
    if m:
        weight_line = m.group(1).strip()
        # Grab first number (lbs value) from the line
        m2 = re.search(r'([\d.]+)\s*lbs', weight_line, re.IGNORECASE)
        if m2:
            return m2.group(1)
        return ""

    # Fallback: any "N lbs" in text
    m3 = re.search(r'weight\*?\s*:\s*([\d.]+)\s*lbs', full_text, re.IGNORECASE)
    if m3:
        return m3.group(1)
    return ""


def parse_socket(full_text: str) -> str:
    """Extract just the socket type code (E26, E12, G4, etc.)."""
    # Look for "Light source:" or "Lamping:" lines and extract socket code
    m = re.search(r'Light\s*source\s*:\s*(.+)', full_text, re.IGNORECASE)
    if m:
        line = m.group(1).strip()
        # Extract socket code like E12, E26, G4, G9, GU10, etc.
        code = re.search(r'\b([EG](?:U)?\d+[a-z]?)\b', line, re.IGNORECASE)
        if code:
            return code.group(1)

    m2 = re.search(r'Lamping\s*:\s*(.+)', full_text, re.IGNORECASE)
    if m2:
        line = m2.group(1).strip()
        code = re.search(r'\b([EG](?:U)?\d+[a-z]?)\b', line, re.IGNORECASE)
        if code:
            return code.group(1)

    # Fallback: any socket code mention
    m3 = re.search(r'\b([EG](?:U)?\d+[a-z]?)\s+(?:medium\s+|candelabra\s+|bi-?pin\s+)?base', full_text, re.IGNORECASE)
    if m3:
        return m3.group(1)
    return ""


def parse_finish(full_text: str) -> str:
    """
    FIX BUG 3: Properly handle section boundaries.
    Collects lines under GLASS FINISH, METAL FINISH, FINISH, or CANOPY FINISHES headers.
    Stops when: next ALL-CAPS header, or line is too long / doesn't look like a finish option.
    """
    glass_lines = []
    metal_lines = []
    other_finish_lines = []
    current_bucket = None

    finish_headers = re.compile(
        r'^(GLASS\s+FINISH(?:ES)?|METAL\s+FINISH(?:ES)?|CANOPY\s+FINISH(?:ES)?|FINISH(?:ES)?)$',
        re.IGNORECASE
    )
    allcaps_header = re.compile(r'^[A-Z][A-Z\s]+$')

    # A finish option line is typically short: "Brushed Brass (CM-008)" or "Satin nickel"
    # If a line is >60 chars or starts with "When", "This", "Please", etc. → not a finish
    def looks_like_finish_option(line):
        if len(line) > 60:
            return False
        if re.match(r'^(When|This|If|Please|Pendant|Rectangular|See)\b', line, re.IGNORECASE):
            return False
        return True

    for line in full_text.splitlines():
        l = line.strip()
        if not l or l == "\xa0":
            continue

        if finish_headers.match(l):
            upper = l.upper()
            if "GLASS" in upper:
                current_bucket = glass_lines
            elif "METAL" in upper:
                current_bucket = metal_lines
            elif "CANOPY" in upper:
                current_bucket = other_finish_lines
            else:
                current_bucket = other_finish_lines
            continue

        if current_bucket is not None:
            # Stop on all-caps header that isn't a finish header
            if allcaps_header.match(l) and len(l) > 3 and not finish_headers.match(l):
                current_bucket = None
                continue
            # Stop if line doesn't look like a finish option
            if not looks_like_finish_option(l):
                current_bucket = None
                continue
            current_bucket.append(l)

    parts = []
    all_lines = glass_lines + metal_lines + other_finish_lines
    if all_lines:
        return ", ".join(all_lines)
    return ""


def parse_canopy(full_text: str) -> str:
    """
    Extract only canopy dimensions:
      - Stem diameter (e.g. 5/8")
      - Canopy Ø and height (e.g. 4.5" Ø x .07")
    """
    dims = []

    # Extract stem diameter: "N/N" diameter" or "N" diameter stem"
    m_stem = re.search(r'(\d+/\d+)\s*"\s*(?:diameter\s+)?(?:stem|rod)', full_text, re.IGNORECASE)
    if m_stem:
        dims.append(f'{m_stem.group(1)}"')

    # Extract canopy dimensions from "Canopy dimensions: X" or "Round Canopy: X" or "Canopy: X Ø..."
    m_cdim = re.search(r'Canopy\s+dimensions?\s*:\s*(.+?)(?:\n|$)', full_text, re.IGNORECASE)
    if m_cdim:
        dims.append(m_cdim.group(1).strip())

    # "Round Canopy: 4.5" Ø x 0.63""
    for m_rc in re.finditer(r'(?:Round|Square)\s+Canopy\s*:\s*(.+?)(?:\n|$)', full_text, re.IGNORECASE):
        val = m_rc.group(1).strip()
        if val not in dims:
            dims.append(val)

    # Extract from CANOPY section if it has dimension-like lines
    in_canopy = False
    section_header = re.compile(r'^[A-Z][A-Z\s]+$')
    for line in full_text.splitlines():
        l = line.strip()
        if not l:
            continue
        if re.match(r'^CANOPY$', l, re.IGNORECASE):
            in_canopy = True
            continue
        if in_canopy:
            if section_header.match(l) and len(l) > 3:
                break
            # Only grab lines that contain dimension-like data (numbers + " or Ø)
            if re.search(r'\d+(?:\.\d+)?\s*"', l) or 'Ø' in l:
                if l not in dims:
                    dims.append(l)

    return ", ".join(dims) if dims else ""


# ─────────────────────────────────────────────────────────────────────────────
# Detail page scraper
# ─────────────────────────────────────────────────────────────────────────────

async def scrape_product_details(page, url: str, retries: int = 3) -> dict:
    result = {
        "Product Family ID": "",
        "Description": "",
        "Weight": "",
        "Width": "", "Depth": "", "Diameter": "", "Height": "",
        "Finish": "", "Socket": "", "Canopy": "",
    }
    for attempt in range(1, retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=90000)
            await page.wait_for_selector('[data-hook="description"]', timeout=30000)

            paras = await page.query_selector_all('[data-hook="description"] p')
            texts = []
            for p_el in paras:
                t = (await p_el.inner_text()).strip()
                if t and t != "\xa0":
                    texts.append(t)

            if texts:
                result["Description"] = texts[0]

            full_text = "\n".join(texts)

            # FIX BUG 2: Use negative lookbehind to skip "Canopy dimensions:"
            # Only match "Dimensions:" not preceded by "Canopy "
            dim_match = re.search(
                r'(?<!Canopy\s)Dimensions?\*?\s*:\s*(.+)',
                full_text, re.IGNORECASE
            )
            if dim_match:
                dims = parse_dimensions(dim_match.group(1).strip())
                result["Height"]   = dims["Height"]
                result["Width"]    = dims["Width"]
                result["Depth"]    = dims["Depth"]
                result["Diameter"] = dims["Diameter"]

            result["Weight"]  = parse_weight(full_text)
            result["Finish"]  = parse_finish(full_text)
            result["Socket"]  = parse_socket(full_text)
            result["Canopy"]  = parse_canopy(full_text)
            break  # success

        except Exception as e:
            if attempt < retries:
                print(f"    Retry {attempt}/{retries} [{url}]: {e}")
                await asyncio.sleep(3)
            else:
                print(f"    Failed after {retries} attempts [{url}]: {e}")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    ("Product URL",       60),
    ("Image URL",         60),
    ("Product Name",      38),
    ("SKU",               16),
    ("Product Family ID", 28),
    ("Description",       70),
    ("Weight",            14),
    ("Width",             12),
    ("Depth",             12),
    ("Diameter",          12),
    ("Height",            12),
    ("Finish",            50),
    ("Socket",            35),
    ("Canopy",            60),
]


async def main():
    wb_in = load_workbook(INPUT_FILE)
    ws_in = wb_in.active

    in_headers = [ws_in.cell(row=1, column=c).value for c in range(1, ws_in.max_column + 1)]

    def col_idx(name):
        try:
            return in_headers.index(name) + 1
        except ValueError:
            return None

    url_col  = col_idx("Product URL")
    img_col  = col_idx("Image URL")
    name_col = col_idx("Product Name")
    sku_col  = col_idx("SKU")

    rows = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append({
                "Product URL":  row[url_col  - 1] if url_col  else "",
                "Image URL":    row[img_col  - 1] if img_col  else "",
                "Product Name": row[name_col - 1] if name_col else "",
                "SKU":          row[sku_col  - 1] if sku_col  else "",
            })

    print(f"Loaded {len(rows)} products from {INPUT_FILE}\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        detail_page = await context.new_page()

        for idx, row in enumerate(rows, start=1):
            url  = row["Product URL"]
            name = row["Product Name"]

            row["Product Family ID"] = get_family_id(name) if name else ""

            if url:
                print(f"  [{idx:03d}] {name}")
                details = await scrape_product_details(detail_page, url)
                row.update(details)
            else:
                print(f"  [{idx:03d}] {name} -> no URL, skipping")

        await browser.close()

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Chandeliers"

    for ci, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws_out.cell(row=1, column=ci, value=header)
        cell.font = Font(bold=True)
        ws_out.column_dimensions[get_column_letter(ci)].width = width

    for ri, row in enumerate(rows, start=2):
        for ci, (key, _) in enumerate(COLUMNS, start=1):
            ws_out.cell(row=ri, column=ci, value=row.get(key, ""))

    wb_out.save(OUTPUT_FILE)
    print(f"\nDone! {len(rows)} products saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main())