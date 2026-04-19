"""
Web Scraper for Siemon & Salazar - Chandeliers & Pendants Category
URLs:
  https://www.siemonandsalazar.com/lights?Category=Chandeliers
  https://www.siemonandsalazar.com/lights?Category=Pendants

Output: chandeliers_Pendants.xlsx

Requirements:
  pip install playwright openpyxl
  playwright install chromium
"""

import asyncio
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

TARGET_URLS = [
    "https://www.siemonandsalazar.com/lights?Category=Sconces",
    #"https://www.siemonandsalazar.com/lights?Category=Multiport%2520Canopy",
]

CATEGORY_CODE = "SC"
OUTPUT_FILE   = "SiemonSalazar_Sconces.xlsx"


def generate_sku(product_name: str, index: int, category_code: str = "CH") -> str:
    words = product_name.strip().split()
    vendor = words[0] if words else "UNK"
    vendor_code = re.sub(r"[^A-Za-z]", "", vendor)[:3].upper().ljust(3, "X")
    return f"{vendor_code}{category_code}{str(index).zfill(3)}"


def clean_image_url(src: str) -> str:
    match = re.match(r"(https://static\.wixstatic\.com/media/[^/]+)", src)
    return match.group(1) if match else src


async def scroll_to_bottom(page, pause: float = 2.5, max_scrolls: int = 40):
    prev_height = 0
    for _ in range(max_scrolls):
        await page.evaluate("window.scrollBy(0, document.body.scrollHeight)")
        await asyncio.sleep(pause)
        curr_height = await page.evaluate("document.body.scrollHeight")
        if curr_height == prev_height:
            break
        prev_height = curr_height


def save_to_excel(products: list, filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chandeliers & Pendants"

    headers    = ["Product URL", "Image URL", "Product Name", "SKU", "Category"]
    col_widths = [65, 65, 40, 16, 15]

    # Bold header row
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    for ri, prod in enumerate(products, start=2):
        ws.cell(row=ri, column=1, value=prod["Product URL"])
        ws.cell(row=ri, column=2, value=prod["Image URL"])
        ws.cell(row=ri, column=3, value=prod["Product Name"])
        ws.cell(row=ri, column=4, value=prod["SKU"])
        ws.cell(row=ri, column=5, value=prod["Category"])

    wb.save(filepath)
    print(f"\nSaved to: {filepath}")


async def scrape_chandeliers():
    all_products = []
    seen_urls    = set()   # prevent duplicate products across categories
    global_index = 1       # continuous SKU index across all categories

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        # ── Loop over every category URL ──────────────────────────────────
        for target_url in TARGET_URLS:
            # Extract category name from URL for labelling
            category_name = target_url.split("Category=")[-1]
            print(f"\n{'='*60}")
            print(f"Scraping category : {category_name}")
            print(f"URL               : {target_url}")
            print(f"{'='*60}")

            try:
                await page.goto(target_url, wait_until="networkidle", timeout=60000)
                await page.wait_for_selector(
                    '[data-hook="product-list-grid-item"]', timeout=30000
                )
            except Exception as e:
                print(f"  Failed to load page: {e}")
                continue

            print("Scrolling to load all products...")
            await scroll_to_bottom(page, pause=2.5, max_scrolls=40)
            await asyncio.sleep(3)

            items = await page.query_selector_all('[data-hook="product-list-grid-item"]')
            print(f"Found {len(items)} items on this page")

            for item in items:
                try:
                    # ── Product URL ──────────────────────────────────────
                    link_el = await item.query_selector(
                        '[data-hook="product-item-root"] a[data-hook="product-item-container"]'
                    )
                    product_url = ""
                    if link_el:
                        href = await link_el.get_attribute("href") or ""
                        product_url = (
                            href if href.startswith("http")
                            else "https://www.siemonandsalazar.com" + href
                        )

                    # ── Skip duplicates ──────────────────────────────────
                    if product_url and product_url in seen_urls:
                        print(f"  [SKIP duplicate] {product_url}")
                        continue
                    if product_url:
                        seen_urls.add(product_url)

                    # ── Product Name ─────────────────────────────────────
                    name_el = await item.query_selector('[data-hook="product-item-name"]')
                    product_name = (await name_el.inner_text()).strip() if name_el else ""

                    # ── Image URL ────────────────────────────────────────
                    image_url = ""
                    img_el = await item.query_selector("wow-image img")
                    if img_el:
                        src = await img_el.get_attribute("src") or ""
                        image_url = clean_image_url(src) if src else ""

                    if not image_url:
                        wow_el = await item.query_selector("wow-image")
                        if wow_el:
                            info = await wow_el.get_attribute("data-image-info") or ""
                            m = re.search(r'"uri":"([^"]+)"', info)
                            if m:
                                image_url = (
                                    f"https://static.wixstatic.com/media/{m.group(1)}"
                                )

                    # ── SKU ──────────────────────────────────────────────
                    sku = (
                        generate_sku(product_name, global_index, CATEGORY_CODE)
                        if product_name
                        else f"UNKCH{str(global_index).zfill(3)}"
                    )

                    all_products.append({
                        "Product URL":  product_url,
                        "Image URL":    image_url,
                        "Product Name": product_name,
                        "SKU":          sku,
                        "Category":     category_name,
                    })

                    print(f"  [{global_index:03d}] [{category_name}] {product_name} -> {sku}")
                    global_index += 1

                except Exception as e:
                    print(f"  Error on item: {e}")

        await browser.close()

    # ── Save results ──────────────────────────────────────────────────────
    if all_products:
        save_to_excel(all_products, OUTPUT_FILE)
        print(f"\nDone! {len(all_products)} total products saved to '{OUTPUT_FILE}'")
    else:
        print("No products found.")

    return all_products


if __name__ == "__main__":
    asyncio.run(scrape_chandeliers())