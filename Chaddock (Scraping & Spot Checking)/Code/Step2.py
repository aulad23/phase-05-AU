# -*- coding: utf-8 -*-
# Chaddock Product Detail Scraper – FINAL (Input/Output same like your Excel)

import time
import re
from pathlib import Path
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

WAIT_TIMEOUT = 15

# Input and output files
INPUT_XLSX = Path("Chaddock.xlsx")
OUTPUT_XLSX = Path("Chaddock_details.xlsx")

# -------------------------- Helpers --------------------------

def connect_driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts
    )

def clean_text(text):
    if not text:
        return ""
    return re.sub(r"[\x00-\x1f\x7f]", "", str(text)).strip()

def extract_number(text):
    if not text:
        return ""
    match = re.search(r"(\d+\.?\d*)", str(text))
    return match.group(1) if match else ""

def parse_overall(text):
    dims = {"Width": "", "Depth": "", "Height": ""}
    if not text:
        return dims
    matches = re.findall(r"([WDH])\s*(\d+\.?\d*)", str(text).upper())
    for dim, val in matches:
        if dim == "W":
            dims["Width"] = val
        elif dim == "D":
            dims["Depth"] = val
        elif dim == "H":
            dims["Height"] = val
    return dims

def parse_seat_dimensions(text):
    seat = {"Seat Height": "", "Seat Depth": "", "Seat Width": ""}
    if not text:
        return seat
    matches = re.findall(r"([HDW])\s*(\d+\.?\d*)", str(text).upper())
    for dim, val in matches:
        if dim == "H":
            seat["Seat Height"] = val
        elif dim == "D":
            seat["Seat Depth"] = val
        elif dim == "W":
            seat["Seat Width"] = val
    return seat

# -------------------------- Scraper --------------------------

def extract_product_details(driver, url):
    details = {
        "Product Family Id": "",
        "Description": "",
        "Weight": "",
        "Width": "",
        "Depth": "",
        "Height": "",
        "Diameter": "",
        "Seat Height": "",
        "Seat Depth": "",
        "Seat Width": "",
        "Arm Height": "",
        "Com": "",
        "col": ""
    }

    try:
        driver.get(url)
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.grid_4"))
        )
        time.sleep(0.5)
    except TimeoutException:
        return details

    try:
        desc = driver.find_element(
            By.CSS_SELECTOR,
            "div.grid_4 > div[style*='margin-top:20px']:nth-of-type(2)"
        )
        details["Description"] = clean_text(desc.text)
    except:
        pass

    try:
        overall = driver.find_element(
            By.CSS_SELECTOR,
            "tr#ctl00_ctl00_ChildBodyContent_PageContent_trDimensionsOverall td:nth-of-type(2)"
        )
        details.update(parse_overall(overall.text))
    except:
        pass

    try:
        seat = driver.find_element(
            By.CSS_SELECTOR,
            "tr#ctl00_ctl00_ChildBodyContent_PageContent_trDimensionsSeat td:nth-of-type(2)"
        )
        details.update(parse_seat_dimensions(seat.text))
    except:
        pass

    try:
        arm = driver.find_element(
            By.CSS_SELECTOR,
            "tr#ctl00_ctl00_ChildBodyContent_PageContent_trDimensionsArmHeight td:nth-of-type(2)"
        )
        details["Arm Height"] = extract_number(arm.text)
    except:
        pass

    try:
        dia = driver.find_element(
            By.CSS_SELECTOR,
            "tr#ctl00_ctl00_ChildBodyContent_PageContent_trDiameter td:nth-of-type(2)"
        )
        details["Diameter"] = extract_number(dia.text)
    except:
        pass

    try:
        com = driver.find_element(
            By.CSS_SELECTOR,
            "tr#ctl00_ctl00_ChildBodyContent_PageContent_trComFabric td:nth-of-type(2)"
        )
        details["Com"] = extract_number(com.text)
    except:
        pass

    try:
        leather = driver.find_element(
            By.CSS_SELECTOR,
            "tr#ctl00_ctl00_ChildBodyContent_PageContent_trComLeather td:nth-of-type(2)"
        )
        details["col"] = extract_number(leather.text)
    except:
        pass

    try:
        wt = driver.find_element(
            By.CSS_SELECTOR,
            "tr#ctl00_ctl00_ChildBodyContent_PageContent_trWeight td:nth-of-type(2)"
        )
        details["Weight"] = extract_number(wt.text)
    except:
        pass

    return details

# -------------------------- Main --------------------------

def main():
    if not INPUT_XLSX.exists():
        print("❌ Input file not found")
        return

    wb = load_workbook(INPUT_XLSX)
    driver = connect_driver()

    # ===== Excel layout (same like your screenshot) =====
    HEADER_ROW = 4
    START_ROW = HEADER_ROW + 1

    URL_COL = 3            # C = Product URL
    PRODUCT_NAME_COL = 5   # E = Product Name
    PRODUCT_FAMILY_COL = 7 # G = Product Family (we will set it)

    OUT_START_COL = 8      # H = Description starts here

    output_headers = [
       "Product Family Id", "Description", "Weight", "Width", "Depth", "Diameter", "Height",
        "Seat Height", "Seat Depth", "Seat Width", "Arm Height", "Com", "col"
    ]
    # ================================================

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📄 Processing sheet: {sheet_name}")

            max_row = ws.max_row

            # Write output headers on row 4 from column H (do NOT touch A-G)
            for i, header in enumerate(output_headers):
                ws.cell(row=HEADER_ROW, column=OUT_START_COL + i, value=header)

            for row in range(START_ROW, max_row + 1):
                url = ws.cell(row=row, column=URL_COL).value
                if not url:
                    continue

                url = str(url).strip()
                product_name = ws.cell(row=row, column=PRODUCT_NAME_COL).value or ""
                product_name = str(product_name).strip()

                print(f"[Row {row}] Scraping → {url}")
                data = extract_product_details(driver, url)

                # Write details starting from column H
                for i, key in enumerate(output_headers):
                    ws.cell(row=row, column=OUT_START_COL + i, value=data.get(key, ""))

                # Product Family Id = Product Name before " - " (save in column G)
                family_id = product_name.split(" - ")[0].strip() if product_name else ""
                ws.cell(row=row, column=PRODUCT_FAMILY_COL, value=family_id)

                time.sleep(0.5)

        wb.save(OUTPUT_XLSX)
        print(f"\n✅ FINAL FILE SAVED: {OUTPUT_XLSX}")

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
