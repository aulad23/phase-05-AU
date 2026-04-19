# -*- coding: utf-8 -*-
# Chaddock.com Step-1 Scraper (NiermannWeeks-style I/O)
# Collect: Category, Product URL, Image URL, Product Name, SKU
# Handles: View All + Lazy Load (auto scroll)
# Outputs (same folder as script):
#   1) chaddock_all_products.xlsx  (master list)
#   2) Chaddock.xlsx              (category-wise sheets)

import os
import time
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM
# =========================================================

CATEGORIES = {
    "Nightstands & Accent Tables": [
        "https://chaddock.com/styles?ProdType=Nightstands|Tables%3aAccent+Tables|Tables%3aDrink+Tables|Tables%3aEnd+Tables|Tables%3aSide+Tables"
    ],
    "Coffee & Cocktail Tables": [
        "https://chaddock.com/styles?ProdType=Tables%3aCoffee+Tables|Tables%3aCocktail+Tables|Tables%3aOccasional+Tables"
    ],
    "Dining Tables": [
        "https://chaddock.com/styles?PageSize=2000000000&ProdType=Tables%3aDining+Tables|Tables%3aOval+Dining+Tables|Tables%3aRectangular+Dining+Tables|Tables%3aRound+Dining+Tables"
    ],
    "Consoles": [
        "https://chaddock.com/styles?ProdType=Credenza|Tables%3aconsole+tables&PageIndex=3"
    ],
    "Beds & Headboards": [
        "https://chaddock.com/styles?ProdType=beds"
    ],
    "Desks": [
        "https://chaddock.com/styles?ProdType=Desks"
    ],
    "Dining Chairs": [
        "https://chaddock.com/styles?ProdType=chairs%3aDining+Chairs"
    ],
    "Bar Stools": [
        "https://chaddock.com/styles?ProdType=Stools"
    ],
    "Sofas & Loveseats": [
        "https://chaddock.com/styles?ProdType=Sofas"
    ],
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "chaddock_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "Chaddock.xlsx")

# =========================================================
# SCRAPING SETTINGS
# =========================================================

ROOT_URL = "https://chaddock.com"
WAIT_TIMEOUT = 15

# Lazy-load scrolling config
SCROLL_PAUSE = 0.7
MAX_SCROLL_LOOPS = 80     # safety
NO_GROWTH_LIMIT = 4       # stop after N loops with no new products

# =========================================================
# DRIVER
# =========================================================

def connect_driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts
    )

# =========================================================
# VIEW ALL PAGINATION (Fix click intercepted by GA banner)
# =========================================================

def view_all_items(driver):
    try:
        def hide_ga_banner():
            try:
                driver.execute_script("""
                    var b = document.getElementById('divGABanner');
                    if (b) { b.remove(); }
                """)
            except:
                pass

        def js_click(by, selector, timeout=10):
            el = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, selector))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", el)
            return el

        hide_ga_banner()

        # Step 1: set dropdown to "All"
        ddl_id = "ctl00_ctl00_ChildBodyContent_ContentPlaceHolderFullWidth_RadDataPagerBottom_ctl03_ddlPageSize"
        ddl = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, ddl_id))
        )
        driver.execute_script("""
            arguments[0].value='2000000000';
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
        """, ddl)

        time.sleep(0.8)
        hide_ga_banner()

        # Step 2: click "View All"
        btn_id = "ctl00_ctl00_ChildBodyContent_ContentPlaceHolderFullWidth_RadDataPagerBottom_ctl03_btnViewAll"
        try:
            js_click(By.ID, btn_id, timeout=10)
        except ElementClickInterceptedException:
            hide_ga_banner()
            js_click(By.ID, btn_id, timeout=10)

        # Step 3: wait for first batch
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.grid_3.SearchResults_Container"))
        )
        print("View All applied (if available).")

    except Exception as e:
        print(f"Could not click View All: {e}. Will continue with current page.")

# =========================================================
# LAZY LOAD SCROLL (loads all products)
# =========================================================

def load_all_by_scrolling(driver):
    """
    Scrolls to the bottom repeatedly until:
    - product count stops increasing for NO_GROWTH_LIMIT loops, or
    - MAX_SCROLL_LOOPS reached
    """
    last_count = 0
    no_growth = 0

    for _ in range(MAX_SCROLL_LOOPS):
        cards = driver.find_elements(By.CSS_SELECTOR, "div.grid_3.SearchResults_Container")
        current_count = len(cards)

        if current_count > last_count:
            last_count = current_count
            no_growth = 0
        else:
            no_growth += 1

        # stop if no new products came after a few scrolls
        if no_growth >= NO_GROWTH_LIMIT:
            break

        # scroll down
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE)

    print(f"Lazy-load scroll done. Total cards loaded: {last_count}")

# =========================================================
# SCRAPING LOGIC
# =========================================================

def extract_products(driver, category_name):
    products = []
    seen_urls = set()

    cards = driver.find_elements(By.CSS_SELECTOR, "div.grid_3.SearchResults_Container")

    for card in cards:
        try:
            link_el = card.find_element(By.CSS_SELECTOR, "a")
            product_url = urljoin(ROOT_URL, link_el.get_attribute("href"))
        except:
            product_url = ""

        if not product_url or product_url in seen_urls:
            continue
        seen_urls.add(product_url)

        try:
            img_el = card.find_element(By.CSS_SELECTOR, "img")
            img_url = urljoin(ROOT_URL, img_el.get_attribute("src"))
        except:
            img_url = ""

        try:
            title_el = card.find_element(By.CSS_SELECTOR, "span.CHAD_SearchResult_Title")
            text_lines = title_el.text.split("\n")
            name = text_lines[0].strip() if len(text_lines) > 0 else ""
            sku = text_lines[1].strip() if len(text_lines) > 1 else ""
        except:
            name = ""
            sku = ""

        if product_url and name:
            products.append({
                "Category": category_name,
                "Product URL": product_url,
                "Image URL": img_url,
                "Product Name": name,
                "SKU": sku
            })

    return products

# =========================================================
# OUTPUT SYSTEM
# =========================================================

def build_step1_master_excel(all_products):
    if not all_products:
        empty = pd.DataFrame(columns=["Category", "Product URL", "Image URL", "Product Name", "SKU"])
        empty.to_excel(master_output_file, index=False)
        return empty

    df = pd.DataFrame(all_products)
    df.drop_duplicates(subset=["Product URL"], inplace=True)
    df = df[["Category", "Product URL", "Image URL", "Product Name", "SKU"]]
    df.to_excel(master_output_file, index=False)
    return df

def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        sheet_name = cat[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = "Brand"
        ws["B1"] = "Chaddock"
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)

# =========================================================
# MAIN
# =========================================================

def main():
    driver = connect_driver()
    all_products = []

    try:
        for cat_name, urls in CATEGORIES.items():
            for base_url in urls:
                driver.get(base_url)

                # View All (if exists)
                view_all_items(driver)

                # IMPORTANT: Lazy-load scroll to load everything
                load_all_by_scrolling(driver)

                products = extract_products(driver, cat_name)
                print(f"[{cat_name}] Products scraped: {len(products)}")
                all_products.extend(products)

                time.sleep(1)

    finally:
        driver.quit()

    df_master = build_step1_master_excel(all_products)

    if not df_master.empty:
        build_category_wise_workbook_from_df(df_master)
    else:
        Workbook().save(master_sheets_file)

    print("\nDone.")
    print("Saved:", master_output_file)
    print("Saved:", master_sheets_file)

if __name__ == "__main__":
    main()
