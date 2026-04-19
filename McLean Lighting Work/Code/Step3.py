# =========================================================
# MCLEAN LIGHTING SCRAPER
# - Scraping logic preserved EXACTLY from CODE-A
# - INPUT system + OUTPUT system integrated from CODE-B
# - Production-ready single file
# =========================================================

import os
import time
import requests
import pandas as pd

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# =========================================================
# INPUT SYSTEM (COPIED FROM CODE-B STRUCTURE)
# =========================================================

CATEGORIES = {
    "Lighting": [
        "https://www.mcleanlighting.com/product-category/lighting/lighting-on-hand/",
        "https://www.mcleanlighting.com/product-category/lighting/antique/"
    ],
    "Chandeliers": [
        "https://www.mcleanlighting.com/product-category/lighting/chandeliers/"
    ],

    "Pendants": [
        "https://www.mcleanlighting.com/product-category/lighting/exterior-hanging/",
    ],

    "Sconces": [
        "https://www.mcleanlighting.com/product-category/lighting/exterior-wall-mount/",
        "https://www.mcleanlighting.com/product-category/lighting/sconces/"
    ],

    "Lanterns": [
        "https://www.mcleanlighting.com/product-category/lighting/post-mount/"
    ],

}

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

VENDOR_NAME = "McLean Lighting"
VENDOR_CODE = VENDOR_NAME.replace(" ", "")[:3].upper()  # MCL

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_OUTPUT_FILE = os.path.join(SCRIPT_DIR, "mclean_all_products.xlsx")
CATEGORY_WORKBOOK_FILE = os.path.join(SCRIPT_DIR, "McLeanLighting.xlsx")


# =========================================================
# SCRAPING LOGIC (UNCHANGED FROM CODE-A)
# =========================================================

def get_category_code(category_url):
    slug = category_url.rstrip("/").split("/")[-1]
    clean_slug = slug.replace("-", "")
    return clean_slug[:2].upper()


def scrape_products():
    all_products = []
    product_index = 1

    for category, urls in CATEGORIES.items():
        for url in urls:
            print(f"Scraping: {url}")

            response = requests.get(url, headers=HEADERS)
            soup = BeautifulSoup(response.text, "lxml")

            category_code = get_category_code(url)

            for li in soup.find_all("li", class_="product"):
                a_tag = li.find("a", href=True)
                img_tag = li.find("img")
                h3_tag = li.find("h3")

                product_url = a_tag["href"] if a_tag else ""
                image_url = img_tag["src"] if img_tag else ""
                product_name = h3_tag.get_text(strip=True) if h3_tag else ""

                sku = f"{VENDOR_CODE}{category_code}{product_index}"

                all_products.append({
                    "Category": category,
                    "Product URL": product_url,
                    "Image URL": image_url,
                    "Product Name": product_name,
                    "SKU": sku,
                })

                product_index += 1

            time.sleep(1)

    return all_products


# =========================================================
# OUTPUT SYSTEM (COPIED FROM CODE-B STRUCTURE)
# =========================================================

def build_master_excel(rows):
    columns = [
        "Category",
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
    ]

    df = pd.DataFrame(rows, columns=columns)
    df.drop_duplicates(subset=["Product URL"], inplace=True)

    df.to_excel(MASTER_OUTPUT_FILE, index=False)

    print(f"\nTOTAL PRODUCTS SAVED: {len(df)}")
    print(f"Master file saved: {MASTER_OUTPUT_FILE}\n")

    return df


def build_category_wise_workbook(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for category, urls in CATEGORIES.items():
        df_cat = df[df["Category"] == category]
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=category[:31])

        ws["A1"] = "Brand"
        ws["B1"] = VENDOR_NAME
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(urls)
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat = df_cat.copy()
        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start_row = 4

        for col_idx, col_name in enumerate(df_cat.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col_name).font = bold

        for row_idx, row in enumerate(df_cat.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Hyperlink Product URL
        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=3).value
            cell = ws.cell(row=r, column=4)
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(CATEGORY_WORKBOOK_FILE)
    print(f"Category workbook saved: {CATEGORY_WORKBOOK_FILE}")


# =========================================================
# MAIN ENTRY POINT (FROM CODE-B PATTERN)
# =========================================================

def main():
    rows = scrape_products()

    if not rows:
        print("No products scraped.")
        return

    df = build_master_excel(rows)
    build_category_wise_workbook(df)


if __name__ == "__main__":
    main()
