# =========================================================
# MCLEAN LIGHTING – STEP 2 DETAILED SCRAPER
# INPUT / OUTPUT SYSTEM UPDATED AS REQUESTED
# ❌ SCRAPING + PARSING LOGIC UNCHANGED
# =========================================================

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM (UPDATED)
# =========================================================

INPUT_FILE = "McLeanLighting.xlsx"
OUTPUT_FILE = "McLeanLighting_Detailed.xlsx"

headers = {
    "User-Agent": "Mozilla/5.0"
}

# =========================================================
# READ INPUT EXCEL (CATEGORY-WISE SHEETS)
# Input Format:
# A1: "Brand"      B1: Brand Name Value
# A2: "Link"       B2: Link Value
# Row 3: Empty
# Row 4: Headers (Index, Category, Product URL, Image URL, Product Name, SKU)
# Row 5+: Data
# =========================================================

all_rows = []

wb = load_workbook(INPUT_FILE)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Read brand and link from B1 and B2
    brand = ws["B1"].value
    link = ws["B2"].value

    # Data starts from row 5 (row 4 contains headers)
    # A4: Index, B4: Category, C4: Product URL, D4: Image URL, E4: Product Name, F4: SKU
    for row in ws.iter_rows(min_row=5, values_only=True):
        # row[0] = Index (A column)
        # row[1] = Category (B column)
        # row[2] = Product URL (C column)
        # row[3] = Image URL (D column)
        # row[4] = Product Name (E column)
        # row[5] = SKU (F column)

        if not row[2]:  # Product URL empty
            continue

        all_rows.append({
            "Brand": brand,
            "Link": link,  # Store the link from B2
            "Category": row[1],
            "Product URL": row[2],
            "Image URL": row[3],
            "Product Name": row[4],
            "SKU": row[5],
        })

df = pd.DataFrame(all_rows)


# =========================================================
# SCRAPING / PARSING LOGIC (UNCHANGED)
# =========================================================

def get_product_family_id(product_name):
    if not product_name:
        return ""
    if ',' in product_name:
        return product_name.split(',')[0].strip()
    elif '-' in product_name:
        return product_name.split('-')[0].strip()
    else:
        return product_name.strip()


def extract_dimension(description_html):
    if not description_html:
        return ""

    dimensions = []

    pattern_letter = r'(?:W|H|D|Diam\.?|Dia\.?|Width|Height|Depth|Diameter|Weight)\s*\d+(?:\.\d+)?[″"]?(?:\s*(?:Ibs?|IBS|lbs?))?'
    dimensions.extend(re.findall(pattern_letter, description_html, re.IGNORECASE))

    pattern_number = r'\d+(?:\.\d+)?[″"]?\s*(?:W|H|D|Diam\.?|Dia\.?)'
    dimensions.extend(re.findall(pattern_number, description_html, re.IGNORECASE))

    seen = set()
    unique = []
    for d in dimensions:
        d = d.strip()
        if d not in seen:
            seen.add(d)
            unique.append(d)

    return " ".join(unique[:5])


def parse_dimensions(dimension_str):
    result = dict.fromkeys(["Width", "Height", "Depth", "Diameter", "Weight"], "")

    if not dimension_str:
        return result

    patterns = {
        "Width": [r'(?:W|Width)[\s:]*(\d+(?:\.\d+)?)', r'(\d+(?:\.\d+)?)\s*W'],
        "Height": [r'(?:H|Height)[\s:]*(\d+(?:\.\d+)?)', r'(\d+(?:\.\d+)?)\s*H'],
        "Depth": [r'(?:D|Depth)[\s:]*(\d+(?:\.\d+)?)', r'(\d+(?:\.\d+)?)\s*D'],
        "Diameter": [r'(?:Diam\.?|Dia\.?|Diameter)[\s:]*(\d+(?:\.\d+)?)'],
        "Weight": [r'(?:Weight)[\s:]*(\d+(?:\.\d+)?)']
    }

    for key, pats in patterns.items():
        for pat in pats:
            m = re.search(pat, dimension_str, re.IGNORECASE)
            if m:
                result[key] = m.group(1)
                break

    return result


def extract_list_price(description_html):
    m = re.search(r'\$(\d+(?:,\d+)?)', description_html or "")
    return m.group(1).replace(",", "") if m else ""


def extract_finishes(description_html):
    if not description_html:
        return ""

    soup = BeautifulSoup(description_html, "html.parser")
    text = soup.get_text()

    finishes = []

    if "Finish" in text:
        parts = re.split(r'Finishes?:', text, flags=re.IGNORECASE)
        if len(parts) > 1:
            section = parts[1]
            section = re.split(r'(Configurations|Sizes?|$)', section)[0]
            for line in section.split("\n"):
                line = line.strip()
                if line.startswith("•"):
                    finishes.append(re.sub(r'\s*\(.*?\)', '', line[1:].strip()))

    return ", ".join(finishes)


# =========================================================
# DETAILED SCRAPING
# =========================================================

detailed_rows = []

for i, row in df.iterrows():
    print(f"[{i + 1}/{len(df)}] {row['Product Name']}")

    try:
        r = requests.get(row["Product URL"], headers=headers, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")

        desc_div = soup.find("div", {"itemprop": "description"})
        description = desc_div.get_text("\n", strip=True) if desc_div else ""

        dimension = extract_dimension(description)
        parsed = parse_dimensions(dimension)

        detailed_rows.append({
            "Brand": row["Brand"],
            "Link": row["Link"],  # Include the link
            "Category": row["Category"],
            "Product URL": row["Product URL"],
            "Image URL": row["Image URL"],
            "Product Name": row["Product Name"],
            "SKU": row["SKU"],
            "Product Family": get_product_family_id(row["Product Name"]),
            "Description": description,
            "Weight": parsed["Weight"],
            "Width": parsed["Width"],
            "Depth": parsed["Depth"],
            "Diameter": parsed["Diameter"],
            "Height": parsed["Height"],
            "List Price": extract_list_price(description),
            "Finish": extract_finishes(description),
            "Dimension": dimension
        })

        time.sleep(1)

    except Exception:
        detailed_rows.append({
            "Brand": row["Brand"],
            "Link": row["Link"],  # Include the link
            "Category": row["Category"],
            "Product URL": row["Product URL"],
            "Image URL": row["Image URL"],
            "Product Name": row["Product Name"],
            "SKU": row["SKU"],
            "Product Family": get_product_family_id(row["Product Name"]),
            "Description": "",
            "Weight": "",
            "Width": "",
            "Depth": "",
            "Diameter": "",
            "Height": "",
            "List Price": "",
            "Finish": "",
            "Dimension": ""
        })

# =========================================================
# OUTPUT SYSTEM (CATEGORY-WISE SHEETS)
# Each sheet will have same format as input:
# A1: "Brand", B1: Brand Name
# A2: "Link", B2: Link
# Row 3: Empty
# Row 4: Headers
# Row 5+: Data
# =========================================================

from openpyxl import Workbook

out_df = pd.DataFrame(detailed_rows)

# Group by Brand and Category
grouped = out_df.groupby(['Brand', 'Category'])

wb_out = Workbook()
wb_out.remove(wb_out.active)  # Remove default sheet

for (brand, category), group_df in grouped:
    # Create sheet name (limit to 31 chars for Excel)
    sheet_name = category[:31] if len(category) <= 31 else category[:28] + "..."

    ws = wb_out.create_sheet(title=sheet_name)

    # A1: "Brand", B1: Brand name
    ws["A1"] = "Brand"
    ws["B1"] = brand

    # A2: "Link", B2: Link from input
    ws["A2"] = "Link"
    ws["B2"] = group_df.iloc[0]["Link"]  # Get link from first row of this group

    # Row 3 is empty

    # Row 4: Headers
    headers = [
        "Index",
        "Category",
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
        "Product Family",
        "Description",
        "Weight",
        "Width",
        "Depth",
        "Diameter",
        "Height",
        "List Price",
        "Finish",
        "Dimension"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Row 5+: Data
    for idx, (_, row) in enumerate(group_df.iterrows(), start=1):
        data_row = [
            idx,  # Index
            row["Category"],
            row["Product URL"],
            row["Image URL"],
            row["Product Name"],
            row["SKU"],
            row["Product Family"],
            row["Description"],
            row["Weight"],
            row["Width"],
            row["Depth"],
            row["Diameter"],
            row["Height"],
            row["List Price"],
            row["Finish"],
            row["Dimension"]
        ]

        for col_num, value in enumerate(data_row, start=1):
            ws.cell(row=4 + idx, column=col_num, value=value)

# Save workbook
wb_out.save(OUTPUT_FILE)

print("\n✅ STEP 2 COMPLETE")
print(f"📦 Products processed: {len(out_df)}")
print(f"📑 Sheets created: {len(wb_out.sheetnames)}")
print(f"💾 Output saved: {OUTPUT_FILE}")