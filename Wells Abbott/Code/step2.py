"""
Wells Abbott - Step 2: Product Detail Scraper
==============================================
Input:  wells_abbott_products.xlsx (Step 1 output)
Output: wells_abbott_products_details.xlsx

Extracts per product URL:
  - SKU
  - Product Family Id  (Product Name থেকে প্রথম অংশ, " - " এর আগে)
  - Description
  - Dimensions (raw)
  - Weight, Width, Depth, Diameter, Height, Seat Height, Seat Depth, Seat Width, Arm Height
  - Note

Requirements:
    pip install playwright beautifulsoup4 openpyxl
    playwright install chromium
"""

import re
import sys
import time

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("❌ Run: pip install playwright && playwright install chromium")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("❌ Run: pip install beautifulsoup4")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("❌ Run: pip install openpyxl")
    sys.exit(1)

# ─── CONFIG ───────────────────────────────────────────────────────────────────

INPUT_FILE  = "wells_abbott_Baskets_Planters.xlsx"
OUTPUT_FILE = "wells_abbott_Baskets_Planters_details.xlsx"

# ─── FRACTION → DECIMAL HELPERS ──────────────────────────────────────────────

# Unicode fraction characters → decimal
UNICODE_FRACTIONS = {
    '½': 0.5, '⅓': 1/3, '⅔': 2/3, '¼': 0.25, '¾': 0.75,
    '⅕': 0.2, '⅖': 0.4, '⅗': 0.6, '⅘': 0.8,
    '⅙': 1/6, '⅚': 5/6, '⅛': 0.125, '⅜': 0.375, '⅝': 0.625, '⅞': 0.875,
}


SUPERSCRIPT_DIGITS = {'⁰':'0','¹':'1','²':'2','³':'3','⁴':'4','⁵':'5','⁶':'6','⁷':'7','⁸':'8','⁹':'9'}
SUBSCRIPT_DIGITS   = {'₀':'0','₁':'1','₂':'2','₃':'3','₄':'4','₅':'5','₆':'6','₇':'7','₈':'8','₉':'9'}


def normalize_super_sub_fractions(text: str) -> str:
    """
    Convert superscript⁄subscript fraction patterns to plain text fractions.
    E.g.  13⁹⁄₁₆  →  13 9/16
    """
    # Pattern: optional digits + superscript digits + fraction slash (⁄ U+2044) + subscript digits
    import re
    def _replace(m):
        prefix = m.group(1) or ''
        sup = ''.join(SUPERSCRIPT_DIGITS.get(c, c) for c in m.group(2))
        sub = ''.join(SUBSCRIPT_DIGITS.get(c, c) for c in m.group(3))
        if prefix:
            return f"{prefix} {sup}/{sub}"
        return f"{sup}/{sub}"

    text = re.sub(
        r'(\d+)?([⁰¹²³⁴⁵⁶⁷⁸⁹]+)⁄([₀₁₂₃₄₅₆₇₈₉]+)',
        _replace, text
    )
    return text


def fraction_to_decimal(text: str) -> str:
    """
    Convert fraction strings to decimal.
    Handles:
      - Unicode fractions:  41¾  → 41.75
      - Text fractions:     1 3/4 → 1.75,  31 7/8 → 31.875
      - Plain numbers:      29 → 29,  22.5 → 22.5
    Returns cleaned string with decimal number.
    """
    text = text.strip()

    # Normalize superscript⁄subscript fractions first (e.g. 13⁹⁄₁₆ → 13 9/16)
    text = normalize_super_sub_fractions(text)
    for uf, dec in UNICODE_FRACTIONS.items():
        if uf in text:
            text = text.replace(uf, '')
            # If there's a whole number before the fraction
            text = text.strip()
            if text:
                try:
                    return str(round(float(text) + dec, 4))
                except ValueError:
                    return str(round(dec, 4))
            else:
                return str(round(dec, 4))

    # Match pattern like "31 7/8" (whole + space + fraction)
    m = re.match(r'^(\d+)\s+(\d+)\s*/\s*(\d+)$', text)
    if m:
        whole = int(m.group(1))
        num   = int(m.group(2))
        den   = int(m.group(3))
        return str(round(whole + num / den, 4))

    # Match pattern like "7/8" (fraction only)
    m = re.match(r'^(\d+)\s*/\s*(\d+)$', text)
    if m:
        num = int(m.group(1))
        den = int(m.group(2))
        return str(round(num / den, 4))

    # Already a plain number
    try:
        return str(round(float(text), 4))
    except ValueError:
        return text


# ─── DIMENSION PARSER ────────────────────────────────────────────────────────

def parse_dimensions(dim_text: str) -> dict:
    """
    Parse a raw dimension string into separate fields.

    Supported formats / abbreviations:
      - 29"W  33"D  32"H   (compact with W/D/H suffixes)
      - 32"W x 35.5"D x 31"H
      - 85.25"W x 31"D x 28.75"H | 17"SH
      - OVERALL: 29"W 33"D 32"H | SEAT HEIGHT: 18" | SEAT DEPTH: 22.5"
      - H (seat) 16 inches, width 23.5 inches, depth 29 inches
      - Depth 1 3/4"  Width 31 7/8"  Height 78 3/4"
      - height 41¾in width 21¾in depth 2in
      - Ib / IBS / Ibs / lbs / lb → Weight
      - Diam. / Dia / DIA / dia → Diameter
      - SH → Seat Height,  SD → Seat Depth,  SW → Seat Width,  AH → Arm Height
    """
    result = {
        "Weight":      "",
        "Width":       "",
        "Depth":       "",
        "Diameter":    "",
        "Length":      "",
        "Height":      "",
        "Seat Height": "",
        "Seat Depth":  "",
        "Seat Width":  "",
        "Arm Height":  "",
    }

    if not dim_text or dim_text == "":
        return result

    # Normalise whitespace
    text = dim_text.strip()

    # Normalise smart/curly quotes to straight quotes
    text = text.replace('\u201c', '"').replace('\u201d', '"')  # " "
    text = text.replace('\u2018', "'").replace('\u2019', "'")  # ' '

    # Normalise superscript/subscript fractions (e.g. 13⁹⁄₁₆ → 13 9/16)
    text = normalize_super_sub_fractions(text)

    # If multiple variants separated by known labels like "Rectangular:" / "Waist-shaped:",
    # take only the FIRST variant.
    variant_split = re.split(r'(?:Waist-shaped|Oval|Round|Square|Circular)\s*:', text, flags=re.IGNORECASE)
    if len(variant_split) > 1:
        # Keep everything before the second variant label
        # But we need the first variant — find from start up to second label
        first_label_match = re.match(r'^(.*?)(?:Rectangular|Standard|First)\s*:\s*', text, re.IGNORECASE)
        if first_label_match:
            # strip the label prefix from the first variant
            text_after_label = text[first_label_match.end():]
            # take up to the next variant label
            next_variant = re.search(r'(?:Waist-shaped|Oval|Round|Square|Circular)\s*:', text_after_label, re.IGNORECASE)
            if next_variant:
                text = text_after_label[:next_variant.start()]
            else:
                text = text_after_label
        else:
            # No explicit first label; just take everything before the second variant
            next_variant = re.search(r'(?:Waist-shaped|Oval|Round|Square|Circular)\s*:', text, re.IGNORECASE)
            if next_variant:
                text = text[:next_variant.start()]

    # Also handle "Rectangular:" at the start — strip it
    text = re.sub(r'^(?:Rectangular|Standard|First)\s*:\s*', '', text, flags=re.IGNORECASE)

    # Remove "OVERALL:" prefix
    text = re.sub(r'OVERALL\s*:\s*', '', text, flags=re.IGNORECASE)

    # Remove (cm) / (mm) metric parentheticals
    text = re.sub(r'\([^)]*(?:cm|mm|m)[^)]*\)', '', text, flags=re.IGNORECASE)

    # ── Step 1: Extract labelled multi-word keys first (SEAT HEIGHT, ARM HEIGHT, etc.) ──

    # Pattern for SEAT HEIGHT / SEAT DEPTH / SEAT WIDTH / ARM HEIGHT (full label)
    labelled_patterns = [
        (r'(?:SEAT\s*HEIGHT|seat\s*height)\s*[:\s]*'                          , "Seat Height"),
        (r'(?:SEAT\s*DEPTH|seat\s*depth)\s*[:\s]*'                            , "Seat Depth"),
        (r'(?:SEAT\s*WIDTH|seat\s*width)\s*(?:front|rear|back)?\s*[:\s]*'     , "Seat Width"),
        (r'(?:ARM\s*HEIGHT|arm\s*height)\s*[:\s]*'                            , "Arm Height"),
    ]

    # Value: digits + optional unicode fraction or space-fraction or decimal
    labelled_val_pat = r'(\d+(?:[½⅓⅔¼¾⅕⅖⅗⅘⅙⅚⅛⅜⅝⅞])?(?:\s+\d+/\d+)?(?:\.\d+)?)\s*(?:"|in(?:ches)?|\'|)'

    for pat, key in labelled_patterns:
        m = re.search(pat + labelled_val_pat, text, re.IGNORECASE)
        if m:
            result[key] = fraction_to_decimal(m.group(1))
            # Remove matched portion to avoid double-matching
            text = text[:m.start()] + " " + text[m.end():]

    # ── Step 1b: H (seat) pattern → Seat Height ──
    m = re.search(r'H\s*\(\s*seat\s*\)\s*[:\s]*' + labelled_val_pat, text, re.IGNORECASE)
    if m:
        result["Seat Height"] = fraction_to_decimal(m.group(1))
        text = text[:m.start()] + " " + text[m.end():]

    # ── Step 1c: SH / SD / SW / AH abbreviations (e.g. 17"SH) ──
    abbrev_map = {
        'SH': 'Seat Height', 'SD': 'Seat Depth', 'SW': 'Seat Width', 'AH': 'Arm Height',
    }
    for abbr, key in abbrev_map.items():
        # e.g. 17"SH or 17" SH
        m = re.search(r'(\d+(?:\.\d+)?)\s*"\s*' + abbr + r'\b', text, re.IGNORECASE)
        if m and not result[key]:
            result[key] = fraction_to_decimal(m.group(1))
            text = text[:m.start()] + " " + text[m.end():]
        # e.g. SH: 17" or SH 17 (word boundary before to avoid OAH→AH)
        m = re.search(r'\b' + abbr + r'\s*[:\s]\s*(\d+(?:\.\d+)?)\s*(?:"|in)?', text, re.IGNORECASE)
        if m and not result[key]:
            result[key] = fraction_to_decimal(m.group(1))
            text = text[:m.start()] + " " + text[m.end():]

    # ── Step 2: Weight (Ib, IBS, Ibs, lbs, lb) ──
    m = re.search(r'(\d+(?:\.\d+)?)\s*(?:Ibs|IBS|Ib|lbs?|pounds?)\b', text, re.IGNORECASE)
    if m:
        result["Weight"] = fraction_to_decimal(m.group(1))
        text = text[:m.start()] + " " + text[m.end():]
    else:
        # "Weight: 30" or "weight 30 lbs"
        m = re.search(r'(?:weight)\s*[:\s]*(\d+(?:\.\d+)?)\s*(?:Ibs|IBS|Ib|lbs?|pounds?|"|)?\b', text, re.IGNORECASE)
        if m:
            result["Weight"] = fraction_to_decimal(m.group(1))
            text = text[:m.start()] + " " + text[m.end():]

    # ── Step 3: Diameter (Diam. / Dia / DIA / Diameter) ──
    # Label-first: "Diameter 60" / "Diam. 54"
    m = re.search(r'(?:Diam\.?|Dia\.?|DIA|diameter)\s*[:\s]*(\d+(?:\s+\d+/\d+)?(?:\.\d+)?)\s*(?:"|in(?:ches)?|)', text, re.IGNORECASE)
    if m:
        result["Diameter"] = fraction_to_decimal(m.group(1))
        text = text[:m.start()] + " " + text[m.end():]
    else:
        # Compact: 60"Diameter / 66inDIA / 54"Diam
        m = re.search(r'(\d+(?:\.\d+)?)\s*(?:"|in)\s*(?:Diameter|Diam\.?|DIA)\b', text, re.IGNORECASE)
        if m:
            result["Diameter"] = fraction_to_decimal(m.group(1))
            text = text[:m.start()] + " " + text[m.end():]

    # ── Step 4: Compact format — number"W / number"D / number"H / number"L ──
    # Handles: 29"W  33"D  32"H  or  32"W x 35.5"D x 31"H
    # Number can include unicode fractions or text fractions

    # number + optional fraction + inch mark + letter
    num_pat = r'(\d+(?:\s+\d+/\d+)?(?:[½⅓⅔¼¾⅕⅖⅗⅘⅙⅚⅛⅜⅝⅞])?(?:\.\d+)?)'

    compact_w = re.search(num_pat + r'\s*(?:"|in(?:ches)?|\')\s*W\b', text, re.IGNORECASE)
    compact_d = re.search(num_pat + r'\s*(?:"|in(?:ches)?|\')\s*D\b', text, re.IGNORECASE)
    compact_h = re.search(num_pat + r'\s*(?:"|in(?:ches)?|\')\s*(?:H|OAH)\b', text, re.IGNORECASE)
    compact_l = re.search(num_pat + r'\s*(?:"|in(?:ches)?|\')\s*L\b', text, re.IGNORECASE)

    if compact_w and not result["Width"]:
        result["Width"] = fraction_to_decimal(compact_w.group(1))
    if compact_d and not result["Depth"]:
        result["Depth"] = fraction_to_decimal(compact_d.group(1))
    if compact_h and not result["Height"]:
        result["Height"] = fraction_to_decimal(compact_h.group(1))
    if compact_l and not result["Length"]:
        result["Length"] = fraction_to_decimal(compact_l.group(1))

    # If we found compact format, we're done with W/D/H/L — skip to end
    found_compact = compact_w or compact_d or compact_h or compact_l

    if not found_compact:
        # ── Step 5: Label-first format — Width 31 7/8" or width 23.5 inches ──
        label_map = {
            r'(?<!\bSEAT\s)(?<!\bARM\s)(?:width|\bW\b)\s*(?:\([^)]*\)\s*)?[:\s]*':  "Width",
            r'(?<!\bSEAT\s)(?:depth|\bD\b)\s*(?:\([^)]*\)\s*)?[:\s]*':               "Depth",
            r'(?<!\bSEAT\s)(?<!\bARM\s)(?:height|OAH|\bH\b)\s*(?:\([^)]*\)\s*)?[:\s]*':  "Height",
            r'(?:length|\bL\b)\s*(?:\([^)]*\)\s*)?[:\s]*':                            "Length",
        }

        for pat, key in label_map.items():
            if result[key]:
                continue
            # Value: digits + optional space-fraction or unicode fraction or decimal
            val_pat = r'(\d+(?:\s+\d+/\d+)?(?:[½⅓⅔¼¾⅕⅖⅗⅘⅙⅚⅛⅜⅝⅞])?(?:\.\d+)?)\s*(?:"|in(?:ches)?|\'|)'
            m = re.search(pat + val_pat, text, re.IGNORECASE)
            if m:
                result[key] = fraction_to_decimal(m.group(1))

    # Clean up: strip trailing zeros like 29.0 → 29, but keep 22.5
    for k in result:
        if result[k]:
            try:
                val = float(result[k])
                result[k] = str(int(val)) if val == int(val) else str(val)
            except (ValueError, TypeError):
                pass

    return result


# ─── NOTE PARSER ─────────────────────────────────────────────────────────────

def parse_note(note_text: str) -> dict:
    """
    Parse Note field to extract COM, COL, COT numeric values.

    Examples:
      "COM 4 Yds COL 72 Sq Ft COT 8.5 Yds"
        → COM: 4, COL: 72, COT: 8.5
      "Com 3.5 Yds"
        → COM: 3.5
    """
    result = {"COM": "", "COL": "", "COT": ""}

    if not note_text or note_text == "":
        return result

    text = note_text.strip()

    # COM / Com / com  → followed by optional ":" then a number
    m = re.search(r'\bCOM\s*[:\s]\s*(\d+(?:\.\d+)?)', text, re.IGNORECASE)
    if m:
        result["COM"] = m.group(1)

    # COL / Col / col
    m = re.search(r'\bCOL\s*[:\s]\s*(\d+(?:\.\d+)?)', text, re.IGNORECASE)
    if m:
        result["COL"] = m.group(1)

    # COT / Cot / cot
    m = re.search(r'\bCOT\s*[:\s]\s*(\d+(?:\.\d+)?)', text, re.IGNORECASE)
    if m:
        result["COT"] = m.group(1)

    # Clean trailing .0
    for k in result:
        if result[k]:
            try:
                val = float(result[k])
                result[k] = str(int(val)) if val == int(val) else str(val)
            except (ValueError, TypeError):
                pass

    return result


# ─── OTHER HELPERS ───────────────────────────────────────────────────────────

def get_product_family_id(product_name: str) -> str:
    if not product_name or product_name == "":
        return ""
    return product_name.split(" - ")[0].strip()


def extract_label_value(soup, label: str) -> str:
    """Find <p> containing a bold label and return the text after it."""
    for p in soup.select("p"):
        strong = p.find("strong")
        if strong and label.lower() in strong.get_text(strip=True).lower():
            strong.extract()
            value = p.get_text(separator=" ", strip=True).lstrip(": ").strip()
            if value:
                return value
    return ""


def scrape_detail(page, url: str) -> dict:
    result = {
        "SKU":         "",
        "Description": "",
        "Dimensions":  "",
        "Note":        "",
    }

    try:
        page.goto(url, wait_until="networkidle", timeout=60_000)
        page.wait_for_selector("product-info", timeout=15_000)
    except Exception:
        print(f"  ⚠️  Timeout/error loading: {url}")
        return result

    soup = BeautifulSoup(page.content(), "html.parser")

    # SKU
    sku_tag = soup.select_one("span.variant-sku")
    if sku_tag:
        result["SKU"] = sku_tag.get_text(strip=True)

    # Description
    desc_tag = soup.select_one("div.product__description")
    if desc_tag:
        result["Description"] = desc_tag.get_text(separator=" ", strip=True)

    # Dimensions
    result["Dimensions"] = extract_label_value(soup, "Dimensions")

    # If Dimensions is broken (#REF!) or empty, check if Description has dimension data
    dim = result["Dimensions"]
    desc = result["Description"]
    if (not dim or dim == "" or "#REF" in dim) and desc:
        # Check if description looks like dimension data
        dim_pattern = re.compile(
            r'(?:OVERALL\s*:|(?:\d+(?:\.\d+)?)\s*"\s*[WDH]\b|'
            r'(?:width|depth|height|seat\s*height)\s)',
            re.IGNORECASE,
        )
        if dim_pattern.search(desc):
            result["Dimensions"] = desc
            result["Description"] = ""

    # Note
    result["Note"] = extract_label_value(soup, "Note")

    return result


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print(f"📂 Reading: {INPUT_FILE}")
    try:
        wb_in = openpyxl.load_workbook(INPUT_FILE)
    except FileNotFoundError:
        print(f"❌ File not found: {INPUT_FILE}")
        sys.exit(1)

    ws_in = wb_in.active
    rows  = list(ws_in.iter_rows(values_only=True))

    if not rows:
        print("❌ Excel file is empty.")
        sys.exit(1)

    header = rows[0]
    data   = rows[1:]

    try:
        url_idx  = header.index("Product URL")
        img_idx  = header.index("Image URL")
        name_idx = header.index("Product Name")
    except ValueError as e:
        print(f"❌ Missing column in Excel: {e}")
        sys.exit(1)

    print(f"✅ Found {len(data)} product(s) to process\n")

    all_results = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page = browser.new_page(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            )
        )

        for i, row in enumerate(data, 1):
            product_url  = row[url_idx]  if row[url_idx]  else ""
            image_url    = row[img_idx]  if row[img_idx]  else ""
            product_name = row[name_idx] if row[name_idx] else ""
            family_id    = get_product_family_id(product_name)

            print(f"[{i}/{len(data)}] {product_name}")

            if product_url and product_url != "":
                detail = scrape_detail(page, product_url)
            else:
                detail = {
                    "SKU": "", "Description": "",
                    "Dimensions": "", "Note": "",
                }

            # Parse dimensions into separate fields
            dim_parsed = parse_dimensions(detail["Dimensions"])

            # Parse note for COM / COL / COT
            note_parsed = parse_note(detail["Note"])

            all_results.append({
                "Product URL":       product_url,
                "Image URL":         image_url,
                "Product Name":      product_name,
                "Product Family Id": family_id,
                "SKU":               detail["SKU"],
                "Description":       detail["Description"],
                "Dimensions":        detail["Dimensions"],
                "Weight":            dim_parsed["Weight"],
                "Width":             dim_parsed["Width"],
                "Depth":             dim_parsed["Depth"],
                "Diameter":          dim_parsed["Diameter"],
                "Length":            dim_parsed["Length"],
                "Height":            dim_parsed["Height"],
                "Seat Height":       dim_parsed["Seat Height"],
                "Seat Depth":        dim_parsed["Seat Depth"],
                "Seat Width":        dim_parsed["Seat Width"],
                "Arm Height":        dim_parsed["Arm Height"],
                "Note":              detail["Note"],
                "COM":               note_parsed["COM"],
                "COL":               note_parsed["COL"],
                "COT":               note_parsed["COT"],
            })

            time.sleep(0.5)

        browser.close()

    # Save Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Products"

    columns = [
        "Product URL",
        "Image URL",
        "Product Name",
        "SKU",
        "Product Family Id",
        "Description",
        "Dimensions",
        "Weight",
        "Width",
        "Depth",
        "Diameter",
        "Length",
        "Height",
        "Seat Height",
        "Seat Depth",
        "Seat Width",
        "Arm Height",
        "Note",
        "COM",
        "COL",
        "COT",
    ]
    ws_out.append(columns)

    for r in all_results:
        ws_out.append([r[c] for c in columns])

    wb_out.save(OUTPUT_FILE)
    print(f"\n✅ Done! {len(all_results)} products saved → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()