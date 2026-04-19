# -*- coding: utf-8 -*-
# worldsaway_step2_fill_into_master_sheets_REALIGN_AND_SKU.py

import time
import re
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from openpyxl import load_workbook

# ================= CONFIG =================
MASTER_FILE = "WorldsAway.xlsx"
SAVE_AS     = "WorldsAway_with_details.xlsx"

VENDOR_NAME = "Worlds Away"   # used for auto SKU
HEADLESS = True

TIMEOUT_PAGELOAD = 30
WAIT_BODY_SEC = 12
PER_PAGE_SLEEP = 2

HEADER_ROW = 4
DATA_START_ROW = 5

# ✅ Step-2 final columns (includes Size to avoid shifting)
COLUMN_ORDER = [
    "Index", "Category", "Product URL", "Image URL", "Product Name",
    "Size",
    "SKU", "Product Family ID", "Description", "Weight",
    "Width", "Depth", "Diameter", "Height", "Length",
    "Seat Height", "Seat Depth", "Seat Width", "Arm Height"
]
# =========================================


def setup_driver():
    chromedriver_autoinstaller.install()
    chrome_options = Options()
    if HEADLESS:
        chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(options=chrome_options)
    driver.set_page_load_timeout(TIMEOUT_PAGELOAD)
    return driver


def get_header_map(ws):
    """
    Reads current header row cells and builds mapping name->col
    """
    header_map = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            header_map[str(v).strip()] = c
    return header_map


def write_new_header(ws):
    """
    Writes header row exactly as COLUMN_ORDER and returns new map.
    Does NOT touch data rows.
    """
    # clear old header row cells safely
    max_clear_cols = max(ws.max_column, len(COLUMN_ORDER)) + 10
    for c in range(1, max_clear_cols + 1):
        ws.cell(row=HEADER_ROW, column=c, value=None)

    for i, col in enumerate(COLUMN_ORDER, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=col)

    return {col: i for i, col in enumerate(COLUMN_ORDER, start=1)}


def get_last_data_row(ws, col_url):
    last_row = ws.max_row
    while last_row >= DATA_START_ROW and not ws.cell(row=last_row, column=col_url).value:
        last_row -= 1
    return last_row


def normalize_letters(s: str) -> str:
    if not isinstance(s, str):
        return ""
    return re.sub(r"[^A-Za-z0-9]+", "", s.strip().upper())


def generate_sku_if_missing(vendor_name: str, category_name: str, index_value) -> str:
    v = normalize_letters(vendor_name)[:3]
    c = normalize_letters(category_name)[:2]

    idx = ""
    if index_value is not None:
        idx = str(index_value).strip()
        if re.fullmatch(r"\d+\.0", idx):
            idx = str(int(float(idx)))
    if not idx:
        idx = "0"

    if len(v) < 3:
        v = (v + "XXX")[:3]
    if len(c) < 2:
        c = (c + "XX")[:2]

    return f"{v}{c}{idx}"


def extract_dimensions(text, dim_source_only):
    width = depth = height = diameter = length = ""
    if not isinstance(text, str):
        text = ""
    if not isinstance(dim_source_only, str):
        dim_source_only = ""

    text = text.replace("″", '"').replace("’", "'")

    rules = {
        "width": r'(\d+\.?\d*)\s*"?\s*(W|WIDTH)',
        "depth": r'(\d+\.?\d*)\s*"?\s*(D(?!IAM)|DEPTH)',
        "height": r'(\d+\.?\d*)\s*"?\s*(H|HT|HEIGHT)',
        "diameter": r'(\d+\.?\d*)\s*"?\s*(DIAM|DIA)'
    }

    for key, pattern in rules.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            val = match.group(1).strip()
            if key == "width":
                width = val
            if key == "depth":
                depth = val
            if key == "height":
                height = val
            if key == "diameter":
                diameter = val

    if dim_source_only:
        length_rules = [
            r'(\d+\.?\d*)\s*"?\s*(L|LEN|LENGTH)',
            r'L(?:ENGTH)?[:\s]+(\d+\.?\d*)',
            r'(\d+\.?\d*)\s*"?\s*LONG'
        ]
        for pattern in length_rules:
            m = re.search(pattern, dim_source_only, re.IGNORECASE)
            if m:
                length = m.group(1).strip()
                break

    return width, depth, diameter, height, length


def extract_detail_values(text):
    if not isinstance(text, str):
        text = ""

    def find(pattern):
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else ""

    finish_code = find(r'Finish Sample Code[:\s]*([A-Za-z0-9\-_/]+)')
    if len(finish_code) < 3:
        finish_code = ""

    seat_height = find(r'Seat Height[:\s]*([\d\.]+)')
    seat_depth  = find(r'Seat Depth[:\s]*([\d\.]+)')
    seat_width  = find(r'Seat Width[:\s]*([\d\.]+)')
    arm_height  = find(r'Arm Height[:\s]*([\d\.]+)')

    weight = ""
    m = re.search(r'([\d\.]+)\s*(lb|lbs)', text, re.IGNORECASE)
    if m:
        weight = m.group(1).strip()

    return finish_code, seat_height, seat_depth, seat_width, arm_height, weight


def extract_product_details(driver, url):
    out = {
        "SKU": "",
        "Description": "",
        "Weight": "",
        "Width": "",
        "Depth": "",
        "Diameter": "",
        "Height": "",
        "Length": "",
        "Seat Height": "",
        "Seat Depth": "",
        "Seat Width": "",
        "Arm Height": ""
    }

    try:
        driver.get(url)
    except Exception:
        return out

    time.sleep(PER_PAGE_SLEEP)

    try:
        WebDriverWait(driver, WAIT_BODY_SEC).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except TimeoutException:
        return out

    # SKU
    sku = ""
    try:
        sku = driver.find_element(By.CSS_SELECTOR, ".productView-info-value--sku").text.strip()
    except:
        try:
            m = re.search(r'(SKU|Item Code)[:\s]*([A-Za-z0-9\-_/]+)', driver.page_source, re.IGNORECASE)
            if m:
                sku = m.group(2).strip()
        except:
            sku = ""

    # Description
    try:
        desc = driver.find_element(By.CSS_SELECTOR, "article.productView-description").text.strip()
    except:
        desc = ""

    # Dimensions block (if exists)
    try:
        dim_source = driver.find_element(By.CSS_SELECTOR, ".card-dimensions-standard").text.strip()
    except:
        dim_source = ""

    width, depth, diameter, height, length = extract_dimensions((dim_source + " " + desc), dim_source)
    finish_code, seat_height, seat_depth, seat_width, arm_height, weight = extract_detail_values(desc)

    # If SKU empty but Finish code exists -> use finish code
    if not sku and finish_code:
        sku = finish_code

    out.update({
        "SKU": sku,
        "Description": desc,
        "Weight": weight,
        "Width": width,
        "Depth": depth,
        "Diameter": diameter,
        "Height": height,
        "Length": length,
        "Seat Height": seat_height,
        "Seat Depth": seat_depth,
        "Seat Width": seat_width,
        "Arm Height": arm_height
    })
    return out


def realign_row(ws, old_map, new_map, r):
    """
    ✅ Reads values by old headers and writes them into correct new columns.
    This prevents: Size/Description shifting into SKU etc.
    """
    def read_old(col_name):
        c = old_map.get(col_name)
        return ws.cell(row=r, column=c).value if c else None

    # pull existing values from old layout (Step-1)
    keep = {
        "Index": read_old("Index"),
        "Category": read_old("Category"),
        "Product URL": read_old("Product URL"),
        "Image URL": read_old("Image URL"),
        "Product Name": read_old("Product Name"),
        "Size": read_old("Size"),
        "Description": read_old("Description"),
        "SKU": read_old("SKU"),
    }

    # clear new row area (only within our output columns)
    for col_name, col_idx in new_map.items():
        ws.cell(row=r, column=col_idx, value=None)

    # write back preserved fields into correct places
    for k, v in keep.items():
        if k in new_map and v is not None and str(v).strip() != "":
            ws.cell(row=r, column=new_map[k], value=v)


def main():
    wb = load_workbook(MASTER_FILE)
    driver = setup_driver()

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # ✅ read old header mapping first (Step-1 layout)
            old_map = get_header_map(ws)

            # ✅ write new header and get new mapping (Step-2 layout)
            new_map = write_new_header(ws)

            col_url = new_map["Product URL"]
            col_sku = new_map["SKU"]
            col_cat = new_map["Category"]
            col_idx = new_map["Index"]
            col_name = new_map["Product Name"]

            last_row = get_last_data_row(ws, col_url)
            if last_row < DATA_START_ROW:
                continue

            print(f"\n==============================")
            print(f"📄 Sheet: {sheet_name} | Rows: {last_row - DATA_START_ROW + 1}")
            print("==============================")

            for r in range(DATA_START_ROW, last_row + 1):
                # ✅ First realign existing row data (fix shifting permanently)
                realign_row(ws, old_map, new_map, r)

                product_url  = ws.cell(row=r, column=col_url).value
                product_name = ws.cell(row=r, column=col_name).value
                index_value  = ws.cell(row=r, column=col_idx).value

                if not product_url or not isinstance(product_url, str) or not product_url.startswith("http"):
                    continue

                # ✅ Resume: if SKU already exists (after realign) -> skip
                existing_sku = ws.cell(row=r, column=col_sku).value
                if existing_sku and str(existing_sku).strip():
                    continue

                # Ensure Category filled
                if not ws.cell(row=r, column=col_cat).value:
                    ws.cell(row=r, column=col_cat, value=sheet_name)

                print(f"🔎 {sheet_name} | Row {r} -> {product_url}")

                details = extract_product_details(driver, product_url)

                # Product Family ID = Product Name
                family_id = str(product_name).strip() if product_name else ""

                # ✅ If scraped SKU empty -> generate
                final_sku = (details.get("SKU") or "").strip()
                if not final_sku:
                    final_sku = generate_sku_if_missing(VENDOR_NAME, sheet_name, index_value)

                ws.cell(row=r, column=new_map["SKU"], value=final_sku)
                ws.cell(row=r, column=new_map["Product Family ID"], value=family_id)

                # Note: keep Step-1 Description if scraping desc empty
                scraped_desc = details.get("Description", "")
                if scraped_desc:
                    ws.cell(row=r, column=new_map["Description"], value=scraped_desc)

                ws.cell(row=r, column=new_map["Weight"], value=details.get("Weight", ""))

                ws.cell(row=r, column=new_map["Width"], value=details.get("Width", ""))
                ws.cell(row=r, column=new_map["Depth"], value=details.get("Depth", ""))
                ws.cell(row=r, column=new_map["Diameter"], value=details.get("Diameter", ""))
                ws.cell(row=r, column=new_map["Height"], value=details.get("Height", ""))
                ws.cell(row=r, column=new_map["Length"], value=details.get("Length", ""))

                ws.cell(row=r, column=new_map["Seat Height"], value=details.get("Seat Height", ""))
                ws.cell(row=r, column=new_map["Seat Depth"], value=details.get("Seat Depth", ""))
                ws.cell(row=r, column=new_map["Seat Width"], value=details.get("Seat Width", ""))
                ws.cell(row=r, column=new_map["Arm Height"], value=details.get("Arm Height", ""))

            wb.save(SAVE_AS)
            print(f"✅ Saved progress -> {SAVE_AS}")

    finally:
        driver.quit()
        wb.save(SAVE_AS)

    print(f"\n✅ Done. Final saved: {SAVE_AS}")


if __name__ == "__main__":
    main()
