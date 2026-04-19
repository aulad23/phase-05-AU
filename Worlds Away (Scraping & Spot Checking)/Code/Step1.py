# -*- coding: utf-8 -*-
# worldsaway_MASTER_step1_scroll_multi_category.py
# ✅ Step-1 (multi category scroll)
# ✅ Column order fixed:
#    Category, Product URL, Image URL, Product Name, Size, Description
# ✅ Freeze panes OFF (no row 1-4 freeze)

import time
import os
import re
import pandas as pd
import chromedriver_autoinstaller

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ================= CONFIG =================
BASE_DIR = os.getcwd()

MASTER_OUTPUT_FILE = os.path.join(BASE_DIR, "worldsaway_ALL_categories_step1.xlsx")
MASTER_SHEETS_FILE = os.path.join(BASE_DIR, "WorldsAway.xlsx")

HEADLESS = False           # True করলে background এ চলবে
TIMEOUT = 30
SCROLL_PAUSE = 3
MAX_SCROLL_ROUNDS = 200    # safety limit
START_WAIT_SEC = 5         # initial load wait

CATEGORIES = {
    "Coffee & Cocktail Tables": [
        "https://www.worlds-away.com/collection/tables/coffee-tables/"
    ],
    "Side & End Tables": [
        "https://www.worlds-away.com/collection/casegoods/side-tables/"
    ],
    "Dining Tables": [
        "https://www.worlds-away.com/collection/tables/dining-tables/"
    ],
    "Consoles": [
        "https://www.worlds-away.com/collection/tables/console-tables/"
    ],
    "Beds & Headboards": [
        "https://www.worlds-away.com/collection/beds/"
    ],

   "Desks": [
        "https://www.worlds-away.com/collection/tables/desks/"
    ],
   "Bar Stools": [
        "https://www.worlds-away.com/collection/seating/barstools-counterstools/"
    ],
  "Sofas & Loveseats": [
        "https://www.worlds-away.com/collection/seating/sofas/"
    ],
  "Lounge Chairs": [
        "https://www.worlds-away.com/collection/seating/lounge-chairs/"
    ],
  "Sconces": [
        "https://www.worlds-away.com/collection/lighting/sconces/"
    ],
  "Table Lamps": [
        "https://www.worlds-away.com/collection/lighting/table-lamps/"
    ],
  "Floor Lamps": [
        "https://www.worlds-away.com/collection/lighting/floor-lamps/"
    ],
  "Mirrors": [
        "https://www.worlds-away.com/collection/mirrors/"
    ],
  "Objects": [
        "https://www.worlds-away.com/collection/accessories/decorative-objects/"
    ],
  "Trays": [
        "https://www.worlds-away.com/collection/accessories/trays-containers/"
    ],
  "Wall Decor": [
        "https://www.worlds-away.com/collection/wall-art/"
    ],
}
# =========================================

# ✅ EXACT output column order
STEP1_COLUMNS = [
    "Category",
    "Product URL",
    "Image URL",
    "Product Name",
    "Size",
    "Description",
]

LINK_FONT = Font(color="0563C1", underline="single")
BOLD = Font(bold=True)

# ---------- Driver ----------
def create_driver():
    chromedriver_autoinstaller.install()
    options = Options()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-extensions")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(TIMEOUT)
    return driver


# ---------- Helpers ----------
def safe_sheet_name(name):
    return re.sub(r"[\\/*?:\[\]]", " ", name)[:31]


def scroll_until_all_loaded(driver):
    prev_count = -1
    stable_rounds = 0

    for _ in range(MAX_SCROLL_ROUNDS):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE)

        products = driver.find_elements(By.CSS_SELECTOR, "li.product")
        cur_count = len(products)

        if cur_count == prev_count:
            stable_rounds += 1
        else:
            stable_rounds = 0

        prev_count = cur_count
        if stable_rounds >= 2:
            break

    return driver.find_elements(By.CSS_SELECTOR, "li.product")


def extract_products_from_listing(driver, category_name):
    products = scroll_until_all_loaded(driver)
    print(f"✅ {category_name}: Loaded products -> {len(products)}")

    data = []
    seen_links = set()

    img_selectors = [
        "img.card-image",
        "img.card-img-top",
        "img.attachment-woocommerce_thumbnail",
        "img.wp-post-image",
        "img"
    ]

    for p in products:
        # --- Product URL and Name ---
        try:
            body = p.find_element(By.CSS_SELECTOR, "div.card-body")
            link_elem = body.find_element(By.CSS_SELECTOR, "a")
            product_url = (link_elem.get_attribute("href") or "").strip()
            product_name = (link_elem.text or "").strip()
        except:
            product_url = ""
            product_name = ""

        if not product_url or not product_url.startswith("http"):
            continue

        # --- Image URL ---
        image_url = ""
        for sel in img_selectors:
            try:
                img_elem = p.find_element(By.CSS_SELECTOR, sel)
                src = img_elem.get_attribute("data-src") or img_elem.get_attribute("src")
                if src and src.strip():
                    image_url = src.strip()
                    break
            except:
                continue

        # --- Size ---
        try:
            size_elem = p.find_element(By.CSS_SELECTOR, "div.card-dimensions-standard")
            size_text = (size_elem.text or "").strip()
        except:
            size_text = ""

        # --- Description ---
        try:
            desc_elem = p.find_element(By.CSS_SELECTOR, "div.card-desc")
            desc_text = (desc_elem.text or "").strip()
        except:
            desc_text = ""

        # --- Dedup ---
        if product_url in seen_links:
            continue
        seen_links.add(product_url)

        # ✅ Column order exactly as requested
        data.append({
            "Category": category_name,
            "Product URL": product_url,
            "Image URL": image_url,
            "Product Name": product_name,
            "Size": size_text,
            "Description": desc_text
        })

    df = pd.DataFrame(data)

    # ✅ enforce exact column order
    for c in STEP1_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    df = df[STEP1_COLUMNS]

    return df


def build_step1_master_excel():
    driver = create_driver()
    dfs = []

    try:
        for cat, links in CATEGORIES.items():
            for url in links:
                print(f"\n🔎 Category: {cat}")
                print(f"➡️ URL: {url}")

                try:
                    driver.get(url)
                    time.sleep(START_WAIT_SEC)
                except Exception as e:
                    print(f"❌ Page load error: {e}")
                    continue

                df_cat = extract_products_from_listing(driver, cat)
                if not df_cat.empty:
                    df_cat.drop_duplicates(subset=["Product URL"], inplace=True)
                    dfs.append(df_cat)
                else:
                    print(f"⚠️ No products found for category: {cat}")

    finally:
        driver.quit()

    if not dfs:
        master = pd.DataFrame(columns=STEP1_COLUMNS)
        master.to_excel(MASTER_OUTPUT_FILE, index=False)
        print(f"⚠️ No data collected. Empty master saved: {MASTER_OUTPUT_FILE}")
        return master

    master = pd.concat(dfs, ignore_index=True)
    master.drop_duplicates(subset=["Product URL"], inplace=True)

    # ✅ enforce exact column order
    master = master[STEP1_COLUMNS]

    master.to_excel(MASTER_OUTPUT_FILE, index=False)
    print(f"\n✅ Master saved: {MASTER_OUTPUT_FILE} | Rows: {len(master)}")
    return master


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    ordered_categories = list(CATEGORIES.keys())

    for cat in ordered_categories:
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        df_cat = df_cat[STEP1_COLUMNS]

        ws = wb.create_sheet(title=safe_sheet_name(cat))

        ws["A1"] = "Brand"
        ws["B1"] = "Worlds Away"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = BOLD

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # hyperlink Product Name -> Product URL (optional)
        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        if "Product URL" in headers and "Product Name" in headers:
            col_url = headers["Product URL"]
            col_name = headers["Product Name"]
            for r in range(start + 1, ws.max_row + 1):
                url = ws.cell(row=r, column=col_url).value
                name_cell = ws.cell(row=r, column=col_name)
                if url:
                    name_cell.hyperlink = url
                    name_cell.font = LINK_FONT

        # ✅ Freeze panes OFF
        ws.freeze_panes = None

    wb.save(MASTER_SHEETS_FILE)
    print(f"✅ Final WorldsAway.xlsx created (Category serial order): {MASTER_SHEETS_FILE}")


def main():
    df = build_step1_master_excel()

    if df.empty:
        wb = Workbook()
        wb.save(MASTER_SHEETS_FILE)
        print(f"⚠️ Empty workbook saved: {MASTER_SHEETS_FILE}")
        return

    build_category_wise_workbook_from_df(df)


if __name__ == "__main__":
    main()
