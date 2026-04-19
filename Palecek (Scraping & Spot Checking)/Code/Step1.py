import os
import time
import pandas as pd
from urllib.parse import urljoin
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from webdriver_manager.chrome import ChromeDriverManager

# =========================================================
# INPUT SYSTEM
# =========================================================

VENDOR_NAME = "Palecek"

LIST_URLS = {
    "Bar Stools": [
        "https://www.palecek.com/itembrowser.aspx?action=attributes&itemtype=furniture&custom%20department=furniture&custom%20category=stools&viewall=true",
    ]
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "palecek_bar_stools.xlsx")
master_sheets_file = os.path.join(script_dir, "palecek_bar_stools_category.xlsx")

WAIT_TIMEOUT = 25
ENABLE_DETAIL_BACKFILL = True
DETAIL_BACKFILL_TIMEOUT = 12

# =========================================================
# SCRAPING LOGIC
# =========================================================

def connect_driver() -> webdriver.Chrome:
    opts = Options()
    # opts.add_argument("--headless=new")  # uncomment if headless
    opts.add_argument("--disable-gpu")
    opts.add_argument("--start-maximized")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


def normalize_image_url(src: str) -> str:
    if not src:
        return ""
    src = src.strip()
    if src.startswith("//"):
        return "https:" + src
    if src.startswith("/"):
        return urljoin("https://www.palecek.com", src)
    return src


def safe_text(el) -> str:
    if not el:
        return ""
    return (el.get_attribute("textContent") or "").strip()


def extract_item_strong(card, driver, max_retries=3):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", card)
        time.sleep(0.15)
    except Exception:
        pass

    product_url, img_url, name_text, sku_text = "", "", "", ""

    try:
        a = card.find_element(By.CSS_SELECTOR, "a[href*='iteminformation.aspx']")
        product_url = urljoin("https://www.palecek.com", a.get_attribute("href"))
    except Exception:
        pass

    try:
        img = card.find_element(By.CSS_SELECTOR, "img.ProductThumbnailImg")
        img_url = img.get_attribute("data-src") or img.get_attribute("src") or ""
        img_url = normalize_image_url(img_url)
    except Exception:
        pass

    for _ in range(max_retries):
        try:
            name_el = card.find_element(
                By.CSS_SELECTOR,
                "div.ProductThumbnailDetails p.ProductThumbnailParagraphDescription a"
            )
            sku_el = card.find_element(
                By.CSS_SELECTOR,
                "div.ProductThumbnailDetails p.ProductThumbnailParagraphSkuName a"
            )
            name_text = safe_text(name_el)
            sku_text = safe_text(sku_el)

            if not sku_text:
                try:
                    h3 = card.find_element(By.CSS_SELECTOR, "div.ProductThumbnailDetails h3")
                    sku_text = safe_text(h3)
                except Exception:
                    pass
            if name_text and sku_text:
                break
            time.sleep(0.25)
        except StaleElementReferenceException:
            time.sleep(0.2)
        except Exception:
            time.sleep(0.2)

    return {
        "Product URL": product_url,
        "Image URL": img_url,
        "Product Name": name_text,
        "SKU": sku_text,
    }


def scroll_and_collect(driver, base_url):
    all_data = []
    seen_keys = set()

    driver.get(base_url)
    time.sleep(4)
    print("⬇️ Loading all products (one scroll)...")
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(4)

    cards = driver.find_elements(
        By.CSS_SELECTOR,
        "div.ItemBrowserThumbnailContainer section.ProductThumbnailSection div.ProductThumbnail"
    )

    print(f"🔍 Found {len(cards)} product cards — extracting...")

    for card in cards:
        data = extract_item_strong(card, driver)
        key = data["Product URL"] or (data["Product Name"], data["SKU"])
        if key and key not in seen_keys:
            seen_keys.add(key)
            all_data.append(data)

    print(f"✅ Finished collecting {len(all_data)} unique products.")
    return all_data


def fill_from_detail(driver, url, timeout=DETAIL_BACKFILL_TIMEOUT):
    if not url:
        return ("", "")

    driver.execute_script("window.open(arguments[0], '_blank');", url)
    driver.switch_to.window(driver.window_handles[-1])
    name = ""
    sku = ""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
        )
        for sel in ["h1", "h2", ".ItemName", ".ProductTitle", "#lblItemTitle"]:
            try:
                name = (driver.find_element(By.CSS_SELECTOR, sel)
                        .get_attribute("textContent") or "").strip()
                if name:
                    break
            except Exception:
                pass

        for sel in [".sku", ".ItemNumber", "#lblItemNumber", "[data-sku]", ".item-number"]:
            try:
                sku = (driver.find_element(By.CSS_SELECTOR, sel)
                       .get_attribute("textContent") or "").strip()
                if sku:
                    break
            except Exception:
                pass
    finally:
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

    return (name, sku)


def harvest_from_list_page(driver, url):
    collected = scroll_and_collect(driver, url)

    if ENABLE_DETAIL_BACKFILL:
        for r in collected:
            if (not r["Product Name"] or not r["SKU"]) and r["Product URL"]:
                name, sku = fill_from_detail(driver, r["Product URL"])
                if name and not r["Product Name"]:
                    r["Product Name"] = name
                if sku and not r["SKU"]:
                    r["SKU"] = sku

    return collected

# =========================================================
# OUTPUT SYSTEM
# =========================================================

def build_master_excel(all_products):
    if not all_products:
        cols = ["Index", "Category", "Product URL", "Image URL", "Product Name", "SKU"]
        pd.DataFrame(columns=cols).to_excel(master_output_file, index=False)
        return pd.DataFrame(columns=cols)

    df = pd.DataFrame(all_products)
    df.drop_duplicates(subset=["Product URL"], inplace=True)
    df.insert(0, "Index", range(1, len(df) + 1))
    df = df[["Index", "Category", "Product URL", "Image URL", "Product Name", "SKU"]]
    df.to_excel(master_output_file, index=False)
    return df


def build_category_wise_workbook(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in LIST_URLS.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)

        ws["A1"] = "Brand"
        ws["B1"] = VENDOR_NAME
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(LIST_URLS.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        start = 4
        for j, col in enumerate(df_cat.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_cat.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers_map = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}

        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers_map["Product URL"]).value
            name_cell = ws.cell(row=r, column=headers_map["Product Name"])
            if url:
                name_cell.hyperlink = url
                name_cell.font = link_font

    wb.save(master_sheets_file)

# =========================================================
# MAIN
# =========================================================

def main():
    driver = connect_driver()
    all_products = []
    global_seen = set()

    try:
        for category_name, urls in LIST_URLS.items():
            for url in urls:
                print(f"Harvesting category '{category_name}' from {url}")
                rows = harvest_from_list_page(driver, url)

                for r in rows:
                    key = r.get("Product URL") or (r.get("Product Name"), r.get("SKU"))
                    if key and key not in global_seen:
                        global_seen.add(key)
                        r["Category"] = category_name
                        all_products.append(r)

        df = build_master_excel(all_products)
        if not df.empty:
            build_category_wise_workbook(df)

        print(f"Completed. Total products scraped: {len(df)}")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
