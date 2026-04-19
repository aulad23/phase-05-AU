import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import time
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# ------------------------------
# FILE PATHS (SCRIPT LOCATION)
# ------------------------------
script_folder = os.path.dirname(os.path.abspath(__file__))

# ✅ INPUT: Step-1 workbook (multi-sheet)
input_file = os.path.join(script_folder, "GabriellaWhite.xlsx")

# ✅ OUTPUT: Step-2 workbook (same structure, multi-sheet)
output_file = os.path.join(script_folder, "GabriellaWhite_Final.xlsx")

# ------------------------------
# EXCEL TEMPLATE SETTINGS
# ------------------------------
HEADER_ROW = 4      # Row 4 = headers
START_ROW = 5       # Row 5+ = data rows
META_BRAND_ROW = 1  # A1/B1
META_LINK_ROW = 2   # A2/B2
BLANK_ROW = 3       # row 3 blank

bold_font = Font(bold=True)
link_font = Font(color="0563C1", underline="single")

# ------------------------------
# Columns to add (same as your code)
# ------------------------------
dimension_cols = [
    "Weight", "Width", "Depth", "Diameter", "Length", "Height",
    "Seat Depth", "Seat Height", "Seat Width", "Cushion", "Shade Details", "Base", "Canopy"
]

feature_cols = ["Finish", "Com", "Color", "Pattern"]

ADD_COLS = ["SKU", "Product Family Id", "Description"] + dimension_cols + feature_cols

# ------------------------------
# PRODUCT FAMILY ID EXTRACTION
# ------------------------------
def extract_product_family(product_name):
    """Extracts the base product family name by removing suffixes like '- Dark'."""
    return re.sub(r" - .*", "", str(product_name or ""))

# ------------------------------
# HEADERS (same)
# ------------------------------
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0 Safari/537.36"
}

def clean_dimension(value):
    """
    Keeps numeric values including decimals (e.g., 32.75).
    If multiple values exist (comma-separated), returns only the first one.
    """
    value = str(value or "").strip()

    # keep digits, dot, comma only
    value = re.sub(r"[^0-9\.,]+", "", value)

    # take first if multiple values exist
    return value.split(",")[0].strip()

# ------------------------------
# READ INPUT WORKBOOK (ALL SHEETS)
# ------------------------------
def read_step1_workbook(path_xlsx: str):
    wb = load_workbook(path_xlsx)
    sheets_data = []

    for ws in wb.worksheets:
        # Read meta rows (B1 brand name, B2 link)
        brand_name = ws["B1"].value or ""
        brand_link = ws["B2"].value or ""

        # Read headers from row 4
        headers_row = []
        max_col = ws.max_column
        for c in range(1, max_col + 1):
            headers_row.append(ws.cell(row=HEADER_ROW, column=c).value)

        # If row4 is empty, skip
        if not any(headers_row):
            continue

        # Build rows from row 5 onwards until blank Product URL
        data = []
        for r in range(START_ROW, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue

            # Must have Product URL to be valid
            try:
                url_idx = headers_row.index("Product URL")
                product_url = row_vals[url_idx]
            except ValueError:
                product_url = None

            if product_url is None or str(product_url).strip() == "":
                continue

            data.append(row_vals)

        df = pd.DataFrame(data, columns=headers_row)

        # Normalize required base columns (no logic change, just safety)
        for base_col in ["Index", "Category", "Product URL", "Image URL", "Product Name"]:
            if base_col not in df.columns:
                df[base_col] = ""

        sheets_data.append({
            "sheet_name": ws.title,
            "brand_name": brand_name,
            "brand_link": brand_link,
            "df": df
        })

    return sheets_data

# ------------------------------
# RESTORE PROGRESS FROM OUTPUT (MULTI-SHEET)
# ------------------------------
def restore_progress_if_exists(sheets_data, output_path):
    if not os.path.exists(output_path):
        return sheets_data

    print("Previous output found — restoring progress...")
    wb_out = load_workbook(output_path)

    for item in sheets_data:
        sname = item["sheet_name"]
        if sname not in wb_out.sheetnames:
            continue

        ws = wb_out[sname]

        # Read output headers row 4
        out_headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(out_headers):
            continue

        out_rows = []
        for r in range(START_ROW, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in vals):
                continue
            out_rows.append(vals)

        out_df = pd.DataFrame(out_rows, columns=out_headers)

        # Merge progress by Product URL
        if "Product URL" in out_df.columns and "Product URL" in item["df"].columns:
            out_df = out_df.drop_duplicates(subset=["Product URL"])
            merged = item["df"].merge(
                out_df,
                on="Product URL",
                how="left",
                suffixes=("", "__old")
            )

            # For each column in ADD_COLS, restore from old if exists
            for col in ADD_COLS:
                old_col = f"{col}__old"
                if old_col in merged.columns:
                    # if current is empty, take old
                    cur = merged[col] if col in merged.columns else ""
                    merged[col] = merged[col].where(
                        merged[col].astype(str).str.strip().ne(""),
                        merged[old_col]
                    )
                    merged.drop(columns=[old_col], inplace=True)

            # Also restore any already-existing columns that were in output (optional but safe)
            for col in out_df.columns:
                old_col = f"{col}__old"
                if old_col in merged.columns and col in item["df"].columns:
                    merged[col] = merged[col].where(
                        merged[col].astype(str).str.strip().ne(""),
                        merged[old_col]
                    )
                    merged.drop(columns=[old_col], inplace=True)

            item["df"] = merged

    print("Progress restored!")
    return sheets_data

# ------------------------------
# WRITE OUTPUT WORKBOOK (KEEP TEMPLATE)
# ------------------------------
def write_step2_workbook(sheets_data, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    for item in sheets_data:
        ws = wb.create_sheet(title=item["sheet_name"])
        df = item["df"].copy()

        # Ensure columns exist
        for col in ADD_COLS:
            if col not in df.columns:
                df[col] = ""

        # Apply Product Family Id (same logic)
        df["Product Family Id"] = df["Product Name"].apply(extract_product_family)

        # Reorder: keep your Step-1 order first, then append step-2 columns
        base_order = ["Index", "Category", "Product URL", "Image URL", "Product Name"]
        final_cols = base_order + [c for c in df.columns if c not in base_order]

        df = df[final_cols]

        # Meta rows
        ws["A1"] = "Brand"
        ws["B1"] = item["brand_name"]
        ws["A2"] = "Link"
        ws["B2"] = item["brand_link"]
        ws["B2"].alignment = Alignment(wrap_text=True)

        # Row 3 blank (do nothing)

        # Header row 4
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=HEADER_ROW, column=j, value=col)
            cell.font = bold_font

        # Data rows start row 5
        for i, row in enumerate(df.itertuples(index=False), start=START_ROW):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # Hyperlink Product Name -> Product URL (same style idea as Step-1)
        headers_map = {ws.cell(row=HEADER_ROW, column=j).value: j for j in range(1, ws.max_column + 1)}
        if "Product URL" in headers_map and "Product Name" in headers_map:
            url_col = headers_map["Product URL"]
            name_col = headers_map["Product Name"]
            for r in range(START_ROW, ws.max_row + 1):
                url = ws.cell(row=r, column=url_col).value
                cell = ws.cell(row=r, column=name_col)
                if url:
                    cell.hyperlink = url
                    cell.font = link_font

    wb.save(output_path)

# ------------------------------
# MAIN SCRAPING PROCESS (LOGIC UNCHANGED)
# ------------------------------
sheets_data = read_step1_workbook(input_file)

# Restore progress if output exists
sheets_data = restore_progress_if_exists(sheets_data, output_file)

# Flatten all sheets into one combined dataframe for scraping loop,
# but keep reference so we can write back sheet-wise after scraping.
all_parts = []
sheet_ranges = []  # (sheet_name, start_idx, end_idx)
cursor = 0

for item in sheets_data:
    df = item["df"].copy()

    # Add missing columns (same as your original)
    for col in ["SKU", "Product Family Id", "Description"] + dimension_cols + feature_cols:
        if col not in df.columns:
            df[col] = ""

    # Apply Product Family Id
    df["Product Family Id"] = df["Product Name"].apply(extract_product_family)

    start = cursor
    all_parts.append(df)
    cursor += len(df)
    end = cursor
    sheet_ranges.append((item["sheet_name"], start, end))

df = pd.concat(all_parts, ignore_index=True) if all_parts else pd.DataFrame()

# ------------------------------
# MAIN SCRAPING LOOP (SAME LOGIC)
# ------------------------------
batch_counter = 0
batch_size = 5

for idx, row in df.iterrows():
    if pd.notna(df.at[idx, "SKU"]) and str(df.at[idx, "SKU"]).strip() != "":
        print(f"Skipping {idx+1}: Already scraped.")
        continue

    url = row["Product URL"]
    print(f"\nScraping {idx+1}/{len(df)} → {row['Product Name']}")

    try:
        retries = 3
        while retries > 0:
            try:
                response = requests.get(url, headers=headers, timeout=20)
                if response.status_code == 200:
                    break
                else:
                    print(f"Error fetching {url}: Status Code {response.status_code}")
                    retries -= 1
                    time.sleep(3)
            except requests.RequestException as e:
                print(f"Error fetching {url}: {e}")
                retries -= 1
                time.sleep(3)

        if retries == 0:
            print(f"Failed to fetch {url} after multiple attempts.")
            continue

        soup = BeautifulSoup(response.text, "html.parser")

        # SKU
        sku_tag = soup.find("p", id=lambda x: x and x.startswith("Sku-template"))
        if sku_tag:
            sku = sku_tag.get_text(strip=True).replace("SKU:", "").strip()
            df.at[idx, "SKU"] = sku
            df.at[idx, "Product Family Id"] = row["Product Family Id"]

        # DESCRIPTION
        desc_div = soup.find("div", class_="inline-richtext")
        if desc_div:
            df.at[idx, "Description"] = desc_div.get_text(separator=" ", strip=True)

        # DIMENSIONS & FEATURES
        sections = soup.find_all("div", class_="specs-attributes-section")

        dimensions_map = {
            "Product weight": "Weight",
            "Product width": "Width",
            "Product depth": "Depth",
            "Product diameter": "Diameter",
            "Product length": "Length",
            "Product height": "Height",
            "Seat depth": "Seat Depth",
            "Seat height": "Seat Height",
            "Seat width": "Seat Width",
            "Seat cushion depth": "Cushion",
            "Seat cushion height": "Cushion",
            "Seat cushion width": "Cushion",
            "Seat height with cushion": "Cushion",
            "Seat height without cushion": "Cushion",
            "Shade diffuser depth at bottom": "Shade Details",
            "Shade diffuser depth at top": "Shade Details",
            "Shade diffuser width at bottom": "Shade Details",
            "Shade diffuser width at top": "Shade Details",
            "Base width": "Base",
            "Base diameter": "Base",
            "Base depth": "Base",
            "Canopy diameter": "Canopy",
            "Canopy height": "Canopy",
            "Canopy width": "Canopy"
        }

        features_map = {
            "Finish family": "Finish",
            "Com fabric railroad yardage": "Com",
            "Com fabric up the roll yardage": "Com",
            "Color family": "Color",
            "Pattern": "Pattern"
        }

        temp_dimensions = {col: [] for col in dimension_cols}
        temp_features = {col: [] for col in feature_cols}

        for sec in sections:
            h3 = sec.find("h3")
            if not h3:
                continue

            title = h3.get_text(strip=True).upper()

            spans = sec.find_all("span")
            pairs = [
                (spans[i].get_text(strip=True),
                 spans[i+1].get_text(strip=True) if i+1 < len(spans) else "")
                for i in range(0, len(spans), 2)
            ]

            if "DIMENSIONS" in title:
                for k, v in pairs:
                    if k in dimensions_map:
                        col_name = dimensions_map[k]
                        temp_dimensions[col_name].append(clean_dimension(v))

            elif "FEATURES" in title:
                for k, v in pairs:
                    if k in features_map:
                        col_name = features_map[k]
                        temp_features[col_name].append(v)

        for col, vals in temp_dimensions.items():
            df.at[idx, col] = ", ".join(vals)

        for col, vals in temp_features.items():
            df.at[idx, col] = ", ".join(vals)

        print(f"Done: SKU={df.at[idx,'SKU']}")

        # ------------------------------
        # BATCH SAVE (NOW SAVES WORKBOOK)
        # ------------------------------
        batch_counter += 1
        if batch_counter >= batch_size:
            # write df back to sheets, then save workbook
            for item in sheets_data:
                sname = item["sheet_name"]
                for (sheet_name, s, e) in sheet_ranges:
                    if sheet_name == sname:
                        item["df"] = df.iloc[s:e].copy()
                        break

            write_step2_workbook(sheets_data, output_file)
            print("✔ Autosaved batch (workbook)")
            batch_counter = 0

        time.sleep(1)

    except Exception as e:
        print(f"Error scraping {url}: {e}")

# ------------------------------
# FINAL SAVE (WORKBOOK)
# ------------------------------
for item in sheets_data:
    sname = item["sheet_name"]
    for (sheet_name, s, e) in sheet_ranges:
        if sheet_name == sname:
            item["df"] = df.iloc[s:e].copy()
            break

write_step2_workbook(sheets_data, output_file)

print("\n🎉 ALL DONE — File saved at:")
print(output_file)
