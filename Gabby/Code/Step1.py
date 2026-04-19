# =========================================================
# GABRIELLA WHITE SCRAPER (UPDATED FINAL MASTER)
# - Scraping logic preserved (BS4 + requests)
# - Pagination supports:
#     1) nav-based pagination (old logic)
#     2) query-based pagination for /collections/... (page=2&sort_by=manual)
# - Input / Output system retained
# =========================================================

import os
import time
from urllib.request import Request, urlopen
from urllib.parse import urlparse
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM
# =========================================================

CATEGORIES = {
    "Nightstands": [
        "https://gabby.com/products/bedroom/nightstands",
    ],
    "Coffee & Cocktail Tables": [
        "https://gabby.com/products/indoor-living/coffee-tables",
    ],
    "Side & End Tables": [
        "https://gabby.com/products/indoor-living/side-tables",
    ],
    "Consoles": [
        "https://gabby.com/products/indoor-dining/console-tables",
    ],
    "Beds & Headboards": [
        "https://gabby.com/products/bedroom/upholstered-beds",
    ],
    "Desks": [
        "https://gabby.com/products/indoor-living/desks",
    ],
    "Cabinets": [
        "https://gabby.com/products/indoor-dining/cabinets",
        "https://gabby.com/products/storage/sideboards"
    ],
    "Bookcases": [
        "https://gabby.com/products/storage/bookcases",
    ],
    "Accent Tables": [
        "https://gabby.com/products/indoor-living/occasional-tables",
    ],
    "Bar Carts": [
        "https://gabby.com/products/indoor-dining/bar-carts",
    ],
    "Dressers & Chests": [
        "https://gabby.com/products/bedroom/dressers",
        "https://gabby.com/products/bedroom/chests"
    ],
    "Dining Chairs": [
        "https://gabby.com/products/indoor-dining/dining-chairs",
    ],
    "Bar Stools": [
        "https://gabby.com/products/indoor-dining/bar-counter-stools",
    ],
    "Benches": [
        "https://gabby.com/products/indoor-living/benches",
    ],
    "Ottomans": [
        "https://gabby.com/products/indoor-living/ottomans-stools",
    ],
    "Lounge Chairs": [
        "https://gabby.com/products/indoor-living/occasional-chairs",
    ],
    "Chandeliers": [
        "https://gabby.com/lighting/chandeliers",
    ],
    "Pendants": [
        "https://gabby.com/lighting/pendants",
    ],
    "Sconces": [
        "https://gabby.com/lighting/sconces",
    ],
    "Flush Mount": [
        "https://gabby.com/lighting/flush-mounts",
    ],
    "Table Lamps": [
        "https://gabby.com/lighting/table-lamps",
    ],
    "Floor Lamps": [
        "https://gabby.com/lighting/floor-lamps",
    ],
    "Mirrors": [
        "https://gabby.com/mirror",
    ],

    # ✅ FIX: last 2 categories should be from gabriellawhite.com collections
    "Pillows & Throws": [
        "https://gabriellawhite.com/collections/decorative-accessories?sort_by=manual",
    ],
    "Rugs": [
        "https://gabriellawhite.com/collections/decorative-accessories?sort_by=manual",
    ],
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "gabriella_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "GabriellaWhite.xlsx")

# =========================================================
# PAGINATION HELPERS
# =========================================================

def is_gw_collections_url(u: str) -> bool:
    # This type needs query-based pagination: &page=2
    try:
        p = urlparse(u)
        return ("gabriellawhite.com" in p.netloc) and ("/collections/" in p.path)
    except Exception:
        return False


def get_all_page_urls_nav(collection_url):
    """Old logic: read nav pagination links if exists"""
    urls = [collection_url]

    try:
        req = Request(collection_url, headers=HEADERS)
        html = urlopen(req, timeout=20).read()
    except Exception:
        return urls

    soup = BeautifulSoup(html, "html.parser")
    pagination = soup.find("nav", {"aria-label": "Pagination"})

    if pagination:
        for a in pagination.find_all("a", href=True):
            href = a["href"]
            if "page=" in href:
                # keep original behavior (base domain hardcoded earlier)
                full = "https://gabriellawhite.com" + href if href.startswith("/") else href
                if full not in urls:
                    urls.append(full)

    return urls


def get_all_page_urls_query(collection_url, max_pages=80):
    """
    Query pagination: page-1 is the base url
    page-2+: append &page=2 (or ?page=2 if no query exists)
    Stops when no product card found.
    """
    urls = []
    page = 1
    while page <= max_pages:
        if page == 1:
            url = collection_url
        else:
            if "?" in collection_url:
                # your proven working format
                url = f"{collection_url}&page={page}"
            else:
                url = f"{collection_url}?page={page}"

        # test page
        try:
            req = Request(url, headers=HEADERS)
            html = urlopen(req, timeout=20).read()
        except Exception:
            break

        soup = BeautifulSoup(html, "html.parser")
        items = soup.find_all("li", class_="group/product-card")
        if not items:
            break

        urls.append(url)
        page += 1
        time.sleep(0.6)

    return urls


def get_all_page_urls(collection_url):
    if is_gw_collections_url(collection_url):
        return get_all_page_urls_query(collection_url)
    return get_all_page_urls_nav(collection_url)

# =========================================================
# SCRAPING
# =========================================================

def scrape_collection(collection_url):
    products = []
    page_urls = get_all_page_urls(collection_url)

    for url in page_urls:
        try:
            req = Request(url, headers=HEADERS)
            html = urlopen(req, timeout=20).read()
        except Exception:
            continue

        soup = BeautifulSoup(html, "html.parser")
        items = soup.find_all("li", class_="group/product-card")

        if not items:
            continue

        for item in items:
            product_name = ""
            product_url = ""
            image_url = ""

            h3 = item.find("h3")
            if h3:
                a = h3.find("a", href=True)
                if a:
                    product_name = a.get_text(strip=True)

                    # ✅ correct domain by reading the current page domain
                    if a["href"].startswith("http"):
                        product_url = a["href"]
                    else:
                        page_base = "{0.scheme}://{0.netloc}".format(urlparse(url))
                        product_url = page_base + a["href"]

            img = item.find("img")
            if img:
                image_url = (
                    img.get("data-src")
                    or img.get("src")
                    or img.get("srcset", "").split(" ")[0]
                )
                if image_url and image_url.startswith("//"):
                    image_url = "https:" + image_url

            if product_name and product_url and image_url:
                products.append({
                    "Product URL": product_url,
                    "Image URL": image_url,
                    "Product Name": product_name,
                })

        time.sleep(1)

    return products

# =========================================================
# OUTPUT SYSTEM
# =========================================================

def build_step1_master_excel():
    all_rows = []

    for category, urls in CATEGORIES.items():
        for url in urls:
            rows = scrape_collection(url)
            for r in rows:
                r["Category"] = category
                all_rows.append(r)

    if not all_rows:
        cols = ["Category", "Product URL", "Image URL", "Product Name"]
        df_empty = pd.DataFrame(columns=cols)
        df_empty.to_excel(master_output_file, index=False)
        return df_empty

    df = pd.DataFrame(all_rows)
    df = df[["Category", "Product URL", "Image URL", "Product Name"]]
    df.drop_duplicates(subset=["Product URL"], inplace=True)
    df.to_excel(master_output_file, index=False)
    return df


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)

        ws["A1"] = "Brand"
        ws["B1"] = "Gabriella White"
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)

# =========================================================
# MAIN
# =========================================================

def main():
    df = build_step1_master_excel()
    if df.empty:
        Workbook().save(master_sheets_file)
        return
    build_category_wise_workbook_from_df(df)


if __name__ == "__main__":
    main()
