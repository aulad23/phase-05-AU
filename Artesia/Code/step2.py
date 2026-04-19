import os
import re
import time
import pandas as pd
import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter, Retry
from openpyxl import load_workbook, Workbook

# ======================================================
# CONFIG (INPUT = 1st CODE OUTPUT – SAME STRUCTURE AS IMAGE)
# ======================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_FILE  = os.path.join(SCRIPT_DIR, "ArtesiaCollections.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "ArtesiaCollections_FINAL.xlsx")

HEADER_ROW = 4          # header is on row 4
PRODUCT_URL_COL_INDEX = 2  # Column C (0-based)

DELAY = 0.6
TIMEOUT = 25

# ======================================================
# REQUEST SESSION
# ======================================================
session = requests.Session()
retries = Retry(
    total=5,
    backoff_factor=0.7,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"]
)
session.mount("https://", HTTPAdapter(max_retries=retries))
session.mount("http://", HTTPAdapter(max_retries=retries))
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Accept-Language": "en-US,en;q=0.9",
})

# ======================================================
# HELPERS
# ======================================================
def clean(t):
    if t is None:
        return ""
    return re.sub(r"\s+", " ", str(t)).strip()

def normalize_number_value(val):
    s = clean(val)
    s = s.replace('"', "").replace("“", "").replace("”", "")
    s = re.sub(r"\b(inches|inch|in|ft|cm|mm|lbs|lb|kg|g)\b", "", s, flags=re.I)
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    return m.group(1) if m else ""

def get_soup(url):
    r = session.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    return BeautifulSoup(r.text, "lxml")

def get_tabs_map(soup):
    tabs = {}
    for tab in soup.select('div[data-hook="tab-item"]'):
        lbl = tab.select_one('span[data-hook="tab-item-label"]')
        if lbl:
            label = clean(lbl.text).upper()
            panel = clean(tab.get("aria-controls"))
            if label and panel:
                tabs[label] = panel
    return tabs

def get_panel_text(soup, panel_id):
    if not panel_id:
        return ""
    panel = soup.find(id=panel_id)
    if not panel:
        return ""
    texts = []
    for rt in panel.select('[data-testid="richTextElement"]'):
        t = clean(rt.get_text(" ", strip=True))
        if t:
            texts.append(t)
    return " ".join(dict.fromkeys(texts))

def get_image_url(soup):
    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        return clean(og["content"])
    return ""

def extract_size_fields(soup, panel_id):
    fields = {
        "Weight": "", "Width": "", "Depth": "", "Diameter": "",
        "Length": "", "Height": "", "Arm Height": "", "Seat Height": ""
    }
    if not panel_id:
        return fields

    panel = soup.find(id=panel_id)
    if not panel:
        return fields

    text = clean(panel.get_text(" ", strip=True))
    for k in fields:
        m = re.search(rf"{k}\s*[:\-]?\s*([\d\.]+)", text, re.I)
        if m:
            fields[k] = normalize_number_value(m.group(1))
 
    return fields

def build_size_summary(fields):
    order = ["Weight","Width","Depth","Diameter","Length","Height","Arm Height","Seat Height"]
    return " | ".join(f"{k}: {fields[k]}" for k in order if fields[k])

# ======================================================
# MAIN
# ======================================================
def main():
    in_wb = load_workbook(INPUT_FILE)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for sheet in in_wb.sheetnames:
        in_ws = in_wb[sheet]                     # 🔹 READ INPUT SHEET
        category_link = clean(in_ws["B2"].value) # 🔹 READ B2 ONLY

        # read exactly like your file
        df = pd.read_excel(
            INPUT_FILE,
            sheet_name=sheet,
            header=HEADER_ROW - 1
        )

        df.columns = [clean(c) for c in df.columns]

        # force Product URL = column C
        df.rename(columns={df.columns[PRODUCT_URL_COL_INDEX]: "Product URL"}, inplace=True)

        final_cols = [
            "Index",
            "Category",
            "Product URL",
            "Image URL",
            "Product Name",
            "SKU",
            "Description",
            "Weight",
            "Width",
            "Depth",
            "Diameter",
            "Length",
            "Height",
            "Arm Height",
            "Seat Height",
            "Size",
            "Product care",
        ]

        for c in final_cols:
            if c not in df.columns:
                df[c] = ""

        df["Category"] = sheet

        total = len(df)

        for i, row in df.iterrows():
            url = clean(row["Product URL"])
            if not url.startswith("http"):
                continue

            try:
                soup = get_soup(url)
                tabs = get_tabs_map(soup)

                if not clean(df.at[i, "Image URL"]):
                    df.at[i, "Image URL"] = get_image_url(soup)

                if not clean(df.at[i, "Description"]):
                    df.at[i, "Description"] = get_panel_text(soup, tabs.get("DESCRIPTION"))

                size_fields = extract_size_fields(soup, tabs.get("SIZE"))
                for k, v in size_fields.items():
                    if not clean(df.at[i, k]):
                        df.at[i, k] = v

                if not clean(df.at[i, "Size"]):
                    df.at[i, "Size"] = build_size_summary(size_fields)

                if not clean(df.at[i, "Product care"]):
                    df.at[i, "Product care"] = get_panel_text(soup, tabs.get("PRODUCT CARE"))

                print(f"[{sheet}] {i+1}/{total} DONE")

            except Exception as e:
                print(f"[{sheet}] {i+1}/{total} FAIL: {e}")

            time.sleep(DELAY)

        df = df[final_cols]

        ws = out_wb.create_sheet(title=sheet)

        # top info rows (UNCHANGED, only B2 filled)
        ws["A1"] = "Brand"
        ws["B1"] = "Artesia Collections"
        ws["A2"] = "Link"
        ws["B2"] = category_link   # ✅ FIX APPLIED

        # write header on row 4
        for col_idx, col in enumerate(final_cols, start=1):
            ws.cell(row=HEADER_ROW, column=col_idx, value=col)

        # write data from row 5
        for r_idx, row in enumerate(df.itertuples(index=False), start=HEADER_ROW + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    out_wb.save(OUTPUT_FILE)
    print("\n✅ FINAL OUTPUT CREATED (B2 FIXED):", OUTPUT_FILE)

if __name__ == "__main__":
    main()
