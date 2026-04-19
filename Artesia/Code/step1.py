import os
import re
import time
from urllib.parse import urljoin

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM
# =========================================================

CATEGORIES = {
   "Nightstands": [
        "https://www.artesiacollections.com/dressers-nightstands",
    ],
  "Coffee & Cocktail Tables": [
        "https://www.artesiacollections.com/coffee-tables",
    ],
  "Side & End Tables": [
        "https://www.artesiacollections.com/side-tables",
    ],
    "Dining Tables": [
        "https://www.artesiacollections.com/diningtables",
    ],

    "Consoles": [
        "https://www.artesiacollections.com/consoles",
    ],
    "Beds & Headboards": [
        "https://www.artesiacollections.com/beds",
    ],

    "Cabinets": [
        "https://www.artesiacollections.com/cabinets1",
        "https://www.artesiacollections.com/credenzas-sideboards",
        "https://www.artesiacollections.com/media-cabinets"
    ],
    "Dining Chairs": [
        "https://www.artesiacollections.com/dining-chairs",
    ],

    "Bar Stools": [
        "https://www.artesiacollections.com/bar-counter-stools",
        "https://www.artesiacollections.com/stools"
    ],

    "Lounge Chairs": [
        "https://www.artesiacollections.com/occasional-chairs",
    ],

    "Benches": [
        "https://www.artesiacollections.com/benches",
    ],

    "Lighting": [
        "https://www.artesiacollections.com/lighting",
    ],

    "Mirrors": [
        "https://www.artesiacollections.com/mirrors",
    ],

    "Pillows & Throws": [
        "https://www.artesiacollections.com/copy-of-accents-cms",
    ],

    "Vases": [
        "https://www.artesiacollections.com/planters-vases",
    ],
  "Wall Decor": [
        "https://www.artesiacollections.com/wall-decor",
    ],

}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "artesia_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "ArtesiaCollections.xlsx")

# =========================================================
# HELPER FUNCTIONS
# =========================================================

def clean(s):
    if not s:
        return None
    s = re.sub(r"\s+", " ", str(s)).strip()
    return s or None


def extract_image_url(img_el):
    for attr in ["src", "data-src", "data-image-src", "data-original"]:
        v = img_el.get_attribute(attr)
        if v and v.strip():
            return v.strip()

    style = img_el.get_attribute("style") or ""
    m = re.search(r'url\(["\']?(.*?)["\']?\)', style)
    if m:
        return m.group(1)

    wix_img = img_el.get_attribute("wix:image") or img_el.get_attribute("data-wix-image")
    if wix_img:
        return wix_img

    return None


def pick_name_from_text(block_text):
    """Extract product name from a block of text, ignoring SKUs."""
    if not block_text:
        return None
    lines = [clean(x) for x in str(block_text).split("\n")]
    lines = [x for x in lines if x]
    filtered = []
    for ln in lines:
        if re.search(r"\bsku\b", ln, re.I):
            continue
        if re.fullmatch(r"[A-Z0-9\-_/]{3,}", ln):
            continue
        filtered.append(ln)
    return filtered[0] if filtered else None


def pick_sku_from_text(block_text):
    """Extract all possible SKUs from text block."""
    if not block_text:
        return []
    skus = []

    for m in re.finditer(r"\bSKU\b\s*[:#]?\s*([A-Z0-9][A-Z0-9\-_/]{2,})", block_text, re.I):
        sku = clean(m.group(1))
        if sku:
            skus.append(sku)

    if not skus:
        tokens = re.findall(r"\b[A-Z0-9][A-Z0-9\-_/]{2,}\b", block_text.upper())
        noise = {"USD", "ADD", "CART", "SHOP", "VIEW", "MORE", "PRICE"}
        tokens = [t for t in tokens if t not in noise and re.search(r"\d", t)]
        skus.extend(tokens)

    return skus


def safe_skus(skus, product_name=None):
    """Filter SKUs to remove duplicates, non-digit ones, or those matching the product name."""
    filtered = []
    for sku in skus:
        if not sku:
            continue
        sku_c = sku.strip().upper()
        name_c = (product_name or "").strip().upper()
        if sku_c and name_c and (sku_c == name_c or sku_c in name_c):
            continue
        if not re.search(r"\d", sku_c):
            continue
        if sku_c not in filtered:
            filtered.append(sku_c)
    return filtered if filtered else None


# =========================
# FINAL SKU + NAME FIX HELPERS
# =========================

# SKU token like: A2-DT1-120-BLK (must contain at least one hyphen)
SKU_TOKEN_RE = re.compile(r"\b[A-Z0-9]+(?:-[A-Z0-9]+)+\b")

def extract_skus_anywhere(text: str):
    """Find SKU-like tokens anywhere (even if merged with name)."""
    if not text:
        return []
    t = str(text).upper()
    tokens = SKU_TOKEN_RE.findall(t)
    tokens = [x for x in tokens if re.search(r"\d", x)]  # must have digit
    out = []
    for x in tokens:
        if x not in out:
            out.append(x)
    return out

def strip_skus_from_text(text: str):
    """Remove detected SKUs from text -> return clean possible product name."""
    if not text:
        return None
    t = str(text).strip()
    t2 = SKU_TOKEN_RE.sub(" ", t)
    t2 = clean(t2)
    if not t2:
        return None
    if not re.search(r"[A-Za-z]", t2):  # must contain letters to be a name
        return None
    return t2


def scroll_until_loaded(driver, pause=1.0, max_rounds=25):
    last_count = 0
    rounds = 0
    while rounds < max_rounds:
        driver.find_elements(By.CSS_SELECTOR, "div.wixui-repeater__item")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause)
        items2 = driver.find_elements(By.CSS_SELECTOR, "div.wixui-repeater__item")
        if len(items2) == last_count:
            break
        last_count = len(items2)
        rounds += 1


def make_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1400,900")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

# =========================================================
# SCRAPING LOGIC
# =========================================================

def scrape_category(category, urls):
    driver = make_driver()
    wait = WebDriverWait(driver, 20)
    results = []
    seen = set()

    try:
        for url in urls:
            driver.get(url)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            scroll_until_loaded(driver)

            items = driver.find_elements(By.CSS_SELECTOR, "div.wixui-repeater__item")

            for item in items:
                product_url = None
                image_url = None
                product_name = None
                sku_list = []

                # Product URL
                try:
                    a = item.find_element(By.CSS_SELECTOR, "a[href]")
                    href = a.get_attribute("href")
                    if href:
                        product_url = urljoin("https://www.artesiacollections.com", href) if href.startswith("/") else href
                except:
                    pass

                # Image URL
                try:
                    img = item.find_element(By.CSS_SELECTOR, "img")
                    image_url = extract_image_url(img)
                except:
                    pass

                # =========================
                # RichText parsing (FINAL FIX)
                # =========================
                try:
                    rich_elems = item.find_elements(By.CSS_SELECTOR, '[data-testid="richTextElement"]')

                    for el in rich_elems:
                        raw = el.text or ""
                        raw_clean = clean(raw)
                        if not raw_clean:
                            continue

                        # 1) Extract SKUs anywhere in this block (even if merged with name)
                        found_skus = extract_skus_anywhere(raw_clean)
                        if found_skus:
                            for s in found_skus:
                                if s not in sku_list:
                                    sku_list.append(s)

                        # 2) Remove SKUs -> remaining is product name (if valid)
                        if not product_name:
                            name_candidate = strip_skus_from_text(raw_clean)
                            if name_candidate:
                                product_name = name_candidate

                except:
                    pass

                # Fallbacks (extra safety)
                if not sku_list:
                    fallback = pick_sku_from_text(clean(item.text) or "")
                    fallback_safe = safe_skus(fallback)
                    if fallback_safe:
                        for s in fallback_safe:
                            if s not in sku_list:
                                sku_list.append(s)

                if not product_name:
                    product_name = pick_name_from_text(clean(item.text) or "")

                if product_url and product_name:
                    key = (product_url, product_name)
                    if key in seen:
                        continue
                    seen.add(key)

                    results.append({
                        "Category": category,
                        "Product URL": product_url,
                        "Image URL": image_url,
                        "Product Name": product_name,
                        "SKU": ", ".join(sku_list) if sku_list else None
                    })

    finally:
        driver.quit()

    return pd.DataFrame(results)

# =========================================================
# OUTPUT SYSTEM
# =========================================================

def build_step1_master_excel():
    dfs = []
    for cat, links in CATEGORIES.items():
        df = scrape_category(cat, links)
        if not df.empty:
            dfs.append(df)

    if not dfs:
        cols = ["Category", "Product URL", "Image URL", "Product Name", "SKU"]
        empty = pd.DataFrame(columns=cols)
        empty.to_excel(master_output_file, index=False)
        return empty

    master = pd.concat(dfs, ignore_index=True)
    master.drop_duplicates(subset=["Product URL"], inplace=True)
    master.to_excel(master_output_file, index=False)
    return master


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)

        ws["A1"] = "Brand"
        ws["B1"] = "Artesia Collections"
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)


def main():
    df = build_step1_master_excel()
    if df.empty:
        Workbook().save(master_sheets_file)
        return
    build_category_wise_workbook_from_df(df)


if __name__ == "__main__":
    main()
