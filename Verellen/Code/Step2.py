"""
Verellen Detail Scraper - v10
==============================
v9 fixes retained.

v10 Fix: Sub-fields like Seat Depth, Seat Height, Arm Height, Width Between Arms
         were NOT being captured because they live in a different HTML structure
         than .col-md-5 rows inside imperial-wrapper.

Solution: Added text-based regex extraction (extract_subfields_from_text) that
          parses the FULL text of the entire dimension section for known patterns
          like "Seat Depth  18"" or "Arm Height: 29"". This works regardless
          of HTML structure — if the text is visible, it gets captured.

All sub-field values are numeric only: "29" D" → 29
"""

import time
import re
import os
import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "verellen_Outdoor_Seating.xlsx"
OUTPUT_FILE = "verellen_Outdoor_Seating_details.xlsx"
HEADLESS    = True

FIXED_COLS = [
    "Manufacturer",
    "Source",
    "Image URL",
    "Product Name",
    "SKU",
    "Product Family Id",
    "Description",
    "Dimension",
    "Weight",
    "Width",
    "Depth",
    "Diameter",
    "Length",
    "Height",
    "Finish",
    "Fabric",
    "Leather",
    "OVERALL DIMENSIONS",
    "Overhang",
    "Tearsheet Link",
]

SKU_OVERRIDE_MAP = [
    ("barcelona", "24",  "BRC 82700"),
    ("barcelona", "30",  "BRC 82800"),
    ("brisbane",  "",    "BRI 82191"),
    ("georgina",  "",    "GGA 88000"),
    ("giaco",     "22",  "GAC 82770"),
    ("giaco",     "18",  "GAC 81800"),
    ("lago",      "27",  "LGO 82700"),
    ("lago",      "31",  "LGO 82710"),
    ("leon",      "35",  "LEO 83500"),
    ("leon",      "29",  "LEO 83510"),
    ("menorca",   "",    "MEN 83700"),
]

OVERHANG_MAP = {
    "barcelona": '8.125"',
    "menorca":   '4" Long Side, Flush - Short Side',
}


def setup_driver():
    options = Options()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.page_load_strategy = "eager"
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(45)
    return driver


def safe_get(driver, url, max_retries=2):
    for attempt in range(max_retries + 1):
        try:
            driver.get(url)
            return driver, True
        except TimeoutException:
            print(f"    [timeout] attempt {attempt+1}")
            try:
                driver.execute_script("window.stop();")
                time.sleep(2)
                body = driver.find_element(By.TAG_NAME, "body").text
                if len(body) > 100:
                    return driver, True
            except:
                pass
            if attempt < max_retries:
                try: driver.quit()
                except: pass
                time.sleep(3)
                driver = setup_driver()
            else:
                try:
                    driver.execute_script("window.stop();")
                    return driver, True
                except:
                    return driver, False
        except WebDriverException as e:
            print(f"    [WebDriver error] {str(e)[:80]}")
            if attempt < max_retries:
                try: driver.quit()
                except: pass
                time.sleep(3)
                driver = setup_driver()
            else:
                return driver, False
    return driver, False


def wait_and_scroll(driver, retries=2):
    for attempt in range(retries + 1):
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, ".pdp-heading, .product-details, h3, .name-wrapper")
                )
            )
            break
        except:
            if attempt < retries:
                time.sleep(4)
            else:
                time.sleep(6)
    try:
        total = driver.execute_script("return document.body.scrollHeight")
        step  = driver.execute_script("return window.innerHeight")
        pos   = 0
        while pos < total:
            pos += step
            driver.execute_script(f"window.scrollTo(0, {pos});")
            time.sleep(0.5)
            total = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.8)
    except:
        pass


def safe_click(driver, css_or_el, wait_sec=3):
    try:
        if isinstance(css_or_el, str):
            el = WebDriverWait(driver, wait_sec).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, css_or_el))
            )
        else:
            el = css_or_el
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", el)
        time.sleep(1.2)
        return True
    except:
        return False


def get_image_url(driver) -> str:
    try:
        el = driver.find_element(By.CSS_SELECTOR, "link[rel='preload'].pdp-image")
        h = el.get_attribute("href") or ""
        if h: return h
    except: pass
    for sel in [
        ".pdp-left-content img[itemprop='URL']",
        ".pdp-left-content .pdp-image",
        ".pdp-left-content img",
        "img.pdp-image",
        "img[src*='magento']",
        "img[src*='verellen']",
    ]:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            src = el.get_attribute("src") or el.get_attribute("data-src") or ""
            if src and src.startswith("http"):
                return src
        except: pass
    return ""


def get_sku(driver) -> str:
    for sel in [".name-wrapper p", ".heading-wrapper p", ".pdp-sku", "[class*='sku']"]:
        try:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in els:
                t = el.text.strip()
                if t and len(t) < 30:
                    return t
        except: pass
    try:
        result = driver.execute_script("""
            var all = document.querySelectorAll('p, span, div');
            var pat = /^[A-Z]{2,6}\\s+[A-Z0-9]{2,8}$/;
            for (var i=0; i<all.length; i++) {
                var t = (all[i].innerText || '').trim();
                if (pat.test(t)) return t;
            }
            return '';
        """)
        if result: return result.strip()
    except: pass
    return ""


def resolve_sku(raw_sku: str, url: str, dimension_str: str) -> str:
    url_lower = url.lower()
    dim_lower = (dimension_str or "").lower()
    for url_kw, dim_kw, correct_sku in SKU_OVERRIDE_MAP:
        if url_kw in url_lower:
            if dim_kw == "" or dim_kw in dim_lower:
                return correct_sku
    return raw_sku


def resolve_overhang(url: str) -> str:
    url_lower = url.lower()
    for kw, val in OVERHANG_MAP.items():
        if kw in url_lower:
            return val
    return ""


def get_description(driver) -> str:
    for sel in [".pdp-description", "[class*='pdp-description']", ".product-description"]:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            t = el.text.strip()
            if t: return t
        except: pass
    return ""


def get_finishes_fabric_leather(driver) -> dict:
    result = {"Finish": "", "Fabric": "", "Leather": ""}
    links = []
    for sel in [".fabrics-link p", ".fabrics-link a", "p[style*='text-align: right']"]:
        try:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in els:
                txt = el.text.strip().upper()
                if any(kw in txt for kw in ["FINISH", "FABRIC", "LEATHER", "HOUSE"]):
                    links.append((el, txt))
        except: pass

    if not links:
        try:
            for tag in ["p", "a", "span"]:
                els = driver.find_elements(By.TAG_NAME, tag)
                for el in els:
                    try:
                        txt = el.text.strip().upper()
                        if len(txt) < 60 and any(kw in txt for kw in [
                            "IN-HOUSE FINISH", "VIEW FINISH", "VIEW FABRIC",
                            "VIEW LEATHER", "FABRIC OPTION", "LEATHER OPTION"
                        ]):
                            links.append((el, txt))
                    except: pass
                if links: break
        except: pass

    for link_el, link_text in links:
        try:
            category = "Finish"
            if "FABRIC"  in link_text: category = "Fabric"
            elif "LEATHER" in link_text: category = "Leather"
            safe_click(driver, link_el)
            time.sleep(1.5)
            if category == "Finish":
                try:
                    for tsel in [".modal-title", ".modal-header h5", ".modal-header h4"]:
                        mt = driver.find_element(By.CSS_SELECTOR, tsel).text.strip().upper()
                        if "FABRIC"  in mt: category = "Fabric"
                        elif "LEATHER" in mt: category = "Leather"
                        break
                except: pass
            names = []
            for name_sel in [".modal-body .product-name", ".product-container .product-name", ".modal .product-name"]:
                els = driver.find_elements(By.CSS_SELECTOR, name_sel)
                if els:
                    for el in els:
                        t = el.text.strip()
                        if t: names.append(t.upper())
                    break
            if names:
                existing = result[category]
                if existing:
                    existing_set = set(existing.split(", "))
                    new_names = [n for n in names if n not in existing_set]
                    if new_names:
                        result[category] = existing + ", " + ", ".join(new_names)
                else:
                    result[category] = ", ".join(names)
            try:
                close = driver.find_element(By.CSS_SELECTOR,
                    ".modal .close, button[data-dismiss='modal'], .modal-header button")
                driver.execute_script("arguments[0].click();", close)
                time.sleep(0.5)
            except:
                try:
                    from selenium.webdriver.common.keys import Keys
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    time.sleep(0.5)
                except: pass
        except Exception as e:
            print(f"    [finish err] {e}")
    return result


FIELD_RE = {
    "Length":   re.compile(r'(\d+(?:\.\d+)?)\s*["\u2033]?\s*L(?:[^a-zA-Z]|$)', re.I),
    "Width":    re.compile(r'(\d+(?:\.\d+)?)\s*["\u2033]?\s*W(?:[^a-zA-Z]|$)', re.I),
    "Height":   re.compile(r'(\d+(?:\.\d+)?)\s*["\u2033]?\s*H(?:[^a-zA-Z]|$)', re.I),
    "Depth":    re.compile(r'(\d+(?:\.\d+)?)\s*["\u2033]?\s*D(?:[^a-zA-Z]|$)', re.I),
    "Diameter": re.compile(r'(\d+(?:\.\d+)?)\s*["\u2033]?\s*(?:Dia(?:m(?:eter)?)?|DIA|Diameter)', re.I),
    "Weight":   re.compile(r'(\d+(?:\.\d+)?)\s*(?:lbs?|Ibs?|LBS?)', re.I),
}

LABEL_BEFORE_RE = {
    "Width":  re.compile(r'\bW\s+(\d+(?:\.\d+)?)\s*["\u2033]', re.I),
    "Depth":  re.compile(r'\bD\s+(\d+(?:\.\d+)?)\s*["\u2033]', re.I),
    "Height": re.compile(r'\bH\s+(\d+(?:\.\d+)?)\s*["\u2033]', re.I),
    "Length": re.compile(r'\bL\s+(\d+(?:\.\d+)?)\s*["\u2033]', re.I),
}


def parse_dim_fields(text: str) -> dict:
    result = {k: "" for k in ["Length", "Width", "Height", "Depth", "Diameter", "Weight"]}
    for field, pat in FIELD_RE.items():
        m = pat.search(text)
        if m: result[field] = m.group(1)
    for field, pat in LABEL_BEFORE_RE.items():
        if not result[field]:
            m = pat.search(text)
            if m: result[field] = m.group(1)
    return result


def fill_missing_from_overall(fields: dict, overall_str: str) -> dict:
    if not overall_str: return fields
    parsed = parse_dim_fields(overall_str)
    for key in ["Width", "Depth", "Height", "Length", "Diameter"]:
        if not fields.get(key) and parsed.get(key):
            fields[key] = parsed[key]
    return fields


_NUM_RE = re.compile(r'(\d+(?:\.\d+)?)')

def extract_number(text: str) -> str:
    if not text:
        return text
    text = text.strip()
    m = _NUM_RE.search(text)
    if m:
        val = m.group(1)
        try:
            f = float(val)
            return str(int(f)) if f == int(f) else val
        except:
            return val
    return text


# ★ v10: Text-based regex to find ALL known sub-field patterns in dimension text
#   Matches: "Seat Depth  18"" or "Seat Depth: 18"" or "Seat Depth 18" D"
#   Returns dict like {"Seat Depth": "18", "Arm Height": "29"}
_SUBFIELD_RE = re.compile(
    r'(Seat\s*Depth|Seat\s*Height|Seat\s*Width|'
    r'Arm\s*Height|Arm\s*Width|Arm\s*Depth|'
    r'Width\s*Between\s*Arms|'
    r'Interior\s*Depth|Interior\s*Width|'
    r'Exterior\s*Depth|Exterior\s*Width|'
    r'Back\s*Height|Back\s*Width|'
    r'Base\s*Height|Leg\s*Height|Leg\s*Dia(?:meter)?|'
    r'Top\s*Thick(?:ness)?|Thickness|'
    r'Height\s*to\s*Apron|Clearance|'
    r'Footrest(?:\s*Height)?|Stretcher\s*Height|'
    r'Cushion\s*Height|Cushion\s*Depth|'
    r'Rail\s*Height|Spring\s*Height)'
    r'\s*:?\s*'
    r'(\d+(?:\.\d+)?)\s*["\u2033]?',
    re.I
)

def extract_subfields_from_text(text: str) -> dict:
    """Parse dimension section text for ALL known sub-field patterns.
    Works regardless of HTML structure — pure text matching."""
    result = {}
    for m in _SUBFIELD_RE.finditer(text):
        label = m.group(1).strip()
        # Normalize: "seat  depth" → "Seat Depth"
        label = re.sub(r'\s+', ' ', label).title()
        value = m.group(2)
        # Store numeric only
        try:
            f = float(value)
            value = str(int(f)) if f == int(f) else value
        except:
            pass
        if label not in result:
            result[label] = value
    return result


EXTRA_LABEL_RE = re.compile(
    r'Base\s*Height|Top\s*Thick|Thickness|Height\s*to\s*Apron|'
    r'Interior\s*Depth|Interior\s*Width|Exterior\s*Depth|Exterior\s*Width|'
    r'Seat\s*Depth|Seat\s*Height|Seat\s*Width|'
    r'Arm\s*Width|Arm\s*Height|Arm\s*Depth|'
    r'Width\s*Between\s*Arms|Between\s*Arms|'
    r'Shelf|Clearance|'
    r'Opening|Inside|Adjustable|Drawer|Leg\s*Height|'
    r'Leg\s*Dia|Rail|Stretcher|Footrest|Rungs|Slat|'
    r'Back\s*Height|Back\s*Width|Cushion|Spring|Frame|Apron',
    re.I
)


def get_dimensions(driver) -> list:
    expanded = False
    try:
        btns = driver.find_elements(By.CSS_SELECTOR, "button.btn-gray, .dimensions-btn-wrapper button")
        for btn in btns:
            if "DIMENSION" in btn.text.upper():
                safe_click(driver, btn)
                expanded = True
                break
    except: pass

    if not expanded:
        try:
            for btn in driver.find_elements(By.TAG_NAME, "button"):
                try:
                    if "DIMENSION" in btn.text.upper():
                        safe_click(driver, btn)
                        expanded = True
                        break
                except: pass
        except: pass

    time.sleep(1.2)

    imperial = None
    for sel in [".imperial-wrapper", "[class*='imperial']", ".dimensions-content"]:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            if el.text.strip():
                imperial = el
                break
        except:
            imperial = None

    if not imperial:
        print("    [dim] imperial-wrapper not found")
        return []

    # ★ v10: Get the FULL text of the ENTIRE dimension section
    #   (parent of imperial-wrapper, or even grandparent)
    #   This captures ALL sub-fields regardless of HTML nesting
    dim_section_text = ""
    try:
        # Try parent container first (the section that holds imperial + metric)
        parent = imperial.find_element(By.XPATH, "..")
        dim_section_text = parent.text.strip()
        if not dim_section_text:
            dim_section_text = imperial.text.strip()
        # Also try grandparent (the whole dimension panel)
        try:
            grandparent = parent.find_element(By.XPATH, "..")
            gp_text = grandparent.text.strip()
            if len(gp_text) > len(dim_section_text):
                dim_section_text = gp_text
        except:
            pass
    except:
        dim_section_text = imperial.text.strip()

    print(f"    [dim] Full section text ({len(dim_section_text)} chars): {dim_section_text[:120]}...")

    # ★ v10: Extract sub-fields from FULL text using regex
    text_subfields = extract_subfields_from_text(dim_section_text)
    if text_subfields:
        print(f"    [dim] ★ Text-extracted sub-fields: {text_subfields}")

    # Row-based extraction (col-md-5 approach from v9)
    rows = []
    seen_row_ids = set()

    try:
        col_md5_els = imperial.find_elements(By.CSS_SELECTOR, ".col-md-5")
        for col_el in col_md5_els:
            try:
                parent_row = col_el.find_element(By.XPATH, "..")
                row_id = id(parent_row)
                if row_id not in seen_row_ids:
                    seen_row_ids.add(row_id)
                    rows.append(parent_row)
            except:
                pass
    except:
        pass

    try:
        direct_divs = imperial.find_elements(By.XPATH, "./div")
        for div in direct_divs:
            div_id = id(div)
            if div_id not in seen_row_ids:
                text = div.text.strip()
                if text:
                    seen_row_ids.add(div_id)
                    rows.append(div)
    except:
        pass

    # ★ v10: Also search PARENT container for col-md-5 rows (outside imperial-wrapper)
    try:
        parent = imperial.find_element(By.XPATH, "..")
        parent_col5_els = parent.find_elements(By.CSS_SELECTOR, ".col-md-5")
        for col_el in parent_col5_els:
            try:
                parent_row = col_el.find_element(By.XPATH, "..")
                row_id = id(parent_row)
                if row_id not in seen_row_ids:
                    seen_row_ids.add(row_id)
                    rows.append(parent_row)
            except:
                pass
    except:
        pass

    # ★ v10: Also try grandparent
    try:
        grandparent = imperial.find_element(By.XPATH, "../..")
        gp_col5_els = grandparent.find_elements(By.CSS_SELECTOR, ".col-md-5")
        for col_el in gp_col5_els:
            try:
                parent_row = col_el.find_element(By.XPATH, "..")
                row_id = id(parent_row)
                if row_id not in seen_row_ids:
                    seen_row_ids.add(row_id)
                    rows.append(parent_row)
            except:
                pass
    except:
        pass

    print(f"    [dim] Found {len(rows)} dimension rows")

    main_variants = []
    extra_fields  = {}

    STD_KEYS = {"variant_label", "dimension_full",
                "Length", "Width", "Height", "Depth", "Diameter", "Weight"}

    for row in rows:
        full_text = row.text.strip()
        if not full_text: continue

        left_els  = row.find_elements(By.CSS_SELECTOR, ".col-md-5")
        right_els = row.find_elements(By.CSS_SELECTOR, ".col-md:not(.col-md-5)")

        label   = left_els[0].text.strip()  if left_els  else ""
        dim_str = right_els[0].text.strip() if right_els else full_text

        if not dim_str: continue

        has_main = bool(re.search(
            r'\d+\s*["\u2033]?\s*[xX×]\s*\d+'
            r'|\d+\s*["\u2033]?\s*[WDHL](?:\s|$|[^a-z])'
            r'|[WDHL]\s+\d+\s*["\u2033]',
            dim_str, re.I
        ))

        is_extra = bool(EXTRA_LABEL_RE.search(label or full_text))

        if is_extra or (label and not has_main and not re.search(r'[xX×]', dim_str)):
            clean_label = re.sub(r'[\s:]+$', '', label or full_text).strip()
            if clean_label:
                extra_fields[clean_label] = extract_number(dim_str)
                print(f"    [dim] row-extra: {clean_label} = {dim_str} → {extra_fields[clean_label]}")
        else:
            fields = parse_dim_fields(dim_str)
            main_variants.append({
                "variant_label":  label,
                "dimension_full": dim_str,
                **fields,
            })

    # ★ v10: Also extract from JS — look for ALL dimension text including
    #   sub-fields that might be in different containers
    try:
        js_subfields = driver.execute_script("""
            var result = {};
            // Look for any element whose text matches sub-field patterns
            var labels = [
                'Seat Depth', 'Seat Height', 'Seat Width',
                'Arm Height', 'Arm Width', 'Arm Depth',
                'Width Between Arms',
                'Interior Depth', 'Exterior Depth',
                'Back Height', 'Back Width',
                'Base Height', 'Leg Height',
                'Footrest Height', 'Cushion Height'
            ];
            // Strategy 1: Find col-md-5 elements with these labels
            document.querySelectorAll('.col-md-5, .col-md, .col').forEach(function(el) {
                var t = (el.innerText || '').trim();
                for (var i = 0; i < labels.length; i++) {
                    if (t.toLowerCase().indexOf(labels[i].toLowerCase()) >= 0) {
                        // Find sibling with value
                        var sib = el.nextElementSibling;
                        if (sib) {
                            var v = (sib.innerText || '').trim();
                            if (v) result[labels[i]] = v;
                        }
                        // Or parent's other child
                        if (!result[labels[i]]) {
                            var par = el.parentElement;
                            if (par) {
                                var kids = par.children;
                                for (var j = 0; j < kids.length; j++) {
                                    if (kids[j] !== el) {
                                        var kv = (kids[j].innerText || '').trim();
                                        if (kv && /\\d/.test(kv)) {
                                            result[labels[i]] = kv;
                                            break;
                                        }
                                    }
                                }
                            }
                        }
                        break;
                    }
                }
            });
            // Strategy 2: Find any element with these exact labels as text
            var allEls = document.querySelectorAll('p, div, span, td, th, li, dt, dd');
            for (var i = 0; i < allEls.length; i++) {
                var text = (allEls[i].innerText || '').trim();
                for (var j = 0; j < labels.length; j++) {
                    if (!result[labels[j]] && text.toLowerCase() === labels[j].toLowerCase()) {
                        var sib = allEls[i].nextElementSibling;
                        if (sib) {
                            var v = (sib.innerText || '').trim();
                            if (v && /\\d/.test(v)) result[labels[j]] = v;
                        }
                        if (!result[labels[j]]) {
                            var par = allEls[i].parentElement;
                            if (par) {
                                var full = par.innerText.replace(text, '').trim();
                                if (full && /\\d/.test(full)) result[labels[j]] = full;
                            }
                        }
                    }
                }
            }
            return result;
        """)
        if js_subfields:
            print(f"    [dim] ★ JS-extracted sub-fields: {js_subfields}")
            for k, v in js_subfields.items():
                if k and v:
                    nk = re.sub(r'\s+', ' ', k).strip().title()
                    nv = extract_number(v)
                    if nk not in extra_fields:
                        extra_fields[nk] = nv
    except Exception as e:
        print(f"    [dim] JS extraction err: {e}")

    # ★ v10: Merge text-extracted sub-fields (highest priority fallback)
    for k, v in text_subfields.items():
        if k not in extra_fields:
            extra_fields[k] = v

    # Attach extra_fields to all variants
    for v in main_variants:
        for ek, ev in extra_fields.items():
            if ek not in v:
                v[ek] = ev

    seen_dims = set()
    unique_variants = []
    for v in main_variants:
        key = v["dimension_full"].strip().upper()
        if key and key not in seen_dims:
            seen_dims.add(key)
            unique_variants.append(v)

    if not unique_variants:
        full_text = imperial.text.strip()
        if full_text:
            fields = parse_dim_fields(full_text)
            fallback = {"variant_label": "", "dimension_full": full_text, **fields}
            for ek, ev in extra_fields.items():
                fallback[ek] = ev
            unique_variants.append(fallback)

    if extra_fields:
        print(f"    [dim] ★ FINAL extra sub-fields: {extra_fields}")

    return unique_variants


def get_details(driver) -> dict:
    result = {}

    already_open = False
    try:
        minus_icons = driver.find_elements(By.CSS_SELECTOR,
            'svg[icon="minusIcon"], .details-btn-wrapper svg[icon="minusIcon"]')
        if minus_icons and any(m.is_displayed() for m in minus_icons):
            already_open = True
            print("    [details] already open (minus icon visible)")
    except: pass

    details_btn = None
    if not already_open:
        try:
            for btn in driver.find_elements(By.CSS_SELECTOR,
                    ".details-btn-wrapper button, button.btn-gray"):
                txt = btn.text.strip().upper()
                if "DETAIL" in txt and "DIMENSION" not in txt:
                    details_btn = btn
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", btn)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", btn)
                    print("    [details] clicked DETAILS button")
                    break
        except: pass

    appeared = False
    for wait_sel in [".details-header-text", ".pt-3.pl-3 .mb-3", ".pt-3 .mb-3"]:
        try:
            WebDriverWait(driver, 8).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, wait_sel))
            )
            appeared = True
            print(f"    [details] React rendered: '{wait_sel}' found")
            break
        except:
            pass

    if not appeared:
        print("    [details] WebDriverWait timeout — trying anyway after 3s")
        time.sleep(3)

    try:
        debug_info = driver.execute_script("""
            return {
                mb3_count:      document.querySelectorAll('.mb-3').length,
                hdr_count:      document.querySelectorAll('.details-header-text').length,
                pt3_count:      document.querySelectorAll('.pt-3.pl-3').length,
                pb5_count:      document.querySelectorAll('.pb-5').length,
                btn_count:      document.querySelectorAll('.details-btn-wrapper').length,
                minus_count:    document.querySelectorAll('svg[icon="minusIcon"]').length,
                plus_count:     document.querySelectorAll('svg[icon="plusIcon"]').length,
            };
        """)
        print(f"    [details] DOM: {debug_info}")
    except: pass

    try:
        hdrs = driver.find_elements(By.CSS_SELECTOR, ".details-header-text")
        print(f"    [details] 5A: {len(hdrs)} .details-header-text found")
        for h_el in hdrs:
            try:
                h = h_el.text.strip().rstrip(":").strip()
                v = ""
                try:
                    sib = h_el.find_element(By.XPATH, "following-sibling::p[1]")
                    v = sib.text.strip()
                except: pass
                if not v:
                    try:
                        parent = h_el.find_element(By.XPATH, "..")
                        v = parent.text.replace(h_el.text, "").strip().lstrip(":").strip()
                    except: pass
                if h and v and h != v:
                    result[h] = v
                    print(f"    [details] ✓5A {h}: {v[:60]}")
            except: pass
    except Exception as e:
        print(f"    [details] 5A err: {e}")

    if not result:
        try:
            for sec in driver.find_elements(By.CSS_SELECTOR, ".mb-3"):
                try:
                    ps = sec.find_elements(By.TAG_NAME, "p")
                    if len(ps) >= 2:
                        h = ps[0].text.strip().rstrip(":").strip()
                        v = " ".join(p.text.strip() for p in ps[1:] if p.text.strip())
                        if h and v and h != v and len(h) < 80:
                            result[h] = v
                            print(f"    [details] ✓5B {h}: {v[:60]}")
                except: pass
        except Exception as e:
            print(f"    [details] 5B err: {e}")

    if not result:
        try:
            pairs = driver.execute_script("""
                var out = [];
                document.querySelectorAll('.details-header-text').forEach(function(h) {
                    var k = h.textContent.trim().replace(/:$/, '').trim();
                    var v = '';
                    var sib = h.nextElementSibling;
                    if (sib) v = sib.textContent.trim();
                    if (!v) {
                        var par = h.parentElement;
                        if (par) {
                            var allP = par.querySelectorAll('p');
                            if (allP.length >= 2) v = allP[allP.length-1].textContent.trim();
                        }
                    }
                    if (k && v && k !== v) out.push(k + '|||' + v);
                });
                if (out.length === 0) {
                    document.querySelectorAll('.mb-3').forEach(function(sec) {
                        var ps = sec.querySelectorAll('p');
                        if (ps.length >= 2) {
                            var k = ps[0].textContent.trim().replace(/:$/, '').trim();
                            var v = ps[ps.length-1].textContent.trim();
                            if (k && v && k !== v && k.length < 80) out.push(k + '|||' + v);
                        }
                    });
                }
                return out;
            """)
            for pair in (pairs or []):
                parts = pair.split("|||", 1)
                if len(parts) == 2:
                    k, v = parts[0].strip(), parts[1].strip()
                    if k and v:
                        result[k] = v
            if result:
                print(f"    [details] ✓5C JS: {len(result)} pairs")
        except Exception as e:
            print(f"    [details] 5C err: {e}")

    if not result:
        try:
            source = driver.page_source
            has_hdr = 'details-header-text' in source
            has_txt = 'details-text' in source
            print(f"    [details] 5D HTML: has header={has_hdr}, has text={has_txt}")
            if has_hdr:
                header_vals = re.findall(
                    r'class="details-header-text"[^>]*>\s*([^<]+?)\s*</p>', source)
                text_vals = re.findall(
                    r'class="details-text[^"]*"[^>]*>([\s\S]*?)</p>', source)
                for h, t in zip(header_vals, text_vals):
                    h = h.strip().rstrip(":").strip()
                    t = re.sub(r'<[^>]+>', ' ', t).strip()
                    t = re.sub(r'\s+', ' ', t).strip()
                    if h and t:
                        result[h] = t
                        print(f"    [details] ✓5D {h}: {t[:60]}")
        except Exception as e:
            print(f"    [details] 5D err: {e}")

    if not result:
        print("    [details] ⚠ STILL empty — last resort retry")
        try:
            for btn in driver.find_elements(By.CSS_SELECTOR, ".details-btn-wrapper button"):
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(4)
                for h_el in driver.find_elements(By.CSS_SELECTOR, ".details-header-text"):
                    try:
                        h = h_el.text.strip().rstrip(":").strip()
                        sib = h_el.find_element(By.XPATH, "following-sibling::p[1]")
                        v = sib.text.strip()
                        if h and v:
                            result[h] = v
                    except: pass
                if result:
                    print(f"    [details] ✓ last-resort: {len(result)} pairs")
                break
        except: pass
    else:
        print(f"    [details] ✓ Total {len(result)} fields: {list(result.keys())}")

    return result


def extract_overall_dimensions(details: dict, dimension_str: str) -> str:
    for key in details:
        if re.search(r'overall\s*dim', key, re.I):
            return details[key]
    return dimension_str.strip() if dimension_str else ""


def get_product_family_id(product_name: str) -> str:
    if not product_name: return ""
    parts = re.split(r"[\s\-\._]+", product_name.strip(), maxsplit=1)
    return parts[0].strip()


def get_tearsheet_link(driver) -> str:
    for sel in ["a[href*='tearsheet']", "a[href*='pdf']", "a[href*='download']"]:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            href = el.get_attribute("href") or ""
            if href: return href
        except: pass
    return ""


def scrape_product(driver, url: str, product_name: str, manufacturer: str):
    print(f"  → {product_name}")
    rows = []

    try:
        driver, ok = safe_get(driver, url)
        if not ok:
            print(f"    [SKIP] page load failed")
            rows.append(_error_row(manufacturer, product_name, url))
            return driver, rows

        wait_and_scroll(driver)

        image_url   = get_image_url(driver)
        raw_sku     = get_sku(driver)
        description = get_description(driver)

        try:
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.5)
        except: pass

        details     = get_details(driver)
        dimensions  = get_dimensions(driver)
        ffl         = get_finishes_fabric_leather(driver)
        tearsheet   = get_tearsheet_link(driver)
        overhang    = resolve_overhang(url)
        family_id   = get_product_family_id(product_name)

        STD_KEYS = {"variant_label", "dimension_full",
                    "Length", "Width", "Height", "Depth", "Diameter", "Weight"}

        if dimensions:
            seen = set()
            for variant in dimensions:
                dim_key = variant["dimension_full"].strip().upper()
                if dim_key in seen: continue
                seen.add(dim_key)

                sku         = resolve_sku(raw_sku, url, variant["dimension_full"])
                overall_dim = extract_overall_dimensions(details, variant["dimension_full"])

                dim_fields = {
                    "Width":    variant.get("Width",    ""),
                    "Depth":    variant.get("Depth",    ""),
                    "Diameter": variant.get("Diameter", ""),
                    "Length":   variant.get("Length",   ""),
                    "Height":   variant.get("Height",   ""),
                }
                dim_fields = fill_missing_from_overall(dim_fields, overall_dim)

                row = {
                    "Manufacturer":       manufacturer,
                    "Source":             url,
                    "Image URL":          image_url,
                    "Product Name":       product_name,
                    "SKU":                sku,
                    "Product Family Id":  family_id,
                    "Description":        description,
                    "Dimension":          variant["dimension_full"],
                    "OVERALL DIMENSIONS": overall_dim,
                    "Weight":             variant.get("Weight", ""),
                    "Finish":             ffl["Finish"],
                    "Fabric":             ffl["Fabric"],
                    "Leather":            ffl["Leather"],
                    "Tearsheet Link":     tearsheet,
                    "Overhang":           overhang,
                    **dim_fields,
                }

                for key, val in details.items():
                    if not re.search(r'overall\s*dim', key, re.I):
                        if key not in row:
                            row[key] = val

                for k, v in variant.items():
                    if k not in STD_KEYS and k not in row:
                        row[k] = v

                print(f"    SKU={sku}  dim={variant['dimension_full'][:35]}")
                rows.append(row)

        else:
            sku         = resolve_sku(raw_sku, url, "")
            overall_dim = extract_overall_dimensions(details, "")
            dim_fields  = fill_missing_from_overall(
                {k: "" for k in ["Width", "Depth", "Diameter", "Length", "Height"]},
                overall_dim
            )

            row = {
                "Manufacturer":       manufacturer,
                "Source":             url,
                "Image URL":          image_url,
                "Product Name":       product_name,
                "SKU":                sku,
                "Product Family Id":  family_id,
                "Description":        description,
                "Dimension":          "",
                "OVERALL DIMENSIONS": overall_dim,
                "Weight":             "",
                "Finish":             ffl["Finish"],
                "Fabric":             ffl["Fabric"],
                "Leather":            ffl["Leather"],
                "Tearsheet Link":     tearsheet,
                "Overhang":           overhang,
                **dim_fields,
            }
            for key, val in details.items():
                if not re.search(r'overall\s*dim', key, re.I):
                    if key not in row:
                        row[key] = val

            print(f"    SKU={sku}  [no dimensions scraped]")
            rows.append(row)

    except Exception as e:
        print(f"    [ERROR] {e}")
        traceback.print_exc()
        rows.append(_error_row(manufacturer, product_name, url))

    return driver, rows


def _error_row(manufacturer, product_name, url):
    row = {col: "" for col in FIXED_COLS}
    row["Manufacturer"]      = manufacturer
    row["Product Name"]      = product_name
    row["Source"]            = url
    row["Product Family Id"] = get_product_family_id(product_name)
    row["Overhang"]          = resolve_overhang(url)
    return row


def save_output_excel(all_rows: list, filename: str):
    if not all_rows:
        print("No rows to save.")
        return

    seen_keys = set()
    deduped = []
    for row in all_rows:
        sku  = str(row.get("SKU", "")).strip()
        name = str(row.get("Product Name", "")).strip()
        key  = (sku, name)
        if sku and name and key in seen_keys:
            print(f"    [dedup] SKU={sku} + '{name}' duplicate — row skipped")
            continue
        if sku and name:
            seen_keys.add(key)
        deduped.append(row)
    removed = len(all_rows) - len(deduped)
    if removed:
        print(f"    [dedup] {removed} duplicate row(s) removed (same SKU + Product Name)")
    all_rows = deduped

    extra_cols = []
    seen_extra = set()
    for row in all_rows:
        for k in row:
            if k not in FIXED_COLS and k not in seen_extra:
                extra_cols.append(k)
                seen_extra.add(k)

    final_cols = FIXED_COLS + extra_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Details"

    bdr = Border(
        left   = Side(style='thin', color='D0D0D0'),
        right  = Side(style='thin', color='D0D0D0'),
        top    = Side(style='thin', color='D0D0D0'),
        bottom = Side(style='thin', color='D0D0D0'),
    )
    hdr_bdr = Border(
        left   = Side(style='thin',   color='BBBBBB'),
        right  = Side(style='thin',   color='BBBBBB'),
        top    = Side(style='thin',   color='BBBBBB'),
        bottom = Side(style='medium', color='888888'),
    )
    NUM_COLS = {"Width", "Depth", "Diameter", "Length", "Height", "Weight",
                "Seat Depth", "Seat Height", "Seat Width",
                "Arm Width", "Arm Height", "Arm Depth",
                "Width Between Arms", "Interior Depth", "Exterior Depth",
                "Interior Width", "Exterior Width",
                "Back Height", "Back Width", "Base Height",
                "Leg Height", "Leg Dia", "Leg Diameter",
                "Footrest Height", "Cushion Height", "Cushion Depth",
                "Rail Height", "Stretcher Height", "Top Thickness", "Thickness",
                "Clearance", "Height To Apron"}
    WIDTHS   = {
        "Manufacturer":14, "Product Name":22, "SKU":13,
        "Product Family Id":15, "Source":36, "Image URL":36,
        "Description":48, "Dimension":26, "OVERALL DIMENSIONS":26,
        "Weight":8, "Width":8, "Depth":8, "Diameter":9,
        "Length":8, "Height":8, "Finish":52,
        "Fabric":18, "Leather":18, "Overhang":22, "Tearsheet Link":26,
    }

    for ci, col in enumerate(final_cols, 1):
        c = ws.cell(row=1, column=ci)
        c.value     = col
        c.font      = Font(name='Calibri', bold=True, size=10, color='000000')
        c.fill      = PatternFill('solid', fgColor='EFEFEF')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = hdr_bdr
    ws.row_dimensions[1].height = 28

    for ri, row_data in enumerate(all_rows, 2):
        for ci, col in enumerate(final_cols, 1):
            val = row_data.get(col, "")
            if val is None or str(val) == "nan": val = ""

            if col in NUM_COLS and val != "":
                try:
                    f = float(val)
                    val = int(f) if f == int(f) else f
                except: pass

            c = ws.cell(row=ri, column=ci)
            c.value  = val if val != "" else None
            c.border = bdr

            if col in NUM_COLS:
                c.font      = Font(name='Calibri', size=10)
                c.alignment = Alignment(horizontal='center', vertical='top')
            elif col == 'SKU':
                c.font      = Font(name='Calibri', size=10, bold=True)
                c.alignment = Alignment(horizontal='left', vertical='top')
            elif col == 'Source' and val:
                c.hyperlink = str(val)
                c.font      = Font(name='Calibri', size=10,
                                   color='0563C1', underline='single')
                c.alignment = Alignment(horizontal='left', vertical='top')
            else:
                c.font      = Font(name='Calibri', size=10)
                c.alignment = Alignment(horizontal='left', vertical='top',
                                        wrap_text=True)
        ws.row_dimensions[ri].height = 55

    for ci, col in enumerate(final_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(col, 26)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(final_cols))}1"

    wb.save(filename)
    print(f"\n✅  Saved : {filename}")
    print(f"    Rows   : {len(all_rows)}")
    print(f"    Fixed  : {len(FIXED_COLS)}")
    print(f"    Dynamic: {len(extra_cols)} → {extra_cols}")
    print(f"    Total  : {len(final_cols)} columns")


def read_input_excel(filename: str) -> list:
    wb = load_workbook(filename)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    products = []
    seen_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        url  = (item.get("Source") or "").strip()
        name = (item.get("Product Name") or "").strip()
        mfr  = (item.get("Manufacturer") or "Verellen").strip()
        if url and name and url not in seen_urls:
            seen_urls.add(url)
            products.append({"url": url, "name": name, "manufacturer": mfr})
    return products


def main():
    print("Verellen Detail Scraper v10")
    print(f"Headless: {HEADLESS}")
    print("=" * 55)

    if not os.path.exists(INPUT_FILE):
        print(f"❌ Input file not found: {INPUT_FILE}")
        return

    products = read_input_excel(INPUT_FILE)
    print(f"Loaded {len(products)} unique products from {INPUT_FILE}\n")

    driver   = setup_driver()
    all_rows = []

    try:
        for i, p in enumerate(products, 1):
            print(f"[{i}/{len(products)}]", end=" ")
            driver, rows = scrape_product(driver, p["url"], p["name"], p["manufacturer"])
            all_rows.extend(rows)
            time.sleep(1.5)

    except KeyboardInterrupt:
        print("\n⚠ Interrupted — saving partial results...")
    except Exception as e:
        print(f"\n[FATAL] {e}")
        traceback.print_exc()
    finally:
        try: driver.quit()
        except: pass

    save_output_excel(all_rows, OUTPUT_FILE)


if __name__ == "__main__":
    main()