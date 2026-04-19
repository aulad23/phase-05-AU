# ==============================
# Integrated Scraper
# CODE-A Scraping Logic
# CODE-B Input / Output System
# ==============================

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import pandas as pd
import os
import re
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, urlunparse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ================= CONFIG (FROM CODE-B) =================
BASE_DIR = os.getcwd()

MASTER_OUTPUT_FILE = os.path.join(BASE_DIR, "hollyhunt_ALL_categories.xlsx")
MASTER_SHEETS_FILE = os.path.join(BASE_DIR, "HollyHunt.xlsx")

DEFAULT_MAX_PAGES = 200

BASE_URL = "https://www.hollyhunt.com"

CATEGORIES = {
    "Nightstands": [
        "https://www.hollyhunt.com/products/furniture/bedside-tables-dressers"
    ],
    "Coffee & Cocktail Tables": [
        "https://www.hollyhunt.com/products/furniture/cocktail-tables"
    ],
    "Side & End Tables": [
        "https://www.hollyhunt.com/products/furniture/side-drink-tables"
    ],
    "Dining Tables": [
        "https://www.hollyhunt.com/products/furniture/dining-tables"
    ],
    "Consoles": [
        "https://www.hollyhunt.com/products/furniture/desks-consoles"
    ],
    "Beds & Headboards": [
        "https://www.hollyhunt.com/products/furniture/beds"
    ],
    "Cabinets": [
        "https://www.hollyhunt.com/products/furniture/cabinets-bookcases"
    ],
    "Dining Chairs": [
        "https://www.hollyhunt.com/products/furniture/dining-chairs"
    ],
    "Bar Stools": [
        "https://www.hollyhunt.com/products/furniture/counter-bar-stools"
    ],
    "Sofas & Loveseats": [
        "https://www.hollyhunt.com/products/furniture/sofas-sectionals?prod%5BrefinementList%5D%5Bpv_field_sku_product_type_filter%5D%5B0%5D=Sofas"
    ],
    "Sectionals": [
        "https://www.hollyhunt.com/products/furniture/sofas-sectionals?prod%5BrefinementList%5D%5Bpv_field_sku_product_type_filter%5D%5B0%5D=Available%20as%20sectional"
    ],
    "Lounge Chairs": [
        "https://www.hollyhunt.com/products/furniture/lounge-chairs"
    ],
    "Benches": [
        "https://www.hollyhunt.com/products/furniture/benches-ottomans"
    ],

}
# =======================================================

# ============== Selenium Setup (CODE-A UNCHANGED) ==============
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=chrome_options
)

# ============== Scraping Logic (CODE-A UNCHANGED) ==============
def scrape_page(page_url):
    driver.get(page_url)
    time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    products = []

    for card in soup.select("div.e-chip--product"):
        name_tag = card.select_one(".e-chip__row-1 a")
        product_name = name_tag.get_text(strip=True) if name_tag else None
        product_url = urljoin(BASE_URL, name_tag["href"]) if name_tag else None

        img_tag = card.select_one(".e-chip__image img")
        image_url = urljoin(BASE_URL, img_tag["src"]) if img_tag else None

        products.append({
            "Product URL": product_url,
            "Image URL": image_url,
            "Product Name": product_name
        })

    return products

def scrape_category(category, links):
    all_products = []

    for base_url in links:
        for page_number in range(1, DEFAULT_MAX_PAGES + 1):
            page_url = (
                base_url if page_number == 1
                else f"{base_url}?prod%5Bpage%5D={page_number}"
            )

            products = scrape_page(page_url)
            if not products:
                break

            for p in products:
                p["Category"] = category
            all_products.extend(products)

    df = pd.DataFrame(all_products)
    if not df.empty:
        df.drop_duplicates(subset=["Product URL"], inplace=True)
    return df

# ================= OUTPUT SYSTEM (FROM CODE-B) =================
def build_step1_master_excel():
    dfs = []

    for cat, links in CATEGORIES.items():
        df = scrape_category(cat, links)
        if not df.empty:
            dfs.append(df)
        else:
            print(f"No products found for category: {cat}")

    if not dfs:
        empty_cols = ["Category", "Product URL", "Image URL", "Product Name"]
        master = pd.DataFrame(columns=empty_cols)
        master.to_excel(MASTER_OUTPUT_FILE, index=False)
        return master

    master = pd.concat(dfs, ignore_index=True)
    master = master[["Category", "Product URL", "Image URL", "Product Name"]]
    master.to_excel(MASTER_OUTPUT_FILE, index=False)
    return master

def safe_sheet_name(name):
    return re.sub(r"[\\/*?:\[\]]", " ", name)[:31]

def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=safe_sheet_name(cat))

        ws["A1"] = "Brand"
        ws["B1"] = "Holly Hunt"
        ws["A2"] = "Link"
        ws["B2"] = "\n".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_cat.insert(0, "Index", range(1, len(df_cat) + 1))

        start_row = 4
        for col_idx, col_name in enumerate(df_cat.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col_name).font = bold

        for row_idx, row in enumerate(df_cat.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        headers = {ws.cell(row=start_row, column=c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(start_row + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

        ws.freeze_panes = None

    wb.save(MASTER_SHEETS_FILE)

# ================= MAIN ENTRY POINT (FROM CODE-B) =================
def main():
    df = build_step1_master_excel()
    if df.empty:
        wb = Workbook()
        wb.save(MASTER_SHEETS_FILE)
        return

    build_category_wise_workbook_from_df(df)

    driver.quit()

if __name__ == "__main__":
    main()
