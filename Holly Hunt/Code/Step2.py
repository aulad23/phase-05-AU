# ==============================
# Holly Hunt — Step-2 (FINAL FIXED v2)
# ✅ INPUT:  HollyHunt.xlsx (multi-sheet, Step-1 format)
# ✅ OUTPUT: HollyHunt_details.xlsx (same multi-sheet format)
# ✅ FIX 1: Improved SKU extraction (handles "Item no. LEN0-TB" + "VK 120 A" format)
# ✅ FIX 2: Duplicate URL handling - processes same URL in different categories
# ==============================

import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import os
import re

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

BASE_URL = "https://www.hollyhunt.com"

# =========================
# ✅ INPUT / OUTPUT (Step-1 style)
# =========================
BASE_DIR = os.getcwd()
INPUT_FILE = os.path.join(BASE_DIR, "HollyHunt.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "HollyHunt_details.xlsx")

HEADER_ROW = 4
START_ROW = 5

# ✅ FIX: Changed to track (sheet_name, url) pairs instead of just urls
processed_pairs_file = os.path.join(BASE_DIR, "processed_pairs.txt")

# =========================
# Selenium Setup (UNCHANGED)
# =========================
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=chrome_options
)

WAIT = WebDriverWait(driver, 18)


# ---------------- SKU HELPERS (IMPROVED v2) ---------------- #
def _normalize_ws(text: str) -> str:
    """Normalize whitespace"""
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _clean_sku(s: str):
    if not s:
        return None
    s = _normalize_ws(s)
    # trim trailing punctuation
    s = re.sub(r"[.,;:)\]]+$", "", s).strip()
    return s if len(s) > 1 else None


def _extract_sku_from_text(text: str):
    """
    Improved SKU extraction that handles multiple formats including spaces:
    - "Item no. LEN0-TB"
    - "Item no: LEN0-TB"
    - "Item no. VK 120 A"  ✅ (space separated SKU)
    - "SKU: VK 120 A"
    """
    if not text:
        return None

    # clean separators
    text = text.replace('·', ' ').replace('•', ' ')
    t = _normalize_ws(text)

    # ✅ Allow spaces and "/" in SKU
    # Start with alnum, allow alnum/space/hyphen/slash, end with alnum
    sku_capture = r"([A-Z0-9][A-Z0-9\/\-\s]*[A-Z0-9])"

    patterns = [
        rf"Item\s*no\.?\s*[:\-]?\s*{sku_capture}",
        rf"SKU\s*[:#]?\s*{sku_capture}",
    ]

    for pattern in patterns:
        m = re.search(pattern, t, flags=re.IGNORECASE)
        if m:
            sku = _clean_sku(m.group(1))
            if sku:
                return sku

    return None


def _extract_sku_from_html(html: str):
    """
    Extract SKU from raw HTML/JSON structures
    """
    if not html:
        return None

    # 1) JSON-ish sku key
    m = re.search(r'"sku"\s*:\s*"([^"]+)"', html, flags=re.IGNORECASE)
    if m:
        val = _clean_sku(m.group(1))
        if val:
            return val

    # 2) Sometimes "SKU" key
    m = re.search(r'"SKU"\s*:\s*"([^"]+)"', html, flags=re.IGNORECASE)
    if m:
        val = _clean_sku(m.group(1))
        if val:
            return val

    # 3) Item no appears inside html (clean middot first)
    html_clean = html.replace('·', ' ').replace('•', ' ')
    sku_capture = r"([A-Z0-9][A-Z0-9\/\-\s]*[A-Z0-9])"

    m = re.search(rf"Item\s*no\.?\s*[:\-]?\s*{sku_capture}", html_clean, flags=re.IGNORECASE)
    if m:
        sku = _clean_sku(m.group(1))
        if sku:
            return sku

    return None


# ---------------- PRODUCT DETAILS SCRAPER (UNCHANGED except SKU uses new helpers) ---------------- #
def scrape_product_details(product_url):
    driver.get(product_url)

    # ✅ wait until page is rendered enough (breadcrumb OR headline)
    try:
        WAIT.until(
            EC.any_of(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".breadcrumb")),
                EC.presence_of_element_located((By.CSS_SELECTOR, ".c-pdp-hero__headline h2")),
            )
        )
    except:
        pass

    time.sleep(1.5)

    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")
    product_details = {}

    # =========================
    # ✅ SKU (IMPROVED EXTRACTION v2)
    # =========================
    sku = None

    # A) Try breadcrumb first
    bc = soup.select_one(".breadcrumb")
    if bc:
        bc_text = bc.get_text(" ", strip=True)
        sku = _extract_sku_from_text(bc_text)

    # B) If not found, try individual breadcrumb spans
    if not sku:
        for sp in soup.select(".breadcrumb span"):
            candidate = _extract_sku_from_text(sp.get_text(" ", strip=True))
            if candidate:
                sku = candidate
                break

    # C) Try headline area
    if not sku:
        headline = soup.select_one(".c-pdp-hero__headline")
        if headline:
            sku = _extract_sku_from_text(headline.get_text(" ", strip=True))

    # D) Try full visible body text
    if not sku:
        try:
            body_text = soup.get_text(" ", strip=True)
            sku = _extract_sku_from_text(body_text)
        except:
            pass

    # E) LAST fallback: scan raw HTML / scripts
    if not sku:
        sku = _extract_sku_from_html(html)

    product_details["SKU"] = sku

    # Product Family Id
    name_tag = soup.select_one(".c-pdp-hero__headline h2")
    product_details["Product Family Id"] = (
        name_tag.get_text(strip=True) if name_tag else None
    )

    # Description
    description_tag = soup.select_one(".dd-content.p-description")
    product_details["Description"] = (
        description_tag.get_text(separator=" ", strip=True)
        if description_tag else None
    )

    # Details
    details_tag = soup.select(".c-pdp-hero__details .p-details li .value")
    details = [d.get_text(strip=True) for d in details_tag]
    product_details["Details"] = ", ".join(details) if details else None

    # Finish
    finishes_tag = soup.select(".c-product-download-items__item .e-chip__row-1")
    finishes = [f.get_text(strip=True) for f in finishes_tag]
    product_details["Finish"] = ", ".join(finishes) if finishes else None

    # Dimensions
    dimension_tag = soup.select_one(".c-pdp-hero__details .detailed-dimensions .value")
    if dimension_tag:
        dimension_text = dimension_tag.get_text(strip=True)
        product_details.update(extract_dimensions(dimension_text))
    else:
        product_details["Dimension"] = None
        product_details["Width"] = None
        product_details["Depth"] = None
        product_details["Height"] = None
        product_details["Length"] = None
        product_details["Diameter"] = None
        product_details["Weight"] = None

    return product_details


# ---------------- DIMENSION EXTRACTOR (UNCHANGED) ---------------- #
def extract_dimensions(dimension_text):
    dimension_map = {
        "Width": None,
        "Depth": None,
        "Height": None,
        "Length": None,
        "Diameter": None,
        "Weight": None,
        "Dimension": dimension_text
    }

    pattern = r'(\d+\.?\d*)\s*(diameter|dia|w|h|l|d\b|weight|lb|lbs)'
    matches = re.findall(pattern, dimension_text, flags=re.IGNORECASE)

    for value, unit in matches:
        value = float(value)
        unit = unit.lower()

        if unit == "w":
            dimension_map["Width"] = value
        elif unit == "d":
            dimension_map["Depth"] = value
        elif unit == "h":
            dimension_map["Height"] = value
        elif unit == "l":
            dimension_map["Length"] = value
        elif unit in ["dia", "diameter"]:
            dimension_map["Diameter"] = value
        elif unit in ["weight", "lb", "lbs"]:
            dimension_map["Weight"] = value

    return dimension_map


# =========================
# Excel I/O Helpers (UNCHANGED)
# =========================
def safe_sheet_name(name):
    return re.sub(r"[\\/*?:\[\]]", " ", str(name))[:31]


def read_sheet_to_df_and_meta(ws):
    brand = ws["B1"].value
    link = ws["B2"].value

    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=HEADER_ROW, column=c).value)

    while headers and headers[-1] is None:
        headers.pop()

    data = []
    for r in range(START_ROW, ws.max_row + 1):
        row_vals = []
        empty = True
        for c in range(1, len(headers) + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(v)
            if v not in (None, ""):
                empty = False
        if empty:
            continue
        data.append(row_vals)

    df = pd.DataFrame(data, columns=headers)

    required = ["Product URL", "Image URL", "Product Name", "Category"]
    for col in required:
        if col not in df.columns:
            df[col] = None

    if "Index" not in df.columns and "index" in df.columns:
        df.rename(columns={"index": "Index"}, inplace=True)
    if "Index" not in df.columns:
        df.insert(0, "Index", range(1, len(df) + 1))

    return df, brand, link


def write_df_to_sheet(wb, sheet_name, df, brand, link):
    ws = wb.create_sheet(title=safe_sheet_name(sheet_name))

    bold = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    ws["A1"] = "Brand"
    ws["B1"] = brand if brand is not None else ""
    ws["A2"] = "Link"
    ws["B2"] = link if link is not None else ""
    ws["B2"].alignment = Alignment(wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
        cell.font = bold

    for row_idx, row in enumerate(df.itertuples(index=False), start=START_ROW):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    headers_map = {ws.cell(row=HEADER_ROW, column=c).value: c for c in range(1, ws.max_column + 1)}
    url_col = headers_map.get("Product URL")
    name_col = headers_map.get("Product Name")
    if url_col and name_col:
        for r in range(START_ROW, ws.max_row + 1):
            url = ws.cell(row=r, column=url_col).value
            name_cell = ws.cell(row=r, column=name_col)
            if url:
                name_cell.hyperlink = url
                name_cell.font = link_font

    ws.freeze_panes = None


# =========================
# MAIN (IMPROVED DUPLICATE HANDLING)
# =========================
def main():
    processed_pairs = set()
    if os.path.exists(processed_pairs_file):
        with open(processed_pairs_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and "|" in line:
                    processed_pairs.add(line)

    in_wb = load_workbook(INPUT_FILE)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for sheet_name in in_wb.sheetnames:
        ws = in_wb[sheet_name]
        df, brand, link = read_sheet_to_df_and_meta(ws)

        product_urls = df["Product URL"].dropna().astype(str).tolist()
        all_products_details = []
        batch_size = 5

        for i in range(0, len(product_urls), batch_size):
            current_batch = product_urls[i:i + batch_size]
            print(f"[{sheet_name}] Processing batch {i // batch_size + 1}...")

            batch_details = []

            for product_url in current_batch:
                pair_key = f"{sheet_name}|{product_url}"

                if pair_key in processed_pairs:
                    print(f"[{sheet_name}] Skipping {product_url}, already processed in this sheet.")
                    row_data = df[df["Product URL"].astype(str) == str(product_url)].iloc[0].to_dict()
                    batch_details.append(row_data)
                    continue

                print(f"[{sheet_name}] Scraping details for {product_url}...")
                product_details = scrape_product_details(product_url)

                product_data = df[df["Product URL"].astype(str) == str(product_url)].iloc[0].to_dict()
                product_data.update(product_details)
                batch_details.append(product_data)

                processed_pairs.add(pair_key)

            all_products_details.extend(batch_details)

            with open(processed_pairs_file, "w", encoding="utf-8") as f:
                f.write("\n".join(sorted(processed_pairs)))

        final_df = pd.DataFrame(all_products_details) if all_products_details else df.copy()

        preferred_prefix = ["Index", "Category", "Product URL", "Image URL", "Product Name"]
        preferred_suffix = [
            "SKU", "Product Family Id", "Description", "Weight", "Width", "Depth",
            "Diameter", "Length", "Height", "Finish", "Dimension", "Details"
        ]

        existing_cols = list(final_df.columns)
        col_order = []

        for c in preferred_prefix:
            if c in existing_cols and c not in col_order:
                col_order.append(c)

        for c in existing_cols:
            if c not in col_order and c not in preferred_suffix:
                col_order.append(c)

        for c in preferred_suffix:
            if c in existing_cols and c not in col_order:
                col_order.append(c)

        final_df = final_df[col_order]

        write_df_to_sheet(out_wb, sheet_name, final_df, brand, link)

    out_wb.save(OUTPUT_FILE)
    driver.quit()
    print(f"\n✅ Step-2 done! Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
