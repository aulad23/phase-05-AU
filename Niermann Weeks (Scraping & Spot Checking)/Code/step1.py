import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# INPUT SYSTEM (COPIED FROM CODE-B)
# =========================================================

CATEGORIES = {
    "Coffee & Cocktail Tables" : [
        "https://niermannweeks.com/product-type/coffee/"
    ],
  "Side & End Tables": [
        "https://niermannweeks.com/product-type/occasional/",
    ],
  "Dining Tables": [
        "https://niermannweeks.com/product-type/dining/",
    ],

    "Consoles": [
        "https://niermannweeks.com/product-type/consoles/",
    ],

    "Beds & Headboards": [
        "https://niermannweeks.com/product-type/beds/",
    ],
    "Chandeliers": [
        "https://niermannweeks.com/product-type/chandeliers/",
    ],

    "Sconces": [
        "https://niermannweeks.com/product-type/sconces/",
    ],

    "Flush Mount": [
        "https://niermannweeks.com/product-type/ceiling-fixtures/",
    ],
    "Mirrors": [
        "https://niermannweeks.com/product-type/mirrors/",
    ],
}

script_dir = os.path.dirname(os.path.abspath(__file__))
master_output_file = os.path.join(script_dir, "niermannweeks_all_products.xlsx")
master_sheets_file = os.path.join(script_dir, "NiermannWeeks.xlsx")

# =========================================================
# SCRAPING LOGIC (UNCHANGED FROM CODE-A)
# =========================================================

def scrape_category(category, urls):
    all_products = []

    for base_url in urls:
        page = 1
        while True:
            url = base_url if page == 1 else f"{base_url}page/{page}/"
            response = requests.get(url)
            if response.status_code != 200:
                break

            soup = BeautifulSoup(response.text, "html.parser")
            products = soup.find_all("li", class_="nw-product")
            if not products:
                break

            for product in products:
                link_tag = product.find("a")
                product_url = link_tag["href"] if link_tag else None

                img_tag = product.find("img")
                image_url = img_tag["src"] if img_tag else None

                name_tag = product.find("h3")
                product_name = name_tag.get_text(strip=True) if name_tag else None

                if product_url and image_url and product_name:
                    all_products.append({
                        "Category": category,
                        "Product URL": product_url,
                        "Image URL": image_url,
                        "Product Name": product_name
                    })

            page += 1

    return pd.DataFrame(all_products)

# =========================================================
# OUTPUT SYSTEM (COPIED FROM CODE-B)
# =========================================================

def build_step1_master_excel():
    dfs = []
    for cat, links in CATEGORIES.items():
        df = scrape_category(cat, links)
        if not df.empty:
            dfs.append(df)

    if not dfs:
        empty = pd.DataFrame(columns=["Category", "Product URL", "Image URL", "Product Name"])
        empty.to_excel(master_output_file, index=False)
        return empty

    master = pd.concat(dfs, ignore_index=True)
    master.drop_duplicates(subset=["Product URL"], inplace=True)
    master.to_excel(master_output_file, index=False)
    return master


def build_category_wise_workbook_from_df(df):
    wb = Workbook()
    wb.remove(wb.active)

    link_font = Font(color="0563C1", underline="single")
    bold = Font(bold=True)

    for cat in CATEGORIES.keys():
        df_cat = df[df["Category"] == cat].copy()
        if df_cat.empty:
            continue

        ws = wb.create_sheet(title=cat)

        ws["A1"] = "Brand"
        ws["B1"] = "Niermann Weeks"
        ws["A2"] = "Link"
        ws["B2"] = ", ".join(CATEGORIES.get(cat, []))
        ws["B2"].alignment = Alignment(wrap_text=True)

        df_out = df_cat.copy()
        df_out.insert(0, "Index", range(1, len(df_out) + 1))

        start = 4
        for j, col in enumerate(df_out.columns, start=1):
            ws.cell(row=start, column=j, value=col).font = bold

        for i, row in enumerate(df_out.itertuples(index=False), start=start + 1):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        headers = {ws.cell(row=start, column=j).value: j for j in range(1, ws.max_column + 1)}
        for r in range(start + 1, ws.max_row + 1):
            url = ws.cell(row=r, column=headers["Product URL"]).value
            cell = ws.cell(row=r, column=headers["Product Name"])
            if url:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(master_sheets_file)

# =========================================================
# MAIN ENTRY POINT (FROM CODE-B)
# =========================================================

def main():
    df = build_step1_master_excel()
    if df.empty:
        Workbook().save(master_sheets_file)
        return
    build_category_wise_workbook_from_df(df)

if __name__ == "__main__":
    main()
