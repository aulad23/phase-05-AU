import requests
from bs4 import BeautifulSoup
import os
import re
import time
from openpyxl import load_workbook
from openpyxl.styles import Font

# =========================================================
# PATHS
# =========================================================
script_dir = os.path.dirname(os.path.abspath(__file__))
input_path = os.path.join(script_dir, "NiermannWeeks.xlsx")
output_path = os.path.join(script_dir, "NiermannWeeks_Detailed.xlsx")

# =========================================================
# LOAD WORKBOOK
# =========================================================
wb = load_workbook(input_path)

# FIXED STRUCTURE
HEADER_ROW = 4
START_ROW = 5
PRODUCT_URL_COL = 3     # Column C
IMAGE_INPUT_COL = 6     # Column F

# =========================================================
# OUTPUT HEADERS
# =========================================================
OUTPUT_HEADERS = [
    "SKU",
    "Product Family Id",
    "Description",
    "List Price",
    "Weight",
    "Width",
    "Depth",
    "Diameter",
    "Height",
    "COM",
]

bold = Font(bold=True)

# =========================================================
# REQUEST HEADERS
# =========================================================
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# =========================================================
# PROCESS EACH SHEET ONE BY ONE (Coffee table -> Mirrors -> ...)
# =========================================================
total_count = 0

for ws in wb.worksheets:
    print(f"\n==============================")
    print(f"📄 Working Sheet: {ws.title}")
    print(f"==============================")

    # Output columns start for THIS sheet
    output_start_col = ws.max_column + 1

    # Write output headers (only if empty)
    for i, h in enumerate(OUTPUT_HEADERS):
        cell = ws.cell(row=HEADER_ROW, column=output_start_col + i)
        if not cell.value:
            cell.value = h
            cell.font = bold

    row = START_ROW
    count = 0

    # =========================================================
    # SCRAPING LOOP (LOGIC UNCHANGED)
    # =========================================================
    while True:
        url = ws.cell(row=row, column=PRODUCT_URL_COL).value
        if not url:
            break

        print(f"[{count + 1}] Scraping: {url}")

        product_name = ws.cell(row=row, column=5).value or ""
        product_family_id = product_name
        image_url = ws.cell(row=row, column=IMAGE_INPUT_COL).value or ""

        description = ""
        sku = ""
        width = depth = diameter = height = ""
        weight = ""
        list_price = ""
        com = ""

        try:
            r = requests.get(url, timeout=15, headers=HEADERS)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")

            h1 = soup.find("h1")
            if h1:
                product_family_id = h1.get_text(strip=True)

            blocks = soup.find_all("div", class_="col grid_6_of_12")
            for block in blocks:
                header = block.find("h4", class_="product-info")
                if not (header and "Product Info" in header.get_text()):
                    continue

                for p in block.find_all("p"):
                    txt = p.get_text(" ", strip=True)
                    if not txt:
                        continue

                    if not p.find("strong") and not description:
                        description = txt

                    m = re.search(r"\bCOM\b\s+(\d*\.?\d+)\s*YARD", txt, re.I)
                    if m and not com:
                        com = m.group(1)

                    if p.find("strong"):
                        if not sku:
                            m = re.search(r"ITEM\s*#\s*[:\-]?\s*([A-Z0-9\-]+)", txt, re.I)
                            if m:
                                sku = m.group(1)

                        t = txt.upper().replace("\xa0", " ")
                        if not width:
                            m = re.search(r"(\d*\.?\d+)\s*W\b", t)
                            if m:
                                width = m.group(1)
                        if not depth:
                            m = re.search(r"(\d*\.?\d+)\s*D(?!IA)\b", t)
                            if m:
                                depth = m.group(1)
                        if not diameter:
                            m = re.search(r"(\d*\.?\d+)\s*DIA\b", t)
                            if m:
                                diameter = m.group(1)
                        if not height:
                            m = re.search(r"(\d*\.?\d+)\s*H\b", t)
                            if m:
                                height = m.group(1)
                        if not list_price:
                            m = re.search(r"\$[\d,]+", txt)
                            if m:
                                list_price = m.group(0)
                break

        except Exception as e:
            print(f"ERROR: {e}")

        values = [
            sku,
            product_family_id,
            description,
            list_price,
            weight,
            width,
            depth,
            diameter,
            height,
            com,
        ]

        for i, v in enumerate(values):
            ws.cell(row=row, column=output_start_col + i, value=v)

        row += 1
        count += 1
        total_count += 1
        time.sleep(1)

    print(f"✅ Sheet Done: {ws.title} | Products: {count}")

# =========================================================
# SAVE
# =========================================================
wb.save(output_path)
print(f"\n✅ Completed Total {total_count} products")
print(f"📁 Saved to: {output_path}")
