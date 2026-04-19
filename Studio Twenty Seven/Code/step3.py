import time
import re
import os
import requests
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# ============================================================
#  শুধু এই দুই লাইন পরিবর্তন করুন
# ============================================================
INPUT_FILE = "chandeliers.xlsx"
OUTPUT_FILE = "chandeliers-final.xlsx"
TEARSHEET_DIR = "tearsheets"  # PDF গুলো এই ফোল্ডারে সেভ হবে
# ============================================================

os.makedirs(TEARSHEET_DIR, exist_ok=True)

# ── PDF text extraction ──────────────────────────────────────
try:
    import pdfplumber

    PDF_BACKEND = "pdfplumber"
except ImportError:
    try:
        import pypdf

        PDF_BACKEND = "pypdf"
    except ImportError:
        PDF_BACKEND = None
        print("⚠️  pdfplumber বা pypdf ইনস্টল নেই — PDF parsing বন্ধ থাকবে।")
        print("   চালু করতে:  pip install pdfplumber")


def extract_text_from_pdf(pdf_path: str) -> str:
    """PDF থেকে পুরো text বের করে।"""
    text = ""
    if PDF_BACKEND == "pdfplumber":
        import pdfplumber
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"
        except Exception as e:
            print(f"    ⚠️ pdfplumber error: {e}")
    elif PDF_BACKEND == "pypdf":
        import pypdf
        try:
            reader = pypdf.PdfReader(pdf_path)
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
        except Exception as e:
            print(f"    ⚠️ pypdf error: {e}")
    return text.strip()


def parse_pdf_data(pdf_text: str) -> dict:
    """
    PDF text থেকে শুধু সেই fields বের করে
    যেগুলো web scraping থেকে পাওয়া যায় না।

    ❌ বাদ দেওয়া হয়েছে (web থেকে আগেই পাচ্ছি):
        Weight, Dimensions, Socket, Wattage

    ✅ শুধু নতুন unique data নেওয়া হচ্ছে:
        Material, Finish, Shade, Bulb Type,
        Country, UL/Certification, Collection, Designer
    """
    data = {
        "PDF_Material": "",  # শরীর কী দিয়ে তৈরি
        "PDF_Finish": "",  # রঙ / finish
        "PDF_Shade": "",  # shade / diffuser / glass
        "PDF_Bulb_Type": "",  # LED, Incandescent ইত্যাদি
        "PDF_Country": "",  # Country of origin
        "PDF_UL_Listed": "",  # UL / ETL certification
        "PDF_Collection": "",  # Collection নাম
        "PDF_Designer": "",  # Designer নাম
        "PDF_Raw_Text": "",  # প্রথম ১৫০০ char (debugging)
    }

    if not pdf_text:
        return data

    data["PDF_Raw_Text"] = pdf_text[:1500].replace("\n", " | ")

    # key: value লাইন parse করো
    kv_map = {}
    for line in pdf_text.splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            k = k.strip().upper()
            v = v.strip()
            if k and v:
                kv_map[k] = v

    def find(keys):
        for k in keys:
            for mk, mv in kv_map.items():
                if k.upper() in mk:
                    return mv
        return ""

    data["PDF_Material"] = find(["MATERIAL", "BODY MATERIAL", "FRAME MATERIAL", "STRUCTURE"])
    data["PDF_Finish"] = find(["FINISH", "COLOR", "COLOUR", "POWDER COAT"])
    data["PDF_Shade"] = find(["SHADE", "DIFFUSER", "GLASS", "LENS", "GLOBE"])
    data["PDF_Bulb_Type"] = find(["BULB TYPE", "LAMP TYPE", "LIGHT SOURCE", "LAMP INCLUDED"])
    data["PDF_Country"] = find(["COUNTRY", "ORIGIN", "MADE IN", "COUNTRY OF ORIGIN"])
    data["PDF_UL_Listed"] = find(["UL", "ETL", "LISTED", "CERTIFICATION", "RATED FOR"])
    data["PDF_Collection"] = find(["COLLECTION", "SERIES", "LINE"])
    data["PDF_Designer"] = find(["DESIGNER", "DESIGNED BY", "ARCHITECT"])

    return data


def download_tearsheet(url: str, sku: str) -> tuple[str, str]:
    """
    Tearsheet PDF ডাউনলোড করে।
    Returns: (local_pdf_path, status_message)
    """
    if not url:
        return "", "No URL"

    # ফাইলের নাম তৈরি করো SKU দিয়ে
    safe_sku = re.sub(r'[\\/*?:"<>|]', "_", str(sku)) if sku else "unknown"
    filename = f"{safe_sku}_tearsheet.pdf"
    filepath = os.path.join(TEARSHEET_DIR, filename)

    # আগে ডাউনলোড হয়ে থাকলে আবার করার দরকার নেই
    if os.path.exists(filepath) and os.path.getsize(filepath) > 1000:
        print(f"    ♻️  Already downloaded: {filename}")
        return filepath, "Cached"

    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        }
        resp = requests.get(url, headers=headers, timeout=30, stream=True)
        resp.raise_for_status()

        # Content-Type চেক — PDF কিনা নিশ্চিত করা
        ct = resp.headers.get("Content-Type", "")
        if "pdf" not in ct.lower() and not url.lower().endswith(".pdf"):
            print(f"    ⚠️  Not a PDF (Content-Type: {ct})")
            return "", f"Not PDF ({ct})"

        with open(filepath, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)

        size_kb = os.path.getsize(filepath) / 1024
        print(f"    ✅ Downloaded: {filename} ({size_kb:.1f} KB)")
        return filepath, f"OK ({size_kb:.1f} KB)"

    except requests.exceptions.Timeout:
        print(f"    ⚠️ Timeout downloading tearsheet")
        return "", "Timeout"
    except Exception as e:
        print(f"    ⚠️ Download error: {e}")
        return "", f"Error: {e}"


# ────────────────────────────────────────────────────────────
#  নিচের সব ফাংশন আগের মতোই (কোনো পরিবর্তন নেই)
# ────────────────────────────────────────────────────────────

def init_driver():
    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    return webdriver.Chrome(options=opts)


def parse_dimensions(dim_str):
    h = w = d = dia = length = sh = sd = ""
    if not dim_str:
        return h, w, d, dia, length, sh, sd
    h_m = re.search(r"(?<!\bS)H\s*([\d.]+)", dim_str)
    w_m = re.search(r"(?<!\bS)W\s*([\d.]+)", dim_str)
    d_m = re.search(r"(?<!\bS)D\s*([\d.]+)", dim_str)
    dia_m = re.search(r"Ø\s*([\d.]+)", dim_str)
    l_m = re.search(r"L\s*([\d.]+)", dim_str)
    sh_m = re.search(r"SH\s*([\d.]+)", dim_str)
    sd_m = re.search(r"SD\s*([\d.]+)", dim_str)
    if h_m:   h = h_m.group(1)
    if w_m:   w = w_m.group(1)
    if d_m:   d = d_m.group(1)
    if dia_m: dia = dia_m.group(1)
    if l_m:   length = l_m.group(1)
    if sh_m:  sh = sh_m.group(1)
    if sd_m:  sd = sd_m.group(1)
    return h, w, d, dia, length, sh, sd


SKIP_DETAILS = ["STUDIOTWENTYSEVEN", "PAYMENT", "SHIPPING", "VISA", "MASTERCARD",
                "AMERICAN EXPRESS", "DISCOVER", "APPLE PAY", "White Glove",
                "Standard Delivery", "RETURNS"]


def parse_socket_wattage(details_str):
    socket = ""
    wattage = ""
    if not details_str:
        return socket, wattage
    parts = [p.strip() for p in details_str.split("|")]
    socket_parts = []
    for p in parts:
        if re.search(r"E\d+(?:/\d+)?\s*BASE", p, re.IGNORECASE):
            socket_parts.append(p.strip())
        elif re.search(r"E\d+(?:/\d+)?\s*SOCKET", p, re.IGNORECASE):
            m = re.search(r"(E\d+(?:/\d+)?)", p, re.IGNORECASE)
            if m:
                socket_parts.append(m.group(1).strip())
        elif re.search(r"G[\d.]+\s*BULBS?", p, re.IGNORECASE):
            socket_parts.append(p.strip())
    socket = ", ".join(socket_parts)
    for p in parts:
        wm = re.search(r"(\d+\s*X\s*\d+\s*W?)\b", p, re.IGNORECASE)
        if wm:
            wattage = wm.group(1).strip()
            break
    if not wattage:
        for p in parts:
            wm = re.search(r"\b(\d+)\s*W\b", p, re.IGNORECASE)
            if wm:
                wattage = wm.group(0).strip()
                break
    return socket, wattage


def scrape_product_page(driver, url):
    driver.get(url)
    time.sleep(2)

    # ── Description ──
    desc_lines = []
    try:
        block_text = driver.find_element(By.CSS_SELECTOR, "div.block-text")
        raw = block_text.text
        for line in raw.splitlines():
            line = line.strip()
            if not line:
                continue
            skip = ["USD", "Estimated Shipping", "Estimated Production",
                    "Price", "Payment", "Shipping", "Assistance", "Download"]
            if any(k.lower() in line.lower() for k in skip):
                continue
            desc_lines.append(line)
    except Exception as e:
        print(f"  ⚠️ Description error: {e}")
    description = " | ".join(desc_lines)

    # ── Weight ──
    weight = ""
    wm = re.search(r"(\d+[\.,]?\d*)\s*(lbs?|kg)", description, re.IGNORECASE)
    if wm:
        weight = wm.group(0)

    # ── Dimension ──
    dimension = ""
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(driver.page_source, "html.parser")
        block = soup.select_one("div.block-text")
        if block:
            table = block.find("table")
            if table:
                for row in table.find_all("tr"):
                    cells = [td.get_text(strip=True) for td in row.find_all("td")]
                    if cells and cells[0].upper() == "IN":
                        parts = [c for c in cells if c.strip()]
                        dimension = " ".join(parts[1:])
                        break
            if not dimension:
                raw_text = block.get_text(" ")
                m = re.search(
                    r"IN\s*\|?\s*(H\s*[\d.]+[\s\|]*(?:W\s*[\d.]+)?[\s\|]*(?:D\s*[\d.]+)?[\s\|]*(?:Ø\s*[\d.]+)?[\s\|]*(?:L\s*[\d.]+)?)",
                    raw_text, re.IGNORECASE
                )
                if m:
                    dimension = m.group(1).strip()
    except Exception as e:
        print(f"  ⚠️ Dimension error: {e}")

    height, width, depth, diameter, length, seat_height, seat_depth = parse_dimensions(dimension)

    # ── Details ──
    details = ""
    details_links = driver.find_elements(By.CSS_SELECTOR, "a[data-drawer='pdp-description-drawer']")
    if details_links:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", details_links[0])
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", details_links[0])
            time.sleep(2)
            drawers = driver.find_elements(By.CSS_SELECTOR, "div.drawer-body")
            for dw in drawers:
                txt = dw.text.strip()
                if txt:
                    lines = [l.strip() for l in txt.splitlines() if l.strip()]
                    lines = [l for l in lines if not any(k.lower() in l.lower() for k in SKIP_DETAILS)]
                    if lines:
                        details = " | ".join(lines)
                        print(f"  ✅ Details found ({len(lines)} lines)")
                        break
            if not details:
                try:
                    drawer_el = driver.find_element(By.ID, "pdp-description-drawer")
                    txt = drawer_el.text.strip()
                    if txt:
                        lines = [l.strip() for l in txt.splitlines() if l.strip()]
                        lines = [l for l in lines if not any(k.lower() in l.lower() for k in SKIP_DETAILS)]
                        if lines:
                            details = " | ".join(lines)
                            print(f"  ✅ Details found via fallback ({len(lines)} lines)")
                except:
                    pass
            if not details:
                print("  ℹ️ Details drawer opened but was empty")
        except Exception as e:
            print(f"  ⚠️ Details read error: {e}")
    else:
        print("  ℹ️ No Details link on this page — skipping")

    # ── Tearsheet URL ──
    tearsheet_url = ""
    dl_links = driver.find_elements(By.XPATH, "//a[contains(text(),'Download Product Info')]")
    if dl_links:
        tearsheet_url = dl_links[0].get_attribute("href")
        print(f"  ✅ Tearsheet URL found: {tearsheet_url[:60]}...")
    else:
        print("  ℹ️ No Tearsheet link on this page — skipping")

    # ── Socket & Wattage ──
    socket, wattage = parse_socket_wattage(details)
    if socket:
        print(f"  ✅ Socket: {socket}")
    if wattage:
        print(f"  ✅ Wattage: {wattage}")

    return (description, weight, dimension, height, width, depth, diameter,
            length, seat_height, seat_depth, details, tearsheet_url, socket, wattage)


# ── Read Input Excel ──
wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in.active

headers_in = [ws_in.cell(1, c).value for c in range(1, ws_in.max_column + 1)]
url_col = headers_in.index("Product URL") + 1
image_col = headers_in.index("Image URL") + 1
name_col = headers_in.index("Product Name") + 1
sku_col = headers_in.index("SKU") + 1
price_col = headers_in.index("List Price (USD)") + 1

# ── Output Excel — PDF columns সহ ──
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.append([
    # ── মূল columns ──
    "Product URL", "Image URL", "Product Name", "SKU",
    "Product Family Id", "Description", "Weight",
    "Dimension", "Width", "Depth", "Diameter", "Length", "Height",
    "Seat Height", "Seat Depth",
    "List Price", "Details", "Socket", "Wattage",
    # ── Tearsheet ──
    "Tearsheet Link",
    # ── PDF থেকে শুধু নতুন unique data ──
    "Material", "Finish", "Shade",
    "Bulb Type", "Country",
    "UL Listed", "Collection", "Designer",
    "Raw Text",
])

driver = init_driver()
total = ws_in.max_row - 1

for row_idx in range(2, ws_in.max_row + 1):
    product_url = ws_in.cell(row_idx, url_col).value
    image_url = ws_in.cell(row_idx, image_col).value
    product_name = ws_in.cell(row_idx, name_col).value
    sku = ws_in.cell(row_idx, sku_col).value
    price = ws_in.cell(row_idx, price_col).value

    print(f"\n[{row_idx - 1}/{total}] {product_name}")
    print(f"  URL: {product_url}")

    try:
        (desc, weight, dimension, height, width, depth, diameter,
         length, seat_height, seat_depth, details,
         tearsheet_url, socket, wattage) = scrape_product_page(driver, product_url)
    except Exception as e:
        print(f"  ⚠️ Fatal scrape error: {e}")
        (desc, weight, dimension, height, width, depth, diameter,
         length, seat_height, seat_depth, details,
         tearsheet_url, socket, wattage) = ("",) * 14

    # ── Tearsheet ডাউনলোড ও PDF parse ──
    pdf_data = {k: "" for k in [
        "PDF_Material", "PDF_Finish", "PDF_Shade",
        "PDF_Bulb_Type", "PDF_Country",
        "PDF_UL_Listed", "PDF_Collection", "PDF_Designer",
        "PDF_Raw_Text",
    ]}
    dl_status = ""

    if tearsheet_url:
        pdf_path, dl_status = download_tearsheet(tearsheet_url, sku)
        if pdf_path and PDF_BACKEND:
            print(f"    📄 Parsing PDF...")
            pdf_text = extract_text_from_pdf(pdf_path)
            if pdf_text:
                pdf_data = parse_pdf_data(pdf_text)
                print(f"    ✅ PDF parsed — {len(pdf_text)} chars extracted")
            else:
                dl_status += " | PDF empty/unreadable"
                print(f"    ⚠️ PDF text extraction returned empty")
    else:
        dl_status = "No tearsheet URL"

    ws_out.append([
        product_url, image_url, product_name, sku,
        product_name,
        desc, weight,
        dimension, width, depth, diameter, length, height,
        seat_height, seat_depth,
        price, details, socket, wattage,
        # Tearsheet
        tearsheet_url,
        # PDF — শুধু নতুন unique data
        pdf_data["PDF_Material"],
        pdf_data["PDF_Finish"],
        pdf_data["PDF_Shade"],
        pdf_data["PDF_Bulb_Type"],
        pdf_data["PDF_Country"],
        pdf_data["PDF_UL_Listed"],
        pdf_data["PDF_Collection"],
        pdf_data["PDF_Designer"],
        pdf_data["PDF_Raw_Text"],
    ])
    time.sleep(0.5)

driver.quit()
wb_out.save(OUTPUT_FILE)
print(f"\n✅ Done! → {OUTPUT_FILE}")
print(f"📁 Tearsheet PDFs saved in: ./{TEARSHEET_DIR}/")
