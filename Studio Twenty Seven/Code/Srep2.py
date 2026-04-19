import time
import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# ============================================================
#  শুধু এই দুই লাইন পরিবর্তন করুন
# ============================================================
INPUT_FILE  = "Objects.xlsx"
OUTPUT_FILE = "Objects-final.xlsx"
# ============================================================

def init_driver():
    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    return webdriver.Chrome(options=opts)

def parse_dimensions(dim_str):
    h = w = d = dia = length = sh = sd = ""
    if not dim_str:
        return h, w, d, dia, length, sh, sd
    h_m   = re.search(r"(?<!\bS)H\s*([\d.]+)", dim_str)
    w_m   = re.search(r"(?<!\bS)W\s*([\d.]+)", dim_str)
    d_m   = re.search(r"(?<!\bS)D\s*([\d.]+)", dim_str)
    dia_m = re.search(r"Ø\s*([\d.]+)", dim_str)
    l_m   = re.search(r"L\s*([\d.]+)", dim_str)
    sh_m  = re.search(r"SH\s*([\d.]+)", dim_str)
    sd_m  = re.search(r"SD\s*([\d.]+)", dim_str)
    if h_m:   h      = h_m.group(1)
    if w_m:   w      = w_m.group(1)
    if d_m:   d      = d_m.group(1)
    if dia_m: dia    = dia_m.group(1)
    if l_m:   length = l_m.group(1)
    if sh_m:  sh     = sh_m.group(1)
    if sd_m:  sd     = sd_m.group(1)
    return h, w, d, dia, length, sh, sd

SKIP_DETAILS = ["STUDIOTWENTYSEVEN", "PAYMENT", "SHIPPING", "VISA", "MASTERCARD",
                "AMERICAN EXPRESS", "DISCOVER", "APPLE PAY", "White Glove",
                "Standard Delivery", "RETURNS"]

def parse_socket_wattage(details_str):
    """Details string থেকে Socket আর Wattage বের করে।"""
    socket = ""
    wattage = ""
    if not details_str:
        return socket, wattage

    parts = [p.strip() for p in details_str.split("|")]

    socket_parts = []
    for p in parts:
        # E26 BASE, E26/27 BASE, E14 BASE ইত্যাদি
        if re.search(r"E\d+(?:/\d+)?\s*BASE", p, re.IGNORECASE):
            socket_parts.append(p.strip())
        # E27 SOCKET, E14 SOCKET ইত্যাদি
        elif re.search(r"E\d+(?:/\d+)?\s*SOCKET", p, re.IGNORECASE):
            # শুধু E27 part নিব, "SOCKET" বাদ
            m = re.search(r"(E\d+(?:/\d+)?)", p, re.IGNORECASE)
            if m:
                socket_parts.append(m.group(1).strip())
        # G16.5 BULBS, G9 BULBS ইত্যাদি
        elif re.search(r"G[\d.]+\s*BULBS?", p, re.IGNORECASE):
            socket_parts.append(p.strip())
    socket = ", ".join(socket_parts)

    # Wattage — দুই ধরনের pattern:
    # 1) Multi-bulb: "8X 8", "8 X 8W", "12 X 5W"
    # 2) Single: "100W", "60W"
    for p in parts:
        wm = re.search(r"(\d+\s*X\s*\d+\s*W?)\b", p, re.IGNORECASE)
        if wm:
            wattage = wm.group(1).strip()
            break
    if not wattage:
        for p in parts:
            wm = re.search(r"\b(\d+)\s*W\b", p, re.IGNORECASE)
            if wm:
                wattage = wm.group(0).strip()
                break

    return socket, wattage


def scrape_product_page(driver, url):
    driver.get(url)
    time.sleep(2)

    # ── Description ──
    desc_lines = []
    try:
        block_text = driver.find_element(By.CSS_SELECTOR, "div.block-text")
        raw = block_text.text
        for line in raw.splitlines():
            line = line.strip()
            if not line:
                continue
            skip = ["USD", "Estimated Shipping", "Estimated Production",
                    "Price", "Payment", "Shipping", "Assistance", "Download"]
            if any(k.lower() in line.lower() for k in skip):
                continue
            desc_lines.append(line)
    except Exception as e:
        print(f"  ⚠️ Description error: {e}")
    description = " | ".join(desc_lines)

    # ── Weight ──
    weight = ""
    wm = re.search(r"(\d+[\.,]?\d*)\s*(lbs?|kg)", description, re.IGNORECASE)
    if wm:
        weight = wm.group(0)

    # ── Dimension — from page source table IN row ──
    dimension = ""
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(driver.page_source, "html.parser")
        block = soup.select_one("div.block-text")
        if block:
            table = block.find("table")
            if table:
                for row in table.find_all("tr"):
                    cells = [td.get_text(strip=True) for td in row.find_all("td")]
                    if cells and cells[0].upper() == "IN":
                        parts = [c for c in cells if c.strip()]
                        dimension = " ".join(parts[1:])
                        break
            if not dimension:
                raw_text = block.get_text(" ")
                m = re.search(
                    r"IN\s*\|?\s*(H\s*[\d.]+[\s\|]*(?:W\s*[\d.]+)?[\s\|]*(?:D\s*[\d.]+)?[\s\|]*(?:Ø\s*[\d.]+)?[\s\|]*(?:L\s*[\d.]+)?)",
                    raw_text, re.IGNORECASE
                )
                if m:
                    dimension = m.group(1).strip()
    except Exception as e:
        print(f"  ⚠️ Dimension error: {e}")

    height, width, depth, diameter, length, seat_height, seat_depth = parse_dimensions(dimension)

    # ── Details — click the Details drawer link (if exists) ──
    details = ""
    details_links = driver.find_elements(By.CSS_SELECTOR, "a[data-drawer='pdp-description-drawer']")
    if details_links:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", details_links[0])
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", details_links[0])
            time.sleep(2)

            # drawer-body তে content থাকে — একাধিক drawer থাকতে পারে
            drawers = driver.find_elements(By.CSS_SELECTOR, "div.drawer-body")
            for dw in drawers:
                txt = dw.text.strip()
                if txt:
                    lines = [l.strip() for l in txt.splitlines() if l.strip()]
                    lines = [l for l in lines if not any(k.lower() in l.lower() for k in SKIP_DETAILS)]
                    if lines:
                        details = " | ".join(lines)
                        print(f"  ✅ Details found ({len(lines)} lines)")
                        break

            if not details:
                # Fallback: try #pdp-description-drawer directly
                try:
                    drawer_el = driver.find_element(By.ID, "pdp-description-drawer")
                    txt = drawer_el.text.strip()
                    if txt:
                        lines = [l.strip() for l in txt.splitlines() if l.strip()]
                        lines = [l for l in lines if not any(k.lower() in l.lower() for k in SKIP_DETAILS)]
                        if lines:
                            details = " | ".join(lines)
                            print(f"  ✅ Details found via fallback ({len(lines)} lines)")
                except:
                    pass

            if not details:
                print("  ℹ️ Details drawer opened but was empty")

        except Exception as e:
            print(f"  ⚠️ Details read error: {e}")
    else:
        print("  ℹ️ No Details link on this page — skipping")

    # ── Tearsheet ──
    tearsheet = ""
    dl_links = driver.find_elements(By.XPATH, "//a[contains(text(),'Download Product Info')]")
    if dl_links:
        tearsheet = dl_links[0].get_attribute("href")
    else:
        print("  ℹ️ No Tearsheet link on this page — skipping")

    # ── Socket & Wattage from Details ──
    socket, wattage = parse_socket_wattage(details)
    if socket:
        print(f"  ✅ Socket: {socket}")
    if wattage:
        print(f"  ✅ Wattage: {wattage}")

    return description, weight, dimension, height, width, depth, diameter, length, seat_height, seat_depth, details, tearsheet, socket, wattage


# ── Read Step 1 Excel ──
wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in.active

headers_in = [ws_in.cell(1, c).value for c in range(1, ws_in.max_column + 1)]
url_col   = headers_in.index("Product URL") + 1
image_col = headers_in.index("Image URL") + 1
name_col  = headers_in.index("Product Name") + 1
sku_col   = headers_in.index("SKU") + 1
price_col = headers_in.index("List Price (USD)") + 1

# ── Output Excel ──
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.append([
    "Product URL", "Image URL", "Product Name", "SKU",
    "Product Family Id", "Description", "Weight",
    "Dimension", "Width", "Depth", "Diameter", "Length", "Height",
    "Seat Height", "Seat Depth",
    "List Price", "Details", "Socket", "Wattage", "Tearsheet Link"
])

driver = init_driver()
total = ws_in.max_row - 1

for row_idx in range(2, ws_in.max_row + 1):
    product_url  = ws_in.cell(row_idx, url_col).value
    image_url    = ws_in.cell(row_idx, image_col).value
    product_name = ws_in.cell(row_idx, name_col).value
    sku          = ws_in.cell(row_idx, sku_col).value
    price        = ws_in.cell(row_idx, price_col).value

    print(f"\n[{row_idx-1}/{total}] {product_name}")
    print(f"  URL: {product_url}")

    try:
        desc, weight, dimension, height, width, depth, diameter, length, seat_height, seat_depth, details, tearsheet, socket, wattage = scrape_product_page(driver, product_url)
    except Exception as e:
        print(f"  ⚠️ Fatal error: {e}")
        desc, weight, dimension, height, width, depth, diameter, length, seat_height, seat_depth, details, tearsheet, socket, wattage = "", "", "", "", "", "", "", "", "", "", "", "", "", ""

    ws_out.append([
        product_url, image_url, product_name, sku,
        product_name,
        desc, weight,
        dimension, width, depth, diameter, length, height,
        seat_height, seat_depth,
        price, details, socket, wattage, tearsheet,
    ])
    time.sleep(0.5)

driver.quit()
wb_out.save(OUTPUT_FILE)
print(f"\n✅ Done! → {OUTPUT_FILE}")